"""Characterize parse decisions before extracting a pure parse-plan compiler."""

from __future__ import annotations

from collections import deque
from copy import deepcopy
from dataclasses import FrozenInstanceError, dataclass, fields, replace
from enum import Enum
from functools import partial
from pathlib import Path
from types import SimpleNamespace
from typing import Any, ClassVar

import openpyxl
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal

import messy_xlsx.workbook as workbook_module
from messy_xlsx import MessyWorkbook, SheetConfig, StructureInfo
from messy_xlsx._fallback_signals import (
    _fallback_block_reason,
    _FallbackBlockReason,
)
from messy_xlsx._source import SourceHandle
from messy_xlsx.enums import FormatType, HeaderDetectionMode, HeaderFallback, MergeStrategy
from messy_xlsx.exceptions import StructureError
from messy_xlsx.parsing import ParseOptions
from messy_xlsx.parsing.contracts import OutputMode
from messy_xlsx.parsing.parse_plan import (
    ParsePlan,
    compile_parse_plan,
    requires_structure_analysis,
)


def _structure(**overrides: Any) -> StructureInfo:
    base = StructureInfo(
        data_start_row=1,
        data_end_row=20,
        data_start_col=1,
        data_end_col=4,
        header_row=4,
        header_rows_count=2,
        header_confidence=0.8,
        detected_locale="de_DE",
        suggested_skip_footer=2,
    )
    return replace(base, **overrides)


def _compile_with_structure(config: SheetConfig, structure: StructureInfo) -> ParsePlan:
    return compile_parse_plan(config, structure, "xlsx")


@pytest.mark.parametrize(
    ("config", "structure", "expected_rows"),
    [
        pytest.param(
            SheetConfig(
                header_detection_mode="auto",
                header_confidence_threshold=0.8,
                skip_rows=9,
                header_rows=3,
            ),
            _structure(header_confidence=0.8),
            (3, 2),
            id="auto-accepts-confidence-at-threshold",
        ),
        pytest.param(
            SheetConfig(
                header_detection_mode="auto",
                header_confidence_threshold=0.9,
                header_fallback="first_row",
                skip_rows=9,
                header_rows=3,
            ),
            _structure(header_confidence=0.8),
            (0, 1),
            id="auto-low-confidence-first-row",
        ),
        pytest.param(
            SheetConfig(
                header_detection_mode="auto",
                header_confidence_threshold=0.9,
                header_fallback="none",
                skip_rows=2,
                header_rows=3,
            ),
            _structure(header_confidence=0.8),
            (2, 0),
            id="auto-low-confidence-no-header",
        ),
        pytest.param(
            SheetConfig(header_detection_mode="smart"),
            _structure(),
            (3, 2),
            id="smart-detects-with-default-row-offset",
        ),
        pytest.param(
            SheetConfig(
                header_detection_mode="smart",
                skip_rows=2,
                header_rows=3,
            ),
            _structure(),
            (2, 3),
            id="smart-preserves-explicit-row-offset",
        ),
        pytest.param(
            SheetConfig(
                header_detection_mode="smart",
                header_confidence_threshold=0.9,
                skip_rows=0,
                header_rows=3,
            ),
            _structure(header_confidence=0.8),
            (0, 3),
            id="smart-low-confidence-preserves-config",
        ),
        pytest.param(
            SheetConfig(
                header_detection_mode="manual",
                skip_rows=2,
                header_rows=3,
            ),
            _structure(header_confidence=1.0, header_row=10),
            (2, 3),
            id="manual-always-preserves-config",
        ),
    ],
)
def test_header_decision_matrix(
    config: SheetConfig,
    structure: StructureInfo,
    expected_rows: tuple[int, int],
) -> None:
    effective = _compile_with_structure(config, structure)

    assert (effective.skip_rows, effective.header_rows) == expected_rows


def test_auto_error_fallback_preserves_inputs() -> None:
    config = SheetConfig(
        header_detection_mode="auto",
        header_confidence_threshold=0.9,
        header_fallback="error",
        column_renames={"old": "new"},
    )
    structure = _structure(header_confidence=0.4, header_row=None)
    config_before = deepcopy(config)
    structure_before = deepcopy(structure)

    with pytest.raises(StructureError, match="No header detected with sufficient confidence"):
        _compile_with_structure(config, structure)

    assert config == config_before
    assert structure == structure_before


@pytest.mark.parametrize("mode", ["auto", "smart"])
@pytest.mark.parametrize(
    ("include_hidden", "expected_skip_rows"),
    [(False, 2), (True, 5)],
)
def test_detected_header_offset_accounts_for_removed_hidden_rows(
    mode: str,
    include_hidden: bool,
    expected_skip_rows: int,
) -> None:
    config = SheetConfig(
        header_detection_mode=mode,
        include_hidden=include_hidden,
    )
    structure = _structure(
        header_row=6,
        header_rows_count=1,
        hidden_rows=[1, 3, 5, 8],
    )

    effective = _compile_with_structure(config, structure)

    assert effective.skip_rows == expected_skip_rows


@pytest.mark.parametrize(
    ("config", "structure", "expected_footer"),
    [
        pytest.param(
            SheetConfig(skip_footer=4),
            _structure(
                num_tables=2,
                table_ranges=[{"start_row": 1, "end_row": 8}],
                suggested_skip_footer=2,
            ),
            4,
            id="user-footer-wins",
        ),
        pytest.param(
            SheetConfig(include_hidden=False),
            _structure(
                data_end_row=20,
                num_tables=2,
                table_ranges=[{"start_row": 1, "end_row": 8}],
                hidden_rows=[10, 12, 19],
                suggested_skip_footer=2,
            ),
            9,
            id="first-table-boundary-adjusts-for-hidden-rows",
        ),
        pytest.param(
            SheetConfig(include_hidden=True),
            _structure(
                data_end_row=20,
                num_tables=2,
                table_ranges=[{"start_row": 1, "end_row": 8}],
                hidden_rows=[10, 12, 19],
                suggested_skip_footer=2,
            ),
            12,
            id="first-table-boundary-includes-hidden-rows",
        ),
        pytest.param(
            SheetConfig(),
            _structure(num_tables=1, suggested_skip_footer=3),
            3,
            id="single-table-uses-detected-footer",
        ),
    ],
)
def test_footer_decision_matrix(
    config: SheetConfig,
    structure: StructureInfo,
    expected_footer: int,
) -> None:
    effective = _compile_with_structure(config, structure)

    assert effective.skip_footer == expected_footer


def test_hidden_rows_beyond_data_region_do_not_reduce_multi_table_footer() -> None:
    config = SheetConfig(include_hidden=False)
    structure = _structure(
        data_end_row=20,
        num_tables=2,
        table_ranges=[{"start_row": 1, "end_row": 8}],
        hidden_rows=[10, 12, 99, 101],
        suggested_skip_footer=2,
    )

    effective = _compile_with_structure(config, structure)

    # Only rows 10 and 12 are both after the first table and inside the
    # detected data region: (20 - 8) - 2 = 10.
    assert effective.skip_footer == 10


@pytest.mark.parametrize(
    ("cell_range", "expected_footer"),
    [
        pytest.param(None, 12, id="whole-sheet"),
        pytest.param("", 12, id="empty-range-is-whole-sheet"),
        pytest.param("A1:B8", 0, id="bounded-range-skips-global-footer-evidence"),
    ],
)
def test_bounded_range_suppresses_automatic_sheet_footer_rules(
    cell_range: str | None,
    expected_footer: int,
) -> None:
    config = SheetConfig(cell_range=cell_range)
    structure = _structure(
        data_end_row=20,
        num_tables=2,
        table_ranges=[{"start_row": 1, "end_row": 8}],
        suggested_skip_footer=2,
    )

    plan = _compile_with_structure(config, structure)

    assert plan.skip_footer == expected_footer


def test_bounded_range_preserves_explicit_caller_footer() -> None:
    config = SheetConfig(cell_range="A1:B8", skip_footer=1)
    structure = _structure(suggested_skip_footer=4)

    plan = _compile_with_structure(config, structure)

    assert plan.skip_footer == 1


def test_structure_application_preserves_every_user_option_and_inputs() -> None:
    config = SheetConfig(
        skip_rows=2,
        header_rows=3,
        skip_footer=4,
        cell_range="B2:F20",
        column_renames={"raw": "clean"},
        type_hints={"amount": "currency"},
        auto_detect=True,
        include_hidden=True,
        merge_strategy="first_only",
        locale="fr_FR",
        evaluate_formulas=False,
        drop_regex="^TOTAL$",
        drop_conditions=[{"column": "status", "value": "void"}],
        header_detection_mode="manual",
        header_confidence_threshold=0.85,
        header_fallback="none",
        multi_row_headers=True,
        header_patterns=["name", "amount"],
        normalize=False,
        normalize_dates=False,
        normalize_numbers=False,
        normalize_whitespace=False,
        use_extended_missing_list=True,
        preserve_types=False,
        ensure_type_consistency=False,
        decimal_separator=",",
        thousands_separator=" ",
        sanitize_column_names=False,
    )
    structure = _structure(
        hidden_rows=[1, 10],
        table_ranges=[{"start_row": 2, "end_row": 12}],
        detected_locale="de_DE",
    )
    config_before = deepcopy(config)
    structure_before = deepcopy(structure)

    effective = _compile_with_structure(config, structure)

    assert effective == ParsePlan(
        skip_rows=2,
        header_rows=3,
        skip_footer=4,
        merge_strategy="skip",
        ignore_hidden=False,
        cell_range="B2:F20",
        data_only=False,
        auto_detect_header=False,
        normalize=False,
        decimal_separator=",",
        thousands_separator=" ",
        use_extended_missing_list=True,
        preserve_types=False,
        type_hints=(("amount", "currency"),),
        skip_normalization_steps=("whitespace", "numbers", "dates", "type_coercion"),
        sanitize_column_names=False,
        column_renames=(("raw", "clean"),),
        drop_regex="^TOTAL$",
        drop_conditions=(("status", "void"),),
    )
    assert config == config_before
    assert structure == structure_before


@pytest.mark.parametrize(
    ("configured_locale", "detected_locale", "expected_separators"),
    [
        ("fr_FR", "de_DE", (",", " ")),
        (None, "de_DE", (",", ".")),
    ],
)
def test_explicit_locale_overrides_detected_locale(
    configured_locale: str | None,
    detected_locale: str,
    expected_separators: tuple[str, str],
) -> None:
    config = SheetConfig(locale=configured_locale)
    structure = _structure(detected_locale=detected_locale)

    effective = _compile_with_structure(config, structure)

    assert (effective.decimal_separator, effective.thousands_separator) == expected_separators


class _RecordingRegistry:
    def __init__(self) -> None:
        self.options: ParseOptions | None = None
        self.format_type: str | None = None

    def parse(
        self,
        _source: object,
        *,
        sheet: str | None,
        options: ParseOptions,
        format_type: str,
    ) -> pd.DataFrame:
        assert sheet == "Data"
        self.options = options
        self.format_type = format_type
        return pd.DataFrame({"Raw Value": [" 1 "]})


def _parse_with_recording_registry(
    config: SheetConfig,
    *,
    format_type: str = "xlsx",
    structure: StructureInfo | None = None,
) -> tuple[pd.DataFrame, ParseOptions]:
    workbook = object.__new__(MessyWorkbook)
    registry = _RecordingRegistry()
    workbook._sheet_config = config
    workbook._format_info = SimpleNamespace(format_type=format_type)
    workbook._registry = registry
    workbook._source_handle = SourceHandle(Path(f"input.{format_type}"))
    if structure is not None:
        workbook._analyze_structure = lambda _sheet, _config=None: structure  # type: ignore[method-assign]

    result = workbook._parse_sheet("Data", config)

    assert registry.options is not None
    return result, registry.options


@pytest.mark.parametrize("evaluate_formulas", [True, False])
def test_formula_value_mode_maps_directly_to_backend_data_only(
    evaluate_formulas: bool,
) -> None:
    config = SheetConfig(
        auto_detect=False,
        evaluate_formulas=evaluate_formulas,
        normalize=False,
        sanitize_column_names=False,
    )

    _, options = _parse_with_recording_registry(config)

    assert options.data_only is evaluate_formulas


@pytest.mark.parametrize(
    ("config", "structure", "expected_merge", "expected_ignore_hidden"),
    [
        pytest.param(
            SheetConfig(
                merge_strategy="fill",
                include_hidden=False,
                normalize=False,
                sanitize_column_names=False,
            ),
            _structure(merged_ranges=[], hidden_rows=[], hidden_columns=[]),
            "skip",
            False,
            id="clean-structure-enables-fast-backend-hints",
        ),
        pytest.param(
            SheetConfig(
                merge_strategy="first_only",
                include_hidden=False,
                normalize=False,
                sanitize_column_names=False,
            ),
            _structure(
                merged_ranges=[(1, 1, 1, 2)],
                hidden_rows=[7],
                hidden_columns=[3],
            ),
            MergeStrategy.FIRST_ONLY,
            True,
            id="complex-structure-preserves-openpyxl-hints",
        ),
        pytest.param(
            SheetConfig(
                merge_strategy="fill",
                include_hidden=True,
                normalize=False,
                sanitize_column_names=False,
            ),
            _structure(merged_ranges=[(1, 1, 1, 2)], hidden_rows=[7]),
            MergeStrategy.FILL,
            False,
            id="caller-includes-hidden-content",
        ),
    ],
)
def test_backend_merge_and_hidden_decision_matrix(
    config: SheetConfig,
    structure: StructureInfo,
    expected_merge: str,
    expected_ignore_hidden: bool,
) -> None:
    _, options = _parse_with_recording_registry(config, structure=structure)

    assert options.merge_strategy == expected_merge
    assert options.ignore_hidden is expected_ignore_hidden


@pytest.mark.parametrize(
    ("format_type", "auto_detect", "expected"),
    [
        ("csv", True, True),
        ("tsv", True, True),
        ("txt", True, True),
        ("csv", False, False),
        ("xlsx", False, False),
    ],
)
def test_text_header_detection_hint_is_format_scoped(
    format_type: str,
    auto_detect: bool,
    expected: bool,
) -> None:
    config = SheetConfig(
        auto_detect=auto_detect,
        normalize=False,
        sanitize_column_names=False,
    )

    _, options = _parse_with_recording_registry(config, format_type=format_type)

    assert options.auto_detect_header is expected


def test_legacy_csv_manual_mode_still_enables_automatic_header_hint() -> None:
    config = SheetConfig(
        auto_detect=True,
        header_detection_mode="manual",
        normalize=False,
        sanitize_column_names=False,
    )

    _, options = _parse_with_recording_registry(config, format_type="csv")

    # This is current behavior; unifying header policy is deferred to S09.
    assert options.auto_detect_header is True


def test_parse_options_receive_effective_structure_rows_and_user_range() -> None:
    config = SheetConfig(
        header_detection_mode="auto",
        skip_footer=4,
        cell_range="B2:D10",
        normalize=False,
        sanitize_column_names=False,
    )

    _, options = _parse_with_recording_registry(config, structure=_structure())

    assert (options.skip_rows, options.header_rows, options.skip_footer) == (3, 2, 4)
    assert options.cell_range == "B2:D10"


class _RecordingPipeline:
    instances: ClassVar[list[_RecordingPipeline]] = []

    def __init__(
        self,
        decimal_separator: str | None = None,
        thousands_separator: str | None = None,
        extra_missing_values: list[str] | None = None,
        preserve_linebreaks: bool = False,
        use_extended_missing_list: bool = False,
        preserve_types: bool = True,
    ) -> None:
        self.init_options = {
            "decimal_separator": decimal_separator,
            "thousands_separator": thousands_separator,
            "extra_missing_values": extra_missing_values,
            "preserve_linebreaks": preserve_linebreaks,
            "use_extended_missing_list": use_extended_missing_list,
            "preserve_types": preserve_types,
        }
        self.semantic_hints: dict[str, str] | None = None
        self.skip_steps: list[str] | None = None
        type(self).instances.append(self)

    def normalize(
        self,
        df: pd.DataFrame,
        semantic_hints: dict[str, str] | None = None,
        skip_steps: list[str] | None = None,
    ) -> pd.DataFrame:
        self.semantic_hints = semantic_hints
        self.skip_steps = skip_steps
        return df.copy()


def _capture_pipeline(
    monkeypatch: pytest.MonkeyPatch,
    config: SheetConfig,
    *,
    structure: StructureInfo | None = None,
) -> _RecordingPipeline:
    _RecordingPipeline.instances = []
    monkeypatch.setattr(workbook_module, "NormalizationPipeline", _RecordingPipeline)

    _parse_with_recording_registry(config, structure=structure)

    assert len(_RecordingPipeline.instances) == 1
    return _RecordingPipeline.instances[0]


@pytest.mark.parametrize(
    ("config", "structure", "expected_separators"),
    [
        pytest.param(
            SheetConfig(auto_detect=False, locale="en_US", sanitize_column_names=False),
            None,
            (".", ","),
            id="english-locale",
        ),
        pytest.param(
            SheetConfig(auto_detect=False, locale="de_DE", sanitize_column_names=False),
            None,
            (",", "."),
            id="german-locale",
        ),
        pytest.param(
            SheetConfig(auto_detect=False, locale="fr_FR", sanitize_column_names=False),
            None,
            (",", " "),
            id="french-locale",
        ),
        pytest.param(
            SheetConfig(auto_detect=False, locale="de_CH", sanitize_column_names=False),
            None,
            (".", "'"),
            id="swiss-german-locale",
        ),
        pytest.param(
            SheetConfig(
                auto_detect=False,
                locale="de_DE",
                decimal_separator=".",
                thousands_separator="_",
                sanitize_column_names=False,
            ),
            None,
            (".", "_"),
            id="explicit-separators-win",
        ),
        pytest.param(
            SheetConfig(locale=None, sanitize_column_names=False),
            _structure(detected_locale="de_DE"),
            (",", "."),
            id="detected-locale-fills-missing-locale",
        ),
        pytest.param(
            SheetConfig(auto_detect=False, locale="auto", sanitize_column_names=False),
            None,
            (None, None),
            id="auto-locale-defers-to-normalizer",
        ),
    ],
)
def test_normalization_locale_separator_decisions(
    monkeypatch: pytest.MonkeyPatch,
    config: SheetConfig,
    structure: StructureInfo | None,
    expected_separators: tuple[str | None, str | None],
) -> None:
    pipeline = _capture_pipeline(monkeypatch, config, structure=structure)

    assert (
        pipeline.init_options["decimal_separator"],
        pipeline.init_options["thousands_separator"],
    ) == expected_separators


@pytest.mark.parametrize(
    ("disabled", "expected_skip_steps"),
    [
        (set(), []),
        ({"whitespace"}, ["whitespace"]),
        ({"numbers"}, ["numbers"]),
        ({"dates"}, ["dates"]),
        ({"type_coercion"}, ["type_coercion"]),
        (
            {"whitespace", "numbers", "dates", "type_coercion"},
            ["whitespace", "numbers", "dates", "type_coercion"],
        ),
    ],
)
def test_normalization_toggle_decision_matrix(
    monkeypatch: pytest.MonkeyPatch,
    disabled: set[str],
    expected_skip_steps: list[str],
) -> None:
    config = SheetConfig(
        auto_detect=False,
        normalize_whitespace="whitespace" not in disabled,
        normalize_numbers="numbers" not in disabled,
        normalize_dates="dates" not in disabled,
        ensure_type_consistency="type_coercion" not in disabled,
        type_hints={"Raw Value": "currency"},
        use_extended_missing_list=True,
        preserve_types=False,
        sanitize_column_names=False,
    )

    pipeline = _capture_pipeline(monkeypatch, config)

    assert pipeline.skip_steps == expected_skip_steps
    assert pipeline.semantic_hints == {"Raw Value": "currency"}
    assert pipeline.semantic_hints is not config.type_hints
    assert pipeline.init_options["use_extended_missing_list"] is True
    assert pipeline.init_options["preserve_types"] is False


def test_master_normalization_switch_bypasses_pipeline(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class ForbiddenPipeline:
        def __init__(self, *_args: Any, **_kwargs: Any) -> None:
            raise AssertionError("normalization pipeline must not be constructed")

    monkeypatch.setattr(workbook_module, "NormalizationPipeline", ForbiddenPipeline)
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
        column_renames={"Raw Value": "renamed"},
    )

    result, _ = _parse_with_recording_registry(config)

    assert list(result.columns) == ["renamed"]


def test_legacy_normalize_false_bypasses_row_drop_rules() -> None:
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
        drop_regex=r"\b1\b",
        drop_conditions=[{"column": "Raw Value", "value": " 1 "}],
    )

    result, _ = _parse_with_recording_registry(config)

    # Row filters currently live after the normalization early return. Moving
    # them is a behavior decision deferred to the normalization slice (S15).
    assert result["Raw Value"].tolist() == [" 1 "]


def test_parse_decisions_do_not_mutate_config_or_structure() -> None:
    config = SheetConfig(
        header_detection_mode="auto",
        type_hints={"Raw Value": "number"},
        column_renames={"Raw Value": "value"},
        drop_conditions=[{"column": "Raw Value", "value": "TOTAL"}],
        header_patterns=["value"],
        sanitize_column_names=False,
    )
    structure = _structure(
        merged_ranges=[(1, 1, 1, 2)],
        hidden_rows=[2, 9],
        hidden_columns=[3],
        table_ranges=[{"start_row": 1, "end_row": 10}],
    )
    config_before = deepcopy(config)
    structure_before = deepcopy(structure)

    _parse_with_recording_registry(config, structure=structure)

    assert config == config_before
    assert structure == structure_before


def test_sheet_structure_honors_workbook_header_patterns(tmp_path: Path) -> None:
    path = tmp_path / "sheet-patterns.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["First field", "Second field", "Third field"])
    sheet.append([1, 2, 3])
    source.save(path)
    source.close()

    config = SheetConfig(header_patterns=[r"first field", r"second field", r"third field"])
    with MessyWorkbook(path, sheet_config=config) as workbook:
        workbook_structure = workbook.get_structure("Data")
        sheet_structure = workbook.get_sheet("Data").structure

    assert workbook_structure.header_confidence > 0.8
    assert sheet_structure == workbook_structure


def test_table_without_explicit_config_inherits_workbook_config(tmp_path: Path) -> None:
    path = tmp_path / "table-config.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Raw Column", "Other Column"])
    sheet.append(["  padded  ", 1])
    sheet.append(["plain", 2])
    source.save(path)
    source.close()

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with MessyWorkbook(path, sheet_config=config) as workbook:
        expected = workbook.to_dataframe("Data")
        table = workbook.get_sheet("Data").tables[0]
        actual = table.to_dataframe()

    assert_frame_equal(actual, expected)


def test_each_detected_table_extracts_its_own_rows(tmp_path: Path) -> None:
    path = tmp_path / "two-tables.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["First Item", "First Value"])
    sheet.append(["Alpha", 1])
    sheet.append(["Beta", 2])
    sheet.append([])
    sheet.append([])
    sheet.append(["Second Item", "Second Value"])
    sheet.append(["Gamma", 3])
    sheet.append(["Delta", 4])
    source.save(path)
    source.close()

    with MessyWorkbook(path) as workbook:
        tables = workbook.get_sheet("Data").tables
        frames = [table.to_dataframe() for table in tables]

    assert len(frames) == 2
    assert [frame.empty for frame in frames] == [False, False]
    assert frames[0].iloc[:, 0].tolist() == ["Alpha", "Beta"]
    assert frames[1].iloc[:, 0].tolist() == ["Gamma", "Delta"]


def test_compiled_parse_plan_is_frozen_and_slotted() -> None:
    plan = compile_parse_plan(
        SheetConfig(auto_detect=False),
        None,
        "xlsx",
    )

    assert isinstance(plan, ParsePlan)
    assert "__slots__" in ParsePlan.__dict__
    assert not hasattr(plan, "__dict__")
    with pytest.raises(FrozenInstanceError):
        plan.skip_rows = 99


def test_parse_plan_returns_fresh_parse_options_on_every_call() -> None:
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            skip_rows=2,
            header_rows=3,
            skip_footer=4,
            cell_range="B2:D20",
        ),
        None,
        "xlsx",
    )

    first = plan.to_parse_options()
    second = plan.to_parse_options()

    assert isinstance(first, ParseOptions)
    assert first == second
    assert first is not second
    first.skip_rows = 99
    assert plan.to_parse_options().skip_rows == 2


def test_parse_plan_snapshots_consumed_config_collections() -> None:
    config = SheetConfig(
        header_detection_mode="auto",
        merge_strategy="first_only",
        column_renames={"raw": "clean"},
        type_hints={"amount": "currency"},
        drop_conditions=[{"column": "status", "value": "void"}],
    )
    structure = _structure(
        header_row=6,
        merged_ranges=[(1, 1, 1, 2)],
        hidden_rows=[2, 15],
        hidden_columns=[3],
        num_tables=2,
        table_ranges=[{"start_row": 1, "end_row": 10}],
    )

    plan = compile_parse_plan(config, structure, "xlsx")

    config.column_renames["raw"] = "changed"
    config.column_renames["new"] = "field"
    config.type_hints["amount"] = "text"
    config.type_hints["new"] = "number"
    config.drop_conditions[0]["column"] = "changed"
    config.drop_conditions.append({"column": "new", "value": "x"})
    structure.merged_ranges.clear()
    structure.hidden_rows.extend([3, 99])
    structure.hidden_columns.clear()
    structure.table_ranges[0]["end_row"] = 19
    structure.table_ranges.append({"start_row": 20, "end_row": 20})

    assert plan.column_renames == (("raw", "clean"),)
    assert plan.type_hints == (("amount", "currency"),)
    assert plan.drop_conditions == (("status", "void"),)
    assert plan.skip_rows == 4
    assert plan.skip_footer == 9
    assert plan.merge_strategy == MergeStrategy.FIRST_ONLY
    assert plan.ignore_hidden is True


@pytest.mark.parametrize("batch_size", [None, 0, -1, 1.5, True])
def test_streaming_batch_size_is_validated_during_plan_compilation(
    batch_size: object,
) -> None:
    config = SheetConfig(auto_detect=False)

    with pytest.raises(ValueError, match="batch_size must be >= 1") as captured:
        compile_parse_plan(
            config,
            None,
            "xlsx",
            output_mode=OutputMode.STREAMING,
            batch_size=batch_size,  # type: ignore[arg-type]
        )

    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.CONFIGURATION


def test_materialized_plan_preserves_the_legacy_three_argument_call() -> None:
    legacy = compile_parse_plan(SheetConfig(auto_detect=False), None, "xlsx")
    explicit = compile_parse_plan(
        SheetConfig(auto_detect=False),
        None,
        "xlsx",
        output_mode=OutputMode.MATERIALIZED,
        batch_size=None,
    )

    assert legacy == explicit
    assert legacy.output_mode is OutputMode.MATERIALIZED
    assert legacy.batch_size is None


def test_plan_recursively_snapshots_nested_mutable_configuration_values() -> None:
    nested_hint = {
        "members": ["amount", {"precision": [2, 4]}],
        "labels": {"gross", "net"},
    }
    nested_condition = {
        "groups": [{"names": ["open", "closed"]}],
        "thresholds": ({"minimum": [1, 2]},),
    }
    config = SheetConfig(
        auto_detect=False,
        type_hints={"amount": nested_hint},  # type: ignore[dict-item]
        drop_conditions=[{"column": "metadata", "value": nested_condition}],
    )

    plan = compile_parse_plan(config, None, "xlsx")
    type_hints_before = deepcopy(plan.type_hints)
    conditions_before = deepcopy(plan.drop_conditions)

    nested_hint["members"][1]["precision"].append(8)
    nested_hint["labels"].add("tax")
    nested_condition["groups"][0]["names"].append("pending")
    nested_condition["thresholds"][0]["minimum"].append(3)

    assert plan.type_hints == type_hints_before
    assert plan.drop_conditions == conditions_before


def test_recursive_snapshots_are_deterministic_across_mapping_and_set_order() -> None:
    first = SheetConfig(
        auto_detect=False,
        type_hints={
            "payload": {
                "mapping": {"second": [2], "first": [1]},
                "labels": {"beta", "alpha"},
            }
        },  # type: ignore[dict-item]
    )
    second = SheetConfig(
        auto_detect=False,
        type_hints={
            "payload": {
                "labels": {"alpha", "beta"},
                "mapping": {"first": [1], "second": [2]},
            }
        },  # type: ignore[dict-item]
    )

    assert compile_parse_plan(first, None, "xlsx") == compile_parse_plan(
        second,
        None,
        "xlsx",
    )


@dataclass
class _MutablePayload:
    labels: list[str]
    metadata: dict[str, set[int]]


class _MutableObject:
    def __init__(self, values: list[str]) -> None:
        self.values = values

    def __eq__(self, other: object) -> bool:
        return isinstance(other, _MutableObject) and self.values == other.values


class _MutablePayloadEnum(Enum):
    ITEM: ClassVar[list[str]] = ["initial"]


@dataclass
class _PostInitPayload:
    labels: list[str]

    def __post_init__(self) -> None:
        self.cache = {"initial": list(self.labels)}


@dataclass(eq=False)
class _IdentityEqualityDataclass:
    labels: list[str]


@dataclass
class _DataclassList(list[str]):
    label: str


class _PrivateSlotBase:
    __slots__ = ("__cache",)

    def set_cache(self, values: list[str]) -> None:
        self.__cache = values

    @property
    def cache(self) -> list[str]:
        return self.__cache


@dataclass(slots=True)
class _InheritedPrivateSlotDataclass(_PrivateSlotBase):
    labels: list[str]

    def __post_init__(self) -> None:
        self.set_cache(list(self.labels))


class _IdentityCondition:
    def __init__(self) -> None:
        self.mutable_state = ["initial"]


class _HostileIdentityMetaclass(type):
    hash_calls = 0

    def __hash__(cls) -> int:
        del cls
        _HostileIdentityMetaclass.hash_calls += 1
        return _HostileIdentityMetaclass.hash_calls

    def __eq__(cls, other: object) -> bool:
        del cls, other
        raise AssertionError("plan equality must not delegate to a sentinel metaclass")


class _HostileMetaclassIdentityCondition(metaclass=_HostileIdentityMetaclass):
    pass


class _CustomNewValue:
    def __new__(cls, value: str) -> _CustomNewValue:
        del value
        return super().__new__(cls)

    def __init__(self, value: str) -> None:
        self.value = value

    def __eq__(self, other: object) -> bool:
        return isinstance(other, _CustomNewValue) and self.value == other.value


def test_plan_thaws_fresh_values_with_original_supported_container_kinds() -> None:
    dataclass_value = _MutablePayload(["alpha"], {"codes": {1, 2}})
    object_value = _MutableObject(["original"])
    nested = {
        "list": [1, {"nested": [2]}],
        "tuple": (3, [4]),
        "set": {5, 6},
        "frozenset": frozenset({7, 8}),
        "bytearray": bytearray(b"mutable"),
        "memoryview": memoryview(bytearray(b"view")),
        "dataclass": dataclass_value,
        "object": object_value,
    }
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            type_hints={"payload": nested},  # type: ignore[dict-item]
            drop_conditions=[{"column": "value", "value": object_value}],
        ),
        None,
        "xlsx",
    )
    hash(plan)

    nested["list"][1]["nested"].append(99)
    nested["tuple"][1].append(99)
    nested["set"].add(99)
    nested["bytearray"].extend(b"!")
    nested["memoryview"][0] = ord("X")
    dataclass_value.labels.append("changed")
    dataclass_value.metadata["codes"].add(99)
    object_value.values.append("changed")

    first = plan.thaw_type_hints()["payload"]
    second = plan.thaw_type_hints()["payload"]
    assert isinstance(first, dict)
    assert isinstance(first["list"], list)
    assert isinstance(first["tuple"], tuple)
    assert isinstance(first["set"], set)
    assert isinstance(first["frozenset"], frozenset)
    assert isinstance(first["bytearray"], bytearray)
    assert isinstance(first["memoryview"], memoryview)
    assert first["dataclass"] == _MutablePayload(["alpha"], {"codes": {1, 2}})
    assert first["object"] == _MutableObject(["original"])
    assert first["object"] is not object_value
    assert first["object"] is not second["object"]
    assert plan.thaw_drop_conditions()[0][1] == _MutableObject(["original"])

    first["list"].append("consumer mutation")
    first["dataclass"].labels.append("consumer mutation")
    assert second["list"] == [1, {"nested": [2]}]
    assert second["dataclass"].labels == ["alpha"]


def test_mixed_mapping_keys_have_deterministic_plan_equality_and_hash() -> None:
    first_hints = {
        1: "integer",
        "1": "string",
        (1, frozenset({2, 3})): "nested tuple",
        frozenset({4, 5}): "frozenset",
        None: "none",
    }
    second_hints = dict(reversed(tuple(first_hints.items())))

    first = compile_parse_plan(
        SheetConfig(auto_detect=False, type_hints=first_hints),  # type: ignore[arg-type]
        None,
        "xlsx",
    )
    second = compile_parse_plan(
        SheetConfig(auto_detect=False, type_hints=second_hints),  # type: ignore[arg-type]
        None,
        "xlsx",
    )

    assert first == second
    assert hash(first) == hash(second)
    assert first.thaw_type_hints() == first_hints


def test_streaming_plan_is_stable_when_configuration_mutates_after_creation() -> None:
    nested_hint = {"labels": ["gross"], "codes": {1, 2}}
    config = SheetConfig(
        auto_detect=False,
        type_hints={"amount": nested_hint},  # type: ignore[dict-item]
    )
    plan = compile_parse_plan(
        config,
        None,
        "xlsx",
        output_mode=OutputMode.STREAMING,
        batch_size=100,
    )

    nested_hint["labels"].append("net")
    nested_hint["codes"].add(3)

    assert plan.thaw_type_hints() == {"amount": {"labels": ["gross"], "codes": {1, 2}}}


def test_tagged_snapshots_distinguish_list_tuple_set_and_frozenset() -> None:
    plans = [
        compile_parse_plan(
            SheetConfig(auto_detect=False, type_hints={"value": value}),  # type: ignore[dict-item]
            None,
            "xlsx",
        )
        for value in ([1], (1,), {1}, frozenset({1}))
    ]

    assert len(set(plans)) == 4


def test_mutable_enum_payload_is_rejected_before_a_plan_can_retain_its_alias() -> None:
    with pytest.raises(TypeError, match="mutable Enum configuration value"):
        compile_parse_plan(
            SheetConfig(
                auto_detect=False,
                type_hints={"enum": _MutablePayloadEnum.ITEM},  # type: ignore[dict-item]
            ),
            None,
            "xlsx",
        )


def test_unknown_unsupported_mutable_value_is_rejected_before_backend_io() -> None:
    with pytest.raises(TypeError, match="unsupported mutable configuration value"):
        compile_parse_plan(
            SheetConfig(
                auto_detect=False,
                type_hints={"queue": deque([1, 2])},  # type: ignore[dict-item]
            ),
            None,
            "xlsx",
        )


def test_type_and_function_hints_are_preserved_by_thaw_projection() -> None:
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            type_hints={"type": list, "callable": len},  # type: ignore[dict-item]
        ),
        None,
        "xlsx",
    )

    assert plan.thaw_type_hints() == {"type": list, "callable": len}


def test_dataclass_snapshot_preserves_post_init_state_without_aliases() -> None:
    payload = _PostInitPayload(["alpha"])
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            type_hints={"payload": payload},  # type: ignore[dict-item]
        ),
        None,
        "xlsx",
    )

    payload.labels.append("changed")
    payload.cache["initial"].append("changed")
    first = plan.thaw_type_hints()["payload"]
    second = plan.thaw_type_hints()["payload"]

    assert first.labels == ["alpha"]
    assert first.cache == {"initial": ["alpha"]}
    assert first is not payload
    assert first.cache is not second.cache
    assert hash(plan) == hash(plan)


def test_dataclass_without_generated_equality_still_uses_value_snapshot_policy() -> None:
    payload = _IdentityEqualityDataclass(["alpha"])
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            type_hints={"payload": payload},  # type: ignore[dict-item]
        ),
        None,
        "xlsx",
    )

    payload.labels.append("changed")
    thawed = plan.thaw_type_hints()["payload"]

    assert isinstance(thawed, _IdentityEqualityDataclass)
    assert thawed.labels == ["alpha"]
    assert thawed is not payload


def test_identity_equal_dataclass_drop_condition_filters_same_object_row() -> None:
    condition = _IdentityEqualityDataclass(["target"])
    other = _IdentityEqualityDataclass(["target"])
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )

    comparison_value = plan.thaw_drop_conditions()[0][1]
    frame = pd.DataFrame({"value": pd.Series([condition, other], dtype=object)})
    filtered = frame[frame["value"] != comparison_value].reset_index(drop=True)

    assert comparison_value is condition
    assert plan.drop_conditions[0][1] is not condition
    assert filtered["value"].tolist() == [other]


def test_c_backed_dataclass_is_rejected_before_hidden_state_is_lost() -> None:
    payload = _DataclassList("visible")
    payload.extend(["hidden", "list", "state"])

    with pytest.raises(TypeError, match="opaque mutable configuration value"):
        compile_parse_plan(
            SheetConfig(
                auto_detect=False,
                type_hints={"payload": payload},  # type: ignore[dict-item]
            ),
            None,
            "xlsx",
        )


def test_pure_python_custom_new_value_remains_snapshot_reconstructable() -> None:
    payload = _CustomNewValue("target")
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": payload}],
        ),
        None,
        "xlsx",
    )

    payload.value = "changed"
    thawed = plan.thaw_drop_conditions()[0][1]

    assert thawed == _CustomNewValue("target")
    assert thawed is not payload


def test_dataclass_snapshot_preserves_inherited_name_mangled_private_slot() -> None:
    payload = _InheritedPrivateSlotDataclass(["alpha"])
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            type_hints={"payload": payload},  # type: ignore[dict-item]
        ),
        None,
        "xlsx",
    )

    payload.labels.append("changed")
    payload.cache.append("changed")
    thawed = plan.thaw_type_hints()["payload"]

    assert isinstance(thawed, _InheritedPrivateSlotDataclass)
    assert thawed.labels == ["alpha"]
    assert thawed.cache == ["alpha"]
    assert thawed.cache is not payload.cache


def test_identity_semantic_drop_condition_preserves_legacy_identity() -> None:
    condition = _IdentityCondition()
    other = _IdentityCondition()
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )
    initial_hash = hash(plan)

    condition.mutable_state.append("changed")
    condition.__eq__ = lambda _other: True  # type: ignore[method-assign]
    condition.__hash__ = lambda: 0  # type: ignore[method-assign]
    comparison_value = plan.thaw_drop_conditions()[0][1]
    frame = pd.DataFrame({"value": [condition, other]})
    filtered = frame[frame["value"] != comparison_value].reset_index(drop=True)

    assert filtered["value"].tolist() == [other]
    assert comparison_value is condition
    assert plan.drop_conditions[0][1] is not condition
    assert type(condition).__eq__ is object.__eq__
    assert type(condition).__hash__ is object.__hash__
    assert hash(plan) == initial_hash


def test_identity_drop_token_stabilizes_plan_hash_and_equality_against_class_mutation() -> None:
    condition = _IdentityCondition()
    other = _IdentityCondition()
    first = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )
    same_reference = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )
    different_reference = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": other}],
        ),
        None,
        "xlsx",
    )
    first_hash = hash(first)

    def unstable_hash(_self: object) -> int:
        raise AssertionError("plan hash must not delegate to a caller value")

    try:
        _IdentityCondition.__eq__ = lambda _self, _other: True  # type: ignore[method-assign]
        _IdentityCondition.__hash__ = unstable_hash  # type: ignore[assignment]

        assert hash(first) == first_hash
        assert first == same_reference
        assert first != different_reference
    finally:
        _IdentityCondition.__eq__ = object.__eq__  # type: ignore[method-assign]
        _IdentityCondition.__hash__ = object.__hash__  # type: ignore[assignment]


def test_identity_drop_token_uses_nonvirtual_stable_type_identity() -> None:
    condition = _HostileMetaclassIdentityCondition()
    first = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )
    same_reference = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "value", "value": condition}],
        ),
        None,
        "xlsx",
    )
    _HostileIdentityMetaclass.hash_calls = 0

    assert first == same_reference
    assert hash(first) == hash(same_reference)
    assert hash(first) == hash(first)
    assert first.thaw_drop_conditions()[0][1] is condition


def test_opaque_c_backed_mutable_callable_is_rejected_before_backend_io() -> None:
    opaque = partial(pow, 2)

    with pytest.raises(TypeError, match="opaque mutable configuration value"):
        compile_parse_plan(
            SheetConfig(
                auto_detect=False,
                type_hints={"callable": opaque},  # type: ignore[dict-item]
            ),
            None,
            "xlsx",
        )


def test_parse_plan_preserves_identity_sensitive_drop_condition_values() -> None:
    sentinel = object()
    other = object()

    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            drop_conditions=[{"column": "status", "value": sentinel}],
        ),
        None,
        "xlsx",
    )

    comparison_value = plan.thaw_drop_conditions()[0][1]
    frame = pd.DataFrame({"status": [sentinel, other]})
    filtered = frame[frame["status"] != comparison_value].reset_index(drop=True)

    assert filtered["status"].tolist() == [other]
    assert plan.drop_conditions[0][1] is not sentinel


@pytest.mark.parametrize("format_type", ["xlsx", "xlsm", "xltx", "xltm"])
def test_compile_parse_plan_requires_structure_for_auto_detect_ooxml(
    format_type: str,
) -> None:
    config = SheetConfig(auto_detect=True)

    assert requires_structure_analysis(config, format_type) is True
    with pytest.raises(ValueError):
        compile_parse_plan(config, None, format_type)


@pytest.mark.parametrize(
    ("format_type", "auto_detect", "expected_auto_header"),
    [
        ("xlsx", False, False),
        ("csv", True, True),
        ("xls", True, False),
    ],
)
def test_supplied_structure_is_ignored_when_analysis_is_not_required(
    format_type: str,
    auto_detect: bool,
    expected_auto_header: bool,
) -> None:
    config = SheetConfig(
        auto_detect=auto_detect,
        skip_rows=2,
        header_rows=3,
        skip_footer=4,
        merge_strategy="fill",
        include_hidden=False,
        locale=None,
    )
    irrelevant_structure = _structure(
        header_row=10,
        header_rows_count=2,
        header_confidence=1.0,
        merged_ranges=[],
        hidden_rows=[],
        hidden_columns=[],
        detected_locale="de_DE",
        num_tables=2,
        table_ranges=[{"start_row": 1, "end_row": 5}],
        suggested_skip_footer=99,
    )

    assert requires_structure_analysis(config, format_type) is False
    plan = compile_parse_plan(config, irrelevant_structure, format_type)

    assert (plan.skip_rows, plan.header_rows, plan.skip_footer) == (2, 3, 4)
    assert plan.merge_strategy == MergeStrategy.FILL
    assert plan.ignore_hidden is True
    assert plan.auto_detect_header is expected_auto_header
    assert (plan.decimal_separator, plan.thousands_separator) == (None, None)


def test_compile_parse_plan_treats_enums_and_raw_strings_equally() -> None:
    enum_config = SheetConfig(
        auto_detect=True,
        header_detection_mode=HeaderDetectionMode.AUTO,
        header_fallback=HeaderFallback.NONE,
        merge_strategy=MergeStrategy.FIRST_ONLY,
    )
    string_config = SheetConfig(
        auto_detect=True,
        header_detection_mode="auto",
        header_fallback="none",
        merge_strategy="first_only",
    )
    structure = _structure(header_confidence=0.2)

    assert requires_structure_analysis(enum_config, FormatType.XLSX) is True
    assert requires_structure_analysis(string_config, "xlsx") is True
    assert compile_parse_plan(enum_config, structure, FormatType.XLSX) == compile_parse_plan(
        string_config,
        structure,
        "xlsx",
    )


@pytest.mark.parametrize(
    ("locale", "decimal", "thousands", "expected"),
    [
        ("de_DE", None, None, (",", ".")),
        ("de_DE", "!", None, ("!", None)),
        ("de_DE", None, "_", (None, "_")),
        ("auto", None, None, (None, None)),
    ],
)
def test_compile_parse_plan_partial_separator_and_auto_locale_policy(
    locale: str,
    decimal: str | None,
    thousands: str | None,
    expected: tuple[str | None, str | None],
) -> None:
    config = SheetConfig(
        auto_detect=False,
        locale=locale,
        decimal_separator=decimal,
        thousands_separator=thousands,
    )

    plan = compile_parse_plan(config, None, "xlsx")

    assert (plan.decimal_separator, plan.thousands_separator) == expected


def test_every_sheet_config_field_has_an_explicit_parse_plan_disposition() -> None:
    projected_or_consumed = {
        "skip_rows",
        "header_rows",
        "skip_footer",
        "cell_range",
        "column_renames",
        "type_hints",
        "include_hidden",
        "merge_strategy",
        "locale",
        "evaluate_formulas",
        "drop_regex",
        "drop_conditions",
        "normalize",
        "normalize_dates",
        "normalize_numbers",
        "normalize_whitespace",
        "use_extended_missing_list",
        "preserve_types",
        "ensure_type_consistency",
        "decimal_separator",
        "thousands_separator",
        "sanitize_column_names",
    }
    compiler_controls = {
        "auto_detect",
        "header_detection_mode",
        "header_confidence_threshold",
        "header_fallback",
    }
    detection_only_or_deferred = {
        "header_patterns",
        "multi_row_headers",
    }

    categorized = projected_or_consumed | compiler_controls | detection_only_or_deferred
    assert {field.name for field in fields(SheetConfig)} == categorized
