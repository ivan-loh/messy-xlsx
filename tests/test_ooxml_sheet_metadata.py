"""Tests for lazy, bounded worksheet metadata indexing."""

from __future__ import annotations

import io
import os
from array import array
from collections.abc import Callable
from dataclasses import FrozenInstanceError
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pytest
from openpyxl.utils.cell import get_column_letter

from messy_xlsx._source import SourceHandle
from messy_xlsx.exceptions import FormatError
from messy_xlsx.ooxml import manifest as manifest_module
from messy_xlsx.ooxml.manifest import ManifestReader
from messy_xlsx.ooxml.models import (
    DEFAULT_LIMITS,
    Interval,
    IntervalIndex,
    MergeRange,
    SheetDescriptor,
    StyleManifest,
    WorkbookManifest,
)


def _rewrite_member(path: Path, member: str, transform: Callable[[bytes], bytes]) -> None:
    replacement = path.with_suffix(".replacement.xlsx")
    with ZipFile(path) as source, ZipFile(replacement, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            content = source.read(info.filename)
            target.writestr(info, transform(content) if info.filename == member else content)
    replacement.replace(path)


def _replace_required(content: bytes, old: bytes, new: bytes) -> bytes:
    assert old in content
    return content.replace(old, new)


def _synthetic_sheet_manifest(
    row_numbers: range,
    max_column: int,
):
    rows: list[str] = []
    for row in row_numbers:
        cells = "".join(
            f'<c r="{get_column_letter(column)}{row}"><v>1</v></c>'
            for column in range(1, max_column + 1)
        )
        rows.append(f'<row r="{row}">{cells}</row>')
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><sheetData>' + "".join(rows) + "</sheetData></worksheet>"
    )
    return _parse_synthetic_sheet(xml)


def _parse_synthetic_sheet(
    xml: str,
    number_format_codes: tuple[str, ...] = ("General",),
):
    descriptor = SheetDescriptor(
        name="Data",
        relationship_id="rId1",
        target="xl/worksheets/sheet1.xml",
        state="visible",
    )
    reader = object.__new__(ManifestReader)
    reader._limits = DEFAULT_LIMITS
    reader.workbook = WorkbookManifest(
        workbook_type="xlsx",
        date_system="1900",
        sheets=(descriptor,),
        has_shared_strings=False,
        shared_strings_uncompressed_size=0,
        styles=StyleManifest((), (), number_format_codes),
    )
    return reader._parse_sheet_xml(descriptor, io.BytesIO(xml.encode()))


def test_interval_index_normalizes_ranges_without_expanding_cells() -> None:
    index = IntervalIndex(
        (
            Interval(10, 12),
            Interval(2, 4),
            Interval(4, 9),
            Interval(20, 20),
        )
    )

    assert index.intervals == (Interval(2, 12), Interval(20, 20))
    assert index.contains(2)
    assert index.contains(11)
    assert not index.contains(13)
    assert index.contains(20)
    assert len(index.intervals) == 2


@pytest.mark.parametrize("start,end", [(0, 1), (-1, 3), (3, 2)])
def test_interval_rejects_invalid_one_based_bounds(start: int, end: int) -> None:
    with pytest.raises(ValueError, match="invalid one-based interval"):
        Interval(start, end)


def test_interval_models_are_immutable() -> None:
    interval = Interval(2, 4)

    with pytest.raises(FrozenInstanceError):
        interval.start = 3  # type: ignore[misc]


@pytest.fixture
def metadata_xlsx(tmp_path):
    path = tmp_path / "metadata.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = "anchor"
    worksheet["D9"] = "tail"
    worksheet["D9"].number_format = "#.##0,00"
    worksheet.merge_cells("B2:C3")
    worksheet.row_dimensions[2].hidden = True
    worksheet.row_dimensions[3].hidden = True
    worksheet.row_dimensions[5].hidden = True
    worksheet.column_dimensions.group("B", "D", hidden=True)
    worksheet.column_dimensions["F"].hidden = True
    second = workbook.create_sheet("Other")
    second["C4"] = "other"
    workbook.save(path)
    workbook.close()
    return path


def test_sheet_xml_is_loaded_once_and_lazily_per_sheet(metadata_xlsx) -> None:
    opened: list[str] = []
    with SourceHandle(metadata_xlsx) as source:
        reader = ManifestReader(source, on_member_open=opened.append)

        assert not any(name.startswith("xl/worksheets/") for name in opened)
        first = reader.sheet("Data")
        assert reader.sheet("Data") is first
        assert sum(name.startswith("xl/worksheets/") for name in opened) == 1

        reader.sheet("Other")
        assert sum(name.startswith("xl/worksheets/") for name in opened) == 2


def test_sheet_metadata_compacts_hidden_ranges_and_indexes_dimensions(metadata_xlsx) -> None:
    with SourceHandle(metadata_xlsx) as source:
        sheet = ManifestReader(source).sheet("Data")

    assert sheet.name == "Data"
    assert sheet.declared_dimension == (1, 1, 9, 4)
    assert (sheet.observed_max_row, sheet.observed_max_col) == (9, 4)
    assert sheet.hidden_rows.intervals == (Interval(2, 3), Interval(5, 5))
    assert sheet.hidden_columns.intervals == (Interval(2, 4), Interval(6, 6))
    assert sheet.merged_ranges == (MergeRange(2, 2, 3, 3),)
    assert not hasattr(sheet, "number_format_codes")
    assert not hasattr(sheet, "values")


def test_formula_presence_is_exact_but_coordinate_samples_are_capped(tmp_path) -> None:
    path = tmp_path / "formulas.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in range(1, 301):
        worksheet.cell(row=row, column=8, value=f"=A{row}+1")
    workbook.save(path)
    workbook.close()

    with SourceHandle(path) as source:
        sheet = ManifestReader(source).sheet("Data")

    assert sheet.has_formulas is True
    assert len(sheet.formula_samples) == 256
    assert sheet.formula_samples[0] == "H1"
    assert sheet.formula_samples[-1] == "H256"
    assert (sheet.observed_max_row, sheet.observed_max_col) == (300, 8)


@pytest.mark.parametrize(
    ("row_numbers", "max_column"),
    [
        (range(1, 101), 200),
        (range(1, 40_000, 2), 1),
    ],
)
def test_retained_provenance_is_bounded_by_scoring_shape_not_populated_cells(
    row_numbers: range,
    max_column: int,
) -> None:
    sheet = _synthetic_sheet_manifest(row_numbers, max_column)
    scoring_coordinate_bound = 19 * max_column + 51 * min(21, max_column)

    assert len(sheet.cell_evidence) <= scoring_coordinate_bound
    assert type(sheet.semantic_nonempty_rows).__name__ == "RowBitSet"
    assert len(sheet.semantic_nonempty_rows.bits) == 131_072


def test_sheet_does_not_retain_unique_formats_outside_scoring_window() -> None:
    def styled_sheet(style_count: int):
        formats = (
            "General",
            *(f'0.00"style-{index}"' for index in range(style_count)),
        )
        styled_rows = "".join(
            f'<row r="{row}"><c r="A{row}" s="{style}"><v>1</v></c></row>'
            for style, row in enumerate(range(52, 52 + style_count), start=1)
        )
        xml = (
            '<worksheet xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main"><sheetData>'
            '<row r="1"><c r="A1"><v>1</v></c></row>'
            f"{styled_rows}</sheetData></worksheet>"
        )
        return _parse_synthetic_sheet(xml, formats)

    small = styled_sheet(5)
    large = styled_sheet(1_000)

    assert not hasattr(small, "number_format_codes")
    assert not hasattr(large, "number_format_codes")
    assert len(small.cell_evidence) == len(large.cell_evidence) == 51
    assert small.locale_has_european_format is large.locale_has_european_format is False


@pytest.mark.parametrize(
    ("coordinate", "expected"),
    [("A1", (1, 1)), ("XFD1048576", (1_048_576, 16_384))],
)
def test_cell_coordinate_accepts_exact_excel_bounds(
    coordinate: str,
    expected: tuple[int, int],
) -> None:
    assert manifest_module._coordinate(coordinate, "xl/worksheets/sheet1.xml") == expected


@pytest.mark.parametrize(
    ("coordinate", "expected"),
    [
        ("A1:A1", (1, 1, 1, 1)),
        ("A1:XFD1048576", (1, 1, 1_048_576, 16_384)),
    ],
)
def test_range_coordinate_accepts_exact_excel_bounds(
    coordinate: str,
    expected: tuple[int, int, int, int],
) -> None:
    assert manifest_module._range(coordinate, "xl/worksheets/sheet1.xml") == expected


@pytest.mark.parametrize("coordinate", ["A0", "XFE1", "A1048577"])
def test_cell_coordinate_rejects_lower_and_upper_bounds_with_context(coordinate: str) -> None:
    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        manifest_module._coordinate(coordinate, "xl/worksheets/sheet1.xml")

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == coordinate


@pytest.mark.parametrize("coordinate", ["A0:A1", "0:0", "A1:A0", "A1:XFE1"])
def test_range_coordinate_rejects_lower_and_upper_bounds_with_context(coordinate: str) -> None:
    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        manifest_module._range(coordinate, "xl/worksheets/sheet1.xml")

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == coordinate


def test_range_coordinate_rejects_reversed_bounds_with_context() -> None:
    with pytest.raises(FormatError, match="malformed") as raised:
        manifest_module._range("B2:A1", "xl/worksheets/sheet1.xml")

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == "B2:A1"


@pytest.mark.parametrize(
    "xml_fragment",
    [
        '<dimension ref="B2:A1"/>',
        '<mergeCells><mergeCell ref="B2:A1"/></mergeCells>',
    ],
    ids=["dimension", "merge"],
)
def test_sheet_parser_rejects_reversed_dimension_and_merge_ranges(
    xml_fragment: str,
) -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        f'spreadsheetml/2006/main">{xml_fragment}</worksheet>'
    )

    with pytest.raises(FormatError, match="malformed range coordinate") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == "B2:A1"


def test_sheet_parser_rejects_reversed_hidden_column_range_with_context() -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><cols>'
        '<col min="2" max="1" hidden="1"/>'
        "</cols></worksheet>"
    )

    with pytest.raises(FormatError, match="hidden column range") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context == {
        "member": "xl/worksheets/sheet1.xml",
        "min": 2,
        "max": 1,
    }


@pytest.mark.parametrize(
    ("xml_fragment", "coordinate"),
    [
        ('<dimension ref="A0:A1"/>', "A0:A1"),
        ('<dimension ref="0:0"/>', "0:0"),
        ('<mergeCells><mergeCell ref="A0:A1"/></mergeCells>', "A0:A1"),
        ('<mergeCells><mergeCell ref="0:0"/></mergeCells>', "0:0"),
    ],
)
def test_sheet_parser_rejects_zero_bound_dimensions_and_merges(
    xml_fragment: str,
    coordinate: str,
) -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        f'spreadsheetml/2006/main">{xml_fragment}</worksheet>'
    )

    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == coordinate


def test_sheet_parser_rejects_row_zero_before_bitset_mutation() -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><sheetData>'
        '<row r="0"><c r="A0"><v>1</v></c></row>'
        "</sheetData></worksheet>"
    )

    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"


@pytest.mark.parametrize("row", [0, 1_048_577])
def test_row_bitset_mutation_rejects_out_of_bounds_rows_defensively(row: int) -> None:
    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        manifest_module._set_row_bit(bytearray(131_072), row, "xl/worksheets/sheet1.xml")

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == row


@pytest.mark.parametrize("offset", [-1, 1])
def test_packed_mutation_rejects_out_of_bounds_offsets_defensively(offset: int) -> None:
    with pytest.raises(FormatError, match="packed") as raised:
        manifest_module._set_packed_code(
            bytearray(1),
            offset,
            1,
            "xl/worksheets/sheet1.xml",
            "A0",
        )

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == "A0"


@pytest.mark.parametrize(
    ("rows", "message"),
    [
        (
            '<row r="1"><c r="A1"><v>1</v></c></row><row r="1"><c r="B1"><v>2</v></c></row>',
            "strictly increasing worksheet row",
        ),
        (
            '<row r="2"><c r="A2"><v>1</v></c></row><row r="1"><c r="A1"><v>2</v></c></row>',
            "strictly increasing worksheet row",
        ),
        (
            '<row r="1"><c r="A1"><v>1</v></c><c r="A1"><v>2</v></c></row>',
            "strictly increasing worksheet cell",
        ),
        (
            '<row r="1"><c r="B1"><v>1</v></c><c r="A1"><v>2</v></c></row>',
            "strictly increasing worksheet cell",
        ),
        (
            '<row r="2"><c r="A1"><v>1</v></c></row>',
            "disagrees with enclosing row",
        ),
    ],
)
def test_sheet_parser_rejects_duplicate_out_of_order_and_mismatched_coordinates(
    rows: str,
    message: str,
) -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        f'spreadsheetml/2006/main"><sheetData>{rows}</sheetData></worksheet>'
    )

    with pytest.raises(FormatError, match=message) as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"


@pytest.mark.parametrize(
    ("implicit_rows", "explicit_rows"),
    [
        (
            '<row><c r="A1"><v>1</v></c></row>',
            '<row r="1"><c r="A1"><v>1</v></c></row>',
        ),
        (
            '<row r="1"><c r="A1"><v>1</v></c></row><row><c r="A2"><v>2</v></c></row>',
            '<row r="1"><c r="A1"><v>1</v></c></row><row r="2"><c r="A2"><v>2</v></c></row>',
        ),
        (
            '<row r="4"><c r="A4"><v>1</v></c></row><row><c r="A5"><v>2</v></c></row>',
            '<row r="4"><c r="A4"><v>1</v></c></row><row r="5"><c r="A5"><v>2</v></c></row>',
        ),
        (
            '<row r="1"><c><v>1</v></c></row>',
            '<row r="1"><c r="A1"><v>1</v></c></row>',
        ),
        (
            '<row r="1"><c r="A1"><v>1</v></c><c><f>1+1</f><v>2</v></c></row>',
            '<row r="1"><c r="A1"><v>1</v></c><c r="B1"><f>1+1</f><v>2</v></c></row>',
        ),
        (
            '<row r="1"><c r="B1"><v>1</v></c><c><v>2</v></c>'
            '<c r="E1"><v>3</v></c><c><v>4</v></c></row>',
            '<row r="1"><c r="B1"><v>1</v></c><c r="C1"><v>2</v></c>'
            '<c r="E1"><v>3</v></c><c r="F1"><v>4</v></c></row>',
        ),
        (
            '<row r="3"/><row hidden="1"/>',
            '<row r="3"/><row r="4" hidden="1"/>',
        ),
        (
            '<row r="1"><c r="C1"><v>1</v></c></row><row><c><v>2</v></c></row>',
            '<row r="1"><c r="C1"><v>1</v></c></row><row r="2"><c r="A2"><v>2</v></c></row>',
        ),
    ],
    ids=[
        "first-row",
        "later-row",
        "explicit-gap-then-row",
        "first-cell",
        "subsequent-cell",
        "mixed-cells",
        "hidden-row",
        "column-state-reset",
    ],
)
def test_implicit_rows_and_cells_match_fully_explicit_manifest(
    implicit_rows: str,
    explicit_rows: str,
) -> None:
    prefix = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
    )
    suffix = "</sheetData></worksheet>"

    implicit = _parse_synthetic_sheet(f"{prefix}{implicit_rows}{suffix}")
    explicit = _parse_synthetic_sheet(f"{prefix}{explicit_rows}{suffix}")

    assert implicit == explicit


@pytest.mark.parametrize(
    ("rows", "coordinate"),
    [
        ('<row r="1048576"/><row/>', 1_048_577),
        ('<row r="1"><c r="XFD1"/><c/></row>', "XFE1"),
    ],
    ids=["row", "cell"],
)
def test_implicit_coordinate_overflow_is_rejected_with_context(
    rows: str,
    coordinate: int | str,
) -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        f'spreadsheetml/2006/main"><sheetData>{rows}</sheetData></worksheet>'
    )

    with pytest.raises(FormatError, match="out-of-bounds") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == coordinate


def test_explicit_cell_outside_row_remains_rejected() -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><sheetData><c r="A1"><v>1</v></c>'
        "</sheetData></worksheet>"
    )

    with pytest.raises(FormatError, match="no enclosing row") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"


def test_empty_explicit_cell_reference_is_contextual_format_error() -> None:
    xml = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><sheetData>'
        '<row r="1"><c r=""><v>1</v></c></row>'
        "</sheetData></worksheet>"
    )

    with pytest.raises(FormatError, match="malformed cell coordinate") as raised:
        _parse_synthetic_sheet(xml)

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["coordinate"] == ""


def test_sparse_counter_overflow_is_rejected_as_contextual_format_error() -> None:
    counts = array("H", [65_535])

    with pytest.raises(FormatError, match="sparse") as raised:
        manifest_module._increment_sparse_counts(
            counts,
            {1},
            "xl/worksheets/sheet1.xml",
        )

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert raised.value.context["column"] == 1


def test_excel_upper_coordinate_is_accepted(tmp_path) -> None:
    path = tmp_path / "upper-edge.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "value"
    workbook.save(path)
    workbook.close()
    _rewrite_member(
        path,
        "xl/worksheets/sheet1.xml",
        lambda xml: _replace_required(
            _replace_required(xml, b'ref="A1:A1"', b'ref="XFD1048576:XFD1048576"'),
            b'<row r="1"><c r="A1"',
            b'<row r="1048576"><c r="XFD1048576"',
        ),
    )

    with SourceHandle(path) as source:
        sheet = ManifestReader(source).sheet("Sheet")

    assert sheet.declared_dimension == (1_048_576, 16_384, 1_048_576, 16_384)
    assert (sheet.observed_max_row, sheet.observed_max_col) == (1_048_576, 16_384)


@pytest.mark.parametrize(
    ("setup", "old", "new", "coordinate"),
    [
        ("cell", b'r="A1"', b'r="XFE1"', "XFE1"),
        ("cell", b'r="A1"', b'r="A1048577"', "A1048577"),
        ("dimension", b'ref="A1:A1"', b'ref="A1:XFE1"', "A1:XFE1"),
        ("merge", b'ref="A1:B1"', b'ref="XFE1:XFE1"', "XFE1:XFE1"),
        ("hidden_row", b'r="1" hidden="1"', b'r="1048577" hidden="1"', "1048577"),
        (
            "hidden_col",
            b'customWidth="1" min="1" max="1"',
            b'customWidth="1" min="16385" max="16385"',
            "16385",
        ),
    ],
)
def test_out_of_bounds_coordinates_are_rejected_with_member_context(
    tmp_path: Path,
    setup: str,
    old: bytes,
    new: bytes,
    coordinate: str,
) -> None:
    path = tmp_path / f"{setup}.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "value"
    if setup == "merge":
        worksheet.merge_cells("A1:B1")
    elif setup == "hidden_row":
        worksheet.row_dimensions[1].hidden = True
    elif setup == "hidden_col":
        worksheet.column_dimensions["A"].hidden = True
    workbook.save(path)
    workbook.close()
    _rewrite_member(
        path,
        "xl/worksheets/sheet1.xml",
        lambda xml: _replace_required(xml, old, new),
    )

    with (
        SourceHandle(path) as source,
        pytest.raises(FormatError, match="coordinate") as raised,
    ):
        ManifestReader(source).sheet("Sheet")

    assert raised.value.context["member"] == "xl/worksheets/sheet1.xml"
    assert coordinate in str(raised.value.context)


def test_path_manifest_reader_rejects_equal_size_replacement_with_restored_mtime(
    tmp_path: Path,
) -> None:
    path = tmp_path / "stable.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(path)
    workbook.close()

    with SourceHandle(path) as source:
        reader = ManifestReader(source)
        original = path.stat()
        replacement = tmp_path / "replacement.xlsx"
        replacement.write_bytes(path.read_bytes())
        os.utime(replacement, ns=(original.st_atime_ns, original.st_mtime_ns))
        replacement.replace(path)
        os.utime(path, ns=(original.st_atime_ns, original.st_mtime_ns))

        with pytest.raises(FormatError, match="changed") as raised:
            reader.sheet("Sheet")

    assert raised.value.context["file_path"] == str(path)


def test_path_manifest_reader_rejects_replacement_during_eager_build(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "eager-race.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "old"
    workbook.save(path)
    workbook.close()
    original_identity = path.stat()
    original_build_manifest = manifest_module.build_manifest

    def replace_after_build(source, limits=DEFAULT_LIMITS):
        parsed = original_build_manifest(source, limits)
        replacement = tmp_path / "eager-race-replacement.xlsx"
        replacement.write_bytes(path.read_bytes())
        os.utime(
            replacement,
            ns=(original_identity.st_atime_ns, original_identity.st_mtime_ns),
        )
        replacement.replace(path)
        os.utime(path, ns=(original_identity.st_atime_ns, original_identity.st_mtime_ns))
        assert path.stat().st_size == original_identity.st_size
        assert path.stat().st_mtime_ns == original_identity.st_mtime_ns
        return parsed

    monkeypatch.setattr(manifest_module, "build_manifest", replace_after_build)

    with (
        SourceHandle(path) as source,
        pytest.raises(FormatError, match="source changed") as raised,
    ):
        ManifestReader(source)

    assert raised.value.context == {
        "file_path": str(path),
        "operation": "read worksheet metadata",
    }


def test_path_manifest_reader_rechecks_identity_after_lazy_parse(tmp_path: Path) -> None:
    path = tmp_path / "race.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(path)
    workbook.close()

    def replace_after_member_open(_member: str) -> None:
        replacement = tmp_path / "race-replacement.xlsx"
        replacement.write_bytes(path.read_bytes())
        replacement.replace(path)

    with SourceHandle(path) as source:
        reader = ManifestReader(source, on_member_open=replace_after_member_open)
        with pytest.raises(FormatError, match="changed"):
            reader.sheet("Sheet")
