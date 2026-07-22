"""Contracts for the bounded-row OOXML Arrow reader."""

from __future__ import annotations

import io
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import openpyxl
import pyarrow as pa
import pytest

import messy_xlsx.parsing.coordinates as coordinates_module
import messy_xlsx.parsing.xlsx_streaming as streaming_module
from messy_xlsx import SheetConfig
from messy_xlsx._source import SourceHandle
from messy_xlsx.enums import MergeStrategy
from messy_xlsx.ooxml.manifest import ManifestReader
from messy_xlsx.ooxml.models import SheetManifest
from messy_xlsx.parsing.contracts import OutputMode, ParseMetrics
from messy_xlsx.parsing.coordinates import (
    CoordinateBatch,
    CoordinateCompatibilityError,
    CoordinateTransform,
)
from messy_xlsx.parsing.fallback import FallbackCoordinator
from messy_xlsx.parsing.parse_plan import ParsePlan, compile_parse_plan
from messy_xlsx.parsing.xlsx_streaming import (
    OpenpyxlStreamingReader,
    StreamingWorksheetLayout,
    reader_batches,
)


def _plan(*, batch_size: int = 2, **overrides: Any) -> ParsePlan:
    values: dict[str, Any] = {
        "auto_detect": False,
        "header_rows": 0,
        "normalize": False,
        "sanitize_column_names": False,
        "include_hidden": True,
        "merge_strategy": MergeStrategy.SKIP,
    }
    values.update(overrides)
    return compile_parse_plan(
        SheetConfig(**values),
        structure=None,
        format_type="xlsx",
        output_mode=OutputMode.STREAMING,
        batch_size=batch_size,
    )


def _manifest(source: SourceHandle, sheet: str = "Data") -> SheetManifest:
    return ManifestReader(source).sheet(sheet)


def _open_reader(
    source: SourceHandle,
    manifest: SheetManifest,
    plan: ParsePlan,
    raw_schema: pa.Schema,
    *,
    raw_column_numbers: tuple[int, ...] | None = None,
    metrics: ParseMetrics | None = None,
) -> OpenpyxlStreamingReader:
    transform = CoordinateTransform.from_manifest(manifest)
    layout = StreamingWorksheetLayout.compile(
        manifest,
        plan,
        raw_schema,
        transform,
        raw_column_numbers=raw_column_numbers,
    )
    return OpenpyxlStreamingReader(
        source,
        manifest,
        plan,
        layout,
        transform,
        metrics=metrics,
    )


@pytest.fixture
def streaming_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "streaming.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["name", "amount"])
    for index in range(7):
        sheet.append([f"item-{index}", index])
    workbook.save(path)
    workbook.close()
    return path


def test_reader_schema_matches_every_bounded_batch(streaming_xlsx: Path) -> None:
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    plan = _plan(batch_size=2)
    reader = _open_reader(
        source,
        manifest,
        plan,
        pa.schema([("raw-name", pa.string()), ("raw-amount", pa.string())]),
    )

    batches = list(reader_batches(reader))

    assert batches
    assert all(0 < batch.num_rows <= 2 for batch in batches)
    assert all(batch.schema.equals(reader.schema, check_metadata=True) for batch in batches)
    assert reader.schema.names == ["0", "1"]
    reader.close()
    source.close()


def test_layout_uses_absolute_observed_coordinates_and_auxiliary_merge_anchor(
    tmp_path: Path,
) -> None:
    path = tmp_path / "offset-merge.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["C4"] = "anchor"
    sheet.merge_cells("C4:E6")
    sheet["F7"] = "tail"
    workbook.save(path)
    workbook.close()

    source = SourceHandle(path)
    manifest = _manifest(source)
    plan = _plan(
        batch_size=2,
        cell_range="D5:F7",
        merge_strategy=MergeStrategy.FILL,
    )
    transform = CoordinateTransform.from_manifest(manifest)
    layout = StreamingWorksheetLayout.compile(
        manifest,
        plan,
        pa.schema(
            [
                ("c", pa.string()),
                ("d", pa.string()),
                ("e", pa.string()),
                ("f", pa.string()),
            ]
        ),
        transform,
        raw_column_numbers=(3, 4, 5, 6),
    )

    assert (layout.min_row, layout.min_col, layout.max_row, layout.max_col) == (4, 3, 7, 6)
    assert layout.output_column_numbers == (4, 5, 6)
    reader = OpenpyxlStreamingReader(source, manifest, plan, layout, transform)
    result = pa.Table.from_batches(list(reader_batches(reader)), schema=reader.schema)
    assert result.to_pydict() == {
        "0": ["anchor", "anchor", None],
        "1": ["anchor", "anchor", None],
        "2": [None, None, "tail"],
    }
    source.close()


def test_large_padded_range_is_streamed_without_terminal_giant_allocation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "padded.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "first"
    sheet.row_dimensions[1_000_000].height = 15
    workbook.save(path)
    workbook.close()

    requested_null_lengths: list[int] = []
    real_nulls = coordinates_module.pa.nulls

    def recording_nulls(size: int, *args: Any, **kwargs: Any) -> pa.Array:
        requested_null_lengths.append(size)
        return real_nulls(size, *args, **kwargs)

    monkeypatch.setattr(coordinates_module.pa, "nulls", recording_nulls)
    source = SourceHandle(path)
    manifest = _manifest(source)
    plan = _plan(batch_size=7, cell_range="A1:A1003")
    reader = _open_reader(source, manifest, plan, pa.schema([("a", pa.string())]))

    batches = list(reader_batches(reader))

    assert sum(batch.num_rows for batch in batches) == 1003
    assert all(0 < batch.num_rows <= 7 for batch in batches)
    assert not requested_null_lengths or max(requested_null_lengths) <= 7
    source.close()


def test_hidden_header_footer_and_merge_carry_keep_stable_bounded_output(
    tmp_path: Path,
) -> None:
    path = tmp_path / "framed.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["name", "value"])
    sheet.append(["merged", None])
    sheet.merge_cells("A2:B5")
    for row in range(6, 16):
        sheet.append([f"row-{row}", row])
    sheet.row_dimensions[7].hidden = True
    workbook.save(path)
    workbook.close()

    source = SourceHandle(path)
    manifest = _manifest(source)
    plan = _plan(
        batch_size=2,
        header_rows=1,
        skip_footer=5,
        include_hidden=False,
        merge_strategy=MergeStrategy.FILL,
    )
    reader = _open_reader(
        source,
        manifest,
        plan,
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    batches = list(reader_batches(reader))

    assert batches
    assert all(0 < batch.num_rows <= 2 for batch in batches)
    assert all(batch.schema.equals(reader.schema) for batch in batches)
    assert pa.Table.from_batches(batches).column(0).to_pylist()[:4] == [
        "merged",
        "merged",
        "merged",
        "merged",
    ]
    source.close()


def test_openpyxl_flags_formula_modes_and_single_load(
    streaming_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = openpyxl.load_workbook(streaming_xlsx)
    workbook["Data"]["A2"] = "=1+1"
    workbook.save(streaming_xlsx)
    workbook.close()
    real_load = streaming_module.openpyxl.load_workbook
    calls: list[dict[str, object]] = []

    def recording_load(*args: Any, **kwargs: Any) -> Any:
        calls.append(dict(kwargs))
        return real_load(*args, **kwargs)

    monkeypatch.setattr(streaming_module.openpyxl, "load_workbook", recording_load)
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(evaluate_formulas=False),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )

    values = pa.Table.from_batches(list(reader_batches(reader))).column(0).to_pylist()

    assert values[1] == "=1+1"
    assert calls == [
        {
            "read_only": True,
            "data_only": False,
            "keep_links": False,
            "keep_vba": False,
        }
    ]
    source.close()


def test_seekable_cursor_restored_and_caller_stream_left_open(streaming_xlsx: Path) -> None:
    stream = io.BytesIO(streaming_xlsx.read_bytes())
    stream.seek(11)
    source = SourceHandle(stream, filename="streaming.xlsx")
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    next(reader_batches(reader))
    reader.close()

    assert stream.tell() == 11
    assert stream.closed is False
    source.close()


class _FakeReader:
    def __init__(
        self,
        values: list[pa.RecordBatch | None] | None = None,
        *,
        read_error: BaseException | None = None,
        close_error: BaseException | None = None,
    ) -> None:
        self.schema = pa.schema([("0", pa.int64())])
        self._values: Iterator[pa.RecordBatch | None] = iter(values or [])
        self._read_error = read_error
        self._close_error = close_error
        self.read_calls = 0
        self.close_calls = 0

    def read_next_batch(self) -> pa.RecordBatch | None:
        self.read_calls += 1
        if self._read_error is not None:
            error = self._read_error
            self._read_error = None
            raise error
        return next(self._values, None)

    def close(self) -> None:
        self.close_calls += 1
        if self._close_error is not None:
            raise self._close_error


def test_closable_adapter_can_close_before_first_next_and_at_eof() -> None:
    unopened = _FakeReader([pa.record_batch([[1]], names=["0"]), None])
    adapter = reader_batches(unopened)
    adapter.close()
    adapter.close()
    assert unopened.read_calls == 0
    assert unopened.close_calls == 1

    exhausted = _FakeReader([pa.record_batch([[1]], names=["0"]), None])
    adapter = reader_batches(exhausted)
    assert next(adapter).num_rows == 1
    with pytest.raises(StopIteration):
        next(adapter)
    assert exhausted.close_calls == 1


def test_adapter_preserves_operation_error_and_sanitizes_cleanup_failure() -> None:
    primary = ValueError("operation secret")
    cleanup = OSError("cleanup secret")
    reader = _FakeReader(read_error=primary, close_error=cleanup)

    with pytest.raises(ValueError) as captured:
        next(reader_batches(reader))

    assert captured.value is primary
    assert reader.close_calls == 1
    assert primary.__dict__["backend_context"]["cleanup_failure"] == {"type": "OSError"}
    assert "cleanup secret" not in " ".join(getattr(primary, "__notes__", ()))


def test_invalid_layout_fails_before_openpyxl_load(
    streaming_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls = 0

    def forbidden_load(*_args: Any, **_kwargs: Any) -> Any:
        nonlocal calls
        calls += 1
        raise AssertionError("parser I/O must not start")

    monkeypatch.setattr(streaming_module.openpyxl, "load_workbook", forbidden_load)
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    transform = CoordinateTransform.from_manifest(manifest)

    with pytest.raises(ValueError, match="raw schema width"):
        StreamingWorksheetLayout.compile(
            manifest,
            _plan(),
            pa.schema([("only-one", pa.string())]),
            transform,
        )

    assert calls == 0
    source.close()


def test_reader_does_not_mutate_streaming_metrics(streaming_xlsx: Path) -> None:
    metrics = ParseMetrics()
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
        metrics=metrics,
    )

    list(reader_batches(reader))

    assert metrics.streaming_passes == 0
    source.close()


def test_fallback_coordinator_alone_counts_clean_exhaustion() -> None:
    metrics = ParseMetrics()
    coordinator = FallbackCoordinator(lambda _error: False, metrics=metrics)
    reader = _FakeReader([pa.record_batch([[1]], names=["0"]), None])

    batches = list(coordinator.batches(lambda: reader, lambda: _FakeReader()))

    assert len(batches) == 1
    assert metrics.streaming_passes == 1


class _CountingRows(Iterator[tuple[object, ...]]):
    def __init__(self, source: Iterator[tuple[object, ...]]) -> None:
        self._source = source
        self.next_calls = 0

    def __next__(self) -> tuple[object, ...]:
        self.next_calls += 1
        return next(self._source)


class _TrackingWorksheet:
    def __init__(self, worksheet: object) -> None:
        self._worksheet = worksheet
        self.iter_calls: list[dict[str, object]] = []
        self.rows: _CountingRows | None = None

    def reset_dimensions(self) -> None:
        self._worksheet.reset_dimensions()  # type: ignore[attr-defined]

    def iter_rows(self, **kwargs: object) -> _CountingRows:
        self.iter_calls.append(dict(kwargs))
        rows = _CountingRows(iter(self._worksheet.iter_rows(**kwargs)))  # type: ignore[attr-defined]
        self.rows = rows
        return rows


class _TrackingOpenpyxlWorkbook:
    def __init__(self, workbook: object, sheet_name: str) -> None:
        self._workbook = workbook
        self.sheet = _TrackingWorksheet(workbook[sheet_name])  # type: ignore[index]
        self.close_calls = 0

    def __getitem__(self, name: str) -> _TrackingWorksheet:
        if name != "Data":
            raise KeyError(name)
        return self.sheet

    def close(self) -> None:
        self.close_calls += 1
        self._workbook.close()  # type: ignore[attr-defined]


def test_one_exact_iter_rows_call_and_early_close_stops_bounded_pull(
    streaming_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    real_load = streaming_module.openpyxl.load_workbook
    tracked: list[_TrackingOpenpyxlWorkbook] = []

    def tracking_load(*args: object, **kwargs: object) -> _TrackingOpenpyxlWorkbook:
        wrapped = _TrackingOpenpyxlWorkbook(real_load(*args, **kwargs), "Data")
        tracked.append(wrapped)
        return wrapped

    monkeypatch.setattr(streaming_module.openpyxl, "load_workbook", tracking_load)
    reader = _open_reader(
        source,
        manifest,
        _plan(batch_size=2),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    adapter = reader_batches(reader)

    assert tracked[0].sheet.rows is not None
    assert tracked[0].sheet.rows.next_calls == 0
    assert next(adapter).num_rows == 2
    assert tracked[0].sheet.rows.next_calls == 2
    adapter.close()

    assert tracked[0].sheet.iter_calls == [
        {
            "values_only": True,
            "min_row": 1,
            "max_row": manifest.observed_max_row,
            "min_col": 1,
            "max_col": manifest.observed_max_col,
        }
    ]
    assert tracked[0].close_calls == 1
    assert tracked[0].sheet.rows.next_calls == 2
    source.close()


def test_cached_formula_mode_uses_data_only_true_and_returns_cached_null(
    streaming_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = openpyxl.load_workbook(streaming_xlsx)
    workbook["Data"]["A2"] = "=1+1"
    workbook.save(streaming_xlsx)
    workbook.close()
    real_load = streaming_module.openpyxl.load_workbook
    calls: list[dict[str, object]] = []

    def recording_load(*args: object, **kwargs: object) -> object:
        calls.append(dict(kwargs))
        return real_load(*args, **kwargs)

    monkeypatch.setattr(streaming_module.openpyxl, "load_workbook", recording_load)
    source = SourceHandle(streaming_xlsx)
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(evaluate_formulas=True),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )

    values = pa.Table.from_batches(list(reader_batches(reader))).column(0).to_pylist()

    assert values[1] is None
    assert calls[0]["data_only"] is True
    source.close()


def test_coordinate_schema_preflight_is_idempotent_and_strips_field_metadata() -> None:
    transform = CoordinateTransform(
        hidden_rows=coordinates_module.IntervalIndex(()),
        hidden_columns=coordinates_module.IntervalIndex(()),
        merged_ranges=(),
    )
    plan = _plan()
    raw_schema = pa.schema(
        [pa.field("raw", pa.string(), nullable=False, metadata={b"secret": b"value"})],
        metadata={b"schema": b"metadata"},
    )

    first = transform.prepare_schema(plan, raw_schema, (4,))
    second = transform.prepare_schema(plan, raw_schema, (4,))

    assert first == second
    assert first.output_schema == pa.schema([pa.field("0", pa.string())])
    assert first.output_column_numbers == (4,)


def test_prepared_coordinate_operation_rejects_first_push_schema_drift() -> None:
    transform = CoordinateTransform(
        hidden_rows=coordinates_module.IntervalIndex(()),
        hidden_columns=coordinates_module.IntervalIndex(()),
        merged_ranges=(),
    )
    plan = _plan()
    raw_schema = pa.schema([("raw", pa.string())])
    prepared = transform.prepare_schema(plan, raw_schema, (4,))
    operation = transform.open(plan, prepared)
    drifted = CoordinateBatch(
        batch=pa.record_batch([pa.array(["value"])], names=["changed"]),
        row_numbers=pa.array([5], type=pa.int64()),
        column_numbers=(4,),
    )

    with pytest.raises(
        CoordinateCompatibilityError,
        match="coordinate input schema changed across batches",
    ):
        operation.push(drifted)


def test_empty_all_hidden_and_header_only_streams_retain_precompiled_schema(
    tmp_path: Path,
) -> None:
    path = tmp_path / "edge-streams.xlsx"
    workbook = openpyxl.Workbook()
    empty = workbook.active
    empty.title = "Empty"
    hidden = workbook.create_sheet("Hidden")
    hidden.append(["left", "right"])
    hidden.append(["a", "b"])
    hidden.column_dimensions["A"].hidden = True
    hidden.column_dimensions["B"].hidden = True
    header = workbook.create_sheet("Header")
    header.append(["left", "right"])
    workbook.save(path)
    workbook.close()

    source = SourceHandle(path)
    empty_manifest = _manifest(source, "Empty")
    empty_reader = _open_reader(source, empty_manifest, _plan(), pa.schema([]))
    assert empty_reader.read_next_batch() is None
    assert empty_reader.read_next_batch() is None
    assert empty_reader.schema == pa.schema([])
    empty_reader.close()

    hidden_manifest = _manifest(source, "Hidden")
    hidden_reader = _open_reader(
        source,
        hidden_manifest,
        _plan(include_hidden=False),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    assert list(reader_batches(hidden_reader)) == []
    assert hidden_reader.schema == pa.schema([])

    header_manifest = _manifest(source, "Header")
    header_reader = _open_reader(
        source,
        header_manifest,
        _plan(header_rows=1),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    assert list(reader_batches(header_reader)) == []
    assert header_reader.schema == pa.schema([("0", pa.string()), ("1", pa.string())])
    source.close()


class _NonSeekable:
    def __init__(self, content: bytes) -> None:
        self._content = io.BytesIO(content)
        self.closed = False

    def read(self, size: int = -1) -> bytes:
        return self._content.read(size)

    def close(self) -> None:
        self.closed = True


def test_nonseekable_source_is_replayed_without_transferring_ownership(
    streaming_xlsx: Path,
) -> None:
    caller = _NonSeekable(streaming_xlsx.read_bytes())
    source = SourceHandle(caller, filename="streaming.xlsx")  # type: ignore[arg-type]
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )

    assert sum(batch.num_rows for batch in reader_batches(reader)) == 8
    assert caller.closed is False
    source.close()
    assert caller.closed is False


def test_constructor_failure_restores_seekable_cursor_and_closes_partial_borrow(
    streaming_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    caller = io.BytesIO(streaming_xlsx.read_bytes())
    caller.seek(13)
    source = SourceHandle(caller, filename="streaming.xlsx")
    manifest = _manifest(source)
    transform = CoordinateTransform.from_manifest(manifest)
    layout = StreamingWorksheetLayout.compile(
        manifest,
        _plan(),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
        transform,
    )
    expected = RuntimeError("load failed")
    monkeypatch.setattr(
        streaming_module.openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(expected),
    )

    with pytest.raises(RuntimeError) as captured:
        OpenpyxlStreamingReader(source, manifest, _plan(), layout, transform)

    assert captured.value is expected
    assert caller.tell() == 13
    assert caller.closed is False
    source.close()


def test_adapter_process_cleanup_failure_wins_and_retains_exact_operation_error() -> None:
    operation = ValueError("operation secret")
    cleanup = MemoryError("cleanup secret")
    reader = _FakeReader(read_error=operation, close_error=cleanup)

    with pytest.raises(MemoryError) as captured:
        next(reader_batches(reader))

    assert captured.value is cleanup
    assert cleanup.__dict__["backend_context"]["operation_failure"] == {"type": "ValueError"}
    assert "operation secret" not in " ".join(getattr(cleanup, "__notes__", ()))


def test_zero_column_rechunking_preserves_row_count_across_fragments() -> None:
    rechunker = streaming_module._BatchRechunker(pa.schema([]), 2)
    for row in (1, 2):
        physical = pa.record_batch([pa.nulls(1)], names=["unused"]).select([])
        rechunker.push(
            CoordinateBatch(
                batch=physical,
                row_numbers=pa.array([row], type=pa.int64()),
                column_numbers=(),
            )
        )

    result = rechunker.pop()

    assert result is not None
    assert result.num_columns == 0
    assert result.num_rows == 2


class _FailingRows(Iterator[tuple[object, ...]]):
    def __init__(
        self,
        source: Iterator[tuple[object, ...]],
        fail_after: int,
        error: BaseException,
    ) -> None:
        self._source = source
        self._remaining = fail_after
        self._error = error

    def __next__(self) -> tuple[object, ...]:
        if self._remaining == 0:
            raise self._error
        self._remaining -= 1
        return next(self._source)


@pytest.mark.parametrize("fail_after", [0, 2, 7])
def test_first_middle_and_final_row_pass_errors_close_and_restore_cursor(
    streaming_xlsx: Path,
    fail_after: int,
) -> None:
    caller = io.BytesIO(streaming_xlsx.read_bytes())
    caller.seek(17)
    source = SourceHandle(caller, filename="streaming.xlsx")
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(batch_size=2),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    expected = RuntimeError(f"row failure {fail_after}")
    reader._rows = _FailingRows(reader._rows, fail_after, expected)
    adapter = reader_batches(reader)

    with pytest.raises(RuntimeError) as captured:
        while True:
            next(adapter)

    assert captured.value is expected
    assert reader._workbook is None
    assert reader._backend_context is None
    assert caller.tell() == 17
    assert caller.closed is False
    source.close()


class _FinishFailureOperation:
    def __init__(self, operation: object, error: BaseException) -> None:
        self._operation = operation
        self._error = error

    def push(self, batch: CoordinateBatch) -> tuple[CoordinateBatch, ...]:
        return self._operation.push(batch)  # type: ignore[attr-defined, no-any-return]

    def finish(self) -> tuple[CoordinateBatch, ...]:
        raise self._error


def test_coordinate_finalize_error_closes_and_restores_cursor(streaming_xlsx: Path) -> None:
    caller = io.BytesIO(streaming_xlsx.read_bytes())
    caller.seek(19)
    source = SourceHandle(caller, filename="streaming.xlsx")
    manifest = _manifest(source)
    reader = _open_reader(
        source,
        manifest,
        _plan(batch_size=2),
        pa.schema([("a", pa.string()), ("b", pa.string())]),
    )
    expected = RuntimeError("finalize failed")
    reader._operation = _FinishFailureOperation(reader._operation, expected)  # type: ignore[assignment]

    with pytest.raises(RuntimeError) as captured:
        list(reader_batches(reader))

    assert captured.value is expected
    assert reader._workbook is None
    assert reader._backend_context is None
    assert caller.tell() == 19
    assert caller.closed is False
    source.close()
