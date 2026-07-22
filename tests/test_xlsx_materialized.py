"""Contracts for the fastexcel whole-sheet materialized Arrow path."""

from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any

import fastexcel
import openpyxl
import pandas as pd
import pyarrow as pa
import pytest
from pandas.testing import assert_frame_equal

import messy_xlsx.parsing.xlsx_handler as xlsx_handler_module
from messy_xlsx import MessyWorkbook, SheetConfig
from messy_xlsx._source import SourceHandle
from messy_xlsx._spool import DEFAULT_MEMORY_LIMIT
from messy_xlsx.enums import MergeStrategy
from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.parsing.base_handler import ParseOptions
from messy_xlsx.parsing.contracts import OutputMode, ParseMetrics
from messy_xlsx.parsing.fallback import FallbackCoordinator
from messy_xlsx.parsing.fastexcel_session import FastexcelSession
from messy_xlsx.parsing.legacy_adapter import LegacyDataFrameAdapter
from messy_xlsx.parsing.parse_plan import ParsePlan, compile_parse_plan
from messy_xlsx.parsing.xlsx_handler import (
    XLSXHandler,
    _is_fastexcel_materialized_plan,
    is_fastexcel_compatibility_error,
)
from messy_xlsx.parsing.xlsx_materialized import FastexcelMaterializedReader


def _eligible_config(**overrides: Any) -> SheetConfig:
    values: dict[str, Any] = {
        "auto_detect": False,
        "include_hidden": True,
        "merge_strategy": MergeStrategy.SKIP,
        "evaluate_formulas": True,
        "normalize": False,
        "sanitize_column_names": False,
    }
    values.update(overrides)
    return SheetConfig(**values)


@pytest.fixture
def basic_parse_plan() -> ParsePlan:
    return compile_parse_plan(
        _eligible_config(),
        structure=None,
        format_type="xlsx",
        output_mode=OutputMode.MATERIALIZED,
        batch_size=None,
    )


def _save_rows(path: Path, rows: list[list[Any]], *, title: str = "Data") -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _padded_workbook_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "large.xlsx"
    _save_rows(path, [["Name", "Value"], ["A", 1]])
    with zipfile.ZipFile(path, "a", compression=zipfile.ZIP_STORED) as archive:
        archive.writestr("task7-padding.bin", b"x" * (DEFAULT_MEMORY_LIMIT + 1024))
    content = path.read_bytes()
    assert len(content) > DEFAULT_MEMORY_LIMIT
    return content


def _padded_two_sheet_workbook_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "large-two-sheet.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Name", "Value"])
    first.append(["A", 1])
    second = workbook.create_sheet("Second")
    second.append(["Name", "Value"])
    second.append(["B", 2])
    workbook.save(path)
    workbook.close()
    with zipfile.ZipFile(path, "a", compression=zipfile.ZIP_STORED) as archive:
        archive.writestr("task7-padding.bin", b"x" * (DEFAULT_MEMORY_LIMIT + 1024))
    content = path.read_bytes()
    assert len(content) > DEFAULT_MEMORY_LIMIT
    return content


class _NonSeekableBytes:
    def __init__(self, content: bytes, name: str = "upload.xlsx") -> None:
        self._stream = io.BytesIO(content)
        self.name = name
        self.closed = False

    def read(self, size: int = -1) -> bytes:
        return self._stream.read(size)

    def seekable(self) -> bool:
        return False


class _SessionFake:
    def __init__(self, value: object) -> None:
        self.value = value
        self.calls: list[tuple[str, int]] = []

    def materialize(self, sheet: str, *, skip_rows: int) -> object:
        self.calls.append((sheet, skip_rows))
        return self.value


class _ArrowWrapper:
    def __init__(self, value: object) -> None:
        self.value = value
        self.calls = 0

    def to_arrow(self) -> object:
        self.calls += 1
        return self.value


@pytest.mark.parametrize("backend_kind", ["table", "batch", "wrapper_batch", "wrapper_table"])
def test_reader_materializes_once_and_accepts_each_arrow_shape(
    basic_parse_plan: ParsePlan,
    backend_kind: str,
) -> None:
    batch = pa.record_batch({"value": [1, 2]})
    expected = pa.Table.from_batches([batch])
    wrapper: _ArrowWrapper | None = None
    if backend_kind == "table":
        backend: object = expected
    elif backend_kind == "batch":
        backend = batch
    elif backend_kind == "wrapper_batch":
        wrapper = _ArrowWrapper(batch)
        backend = wrapper
    else:
        wrapper = _ArrowWrapper(expected)
        backend = wrapper

    session = _SessionFake(backend)
    reader = FastexcelMaterializedReader(session, "Data", basic_parse_plan)

    actual = reader.read_table()

    assert actual.equals(expected)
    if backend_kind == "table":
        assert actual is expected
    assert session.calls == [("Data", 0)]
    assert reader._plan is basic_parse_plan
    if wrapper is not None:
        assert wrapper.calls == 1


def test_reader_rejects_non_arrow_wrapper_result_after_one_conversion(
    basic_parse_plan: ParsePlan,
) -> None:
    wrapper = _ArrowWrapper(object())
    session = _SessionFake(wrapper)

    with pytest.raises(TypeError, match="pyarrow"):
        FastexcelMaterializedReader(session, "Data", basic_parse_plan).read_table()

    assert session.calls == [("Data", 0)]
    assert wrapper.calls == 1


def test_only_coordinator_records_successful_materialization(
    basic_parse_plan: ParsePlan,
) -> None:
    metrics = ParseMetrics()
    reader = FastexcelMaterializedReader(
        _SessionFake(pa.table({"value": [1]})),
        "Data",
        basic_parse_plan,
    )

    table = FallbackCoordinator(
        is_fastexcel_compatibility_error,
        metrics=metrics,
    ).materialize(lambda: reader, pytest.fail)

    assert table.to_pydict() == {"value": [1]}
    assert metrics == ParseMetrics(full_materializations=1)


def test_failed_conversion_is_not_counted_as_a_materialization(
    basic_parse_plan: ParsePlan,
) -> None:
    metrics = ParseMetrics()
    reader = FastexcelMaterializedReader(
        _SessionFake(_ArrowWrapper(object())),
        "Data",
        basic_parse_plan,
    )

    with pytest.raises(TypeError, match="pyarrow"):
        FallbackCoordinator(
            is_fastexcel_compatibility_error,
            metrics=metrics,
        ).materialize(lambda: reader, pytest.fail)

    assert metrics == ParseMetrics(failed_attempts=1)


def test_leading_blank_rows_remain_in_raw_arrow_coordinates(
    tmp_path: Path,
    basic_parse_plan: ParsePlan,
) -> None:
    path = tmp_path / "leading-blanks.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A3"] = "Name"
    worksheet["B3"] = "Value"
    worksheet["A4"] = "A"
    worksheet["B4"] = 1
    workbook.save(path)
    workbook.close()

    with SourceHandle(path) as source, FastexcelSession(source) as session:
        table = FastexcelMaterializedReader(session, "Data", basic_parse_plan).read_table()

    assert table.num_rows == 4
    assert table.num_columns == 2
    assert table.column(0).to_pylist()[:3] == [None, None, "Name"]
    assert table.column(1).to_pylist()[:3] == [None, None, "Value"]


@pytest.mark.parametrize(
    ("rows", "expected_rows", "expected_columns", "expected_values"),
    [
        ([], 0, 0, {}),
        (
            [["Name", "Value"]],
            1,
            2,
            {"__UNNAMED__0": ["Name"], "__UNNAMED__1": ["Value"]},
        ),
    ],
)
def test_empty_and_header_only_sheets_preserve_raw_shape(
    tmp_path: Path,
    basic_parse_plan: ParsePlan,
    rows: list[list[Any]],
    expected_rows: int,
    expected_columns: int,
    expected_values: dict[str, list[Any]],
) -> None:
    path = tmp_path / f"shape-{expected_rows}-{expected_columns}.xlsx"
    _save_rows(path, rows)

    with SourceHandle(path) as source, FastexcelSession(source) as session:
        table = FastexcelMaterializedReader(session, "Data", basic_parse_plan).read_table()

    assert table.num_rows == expected_rows
    assert table.num_columns == expected_columns
    assert table.to_pydict() == expected_values


def test_explicit_ordinary_reader_path_never_loads_openpyxl(
    sample_xlsx: Path,
    basic_parse_plan: ParsePlan,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def forbidden(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("openpyxl must not load on the ordinary materialized path")

    monkeypatch.setattr(openpyxl, "load_workbook", forbidden)
    with SourceHandle(sample_xlsx) as source, FastexcelSession(source) as session:
        table = FastexcelMaterializedReader(session, "Data", basic_parse_plan).read_table()

    assert table.num_rows == 4


def test_adapter_calls_table_to_pandas_exactly_once(basic_parse_plan: ParsePlan) -> None:
    expected = pd.DataFrame({"value": [1]})

    class TableSpy:
        def __init__(self) -> None:
            self.calls = 0

        def to_pandas(self) -> pd.DataFrame:
            self.calls += 1
            return expected

    table = TableSpy()

    actual = LegacyDataFrameAdapter().to_dataframe(table, basic_parse_plan)

    assert actual is expected
    assert table.calls == 1


def test_public_explicit_ordinary_path_never_loads_openpyxl(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = _eligible_config()
    adapter_calls = 0
    adapter_original = LegacyDataFrameAdapter.to_dataframe

    def forbidden(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("openpyxl must not load on the classified ordinary path")

    def recording_adapter(
        self: LegacyDataFrameAdapter,
        table: object,
        plan: ParsePlan,
    ) -> pd.DataFrame:
        nonlocal adapter_calls
        adapter_calls += 1
        return adapter_original(self, table, plan)

    monkeypatch.setattr(openpyxl, "load_workbook", forbidden)
    monkeypatch.setattr(LegacyDataFrameAdapter, "to_dataframe", recording_adapter)
    with MessyWorkbook(sample_xlsx, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    assert frame.to_dict(orient="list") == {
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": ["30", "25", "35"],
        "City": ["New York", "Los Angeles", "Chicago"],
    }
    assert adapter_calls == 1


def test_eligible_missing_sheet_keeps_registry_error_boundary(
    sample_xlsx: Path,
) -> None:
    config = _eligible_config()

    with (
        MessyWorkbook(sample_xlsx, sheet_config=config) as workbook,
        pytest.raises(FormatError) as captured,
    ):
        workbook._to_dataframe_compat("Missing")

    assert captured.value.message == "All handlers failed for sample.xlsx"
    assert captured.value.context == {
        "file_path": str(sample_xlsx),
        "detected_format": "xlsx",
        "attempted_formats": ["XLSXHandler"],
    }
    assert captured.value.__cause__ is None
    assert captured.value.__context__ is None
    assert captured.value.__suppress_context__ is False


def test_allowed_primary_failure_closes_operation_before_fallback_and_borrowed_session_lasts(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = _eligible_config()
    events: list[str] = []
    close_original = FastexcelSession.close
    load_original = openpyxl.load_workbook
    operation_close_original = xlsx_handler_module._FastexcelDataFrameReader.close

    def failing_read(_self: FastexcelMaterializedReader) -> pa.Table:
        events.append("fastexcel-read")
        raise fastexcel.CalamineCellError("unsupported cell representation")

    def recording_close(self: FastexcelSession) -> None:
        events.append("fastexcel-close")
        close_original(self)

    def recording_operation_close(
        self: xlsx_handler_module._FastexcelDataFrameReader,
    ) -> None:
        events.append("operation-close")
        operation_close_original(self)

    def recording_load(*args: object, **kwargs: object) -> Any:
        events.append("openpyxl-open")
        return load_original(*args, **kwargs)

    with MessyWorkbook(sample_xlsx, sheet_config=config) as workbook:
        monkeypatch.setattr(FastexcelMaterializedReader, "read_table", failing_read)
        monkeypatch.setattr(FastexcelSession, "close", recording_close)
        monkeypatch.setattr(
            xlsx_handler_module._FastexcelDataFrameReader,
            "close",
            recording_operation_close,
        )
        monkeypatch.setattr(openpyxl, "load_workbook", recording_load)
        frame = workbook._to_dataframe_compat("Data")

    assert frame.shape == (3, 3)
    assert events.count("fastexcel-read") == 1
    assert events.count("fastexcel-close") == 1
    assert events.count("operation-close") == 1
    assert events.count("openpyxl-open") == 1
    assert events.index("operation-close") < events.index("openpyxl-open")
    assert events.index("openpyxl-open") < events.index("fastexcel-close")


def test_classified_shared_session_init_failure_reaches_openpyxl_fallback(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = _eligible_config()
    load_original = openpyxl.load_workbook
    session_attempts = 0
    openpyxl_calls = 0

    def failing_session_factory() -> FastexcelSession:
        nonlocal session_attempts
        session_attempts += 1
        raise fastexcel.CalamineCellError("unsupported workbook cell representation")

    def recording_load(*args: object, **kwargs: object) -> Any:
        nonlocal openpyxl_calls
        openpyxl_calls += 1
        return load_original(*args, **kwargs)

    with MessyWorkbook(sample_xlsx, sheet_config=config) as workbook:
        monkeypatch.setattr(workbook, "_get_fastexcel_session", failing_session_factory)
        monkeypatch.setattr(openpyxl, "load_workbook", recording_load)
        frame = workbook._to_dataframe_compat("Data")

    assert frame.shape == (3, 3)
    assert session_attempts == 1
    assert openpyxl_calls == 1


def test_classified_owned_session_init_failure_reaches_openpyxl_fallback(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    real_load_workbook = openpyxl.load_workbook
    session_attempts = 0
    openpyxl_calls = 0

    def failing_session_factory(_source: SourceHandle) -> FastexcelSession:
        nonlocal session_attempts
        session_attempts += 1
        raise fastexcel.CalamineCellError("unsupported workbook cell representation")

    def recording_load(*args: object, **kwargs: object) -> Any:
        nonlocal openpyxl_calls
        openpyxl_calls += 1
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(xlsx_handler_module, "FastexcelSession", failing_session_factory)
    monkeypatch.setattr(openpyxl, "load_workbook", recording_load)

    frame = XLSXHandler().parse(
        sample_xlsx,
        "Data",
        ParseOptions(merge_strategy=MergeStrategy.SKIP),
    )

    assert frame.shape == (3, 3)
    assert session_attempts == 1
    assert openpyxl_calls == 1


def test_shared_session_permission_failure_keeps_file_error_boundary(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = _eligible_config()
    openpyxl_calls = 0

    def failing_session_factory() -> FastexcelSession:
        raise PermissionError("denied")

    def forbidden_openpyxl(*_args: object, **_kwargs: object) -> None:
        nonlocal openpyxl_calls
        openpyxl_calls += 1
        raise AssertionError("permission failure must not retry through openpyxl")

    with MessyWorkbook(sample_xlsx, sheet_config=config) as workbook:
        monkeypatch.setattr(workbook, "_get_fastexcel_session", failing_session_factory)
        monkeypatch.setattr(openpyxl, "load_workbook", forbidden_openpyxl)
        with pytest.raises(FileError) as captured:
            workbook._to_dataframe_compat("Data")

    assert captured.value.message == f"Permission denied: {sample_xlsx}"
    assert captured.value.context == {
        "file_path": str(sample_xlsx),
        "operation": "open",
    }
    assert openpyxl_calls == 0


def test_workbook_reuses_one_session_and_closes_it_before_owned_spill(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    stream = io.BytesIO(_padded_two_sheet_workbook_bytes(tmp_path))
    workbook = MessyWorkbook(stream, sheet_config=_eligible_config(), filename="large.xlsx")
    read_excel_original = fastexcel.read_excel
    session_close_original = FastexcelSession.close
    source_close_original = SourceHandle.close
    backends: list[object] = []
    events: list[str] = []

    def recording_read_excel(backend: object) -> object:
        backends.append(backend)
        return read_excel_original(backend)

    def recording_session_close(self: FastexcelSession) -> None:
        events.append("session-close")
        session_close_original(self)

    def recording_source_close(self: SourceHandle) -> None:
        events.append("source-close")
        source_close_original(self)

    monkeypatch.setattr(fastexcel, "read_excel", recording_read_excel)
    monkeypatch.setattr(FastexcelSession, "close", recording_session_close)
    try:
        first = workbook._to_dataframe_compat("First")
        second = workbook._to_dataframe_compat("Second")
        session = workbook._fastexcel_session
        assert isinstance(session, FastexcelSession)
        spool = workbook._source_handle._spool
        assert spool is not None
        spill_path = spool._path
        assert spill_path is not None and spill_path.exists()
        monkeypatch.setattr(SourceHandle, "close", recording_source_close)
    finally:
        workbook.close()

    assert first.iloc[0].tolist() == ["A", "1"]
    assert second.iloc[0].tolist() == ["B", "2"]
    assert len(backends) == 1
    assert isinstance(backends[0], Path)
    assert events == ["session-close", "source-close"]
    assert not spill_path.exists()
    assert stream.closed is False


@pytest.mark.parametrize(
    "config",
    [
        _eligible_config(evaluate_formulas=False),
        _eligible_config(merge_strategy=MergeStrategy.FILL),
        _eligible_config(merge_strategy=MergeStrategy.FIRST_ONLY),
        _eligible_config(include_hidden=False),
        _eligible_config(cell_range="A1:B2"),
    ],
    ids=["formula-expression", "merge-fill", "merge-first-only", "hidden", "range"],
)
def test_task8_coordinate_and_formula_features_stay_on_openpyxl(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    config: SheetConfig,
) -> None:
    real_load_workbook = openpyxl.load_workbook
    calls = 0

    def recording_load(*args: object, **kwargs: object) -> Any:
        nonlocal calls
        calls += 1
        return real_load_workbook(*args, **kwargs)

    def forbidden_reader(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("advanced compatibility features must not reach fastexcel")

    monkeypatch.setattr(openpyxl, "load_workbook", recording_load)
    monkeypatch.setattr(FastexcelMaterializedReader, "read_table", forbidden_reader)

    with MessyWorkbook(sample_xlsx, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    assert isinstance(frame, pd.DataFrame)
    assert calls == 1


def test_streaming_plan_is_not_eligible_for_materialized_fastexcel() -> None:
    plan = compile_parse_plan(
        _eligible_config(),
        structure=None,
        format_type="xlsx",
        output_mode=OutputMode.STREAMING,
        batch_size=10,
    )

    assert _is_fastexcel_materialized_plan(plan) is False


@pytest.mark.parametrize(
    "error",
    [
        fastexcel.UnsupportedColumnTypeCombinationError("unsupported types"),
        fastexcel.CannotRetrieveCellDataError("cell data"),
        fastexcel.CalamineCellError("cell representation"),
    ],
)
def test_narrow_compatibility_allowlist_retries_and_counts_exactly_once(
    error: Exception,
) -> None:
    metrics = ParseMetrics()

    class FailingReader:
        def read_table(self) -> pa.Table:
            raise error

    class FallbackReader:
        def read_table(self) -> pa.Table:
            return pa.table({"value": [2]})

    table = FallbackCoordinator(
        is_fastexcel_compatibility_error,
        metrics=metrics,
    ).materialize(FailingReader, FallbackReader)

    assert table.to_pydict() == {"value": [2]}
    assert metrics == ParseMetrics(full_materializations=1, failed_attempts=1)


@pytest.mark.parametrize(
    "error",
    [
        fastexcel.ArrowError("generic arrow"),
        fastexcel.CalamineError("malformed OOXML"),
        fastexcel.FastExcelError("generic fastexcel"),
        fastexcel.InvalidParametersError("invalid configuration"),
        fastexcel.SheetNotFoundError("missing sheet"),
        fastexcel.ColumnNotFoundError("missing column"),
        ValueError("configuration"),
        PermissionError("denied"),
        FileNotFoundError("missing"),
        MemoryError("memory"),
        KeyboardInterrupt(),
        SystemExit(),
    ],
    ids=lambda error: type(error).__name__,
)
def test_non_compatibility_failures_never_retry(error: BaseException) -> None:
    metrics = ParseMetrics()
    fallback_calls = 0

    class FailingReader:
        def read_table(self) -> pa.Table:
            raise error

    def fallback_factory() -> object:
        nonlocal fallback_calls
        fallback_calls += 1
        return pytest.fail("non-compatibility failure must not retry")

    with pytest.raises(type(error)):
        FallbackCoordinator(
            is_fastexcel_compatibility_error,
            metrics=metrics,
        ).materialize(FailingReader, fallback_factory)

    assert fallback_calls == 0
    assert metrics == ParseMetrics(failed_attempts=1)


@pytest.mark.parametrize("source_kind", ["seekable", "nonseekable"])
@pytest.mark.parametrize("read_fails", [False, True], ids=["success", "failure"])
def test_large_spill_survives_session_close_until_source_owner_closes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    basic_parse_plan: ParsePlan,
    source_kind: str,
    read_fails: bool,
) -> None:
    content = _padded_workbook_bytes(tmp_path)
    stream: Any
    if source_kind == "seekable":
        stream = io.BytesIO(content)
    else:
        stream = _NonSeekableBytes(content)
    source = SourceHandle(stream, filename="large.xlsx")
    real_read_excel = fastexcel.read_excel
    backends: list[object] = []

    class BackendFake:
        sheet_names = ("Data",)

        def load_sheet(self, *_args: object, **_kwargs: object) -> pa.RecordBatch:
            if read_fails:
                raise fastexcel.CalamineCellError("unsupported cell")
            return pa.record_batch({"value": [1]})

    def recording_read_excel(backend: object) -> object:
        backends.append(backend)
        if read_fails:
            return BackendFake()
        return real_read_excel(backend)

    monkeypatch.setattr(fastexcel, "read_excel", recording_read_excel)
    session = FastexcelSession(source)
    try:
        reader = FastexcelMaterializedReader(session, "Data", basic_parse_plan)
        if read_fails:
            with pytest.raises(fastexcel.CalamineCellError):
                reader.read_table()
        else:
            assert reader.read_table().num_rows == 2
    finally:
        session.close()

    assert len(backends) == 1
    assert isinstance(backends[0], Path)
    spill_path = backends[0]
    assert spill_path.exists()
    with source.open_path_or_bytes() as replay:
        assert replay == spill_path
        assert spill_path.exists()

    source.close()

    assert not spill_path.exists()
    assert stream.closed is False


@pytest.mark.parametrize("source_kind", ["seekable", "nonseekable"])
def test_large_workbook_initialization_reuses_owned_spill_path(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    source_kind: str,
) -> None:
    content = _padded_workbook_bytes(tmp_path)
    stream: Any
    entry_position: int | None
    if source_kind == "seekable":
        stream = io.BytesIO(content)
        stream.seek(17)
        entry_position = stream.tell()
    else:
        stream = _NonSeekableBytes(content)
        entry_position = None
    real_read_excel = fastexcel.read_excel
    backends: list[object] = []

    def recording_read_excel(backend: object) -> object:
        backends.append(backend)
        return real_read_excel(backend)

    monkeypatch.setattr(fastexcel, "read_excel", recording_read_excel)
    workbook = MessyWorkbook(stream, sheet_config=_eligible_config(), filename="large.xlsx")
    try:
        assert len(backends) == 2
        assert all(isinstance(backend, Path) for backend in backends)
        spill_path = backends[0]
        assert isinstance(spill_path, Path)
        assert backends == [spill_path, spill_path]
        assert workbook._source_handle._spool is not None
        assert workbook._source_handle._spool._path == spill_path
        assert spill_path.exists()
        if entry_position is not None:
            assert stream.tell() == entry_position
        assert stream.closed is False
    finally:
        workbook.close()

    assert not spill_path.exists()
    assert stream.closed is False


@pytest.mark.parametrize("parse_fails", [False, True], ids=["success", "failure"])
def test_large_locally_owned_handler_source_removes_spill(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    parse_fails: bool,
) -> None:
    stream = io.BytesIO(_padded_workbook_bytes(tmp_path))
    real_read_excel = fastexcel.read_excel
    backends: list[object] = []

    def recording_read_excel(backend: object) -> object:
        backends.append(backend)
        return real_read_excel(backend)

    monkeypatch.setattr(fastexcel, "read_excel", recording_read_excel)
    if parse_fails:
        monkeypatch.setattr(
            FastexcelSession,
            "materialize",
            lambda _self, _sheet, *, skip_rows: (_ for _ in ()).throw(
                ValueError(f"parse failed at skip_rows={skip_rows}")
            ),
        )

    if parse_fails:
        with pytest.raises(Exception, match="parse failed"):
            XLSXHandler().parse(
                stream,
                "Data",
                compile_parse_plan(_eligible_config(), None, "xlsx").to_parse_options(),
            )
    else:
        frame = XLSXHandler().parse(
            stream,
            "Data",
            compile_parse_plan(_eligible_config(), None, "xlsx").to_parse_options(),
        )
        assert frame.shape == (1, 2)

    path_backends = [backend for backend in backends if isinstance(backend, Path)]
    assert path_backends
    assert all(not path.exists() for path in path_backends)
    assert stream.closed is False


@pytest.mark.parametrize("parse_fails", [False, True], ids=["success", "failure"])
def test_large_workbook_owner_removes_spill(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    parse_fails: bool,
) -> None:
    stream = io.BytesIO(_padded_workbook_bytes(tmp_path))
    config = _eligible_config()

    with MessyWorkbook(stream, sheet_config=config, filename="large.xlsx") as workbook:
        if parse_fails:
            monkeypatch.setattr(
                FastexcelMaterializedReader,
                "read_table",
                lambda _self: (_ for _ in ()).throw(ValueError("parse failed")),
            )
            with pytest.raises(FormatError) as captured:
                workbook._to_dataframe_compat("Data")
            assert captured.value.message == "All handlers failed for large.xlsx"
            assert captured.value.context == {
                "file_path": "large.xlsx",
                "detected_format": "xlsx",
                "attempted_formats": ["XLSXHandler"],
            }
            assert captured.value.__cause__ is None
            assert captured.value.__context__ is None
        else:
            assert workbook._to_dataframe_compat("Data").shape == (1, 2)
        spool = workbook._source_handle._spool
        assert spool is not None
        spill_path = spool._path
        assert spill_path is not None and spill_path.exists()

    assert not spill_path.exists()
    assert stream.closed is False


def test_transform_order_and_all_thaw_helpers_are_preserved(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "ordering.xlsx"
    _save_rows(path, [["Amount", "Status"], [" 1 ", "keep"], [" 2 ", "drop"]])
    config = _eligible_config(
        normalize=True,
        sanitize_column_names=True,
        column_renames={"amount": "renamed_amount"},
        type_hints={"Amount": "string"},
        drop_conditions=[{"column": "status", "value": "drop"}],
    )
    events: list[str] = []

    adapter_original = LegacyDataFrameAdapter.to_dataframe
    apply_original = XLSXHandler._apply_options
    clean_original = XLSXHandler._clean_excel_data
    type_original = ParsePlan.thaw_type_hints
    rename_original = ParsePlan.thaw_column_renames
    drop_original = ParsePlan.thaw_drop_conditions

    def adapter(self: LegacyDataFrameAdapter, table: object, plan: ParsePlan) -> pd.DataFrame:
        events.append("arrow-to-pandas")
        return adapter_original(self, table, plan)

    def apply(self: XLSXHandler, frame: pd.DataFrame, options: object) -> pd.DataFrame:
        events.append("handler-framing")
        return apply_original(self, frame, options)  # type: ignore[arg-type]

    def clean(self: XLSXHandler, frame: pd.DataFrame, options: object) -> pd.DataFrame:
        events.append("handler-cleaning")
        return clean_original(self, frame, options)  # type: ignore[arg-type]

    def thaw_type(self: ParsePlan) -> dict[Any, Any]:
        events.append("thaw-types")
        return type_original(self)

    def thaw_rename(self: ParsePlan) -> dict[Any, Any]:
        events.append("thaw-renames")
        return rename_original(self)

    def thaw_drop(self: ParsePlan) -> list[tuple[Any, Any]]:
        events.append("thaw-drops")
        return drop_original(self)

    monkeypatch.setattr(LegacyDataFrameAdapter, "to_dataframe", adapter)
    monkeypatch.setattr(XLSXHandler, "_apply_options", apply)
    monkeypatch.setattr(XLSXHandler, "_clean_excel_data", clean)
    monkeypatch.setattr(ParsePlan, "thaw_type_hints", thaw_type)
    monkeypatch.setattr(ParsePlan, "thaw_column_renames", thaw_rename)
    monkeypatch.setattr(ParsePlan, "thaw_drop_conditions", thaw_drop)

    with MessyWorkbook(path, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    assert list(frame.columns) == ["renamed_amount", "status"]
    assert len(frame) == 1
    assert events == [
        "arrow-to-pandas",
        "handler-framing",
        "handler-cleaning",
        "thaw-types",
        "thaw-renames",
        "thaw-drops",
    ]


def test_nested_plan_values_thaw_to_fresh_legacy_containers() -> None:
    nested_hint = {"semantic": ["currency", {"locale": "en_US"}]}
    nested_drop = {"flags": ["x", {"active": True}]}
    plan = compile_parse_plan(
        _eligible_config(
            type_hints={"Amount": nested_hint},
            column_renames={"Amount": "renamed_amount"},
            drop_conditions=[{"column": "Status", "value": nested_drop}],
        ),
        structure=None,
        format_type="xlsx",
    )

    hints = plan.thaw_type_hints()
    renames = plan.thaw_column_renames()
    conditions = plan.thaw_drop_conditions()

    assert hints == {"Amount": nested_hint}
    assert hints["Amount"] is not nested_hint
    assert hints["Amount"]["semantic"] is not nested_hint["semantic"]
    assert renames == {"Amount": "renamed_amount"}
    assert conditions == [("Status", nested_drop)]
    assert conditions[0][1] is not nested_drop
    assert conditions[0][1]["flags"] is not nested_drop["flags"]


def test_normalize_false_still_bypasses_filters_on_fast_path(tmp_path: Path) -> None:
    path = tmp_path / "no-normalize.xlsx"
    _save_rows(path, [["Name", "Status"], ["A", "drop"], ["B", "keep"]])
    config = _eligible_config(
        normalize=False,
        drop_regex="drop",
        drop_conditions=[{"column": "Status", "value": "drop"}],
    )

    with MessyWorkbook(path, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    assert frame.to_dict(orient="list") == {
        "Name": ["A", "B"],
        "Status": ["drop", "keep"],
    }


def test_skip_rows_equal_to_sheet_height_keeps_legacy_zero_by_zero_frame(
    tmp_path: Path,
) -> None:
    path = tmp_path / "skip-height.xlsx"
    _save_rows(path, [["Name", "Value"], ["A", 1]])
    options = ParseOptions(skip_rows=2, merge_strategy=MergeStrategy.SKIP)

    frame = XLSXHandler().parse(path, "Data", options)

    assert frame.shape == (0, 0)
    assert list(frame.columns) == []


def test_nonzero_skip_preserves_backend_type_inference_and_avoids_bound_reader(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "skip-type-inference.xlsx"
    _save_rows(path, [["metadata"], [1], [2]])
    options = ParseOptions(
        skip_rows=1,
        header_rows=0,
        merge_strategy=MergeStrategy.SKIP,
    )

    direct = XLSXHandler().parse(path, "Data", options)

    monkeypatch.setattr(
        FastexcelMaterializedReader,
        "read_table",
        lambda _self: (_ for _ in ()).throw(
            AssertionError("nonzero skip must preserve backend-pushed legacy inference")
        ),
    )
    config = _eligible_config(skip_rows=1, header_rows=0)
    with MessyWorkbook(path, sheet_config=config) as workbook:
        public = workbook._to_dataframe_compat("Data")

    expected = pd.DataFrame({"col_0": pd.Series([1.0, 2.0], dtype="float64")})
    assert_frame_equal(direct, expected)
    assert_frame_equal(public, expected)


@pytest.mark.parametrize("skip_rows", [-1, 3])
def test_skip_rows_outside_backend_bounds_keeps_legacy_invalid_parameter_error(
    tmp_path: Path,
    skip_rows: int,
) -> None:
    path = tmp_path / f"skip-invalid-{skip_rows}.xlsx"
    _save_rows(path, [["Name", "Value"], ["A", 1]])
    options = ParseOptions(skip_rows=skip_rows, merge_strategy=MergeStrategy.SKIP)

    with pytest.raises(FormatError) as captured:
        XLSXHandler().parse(path, "Data", options)

    assert isinstance(captured.value.__cause__, fastexcel.InvalidParametersError)


def test_direct_handler_session_open_failure_keeps_legacy_error_boundary() -> None:
    source = io.BytesIO(b"not an OOXML archive")
    options = ParseOptions(merge_strategy=MergeStrategy.SKIP)

    with pytest.raises(FormatError) as captured:
        XLSXHandler().parse(source, "Data", options)

    assert captured.value.message.startswith("Cannot open Excel file:")
    assert isinstance(captured.value.__cause__, fastexcel.CalamineError)
    assert source.closed is False


def test_duplicate_sanitized_labels_remain_positional(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-sanitized.xlsx"
    _save_rows(
        path,
        [
            ["A B", "A-B", "A_B"],
            ["drop", "keep", "keep"],
            ["keep", "drop", "keep"],
        ],
    )
    config = _eligible_config(
        normalize=True,
        sanitize_column_names=True,
        normalize_dates=False,
        normalize_numbers=False,
        normalize_whitespace=False,
        ensure_type_consistency=False,
        drop_conditions=[{"column": "a_b", "value": "drop"}],
    )

    with MessyWorkbook(path, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    assert list(frame.columns) == ["a_b", "a_b_1", "a_b_2"]
    assert frame.to_dict(orient="records") == [{"a_b": "keep", "a_b_1": "drop", "a_b_2": "keep"}]


def test_all_duplicate_unsanitized_condition_labels_keep_legacy_cell_masking(
    tmp_path: Path,
) -> None:
    path = tmp_path / "all-duplicate-unsanitized.xlsx"
    _save_rows(
        path,
        [
            ["Status", "Status"],
            ["drop", "keep"],
            ["keep", "drop"],
            ["keep", "keep"],
        ],
    )
    config = _eligible_config(
        normalize=True,
        normalize_dates=False,
        normalize_numbers=False,
        normalize_whitespace=False,
        ensure_type_consistency=False,
        drop_conditions=[{"column": "Status", "value": "drop"}],
    )

    with MessyWorkbook(path, sheet_config=config) as workbook:
        frame = workbook._to_dataframe_compat("Data")

    expected = pd.DataFrame(
        [[float("nan"), "keep"], ["keep", float("nan")], ["keep", "keep"]],
        columns=["Status", "Status"],
    )
    assert_frame_equal(frame, expected)


def test_mixed_duplicate_unsanitized_condition_labels_keep_legacy_error(
    tmp_path: Path,
) -> None:
    path = tmp_path / "mixed-duplicate-unsanitized.xlsx"
    _save_rows(
        path,
        [
            ["Status", "Status", "Other"],
            ["drop", "keep", "x"],
            ["keep", "drop", "y"],
        ],
    )
    config = _eligible_config(
        normalize=True,
        normalize_dates=False,
        normalize_numbers=False,
        normalize_whitespace=False,
        ensure_type_consistency=False,
        drop_conditions=[{"column": "Status", "value": "drop"}],
    )

    with (
        MessyWorkbook(path, sheet_config=config) as workbook,
        pytest.raises(ValueError, match="duplicate labels"),
    ):
        workbook._to_dataframe_compat("Data")
