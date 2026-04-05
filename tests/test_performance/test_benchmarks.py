"""Benchmark tests using pytest-benchmark for performance regression tracking.

Run with:
    pytest tests/test_performance/test_benchmarks.py -v --benchmark-only
"""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig

# ============================================================================
# Fixtures
# ============================================================================


@pytest.fixture
def xlsx_1000_rows(temp_dir):
    """Create a 1000-row XLSX file for benchmarking."""
    file_path = temp_dir / "bench_1000.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name", "amount", "date", "status"])
    for i in range(1000):
        ws.append([i, f"Item {i}", round(i * 3.14, 2), "2024-01-15", "active"])
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def xlsx_with_merged(temp_dir):
    """Create an XLSX file with merged cells for benchmarking."""
    file_path = temp_dir / "bench_merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Category", "Item", "Value"])
    for i in range(100):
        ws.append([f"Cat {i // 10}", f"Item {i}", i * 2.5])
    # Merge some category cells
    for start in range(2, 102, 10):
        ws.merge_cells(start_row=start, start_column=1, end_row=start + 9, end_column=1)
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture
def csv_1000_rows(temp_dir):
    """Create a 1000-row CSV file for benchmarking."""
    file_path = temp_dir / "bench_1000.csv"
    lines = ["id,name,amount,date,status"]
    for i in range(1000):
        lines.append(f"{i},Item {i},{round(i * 3.14, 2)},2024-01-15,active")
    file_path.write_text("\n".join(lines))
    return file_path


# ============================================================================
# Parse benchmarks
# ============================================================================


@pytest.mark.slow
class TestParseBenchmarks:
    """Benchmark core parsing operations."""

    def test_parse_1000_rows(self, benchmark, xlsx_1000_rows):
        """Benchmark: parse 1000-row XLSX file."""

        def parse():
            with MessyWorkbook(xlsx_1000_rows) as wb:
                return wb.to_dataframe()

        df = benchmark(parse)
        assert len(df) == 1000

    def test_parse_1000_rows_no_normalization(self, benchmark, xlsx_1000_rows):
        """Benchmark: parse 1000-row XLSX without normalization."""
        config = SheetConfig(normalize=False)

        def parse():
            with MessyWorkbook(xlsx_1000_rows, sheet_config=config) as wb:
                return wb.to_dataframe()

        df = benchmark(parse)
        assert len(df) == 1000

    def test_parse_1000_rows_no_auto_detect(self, benchmark, xlsx_1000_rows):
        """Benchmark: parse 1000-row XLSX without structure detection."""
        config = SheetConfig(auto_detect=False)

        def parse():
            with MessyWorkbook(xlsx_1000_rows, sheet_config=config) as wb:
                return wb.to_dataframe()

        df = benchmark(parse)
        assert len(df) == 1000

    def test_parse_csv_1000_rows(self, benchmark, csv_1000_rows):
        """Benchmark: parse 1000-row CSV file."""

        def parse():
            with MessyWorkbook(csv_1000_rows) as wb:
                return wb.to_dataframe()

        df = benchmark(parse)
        assert len(df) == 1000

    def test_parse_with_merged_cells(self, benchmark, xlsx_with_merged):
        """Benchmark: parse XLSX with merged cells (fill strategy)."""
        config = SheetConfig(merge_strategy="fill")

        def parse():
            with MessyWorkbook(xlsx_with_merged, sheet_config=config) as wb:
                return wb.to_dataframe()

        df = benchmark(parse)
        assert len(df) == 100


# ============================================================================
# Structure detection benchmarks
# ============================================================================


@pytest.mark.slow
class TestStructureDetectionBenchmarks:
    """Benchmark structure analysis operations."""

    def test_structure_detection_1000_rows(self, benchmark, xlsx_1000_rows):
        """Benchmark: structure detection on 1000-row file."""

        def detect():
            with MessyWorkbook(xlsx_1000_rows) as wb:
                return wb.get_structure()

        structure = benchmark(detect)
        assert structure.header_row is not None

    def test_sheet_names_lookup(self, benchmark, xlsx_1000_rows):
        """Benchmark: reading sheet names from XLSX."""

        def get_names():
            with MessyWorkbook(xlsx_1000_rows) as wb:
                return wb.sheet_names

        names = benchmark(get_names)
        assert len(names) >= 1
