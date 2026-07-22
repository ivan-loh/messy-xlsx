"""Tests for bounded worksheet structure evidence and fastexcel reuse."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from itertools import pairwise
from pathlib import Path
from zipfile import ZIP_STORED, ZipFile

import openpyxl
import pandas as pd
import pytest

from messy_xlsx._source import SourceHandle
from messy_xlsx.detection.structure_analyzer import StructureAnalyzer
from messy_xlsx.detection.structure_sampler import (
    SampleWindow,
    StructureEvidence,
    StructureSampler,
    structure_sample_windows,
)
from messy_xlsx.ooxml.manifest import ManifestReader
from messy_xlsx.ooxml.models import IntervalIndex, SheetManifest
from messy_xlsx.parsing.fastexcel_session import FastexcelSession


def _sheet_manifest(*, max_row: int = 2, max_col: int = 2) -> SheetManifest:
    return SheetManifest(
        name="Data",
        target="xl/worksheets/sheet1.xml",
        declared_dimension=(1, 1, max_row, max_col),
        observed_max_row=max_row,
        observed_max_col=max_col,
        hidden_rows=IntervalIndex(()),
        hidden_columns=IntervalIndex(()),
        merged_ranges=(),
        has_formulas=False,
        formula_samples=(),
    )


class RecordingManifestReader:
    def __init__(self, manifest: SheetManifest) -> None:
        self.manifest = manifest
        self.requests: list[str] = []

    def sheet(self, name: str) -> SheetManifest:
        self.requests.append(name)
        return self.manifest


class RecordingExcelReader:
    def __init__(self) -> None:
        self.open_count = 1
        self.requests: list[dict[str, object]] = []
        self.complete_dataframe_count = 0

    def sample_windows(
        self,
        sheet: str,
        windows: tuple[SampleWindow, ...],
        min_column: int,
        max_column: int,
    ) -> StructureEvidence:
        self.requests.append(
            {
                "sheet": sheet,
                "windows": windows,
                "min_column": min_column,
                "max_column": max_column,
            }
        )
        return StructureEvidence(
            row_numbers=(1, 2),
            values=pd.DataFrame([["name", "value"], ["a", 1]], index=(1, 2)),
        )


@dataclass
class RecordingMetrics:
    sample_reads: int = 0


def test_structure_sample_windows_are_sorted_coalesced_and_bounded() -> None:
    windows = structure_sample_windows(1_000_000)
    retained = sum(window.n_rows for window in windows)

    assert windows[0] == SampleWindow(1, 10_000)
    assert retained <= 10_500
    assert all(window.start_row >= 1 and window.n_rows >= 1 for window in windows)
    assert all(
        current.start_row + current.n_rows < following.start_row
        for current, following in pairwise(windows)
    )
    assert windows[-1].start_row + windows[-1].n_rows - 1 == 1_000_000


def test_offset_structure_sample_windows_remain_bounded() -> None:
    windows = structure_sample_windows(1_000_000, start_row=12_000)

    assert windows[0].start_row == 12_000
    assert sum(window.n_rows for window in windows) <= 10_500


def test_sampler_reuses_reader_and_caches_by_sheet_and_pattern_tuple() -> None:
    excel_reader = RecordingExcelReader()
    manifest_reader = RecordingManifestReader(_sheet_manifest())
    metrics = RecordingMetrics()
    sampler = StructureSampler(excel_reader, manifest_reader, metrics=metrics)

    first = sampler.analyze("Data")
    second = sampler.analyze("Data")
    patterned = sampler.analyze("Data", header_patterns=["name"])

    assert first is second
    assert patterned.header_confidence >= first.header_confidence
    assert excel_reader.open_count == 1
    assert len(excel_reader.requests) == 2
    assert manifest_reader.requests == ["Data", "Data"]
    assert metrics.sample_reads == 2
    assert excel_reader.complete_dataframe_count == 0
    request = excel_reader.requests[0]
    assert request["sheet"] == "Data"
    assert sum(window.n_rows for window in request["windows"]) <= 10_500


class RecordingSource:
    def __init__(self) -> None:
        self.enter_count = 0
        self.exit_count = 0

    def open_path_or_bytes(self):
        source = self

        class BackendContext:
            def __enter__(self):
                source.enter_count += 1
                return Path("book.xlsx")

            def __exit__(self, exc_type, exc_value, traceback):
                source.exit_count += 1

        return BackendContext()


class FakeBatch:
    def __init__(self, values: list[list[object]]) -> None:
        self._values = values

    def to_pandas(self) -> pd.DataFrame:
        return pd.DataFrame(self._values)


class FakeExcelReader:
    def __init__(self) -> None:
        self.sheet_names = ["Data"]
        self.requests: list[dict[str, object]] = []

    def load_sheet(self, sheet: str, **kwargs):
        self.requests.append({"sheet": sheet, **kwargs})
        row_count = int(kwargs.get("n_rows", 2))
        return FakeBatch([[row] for row in range(row_count)])


def test_fastexcel_session_uses_integer_windows_and_one_backend_context(monkeypatch) -> None:
    source = RecordingSource()
    backend = FakeExcelReader()
    monkeypatch.setattr(
        "messy_xlsx.parsing.fastexcel_session.fastexcel.read_excel",
        lambda value: backend,
    )

    session = FastexcelSession(source)
    assert source.exit_count == 1
    evidence = session.sample_windows(
        "Data",
        (SampleWindow(3, 2), SampleWindow(10, 1)),
        min_column=1,
        max_column=1,
    )
    session.close()
    session.close()

    assert source.enter_count == 1
    assert source.exit_count == 1
    assert session.sheet_names == ("Data",)
    assert evidence.row_numbers == (3, 4, 10)
    assert [request["skip_rows"] for request in backend.requests] == [2, 9]
    assert [request["n_rows"] for request in backend.requests] == [2, 1]
    assert all(type(request["skip_rows"]) is int for request in backend.requests)
    assert all(type(request["n_rows"]) is int for request in backend.requests)
    assert all(
        all(type(column) is int for column in request["use_columns"])
        for request in backend.requests
    )


def test_fastexcel_session_closes_source_when_backend_initialization_fails(monkeypatch) -> None:
    source = RecordingSource()

    def fail(_value):
        raise RuntimeError("backend failed")

    monkeypatch.setattr(
        "messy_xlsx.parsing.fastexcel_session.fastexcel.read_excel",
        fail,
    )

    with pytest.raises(RuntimeError, match="backend failed"):
        FastexcelSession(source)

    assert source.enter_count == 1
    assert source.exit_count == 1


def test_fastexcel_session_closes_source_when_sheet_metadata_fails(monkeypatch) -> None:
    source = RecordingSource()

    class BrokenReader:
        @property
        def sheet_names(self):
            raise RuntimeError("metadata failed")

    monkeypatch.setattr(
        "messy_xlsx.parsing.fastexcel_session.fastexcel.read_excel",
        lambda _value: BrokenReader(),
    )

    with pytest.raises(RuntimeError, match="metadata failed"):
        FastexcelSession(source)

    assert source.enter_count == 1
    assert source.exit_count == 1


class NonSeekableBytes:
    def __init__(self, content: bytes) -> None:
        self._stream = io.BytesIO(content)

    def read(self, size: int = -1) -> bytes:
        return self._stream.read(size)

    def seekable(self) -> bool:
        return False


def _workbook_bytes(tmp_path: Path, *, disk_spool: bool) -> bytes:
    path = tmp_path / "composition.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["name", "amount"])
    first.append(["alpha", 1])
    second = workbook.create_sheet("Second")
    second.append(["label", "value"])
    second.append(["beta", 2])
    workbook.save(path)
    workbook.close()
    if disk_spool:
        with ZipFile(path, "a", ZIP_STORED) as package:
            package.writestr("unreferenced.bin", b"x" * (9 * 1024 * 1024))
    return path.read_bytes()


@pytest.mark.parametrize(
    "source_kind",
    ["path", "seekable_buffer", "memory_spool", "disk_spool"],
)
def test_fresh_components_analyze_multiple_sheets_without_nested_borrows(
    tmp_path: Path,
    source_kind: str,
) -> None:
    content = _workbook_bytes(tmp_path, disk_spool=source_kind == "disk_spool")
    if source_kind == "path":
        raw_source = tmp_path / "composition.xlsx"
    elif source_kind == "seekable_buffer":
        raw_source = io.BytesIO(content)
        raw_source.seek(7)
    else:
        raw_source = NonSeekableBytes(content)

    with SourceHandle(raw_source) as source:
        opened: list[str] = []
        reader = ManifestReader(source, on_member_open=opened.append)
        assert not any(member.startswith("xl/worksheets/") for member in opened)
        with FastexcelSession(source) as session:
            sampler = StructureSampler(session, reader)
            first = sampler.analyze("First")
            second = sampler.analyze("Second")

        assert first.data_start_row == 1
        assert second.data_start_row == 1
        assert sum(member.startswith("xl/worksheets/") for member in opened) == 2
        with pytest.raises(RuntimeError, match="closed"):
            session.sample_windows(
                "First",
                (SampleWindow(1, 1),),
                min_column=1,
                max_column=1,
            )

    if isinstance(raw_source, io.BytesIO):
        assert raw_source.tell() == 7
        assert raw_source.closed is False


@pytest.mark.parametrize(
    "relative_path",
    [
        "tests/samples/budget_vs_actuals.xlsx",
        (
            "tests/generated_messy/parseable/"
            "messy__preset_financial_statements_summary_sheet__seed_1004__"
            "metadata_preamble_merged_headers_irrelevant_summary_sheet_hidden_rows.xlsx"
        ),
        (
            "tests/generated_messy/parseable/"
            "messy__preset_expense_reports_date_unicode_hidden__seed_1006__"
            "date_noise_unicode_whitespace_noise_hidden_cols_blank_row_noise.xlsx"
        ),
        (
            "tests/generated_messy/parseable/"
            "messy__preset_inventory_multitable_ragged_hidden__seed_1010__"
            "multi_table_sheet_ragged_rows_hidden_cols_blank_row_noise.xlsx"
        ),
        (
            "tests/generated_messy/parseable/"
            "messy__preset_cash_flow_offset_formula_multitable__seed_1008__"
            "offset_table_formula_noise_multi_table_sheet_footer_noise.xlsx"
        ),
    ],
)
def test_sampler_matches_existing_structure_characterization(
    relative_path: str,
) -> None:
    path = Path(relative_path)
    with SourceHandle(path) as source:
        manifest_reader = ManifestReader(source)
        sheet = manifest_reader.workbook.sheets[0].name
        expected = StructureAnalyzer().analyze(path, sheet, force=True)
        with FastexcelSession(source) as session:
            actual = StructureSampler(session, manifest_reader).analyze(sheet)

    assert actual == expected


def _differential_results(path: Path) -> tuple[object, object, SheetManifest]:
    expected = StructureAnalyzer().analyze(path, "Data", force=True)
    with SourceHandle(path) as source:
        manifest_reader = ManifestReader(source)
        manifest = manifest_reader.sheet("Data")
        with FastexcelSession(source) as session:
            actual = StructureSampler(session, manifest_reader).analyze("Data")
    return expected, actual, manifest


def test_mixed_cell_types_preserve_numeric_text_boolean_and_date_provenance(
    tmp_path: Path,
) -> None:
    path = tmp_path / "mixed-types.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["identifier"])
    for value in ("001", "2024", "1e3", "+12.50", "-7.25", True, date(2024, 1, 2), 42):
        worksheet.append([value])
    workbook.save(path)
    workbook.close()

    expected, actual, _manifest = _differential_results(path)

    assert actual == expected


def test_uncached_formulas_beyond_diagnostic_cap_do_not_change_structure(
    tmp_path: Path,
) -> None:
    path = tmp_path / "many-formulas.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in range(1, 301):
        worksheet.cell(row=row, column=1, value=f"={row}+1")
    worksheet.cell(row=301, column=2, value="name")
    worksheet.cell(row=301, column=3, value="amount")
    worksheet.cell(row=302, column=2, value="alpha")
    worksheet.cell(row=302, column=3, value=1)
    workbook.save(path)
    workbook.close()

    expected, actual, manifest = _differential_results(path)

    assert len(manifest.formula_samples) == 256
    assert manifest.has_formulas is True
    assert actual == expected
    assert actual.data_start_row == 301


@pytest.mark.parametrize(
    ("formula_row", "formula_column", "expected_has_formulas"),
    [(50, 10, True), (51, 10, False), (50, 11, False)],
)
def test_legacy_formula_flag_uses_only_first_50_data_rows(
    tmp_path: Path,
    formula_row: int,
    formula_column: int,
    expected_has_formulas: bool,
) -> None:
    path = tmp_path / f"formula-{formula_row}-{formula_column}.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in range(1, 61):
        worksheet.cell(row=row, column=1, value=row)
        worksheet.cell(row=row, column=10, value=row)
    worksheet.cell(row=formula_row, column=formula_column, value="=1+1")
    workbook.save(path)
    workbook.close()

    expected, actual, manifest = _differential_results(path)

    assert manifest.has_formulas is True
    assert expected.has_formulas is expected_has_formulas
    assert actual.has_formulas is expected_has_formulas


def test_distant_number_format_does_not_affect_locale_probe(tmp_path: Path) -> None:
    path = tmp_path / "locale-scope.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = "amount"
    worksheet["A2"] = 1.25
    worksheet["A100"] = 2.5
    worksheet["A100"].number_format = "#.##0,00"
    workbook.save(path)
    workbook.close()

    expected, actual, _manifest = _differential_results(path)

    assert expected.detected_locale == "en_US"
    assert actual == expected


def test_offset_large_region_does_not_treat_unrequested_rows_as_blank(tmp_path: Path) -> None:
    path = tmp_path / "offset-large.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in range(5, 20_001):
        worksheet.cell(row=row, column=1, value=row)
    workbook.save(path)
    workbook.close()

    expected, actual, manifest = _differential_results(path)

    assert expected.blank_rows == []
    assert actual.blank_rows == []
    assert actual == expected
    assert len({cell.row for cell in manifest.cell_evidence}) <= 10_500


def test_data_region_can_start_after_legacy_head_window(tmp_path: Path) -> None:
    path = tmp_path / "deep-offset.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.cell(row=12_000, column=3, value="name")
    worksheet.cell(row=12_001, column=3, value="alpha")
    worksheet.cell(row=12_001, column=4, value=1)
    workbook.save(path)
    workbook.close()

    _expected, actual, _manifest = _differential_results(path)

    assert actual.data_start_row == 12_000
    assert actual.data_start_col == 3
