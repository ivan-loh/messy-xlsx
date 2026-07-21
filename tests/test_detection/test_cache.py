"""Tests for LRUCache and StructureCache."""

import time
from pathlib import Path

import openpyxl

from messy_xlsx.cache import LRUCache, StructureCache
from messy_xlsx.models import StructureInfo

# ============================================================================
# LRUCache Tests
# ============================================================================


class TestLRUCache:
    """Test generic LRU cache."""

    def test_basic_get_put(self):
        cache: LRUCache[str] = LRUCache(maxsize=10)
        cache.put("a", "value_a")
        assert cache.get("a") == "value_a"

    def test_get_missing_key_returns_none(self):
        cache: LRUCache[str] = LRUCache(maxsize=10)
        assert cache.get("nonexistent") is None

    def test_eviction_at_maxsize(self):
        cache: LRUCache[int] = LRUCache(maxsize=3)
        cache.put("a", 1)
        cache.put("b", 2)
        cache.put("c", 3)
        # Adding a 4th should evict "a" (least recently used)
        cache.put("d", 4)
        assert cache.get("a") is None
        assert cache.get("b") == 2
        assert cache.get("c") == 3
        assert cache.get("d") == 4

    def test_access_order_updates_on_get(self):
        cache: LRUCache[int] = LRUCache(maxsize=3)
        cache.put("a", 1)
        cache.put("b", 2)
        cache.put("c", 3)
        # Access "a" to move it to end (most recently used)
        cache.get("a")
        # Now "b" is least recently used
        cache.put("d", 4)
        assert cache.get("b") is None  # evicted
        assert cache.get("a") == 1  # still present

    def test_access_order_updates_on_put(self):
        cache: LRUCache[int] = LRUCache(maxsize=3)
        cache.put("a", 1)
        cache.put("b", 2)
        cache.put("c", 3)
        # Update "a" to move it to end
        cache.put("a", 10)
        cache.put("d", 4)
        assert cache.get("b") is None  # evicted
        assert cache.get("a") == 10  # still present, updated value

    def test_invalidate_existing_key(self):
        cache: LRUCache[str] = LRUCache(maxsize=10)
        cache.put("a", "value")
        assert cache.invalidate("a") is True
        assert cache.get("a") is None

    def test_invalidate_missing_key(self):
        cache: LRUCache[str] = LRUCache(maxsize=10)
        assert cache.invalidate("nonexistent") is False

    def test_invalidate_by_prefix(self):
        cache: LRUCache[int] = LRUCache(maxsize=10)
        cache.put("file1:sheet1", 1)
        cache.put("file1:sheet2", 2)
        cache.put("file2:sheet1", 3)
        removed = cache.invalidate_prefix("file1:")
        assert removed == 2
        assert cache.get("file1:sheet1") is None
        assert cache.get("file1:sheet2") is None
        assert cache.get("file2:sheet1") == 3

    def test_invalidate_prefix_no_match(self):
        cache: LRUCache[int] = LRUCache(maxsize=10)
        cache.put("a", 1)
        removed = cache.invalidate_prefix("b")
        assert removed == 0
        assert cache.get("a") == 1

    def test_len(self):
        cache: LRUCache[int] = LRUCache(maxsize=10)
        assert len(cache) == 0
        cache.put("a", 1)
        assert len(cache) == 1
        cache.put("b", 2)
        assert len(cache) == 2
        cache.invalidate("a")
        assert len(cache) == 1

    def test_contains(self):
        cache: LRUCache[str] = LRUCache(maxsize=10)
        cache.put("a", "value")
        assert "a" in cache
        assert "b" not in cache

    def test_clear(self):
        cache: LRUCache[int] = LRUCache(maxsize=10)
        cache.put("a", 1)
        cache.put("b", 2)
        cache.clear()
        assert len(cache) == 0
        assert cache.get("a") is None


# ============================================================================
# StructureCache Tests
# ============================================================================


def _make_structure_info(**kwargs):
    """Create a minimal StructureInfo for testing."""
    defaults = {
        "data_start_row": 1,
        "data_end_row": 10,
        "data_start_col": 1,
        "data_end_col": 5,
        "header_row": 1,
        "header_rows_count": 1,
        "header_confidence": 0.9,
    }
    defaults.update(kwargs)
    return StructureInfo(**defaults)


class TestStructureCache:
    """Test StructureCache with mtime-based invalidation."""

    def test_put_and_get(self, temp_dir):
        """Basic put/get cycle with a real file."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        # Create a real XLSX file
        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        info = _make_structure_info()
        cache.put(file_path, "Sheet1", info)
        result = cache.get(file_path, "Sheet1")
        assert result is not None
        assert result.data_end_row == 10

    def test_get_returns_none_for_missing(self, temp_dir):
        """Get should return None if key is not cached."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        assert cache.get(file_path, "Sheet1") is None

    def test_mtime_invalidation(self, temp_dir):
        """Cache should miss after file is modified."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        info = _make_structure_info()
        cache.put(file_path, "Sheet1", info)

        # Modify the file (change mtime)
        time.sleep(0.05)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["new data"])
        wb.save(file_path)
        wb.close()

        # Should miss because mtime changed
        assert cache.get(file_path, "Sheet1") is None

    def test_different_sheets_cached_separately(self, temp_dir):
        """Different sheets of same file get independent cache entries."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        info1 = _make_structure_info(data_end_row=10)
        info2 = _make_structure_info(data_end_row=20)
        cache.put(file_path, "Sheet1", info1)
        cache.put(file_path, "Sheet2", info2)

        r1 = cache.get(file_path, "Sheet1")
        r2 = cache.get(file_path, "Sheet2")
        assert r1 is not None and r1.data_end_row == 10
        assert r2 is not None and r2.data_end_row == 20

    def test_analysis_variants_cached_separately(self, temp_dir):
        """Configuration-dependent analysis must not reuse another result."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        plain = _make_structure_info(header_confidence=0.8)
        patterned = _make_structure_info(header_confidence=0.95)
        cache.put(file_path, "Sheet1", plain, variant="")
        cache.put(file_path, "Sheet1", patterned, variant="first field")

        assert cache.get(file_path, "Sheet1", variant="") == plain
        assert cache.get(file_path, "Sheet1", variant="first field") == patterned

    def test_invalidate_file(self, temp_dir):
        """Invalidate should remove all entries for a file."""
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        cache.put(file_path, "Sheet1", _make_structure_info())
        cache.put(file_path, "Sheet2", _make_structure_info())
        removed = cache.invalidate(file_path)
        assert removed == 2
        assert cache.get(file_path, "Sheet1") is None
        assert cache.get(file_path, "Sheet2") is None

    def test_get_nonexistent_file_returns_none(self):
        """Get on a file that doesn't exist should return None."""
        cache = StructureCache(maxsize=10)
        result = cache.get(Path("/nonexistent/file.xlsx"), "Sheet1")
        assert result is None

    def test_put_nonexistent_file_is_noop(self):
        """Put on a file that doesn't exist should silently do nothing."""
        cache = StructureCache(maxsize=10)
        cache.put(Path("/nonexistent/file.xlsx"), "Sheet1", _make_structure_info())
        assert len(cache) == 0

    def test_len(self, temp_dir):
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        assert len(cache) == 0
        cache.put(file_path, "Sheet1", _make_structure_info())
        assert len(cache) == 1

    def test_clear(self, temp_dir):
        cache = StructureCache(maxsize=10)
        file_path = temp_dir / "test.xlsx"

        wb = openpyxl.Workbook()
        wb.save(file_path)
        wb.close()

        cache.put(file_path, "Sheet1", _make_structure_info())
        cache.clear()
        assert len(cache) == 0
