"""Unit tests for StructureAnalyzer."""

import os

import openpyxl

from messy_xlsx import MessyWorkbook
from messy_xlsx.cache import StructureCache
from messy_xlsx.detection import StructureAnalyzer
from messy_xlsx.models import StructureInfo


class TestStructureAnalyzer:
    """Test structure analysis functionality."""

    def test_analyze_simple_structure(self, sample_xlsx):
        """Test analyzing simple file structure."""
        with MessyWorkbook(sample_xlsx) as wb:
            structure = wb.get_structure("Data")

            assert structure is not None
            assert structure.header_row is not None
            assert structure.num_tables >= 1
            assert structure.detected_locale in ["en_US", "de_DE", "unknown"]

    def test_header_detection(self, sample_xlsx):
        """Test header row detection."""
        with MessyWorkbook(sample_xlsx) as wb:
            structure = wb.get_structure("Data")

            assert structure.header_row == 1
            assert structure.header_confidence >= 0.5

    def test_detect_messy_structure(self, messy_xlsx):
        """Test detecting messy file structure."""
        with MessyWorkbook(messy_xlsx) as wb:
            structure = wb.get_structure("Report")

            # Should detect metadata rows before data
            assert structure.header_row > 1

    def test_detect_data_region(self, sample_xlsx):
        """Test data region detection."""
        with MessyWorkbook(sample_xlsx) as wb:
            structure = wb.get_structure("Data")

            assert structure.data_start_row >= 1
            assert structure.data_end_row >= structure.data_start_row
            assert structure.data_start_col >= 1
            assert structure.data_end_col >= structure.data_start_col

    def test_detect_formulas(self, messy_xlsx):
        """Test formula detection on a workbook known to contain formulas."""
        with MessyWorkbook(messy_xlsx) as wb:
            structure = wb.get_structure("Report")

            # messy_xlsx fixture has formulas (=B5+C5, =SUM(...), etc.)
            assert structure.has_formulas is True

    def test_uncached_formula_outside_data_does_not_expand_structure(self, tmp_path):
        """Formula expressions must not replace cached-value structure evidence."""
        file_path = tmp_path / "formula-only-tail.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A100"] = "=1+1"
        workbook.save(file_path)
        workbook.close()

        structure = StructureAnalyzer().analyze(file_path, "Data", force=True)

        assert structure.data_start_row == 1
        assert structure.data_end_row == 1
        assert structure.data_start_col == 1
        assert structure.data_end_col == 1
        assert structure.has_formulas is False

    def test_merged_cell_detection(self, merged_cells_xlsx):
        """Test merged cell detection."""
        with MessyWorkbook(merged_cells_xlsx) as wb:
            structure = wb.get_structure("Data")

            assert len(structure.merged_ranges) > 0

    def test_multi_table_detection(self, multi_table_xlsx):
        """Test multiple table detection."""
        with MessyWorkbook(multi_table_xlsx) as wb:
            structure = wb.get_structure("Data")

            # Should detect 2 tables separated by blank rows
            assert structure.num_tables == 2, f"Expected 2 tables, detected {structure.num_tables}"

    def test_locale_detection(self, european_xlsx):
        """Test locale detection."""
        with MessyWorkbook(european_xlsx) as wb:
            structure = wb.get_structure("Data")

            # Should detect decimal/thousands separators
            assert structure.decimal_separator in [".", ","]
            assert structure.thousands_separator in [".", ",", " ", ""]

    def test_structure_caching(self, sample_xlsx):
        """Test that structure analysis results are cached."""
        StructureCache()

        with MessyWorkbook(sample_xlsx) as wb:
            # First call
            structure1 = wb.get_structure("Data")

            # Second call should use cache
            structure2 = wb.get_structure("Data")

            assert structure1.header_row == structure2.header_row
            assert structure1.num_tables == structure2.num_tables

    def test_file_changed_during_analysis_is_not_cached(self, sample_xlsx, monkeypatch):
        cache = StructureCache()
        cache.put(
            sample_xlsx,
            "Sentinel",
            StructureInfo(1, 1, 1, 1, None, 0, 0.0),
        )
        analyzer = StructureAnalyzer(cache)
        original = analyzer._detect_data_region

        def change_path_during_analysis(worksheet):
            result = original(worksheet)
            stat = sample_xlsx.stat()
            os.utime(
                sample_xlsx,
                ns=(stat.st_atime_ns, stat.st_mtime_ns + 1_000_000_000),
            )
            return result

        monkeypatch.setattr(analyzer, "_detect_data_region", change_path_during_analysis)

        analyzer.analyze(sample_xlsx, "Data", force=True)

        assert len(cache) == 1
