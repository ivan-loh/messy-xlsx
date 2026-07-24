"""Return-gap and traceback-retention contracts for public parser streams."""

from __future__ import annotations

import gc
import sys
import weakref
from pathlib import Path
from types import CodeType
from typing import Any

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig
from messy_xlsx.parsing.streams import (
    BatchStream,
    DataFrameChunkStream,
    SheetStream,
)


@pytest.fixture
def lifecycle_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "stream-lifecycle.xlsx"
    physical = openpyxl.Workbook()
    first = physical.active
    first.title = "First"
    first.append(["value"])
    first.append([1])
    second = physical.create_sheet("Second")
    second.append(["value"])
    second.append([2])
    physical.save(path)
    physical.close()
    return path


def _interrupt_return(
    target_code: CodeType,
    *,
    message: str,
    capture: list[weakref.ReferenceType[Any]] | None = None,
) -> None:
    interrupted = False

    def interrupt(frame: Any, event: str, arg: object) -> Any:
        nonlocal interrupted
        if frame.f_code is target_code and event == "return" and not interrupted:
            interrupted = True
            if capture is not None:
                capture.append(weakref.ref(arg))
            sys.settrace(None)
            frame.f_trace = None
            del frame, arg
            raise MemoryError(message)
        return interrupt

    sys.settrace(interrupt)


def _open_stream(
    workbook: MessyWorkbook,
    api_name: str,
) -> BatchStream | DataFrameChunkStream | SheetStream:
    config = SheetConfig(auto_detect=False)
    if api_name == "iter_batches":
        return workbook.iter_batches("First", batch_size=1, config=config)
    if api_name == "iter_dataframe_chunks":
        return workbook.iter_dataframe_chunks("First", batch_size=1, config=config)
    assert api_name == "iter_sheets"
    return workbook.iter_sheets(config)


def test_stream_operation_return_interruption_is_recovered_by_next_operation(
    lifecycle_xlsx: Path,
) -> None:
    workbook = MessyWorkbook(lifecycle_xlsx)
    target_code = workbook._stream_operation.__func__.__code__
    _interrupt_return(
        target_code,
        message="stream lease return interrupted",
    )
    try:
        with pytest.raises(MemoryError, match="stream lease return interrupted"):
            workbook._stream_operation()
    finally:
        sys.settrace(None)

    try:
        assert workbook.to_arrow("First", SheetConfig(auto_detect=False)).to_pydict() == {
            "value": [1]
        }
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "api_name",
    ["iter_batches", "iter_dataframe_chunks", "iter_sheets"],
)
def test_public_stream_return_interruption_releases_unreachable_child(
    lifecycle_xlsx: Path,
    api_name: str,
) -> None:
    workbook = MessyWorkbook(lifecycle_xlsx)
    target_code = getattr(workbook, api_name).__func__.__code__
    returned: list[weakref.ReferenceType[Any]] = []
    _interrupt_return(
        target_code,
        message=f"{api_name} return interrupted",
        capture=returned,
    )
    try:
        with pytest.raises(MemoryError, match=rf"{api_name} return interrupted"):
            _open_stream(workbook, api_name)
    finally:
        sys.settrace(None)

    gc.collect()
    try:
        assert len(returned) == 1
        assert returned[0]() is None
        assert workbook._active_operation_token is None
        assert workbook._active_stream is None
        assert workbook.to_arrow("First", SheetConfig(auto_detect=False)).to_pydict() == {
            "value": [1]
        }
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "api_name",
    ["iter_batches", "iter_dataframe_chunks", "iter_sheets"],
)
def test_every_public_stream_supports_early_close_and_parent_invalidation(
    lifecycle_xlsx: Path,
    api_name: str,
) -> None:
    workbook = MessyWorkbook(lifecycle_xlsx)
    stream = _open_stream(workbook, api_name)
    stream.close()
    stream.close()
    with pytest.raises(StopIteration):
        next(stream)
    assert workbook.to_arrow("First", SheetConfig(auto_detect=False)).to_pydict() == {"value": [1]}

    reachable = _open_stream(workbook, api_name)
    workbook.close()
    with pytest.raises(RuntimeError, match="MessyWorkbook is closed"):
        next(reachable)
    reachable.close()


def test_iter_sheets_releases_projected_planning_failure_traceback_before_yield(
    lifecycle_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = MessyWorkbook(lifecycle_xlsx)
    original = workbook._get_sheet_manifest

    class TracebackSentinel:
        pass

    holder = [TracebackSentinel()]
    sentinel_ref = weakref.ref(holder[0])

    def fail_first(name: str) -> Any:
        if name == "First":
            traceback_sentinel = holder[0]
            raise ValueError(f"planning failure with {type(traceback_sentinel).__name__}")
        return original(name)

    monkeypatch.setattr(workbook, "_get_sheet_manifest", fail_first)
    stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
    holder.clear()

    try:
        first = next(stream)
        assert first.error is not None
        assert first.error.error_type == "ValueError"
        gc.collect()
        assert sentinel_ref() is None

        second = next(stream)
        assert second.name == "Second"
        assert second.error is None
    finally:
        stream.close()
        workbook.close()
