"""Source-boundary contracts for the S07 source-handle refactor.

The first group records supported public behavior today.  The second group is
deliberately red until S07 supplies one source policy: caller-owned seekable
streams should have their position restored, and non-seekable streams should
be snapshotted without taking ownership.
"""

from __future__ import annotations

import gc
import io
import sys
from pathlib import Path, PureWindowsPath
from typing import Any, ClassVar, Literal

import openpyxl
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal

import messy_xlsx._spool as spool_module
import messy_xlsx.workbook as workbook_module
from messy_xlsx import FormatInfo, MessyWorkbook, SheetConfig, read_excel
from messy_xlsx._fallback_signals import (
    _fallback_block_reason,
    _FallbackBlockReason,
)
from messy_xlsx._source import SourceHandle
from messy_xlsx._spool import DEFAULT_MEMORY_LIMIT, ReplaySpool
from messy_xlsx.detection import FormatDetector, StructureAnalyzer
from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.parsing import CSVHandler, HandlerRegistry, ParseOptions
from messy_xlsx.parsing.base_handler import FormatHandler

SourceFormat = Literal["xlsx", "xls", "csv"]
ByteLikeKind = Literal["bytearray", "memoryview"]


class NamedBytesIO(io.BytesIO):
    """A cloud-upload-like seekable buffer carrying a source name."""

    def __init__(self, content: bytes, name: str):
        super().__init__(content)
        self.name = name


class CountingBytesIO(io.BytesIO):
    """Seekable buffer that records the sizes requested by consumers."""

    def __init__(self, content: bytes):
        super().__init__(content)
        self.read_sizes: list[int] = []

    def read(self, size: int = -1) -> bytes:
        self.read_sizes.append(size)
        return super().read(size)


class ByteLikeBytesIO(io.BytesIO):
    """Seekable binary source returning a supported non-``bytes`` byte view."""

    def __init__(self, content: bytes, name: str, kind: ByteLikeKind):
        super().__init__(content)
        self.name = name
        self.kind = kind

    def read(self, size: int = -1) -> bytearray | memoryview:
        content = super().read(size)
        if self.kind == "bytearray":
            return bytearray(content)
        return memoryview(content)


class CanonicalZeroReadMemoryViewBytesIO(io.BytesIO):
    """Return ``bytes`` only for zero-length reads and views otherwise."""

    def __init__(self, content: bytes, name: str):
        super().__init__(content)
        self.name = name

    def read(self, size: int = -1) -> bytes | memoryview:
        content = super().read(size)
        if size == 0:
            return content
        return memoryview(content)


class RestoreFailureBytesIO(io.BytesIO):
    """Seekable source that can fail only when restoring its entry cursor."""

    def __init__(self, content: bytes):
        super().__init__(content)
        self.fail_restoration = False
        self.restoration_position = 0

    def seek(self, position: int, whence: int = 0) -> int:
        if self.fail_restoration and whence == 0 and position == self.restoration_position:
            raise OSError("restore failed")
        return super().seek(position, whence)


class ProcessRestoreFailureBytesIO(io.BytesIO):
    """Raise a selected process-level failure only on final restoration."""

    def __init__(self, content: bytes, restoration_error: BaseException):
        super().__init__(content)
        self.restoration_error = restoration_error
        self.restoration_position = 0
        self.fail_restoration = False

    def seek(self, position: int, whence: int = 0) -> int:
        if self.fail_restoration and whence == 0 and position == self.restoration_position:
            raise self.restoration_error
        return super().seek(position, whence)


class _FatalRestore(BaseException):
    pass


class NthRestoreFailureBytesIO(io.BytesIO):
    """Fail on one selected restoration to an otherwise unused position."""

    def __init__(self, content: bytes):
        super().__init__(content)
        self.enabled = False
        self.restoration_position = 0
        self.restoration_calls = 0
        self.fail_on_restoration = 0

    def seek(self, position: int, whence: int = 0) -> int:
        if self.enabled and whence == 0 and position == self.restoration_position:
            self.restoration_calls += 1
            if self.restoration_calls == self.fail_on_restoration:
                raise OSError("formula-view restore failed")
        return super().seek(position, whence)


class OneShotSpoolRestoreFailureBytesIO(io.BytesIO):
    """Fail exactly once when spool acquisition restores the caller cursor."""

    def __init__(self, content: bytes, restoration_error: BaseException):
        super().__init__(content)
        self.restoration_error = restoration_error
        self.restoration_position = 0
        self.restoration_calls = 0
        self.enabled = False

    def seek(self, position: int, whence: int = 0) -> int:
        if self.enabled and whence == 0 and position == self.restoration_position:
            self.restoration_calls += 1
            if self.restoration_calls == 2:
                error = self.restoration_error
                raise type(error)(*error.args)
        return super().seek(position, whence)


class NonSeekableStream:
    """Minimal caller-owned read-once byte stream."""

    def __init__(self, content: bytes, name: str):
        self._stream = io.BytesIO(content)
        self.name = name
        self.closed = False
        self.read_calls = 0

    def read(self, size: int = -1) -> bytes:
        self.read_calls += 1
        return self._stream.read(size)

    def close(self) -> None:
        self.closed = True


class FailingNonSeekableStream:
    """Read-once source that fails during eager acquisition."""

    def __init__(self, error: Exception, name: str):
        self.error = error
        self.name = name
        self.closed = False

    def read(self, size: int = -1) -> bytes:
        raise self.error

    def close(self) -> None:
        self.closed = True


class TextNonSeekableStream:
    """Invalid read-once text source used to protect the binary boundary."""

    def __init__(self, name: str):
        self.name = name
        self.closed = False

    def read(self, size: int = -1) -> str:
        return "Name,Value\nA,1\n"

    def close(self) -> None:
        self.closed = True


class ByteLikeNonSeekableStream(NonSeekableStream):
    """Read-once binary source returning bytearray or memoryview."""

    def __init__(self, content: bytes, name: str, kind: ByteLikeKind):
        super().__init__(content, name)
        self.kind = kind

    def read(self, size: int = -1) -> bytearray | memoryview:
        self.read_calls += 1
        content = self._stream.read(size)
        if self.kind == "bytearray":
            return bytearray(content)
        return memoryview(content)


def _xlsx_content() -> bytes:
    target = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["Alpha", 1])
    sheet.append(["Beta", 2])
    workbook.save(target)
    workbook.close()
    return target.getvalue()


def _xls_content() -> bytes:
    xlwt = pytest.importorskip("xlwt", reason="legacy XLS fixture requires xlwt")
    target = io.BytesIO()
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Data")
    sheet.write(0, 0, "Name")
    sheet.write(0, 1, "Value")
    sheet.write(1, 0, "Alpha")
    sheet.write(1, 1, 1)
    sheet.write(2, 0, "Beta")
    sheet.write(2, 1, 2)
    workbook.save(target)
    return target.getvalue()


def _source_content(source_format: SourceFormat) -> bytes:
    if source_format == "xlsx":
        return _xlsx_content()
    if source_format == "xls":
        return _xls_content()
    return b"Name,Value\nAlpha,1\nBeta,2\n"


def _filename(source_format: SourceFormat) -> str:
    return f"source.{source_format}"


def _read_workbook(source: object, filename: str | None = None) -> pd.DataFrame:
    with MessyWorkbook(source, filename=filename) as workbook:  # type: ignore[arg-type]
        return workbook.to_dataframe(sheet="Data" if workbook.format_type != "csv" else None)


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
def test_path_and_seekable_buffer_outputs_match(
    tmp_path: Path,
    source_format: SourceFormat,
) -> None:
    """Paths and BytesIO are equivalent public inputs for supported formats."""
    content = _source_content(source_format)
    path = tmp_path / _filename(source_format)
    path.write_bytes(content)

    from_path = _read_workbook(path)
    source = io.BytesIO(content)
    from_buffer = _read_workbook(source, filename=path.name)

    assert_frame_equal(from_buffer, from_path)
    assert source.closed is False


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
def test_named_seekable_buffer_parses_by_content_and_remains_caller_owned(
    source_format: SourceFormat,
) -> None:
    """A benign ``name`` attribute must not interfere with content detection."""
    content = _source_content(source_format)
    source = NamedBytesIO(content, _filename(source_format))

    result = _read_workbook(source)
    expected = _read_workbook(io.BytesIO(content), filename=_filename(source_format))

    assert_frame_equal(result, expected)
    assert source.closed is False


def test_explicit_filename_hint_supplies_stream_identity() -> None:
    """An explicit filename hint remains visible in the workbook representation."""
    source = NamedBytesIO(_xlsx_content(), "ignored-upload-name.bin")

    with MessyWorkbook(source, filename="orders.xlsx") as workbook:
        assert workbook.format_type == "xlsx"
        assert "orders.xlsx" in repr(workbook)
        assert workbook.source is source
        assert workbook.file_path is None

    assert source.closed is False


@pytest.mark.parametrize(
    "stream_name",
    [Path("uploads/orders.xlsx"), PureWindowsPath("uploads/orders.xlsx")],
    ids=["native", "windows"],
)
def test_path_valued_stream_name_supplies_identity(
    stream_name: Path | PureWindowsPath,
) -> None:
    source = NamedBytesIO(_xlsx_content(), "temporary.bin")
    source.name = stream_name

    with MessyWorkbook(source) as workbook:
        assert workbook.format_type == "xlsx"
        assert repr(str(source.name)) in repr(workbook)

    assert source.closed is False


def test_invalid_pathlike_stream_name_is_ignored() -> None:
    class InvalidPathLike:
        def __fspath__(self) -> str:
            raise OSError("metadata unavailable")

    source = NamedBytesIO(_xlsx_content(), "temporary.bin")
    source.name = InvalidPathLike()  # type: ignore[assignment]

    with MessyWorkbook(source) as workbook:
        assert workbook.format_type == "xlsx"
        assert "<stream>" in repr(workbook)

    assert source.closed is False


def test_csv_validation_uses_a_bounded_probe_and_restores_position() -> None:
    source = CountingBytesIO(b"Name,Value\n" + b"A,1\n" * 100_000)
    source.seek(3)

    is_valid, error = CSVHandler().validate(source)

    assert (is_valid, error) == (True, None)
    assert source.read_sizes == [1024]
    assert source.tell() == 3
    assert source.closed is False


def test_seekable_buffer_supports_repeated_full_passes_when_caller_rewinds() -> None:
    """Current workaround: rewinding permits repeated detect/validate/analyze/parse passes."""
    source = io.BytesIO(_xlsx_content())
    frames: list[pd.DataFrame] = []
    structures = []

    for _ in range(2):
        source.seek(0)
        with MessyWorkbook(source, filename="repeated.xlsx") as workbook:
            structures.append(workbook.get_structure("Data"))
            frames.append(workbook.to_dataframe("Data"))

    assert structures[0] == structures[1]
    assert_frame_equal(frames[0], frames[1])
    assert source.closed is False


def test_parse_failure_keeps_caller_buffer_open_and_reusable() -> None:
    """Failure never transfers ownership of a caller-provided buffer."""
    content = _xlsx_content()
    source = io.BytesIO(content)
    config = SheetConfig(auto_detect=False, cell_range="not:a:range")

    with (
        pytest.raises(FormatError),
        MessyWorkbook(source, filename="broken-range.xlsx", sheet_config=config) as workbook,
    ):
        workbook.to_dataframe("Data")

    assert source.closed is False
    source.seek(0)
    assert source.read(4) == content[:4]


# ---------------------------------------------------------------------------
# Intended S07 policy (red until SourceHandle owns rewinding/snapshotting)
# ---------------------------------------------------------------------------


def test_nonzero_initial_position_does_not_change_xlsx_detection() -> None:
    """S07 policy: detection reads from source start, independent of caller position."""
    content = _xlsx_content()
    source = io.BytesIO(content)
    source.seek(3)

    actual = _read_workbook(source)
    expected = _read_workbook(io.BytesIO(content))

    assert_frame_equal(actual, expected)


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
def test_success_restores_caller_buffer_position(source_format: SourceFormat) -> None:
    """S07 policy: successful use restores a seekable caller stream's entry position."""
    source = io.BytesIO(_source_content(source_format))
    entry_position = 3
    source.seek(entry_position)

    _read_workbook(source, filename=_filename(source_format))

    assert source.closed is False
    assert source.tell() == entry_position


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
@pytest.mark.parametrize("kind", ["bytearray", "memoryview"])
@pytest.mark.parametrize("seekable", [True, False])
def test_supported_byte_like_streams_match_bytes_across_formats(
    source_format: SourceFormat,
    kind: ByteLikeKind,
    seekable: bool,
) -> None:
    content = _source_content(source_format)
    filename = _filename(source_format)
    expected = _read_workbook(io.BytesIO(content), filename=filename)

    if seekable:
        source = ByteLikeBytesIO(content, filename, kind)
        source.seek(3)
    else:
        source = ByteLikeNonSeekableStream(content, filename, kind)

    actual = _read_workbook(source, filename=filename)

    assert_frame_equal(actual, expected)
    assert source.closed is False
    if seekable:
        assert source.tell() == 3
    else:
        # One bounded acquisition consists of the data read and EOF probe;
        # all parser passes replay the spool without touching the caller again.
        assert source.read_calls == 2


def test_memoryview_xlsx_stream_supports_the_no_analysis_fast_path() -> None:
    content = _xlsx_content()
    source = ByteLikeBytesIO(content, "source.xlsx", "memoryview")
    config = SheetConfig(auto_detect=False)

    with MessyWorkbook(source, sheet_config=config) as workbook:
        actual = workbook.to_dataframe("Data")

    expected = _read_workbook(io.BytesIO(content), filename="source.xlsx")
    assert_frame_equal(actual, expected)
    assert source.closed is False


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
def test_backend_probe_uses_a_real_read_for_byte_like_streams(
    source_format: SourceFormat,
) -> None:
    content = _source_content(source_format)
    filename = _filename(source_format)
    source = CanonicalZeroReadMemoryViewBytesIO(content, filename)
    source.seek(3)

    actual = _read_workbook(source, filename=filename)
    expected = _read_workbook(io.BytesIO(content), filename=filename)

    assert_frame_equal(actual, expected)
    assert source.tell() == 3
    assert source.closed is False


def test_failure_restores_caller_buffer_position() -> None:
    """S07 policy: failure restores a seekable caller stream's entry position."""
    source = io.BytesIO(_xlsx_content())
    entry_position = 3
    source.seek(entry_position)
    config = SheetConfig(auto_detect=False, cell_range="not:a:range")

    with (
        pytest.raises(FormatError),
        MessyWorkbook(source, filename="broken-range.xlsx", sheet_config=config) as workbook,
    ):
        workbook.to_dataframe("Data")

    assert source.closed is False
    assert source.tell() == entry_position


def test_repeated_public_reads_do_not_require_caller_rewind() -> None:
    """S07 policy: one caller-owned buffer can be passed repeatedly as-is."""
    source = io.BytesIO(_xlsx_content())

    first = read_excel(source)  # type: ignore[arg-type]
    second = read_excel(source)  # type: ignore[arg-type]

    assert_frame_equal(second, first)
    assert source.closed is False


@pytest.mark.parametrize("source_format", ["xlsx", "xls", "csv"])
def test_non_seekable_stream_is_snapshotted_once_without_taking_ownership(
    source_format: SourceFormat,
) -> None:
    """S07 policy: adapters consume an internal snapshot of read-once streams."""
    content = _source_content(source_format)
    source = NonSeekableStream(content, _filename(source_format))
    expected = _read_workbook(io.BytesIO(content), filename=_filename(source_format))

    actual = _read_workbook(source, filename=_filename(source_format))

    assert_frame_equal(actual, expected)
    assert source.closed is False
    assert source.read_calls == 2


def test_named_buffer_identity_is_used_in_source_errors() -> None:
    """S07 policy: a buffer's name is the fallback identity when no hint is passed."""
    source = NamedBytesIO(b"", "empty-upload.xlsx")

    with pytest.raises(FormatError) as captured:
        MessyWorkbook(source)

    assert captured.value.context["file_path"] == "empty-upload.xlsx"
    assert source.closed is False


def test_text_stream_is_rejected_as_a_binary_source_without_position_leak() -> None:
    source = io.StringIO("Name,Value\nA,1\n")
    source.seek(2)

    with pytest.raises(FormatError, match="Binary source read"):
        MessyWorkbook(source, filename="source.csv")  # type: ignore[arg-type]

    assert source.closed is False
    assert source.tell() == 2


def test_nonseekable_acquisition_error_uses_the_format_error_boundary() -> None:
    source = FailingNonSeekableStream(OSError("read broke"), "failed-upload.xlsx")

    with pytest.raises(FormatError, match="read broke") as captured:
        MessyWorkbook(source)  # type: ignore[arg-type]

    assert captured.value.context["file_path"] == "failed-upload.xlsx"
    assert isinstance(captured.value.__cause__, OSError)
    assert source.closed is False


def test_nonseekable_text_read_uses_explicit_error_identity() -> None:
    source = TextNonSeekableStream("ignored-name.csv")

    with pytest.raises(FormatError, match="Binary source read") as captured:
        MessyWorkbook(source, filename="explicit-name.csv")  # type: ignore[arg-type]

    assert captured.value.context["file_path"] == "explicit-name.csv"
    assert isinstance(captured.value.__cause__, TypeError)
    assert source.closed is False


def test_workbook_initialization_failure_closes_only_its_source_handle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    instances: list[TrackingSourceHandle] = []

    class TrackingSourceHandle(SourceHandle):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.close_calls = 0
            instances.append(self)

        def close(self) -> None:
            self.close_calls += 1
            super().close()

    monkeypatch.setattr(workbook_module, "SourceHandle", TrackingSourceHandle)
    source = NamedBytesIO(b"", "empty-upload.xlsx")

    with pytest.raises(FormatError):
        MessyWorkbook(source)

    assert len(instances) == 1
    assert instances[0].closed is True
    assert instances[0].close_calls == 1
    assert source.closed is False


def test_source_handle_restores_seekable_position_when_consumer_raises() -> None:
    source = io.BytesIO(b"complete source")
    source.seek(5)
    handle = SourceHandle(source, filename="source.bin")

    with (
        pytest.raises(RuntimeError, match="consumer failed"),
        handle.open_binary() as borrowed,
    ):
        assert borrowed is source
        assert borrowed.tell() == 0
        assert borrowed.read(8) == b"complete"
        raise RuntimeError("consumer failed")

    assert source.tell() == 5
    assert source.closed is False


def test_source_handle_rejects_nested_active_borrows() -> None:
    source = io.BytesIO(b"complete source")
    source.seek(5)
    handle = SourceHandle(source)

    with handle.open_binary() as borrowed:
        assert borrowed is source
        with (
            pytest.raises(
                RuntimeError,
                match="SourceHandle already has an active borrow",
            ) as captured,
            handle.open_path_or_bytes(),
        ):
            pass

    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP
    assert source.tell() == 5
    assert source.closed is False


@pytest.mark.parametrize("stage", ["entry", "exit"])
def test_source_borrow_trace_interruption_does_not_poison_later_borrow(
    stage: str,
) -> None:
    source = io.BytesIO(b"payload")
    source.seek(3)
    entry = source.tell()
    handle = SourceHandle(source, filename="data.xlsx")
    target_code = SourceHandle._borrow.__wrapped__.__code__
    entered_body = False
    interrupted = False

    def interrupt_borrow_gap(frame: object, event: str, _arg: object) -> object:
        nonlocal interrupted
        if (
            getattr(frame, "f_code", None) is target_code
            and event == "line"
            and handle._active_borrow
            and not interrupted
            and ((stage == "entry" and not entered_body) or (stage == "exit" and entered_body))
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None  # type: ignore[attr-defined]
            raise MemoryError(f"{stage} borrow transition interrupted")
        return interrupt_borrow_gap

    sys.settrace(interrupt_borrow_gap)
    try:
        with (
            pytest.raises(MemoryError, match=rf"{stage} borrow transition interrupted"),
            handle.open_binary() as borrowed,
        ):
            entered_body = True
            assert borrowed.read(1) == b"p"
    finally:
        sys.settrace(None)

    assert interrupted
    with handle.open_binary() as borrowed:
        assert borrowed.read() == b"payload"
    assert source.tell() == entry
    assert handle._active_borrow is False
    handle.close()


def test_seekable_open_path_or_bytes_spills_without_leaking_cursor() -> None:
    source = io.BytesIO(b"x" * (DEFAULT_MEMORY_LIMIT + 1))
    source.seek(7)
    handle = SourceHandle(source)

    with handle.open_path_or_bytes() as backend:
        assert isinstance(backend, Path)
        path = backend
        assert path.exists()
        assert path.stat().st_mode & 0o777 == 0o600
        assert path.stat().st_size == DEFAULT_MEMORY_LIMIT + 1

    assert source.tell() == 7
    assert source.closed is False
    handle.close()
    assert not path.exists()


@pytest.mark.parametrize(
    "restoration_error",
    [
        OSError("spool cursor restoration interrupted"),
        MemoryError("spool cursor restoration interrupted"),
    ],
    ids=["ordinary", "process"],
)
def test_open_path_or_bytes_retries_failed_spool_cursor_restore_on_close(
    restoration_error: BaseException,
) -> None:
    content = b"complete source payload"
    source = OneShotSpoolRestoreFailureBytesIO(content, restoration_error)
    entry_position = 5
    source.seek(entry_position)
    source.restoration_position = entry_position
    handle = SourceHandle(source)
    source.enabled = True

    with (
        pytest.raises(type(restoration_error), match="restoration interrupted"),
        handle.open_path_or_bytes(),
    ):
        pass

    assert source.tell() == len(content)
    assert handle._pending_restore_position == entry_position
    assert handle._spool is None
    assert source.closed is False

    handle.close()

    assert source.tell() == entry_position
    assert source.restoration_calls == 3
    assert source.closed is False
    assert handle.closed is True


def test_process_spool_cleanup_precedence_preserves_cursor_restore_for_close_retry(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    content = b"complete source payload"
    source = OneShotSpoolRestoreFailureBytesIO(
        content,
        OSError("spool cursor restoration interrupted"),
    )
    entry_position = 7
    source.seek(entry_position)
    source.restoration_position = entry_position
    handle = SourceHandle(source)
    source.enabled = True
    close_calls = 0
    real_close = ReplaySpool.close

    def interrupt_cleanup_once(spool: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        if close_calls == 1:
            raise MemoryError("spool cleanup interrupted")
        real_close(spool)

    monkeypatch.setattr(ReplaySpool, "close", interrupt_cleanup_once)

    with (
        pytest.raises(MemoryError, match="spool cleanup interrupted"),
        handle.open_path_or_bytes(),
    ):
        pass

    assert source.tell() == len(content)
    assert handle._pending_restore_position == entry_position
    assert handle._spool is not None

    handle.close()

    assert source.tell() == entry_position
    assert source.restoration_calls == 3
    assert close_calls == 2
    assert handle._spool is None
    assert source.closed is False


def test_failed_spool_cursor_restore_is_owned_by_source_handle_finalizer() -> None:
    def abandon_failed_handle() -> OneShotSpoolRestoreFailureBytesIO:
        content = b"complete source payload"
        source = OneShotSpoolRestoreFailureBytesIO(
            content,
            OSError("spool cursor restoration interrupted"),
        )
        source.seek(9)
        source.restoration_position = 9
        handle = SourceHandle(source)
        source.enabled = True

        with (
            pytest.raises(OSError, match="restoration interrupted"),
            handle.open_path_or_bytes(),
        ):
            pass

        assert source.tell() == len(content)
        assert handle._pending_restore_position == 9
        return source

    source = abandon_failed_handle()
    gc.collect()

    assert source.tell() == 9
    assert source.restoration_calls == 3
    assert source.closed is False


def test_temporary_spool_failure_becomes_file_error_and_restores_cursor(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(b"x" * (DEFAULT_MEMORY_LIMIT + 1))
    source.seek(9)
    handle = SourceHandle(source, filename="large-upload.xlsx")

    def fail_mkstemp(*_args: object, **_kwargs: object) -> tuple[int, str]:
        raise OSError("injected spool capacity failure")

    monkeypatch.setattr(spool_module.tempfile, "mkstemp", fail_mkstemp)

    with (
        pytest.raises(FileError, match="injected spool capacity failure") as captured,
        handle.open_path_or_bytes(),
    ):
        pass

    assert captured.value.context == {
        "file_path": "large-upload.xlsx",
        "operation": "spool",
    }
    assert isinstance(captured.value.__cause__, OSError)
    assert source.tell() == 9
    assert source.closed is False


def test_consumer_failure_remains_primary_when_cursor_restoration_also_fails() -> None:
    source = RestoreFailureBytesIO(b"complete source")
    source.seek(5)
    source.restoration_position = 5
    handle = SourceHandle(source)
    source.fail_restoration = True

    primary_error = RuntimeError("consumer failed")
    with pytest.raises(RuntimeError, match="consumer failed") as captured, handle.open_binary():
        raise primary_error

    assert captured.value is primary_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "cleanup_failure": {"type": "OSError"}
    }
    assert captured.value.__notes__ == ["cursor restoration also failed: OSError"]
    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP


@pytest.mark.parametrize(
    "restore_error",
    [
        MemoryError("capacity"),
        KeyboardInterrupt(),
        SystemExit(2),
        _FatalRestore(),
        ExceptionGroup("outer", [MemoryError("nested capacity")]),
        BaseExceptionGroup("outer", [KeyboardInterrupt()]),
    ],
)
def test_process_level_cursor_restoration_failure_wins_over_consumer_error(
    restore_error: BaseException,
) -> None:
    source = ProcessRestoreFailureBytesIO(b"complete source", restore_error)
    source.seek(5)
    source.restoration_position = 5
    handle = SourceHandle(source)
    source.fail_restoration = True

    with pytest.raises(type(restore_error)) as captured, handle.open_binary():
        raise RuntimeError("consumer failed")

    assert captured.value is restore_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "operation_failure": {"type": "RuntimeError"}
    }
    assert captured.value.__notes__ == ["source operation also failed: RuntimeError"]
    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP


def test_source_handle_exit_preserves_primary_over_ordinary_close_failure(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    handle = SourceHandle(NonSeekableStream(b"source", "source.bin"))
    primary_error = ValueError("parse failed")
    cleanup_error = OSError("spool unlink failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    def fail_parse() -> None:
        raise primary_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with pytest.raises(ValueError) as captured, handle:
            fail_parse()

    assert captured.value is primary_error
    assert captured.traceback[-1].name == "fail_parse"
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "cleanup_failure": {"type": "OSError"}
    }
    assert "spool unlink failed" not in repr(captured.value.__notes__)
    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP
    handle.close()


@pytest.mark.parametrize(
    "primary_error",
    [MemoryError("capacity"), KeyboardInterrupt(), SystemExit(2), _FatalRestore()],
)
def test_source_handle_exit_preserves_process_primary_over_ordinary_close_failure(
    monkeypatch: pytest.MonkeyPatch,
    primary_error: BaseException,
) -> None:
    handle = SourceHandle(NonSeekableStream(b"source", "source.bin"))
    cleanup_error = OSError("spool unlink failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with pytest.raises(type(primary_error)) as captured, handle:
            raise primary_error

    assert captured.value is primary_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "cleanup_failure": {"type": "OSError"}
    }
    assert "spool unlink failed" not in repr(captured.value.__notes__)
    handle.close()


@pytest.mark.parametrize(
    "cleanup_error",
    [
        MemoryError("capacity"),
        KeyboardInterrupt(),
        SystemExit(2),
        _FatalRestore(),
        ExceptionGroup("outer", [MemoryError("nested capacity")]),
        BaseExceptionGroup("outer", [KeyboardInterrupt()]),
    ],
)
def test_source_handle_exit_process_cleanup_wins_over_ordinary_primary(
    monkeypatch: pytest.MonkeyPatch,
    cleanup_error: BaseException,
) -> None:
    handle = SourceHandle(NonSeekableStream(b"source", "source.bin"))
    primary_error = ValueError("parse failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with pytest.raises(type(cleanup_error)) as captured, handle:
            raise primary_error

    assert captured.value is cleanup_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "operation_failure": {"type": "ValueError"},
        "cleanup_failure": {"type": type(cleanup_error).__name__},
    }
    assert "parse failed" not in repr(captured.value.__notes__)
    handle.close()


def test_source_handle_exit_without_primary_propagates_close_failure(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    handle = SourceHandle(NonSeekableStream(b"source", "source.bin"))
    cleanup_error = OSError("spool unlink failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with pytest.raises(OSError) as captured, handle:
            pass

    assert captured.value is cleanup_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "cleanup_failure": {"type": "OSError"}
    }
    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP
    assert "spool unlink failed" not in repr(captured.value.__notes__)
    handle.close()


def test_source_handle_exit_merges_existing_sanitized_context(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    handle = SourceHandle(NonSeekableStream(b"source", "source.bin"))
    primary_error = ValueError("parse failed")
    primary_error.backend_context = {  # type: ignore[attr-defined]
        "primary_failure": {"type": "PrimaryReaderError"},
        "fallback_failure": {"type": "FallbackReaderError"},
    }
    cleanup_error = OSError("spool unlink failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with pytest.raises(ValueError) as captured, handle:
            raise primary_error

    assert captured.value is primary_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "primary_failure": {"type": "PrimaryReaderError"},
        "fallback_failure": {"type": "FallbackReaderError"},
        "cleanup_failure": {"type": "OSError"},
    }
    handle.close()


def test_registry_owned_source_uses_source_handle_exit_precedence(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry = HandlerRegistry()
    primary_error = ValueError("parse failed")
    cleanup_error = OSError("spool unlink failed")

    def fail_close(_spool: ReplaySpool) -> None:
        raise cleanup_error

    with monkeypatch.context() as patch:
        patch.setattr(ReplaySpool, "close", fail_close)
        with (
            pytest.raises(ValueError) as captured,
            registry._source_handle(NonSeekableStream(b"source", "source.bin")),
        ):
            raise primary_error

    assert captured.value is primary_error
    assert captured.value.backend_context == {  # type: ignore[attr-defined]
        "cleanup_failure": {"type": "OSError"}
    }
    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP


def test_cursor_restoration_failure_is_reported_after_successful_consumption() -> None:
    source = RestoreFailureBytesIO(b"complete source")
    source.seek(5)
    source.restoration_position = 5
    handle = SourceHandle(source)
    source.fail_restoration = True

    with (
        pytest.raises(OSError, match="restore failed") as captured,
        handle.open_binary() as borrowed,
    ):
        assert borrowed.read(4) == b"comp"

    assert _fallback_block_reason(captured.value) is _FallbackBlockReason.SOURCE_OWNERSHIP


def test_structure_analysis_preserves_consumer_error_when_restoration_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class AnalysisWorkbook:
        sheetnames: ClassVar[list[str]] = ["Data"]

        def __getitem__(self, _sheet: str) -> object:
            return object()

        def close(self) -> None:
            pass

    source = RestoreFailureBytesIO(b"content supplied to fake loader")
    source.seek(5)
    source.restoration_position = 5
    handle = SourceHandle(source)
    with handle.open_backend():
        pass
    source.fail_restoration = True

    analyzer = StructureAnalyzer()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: AnalysisWorkbook())

    def fail_analysis(_worksheet: object) -> dict[str, int]:
        raise RuntimeError("analysis failed")

    monkeypatch.setattr(analyzer, "_detect_data_region", fail_analysis)

    with pytest.raises(RuntimeError, match="analysis failed"):
        analyzer.analyze(handle, "Data")


def test_formula_view_cursor_restoration_failure_is_not_silently_ignored() -> None:
    content = _xlsx_content()
    source = NthRestoreFailureBytesIO(content)
    entry_position = len(content) + 10_000
    source.seek(entry_position)
    source.restoration_position = entry_position
    handle = SourceHandle(source)
    source.fail_on_restoration = 3
    source.enabled = True

    with pytest.raises(OSError, match="formula-view restore failed"):
        StructureAnalyzer().analyze(handle, "Data", force=True)

    assert source.restoration_calls == 3
    assert source.tell() != entry_position
    assert source.closed is False


def test_seekable_read_bytes_replays_without_a_permanent_cache() -> None:
    source = CountingBytesIO(b"repeatable")
    source.seek(4)
    handle = SourceHandle(source)

    first = handle.read_bytes()
    second = handle.read_bytes()
    detached_one = handle.detached_binary()
    detached_two = handle.detached_binary()

    assert first == b"repeatable"
    assert second == first
    assert source.read_sizes == [-1, -1, -1, -1]
    assert source.tell() == 4
    assert detached_one is not detached_two
    assert detached_one.read() == detached_two.read() == first
    detached_one.close()
    detached_two.close()


def test_nonseekable_handle_snapshots_once_and_replays_without_ownership() -> None:
    source = NonSeekableStream(b"repeatable", "source.bin")
    handle = SourceHandle(source)
    acquisition_reads = source.read_calls

    first = handle.read_bytes()
    second = handle.read_bytes()
    with handle.open_binary() as replay:
        third = replay.read()

    assert acquisition_reads == 2
    assert source.read_calls == acquisition_reads
    assert first == second == third == b"repeatable"
    assert handle.was_snapshotted is True

    handle.close()
    handle.close()
    assert source.closed is False


def test_nonseekable_spool_handoff_process_failure_closes_acquired_spool(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"repeatable", "source.bin")

    class AcquiredSpool:
        def __init__(self) -> None:
            self.close_calls = 0

        def close(self) -> None:
            self.close_calls += 1

    acquired = AcquiredSpool()

    def acquire_spool(
        _cls: type[ReplaySpool],
        _stream: object,
        _memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> Any:
        if callable(_adopter):
            _adopter(acquired)
        return acquired

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(acquire_spool))
    target_code = SourceHandle._ensure_spool.__code__
    interrupted = False

    def interrupt_after_spool_return(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        handle = frame.f_locals.get("self")
        returned_spool = frame.f_locals.get("spool")
        if (
            frame.f_code is target_code
            and event == "line"
            and (returned_spool is acquired or getattr(handle, "_spool", None) is acquired)
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("spool handoff interrupted")
        return interrupt_after_spool_return

    sys.settrace(interrupt_after_spool_return)
    try:
        with pytest.raises(MemoryError, match="spool handoff interrupted"):
            SourceHandle(source)
    finally:
        sys.settrace(None)

    assert interrupted
    assert acquired.close_calls == 1
    assert source.closed is False


def test_spool_handoff_retains_process_failed_cleanup_for_close_retry(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NamedBytesIO(b"repeatable", "source.bin")
    handle = SourceHandle(source)

    class RetryableSpool:
        def __init__(self) -> None:
            self.close_calls = 0

        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise MemoryError("spool cleanup interrupted")

    acquired = RetryableSpool()

    def acquire_spool(
        _cls: type[ReplaySpool],
        _stream: object,
        _memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> Any:
        if callable(_adopter):
            _adopter(acquired)
        return acquired

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(acquire_spool))
    target_code = SourceHandle._ensure_spool.__code__
    interrupted = False

    def interrupt_after_spool_return(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        if (
            frame.f_code is target_code
            and event == "line"
            and getattr(handle, "_spool", None) is acquired
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("spool handoff interrupted")
        return interrupt_after_spool_return

    sys.settrace(interrupt_after_spool_return)
    try:
        with pytest.raises(MemoryError, match="spool handoff interrupted"):
            handle._ensure_spool()
    finally:
        sys.settrace(None)

    assert acquired.close_calls == 1
    handle.close()
    assert acquired.close_calls == 2
    assert source.closed is False


def test_workbook_owns_prepared_source_handle_before_nonseekable_spool_acquisition(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"not-an-excel-file", "source.xlsx")
    observed: list[bool] = []
    real_from_stream = ReplaySpool.from_stream.__func__

    def acquire_spool(
        cls: type[ReplaySpool],
        stream: object,
        memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> ReplaySpool:
        frame = sys._getframe()
        handle: SourceHandle | None = None
        workbook: MessyWorkbook | None = None
        while frame is not None:
            candidate = frame.f_locals.get("self")
            if type(candidate) is SourceHandle:
                handle = candidate
            elif type(candidate) is MessyWorkbook:
                workbook = candidate
            frame = frame.f_back
        observed.append(
            handle is not None
            and workbook is not None
            and getattr(workbook, "_source_handle", None) is handle
        )
        return real_from_stream(
            cls,
            stream,
            memory_limit,
            _adopter=_adopter,
        )

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(acquire_spool))

    class FailingRegistry(HandlerRegistry):
        _accepts_source_handle = True

        def detect_format(self, *_args: object, **_kwargs: object) -> FormatInfo:
            raise RuntimeError("stop after source acquisition")

    with pytest.raises(RuntimeError, match="stop after source acquisition"):
        MessyWorkbook(source, filename=source.name, registry=FailingRegistry())

    assert observed == [True]
    assert source.closed is False


def test_direct_source_handle_constructor_remains_eager_after_prepare_start_split(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"repeatable", "source.bin")
    calls = 0
    real_from_stream = ReplaySpool.from_stream.__func__

    def acquire_spool(
        cls: type[ReplaySpool],
        stream: object,
        memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> ReplaySpool:
        nonlocal calls
        calls += 1
        return real_from_stream(
            cls,
            stream,
            memory_limit,
            _adopter=_adopter,
        )

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(acquire_spool))
    handle = SourceHandle(source)

    assert calls == 1
    assert handle.was_snapshotted is True
    handle.close()
    assert source.closed is False


def test_workbook_constructor_retries_process_interrupted_spool_cleanup(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"not-an-excel-file", "source.xlsx")
    close_calls = 0
    real_close = ReplaySpool.close

    def fail_close_once(spool: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        if close_calls == 1:
            raise MemoryError("spool cleanup interrupted")
        real_close(spool)

    monkeypatch.setattr(ReplaySpool, "close", fail_close_once)

    class FailingRegistry(HandlerRegistry):
        _accepts_source_handle = True

        def detect_format(self, *_args: object, **_kwargs: object) -> FormatInfo:
            raise RuntimeError("stop after source acquisition")

    with pytest.raises(MemoryError, match="spool cleanup interrupted"):
        MessyWorkbook(source, filename=source.name, registry=FailingRegistry())

    assert close_calls == 2
    assert source.closed is False


# Task 12 final acceptance: direct eager construction and ReplaySpool's final
# restoration/return boundaries are transactionally owned.


@pytest.mark.parametrize(
    ("start_error", "close_error", "expected_type"),
    [
        (MemoryError("start interrupted"), OSError("close failed"), MemoryError),
        (RuntimeError("start failed"), MemoryError("close interrupted"), MemoryError),
    ],
)
def test_direct_source_handle_start_rollback_obeys_cleanup_precedence(
    monkeypatch: pytest.MonkeyPatch,
    start_error: BaseException,
    close_error: BaseException,
    expected_type: type[BaseException],
) -> None:
    source = NonSeekableStream(b"repeatable", "source.bin")

    class AcquiredSpool:
        def __init__(self) -> None:
            self.close_calls = 0

        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise close_error

    acquired = AcquiredSpool()

    def acquire_spool(
        _cls: type[ReplaySpool],
        _stream: object,
        _memory_limit: int = DEFAULT_MEMORY_LIMIT,
        **kwargs: object,
    ) -> Any:
        adopter = kwargs.get("_adopter")
        if callable(adopter):
            adopter(acquired)
        return acquired

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(acquire_spool))
    target_code = SourceHandle.start.__code__
    interrupted = False

    def interrupt_after_spool_adoption(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        handle = frame.f_locals.get("self")
        if (
            frame.f_code is target_code
            and event == "line"
            and getattr(handle, "_spool", None) is acquired
            and not getattr(handle, "_started", False)
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise start_error
        return interrupt_after_spool_adoption

    sys.settrace(interrupt_after_spool_adoption)
    try:
        with pytest.raises(expected_type):
            SourceHandle(source)
    finally:
        sys.settrace(None)

    assert interrupted
    assert acquired.close_calls >= 1
    assert source.closed is False


def test_replay_spool_completed_return_gap_is_closed_by_adopting_handle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NamedBytesIO(b"complete", "source.bin")
    source.seek(3)
    handle = SourceHandle(source)
    close_calls = 0
    real_close = ReplaySpool.close
    target_code = ReplaySpool.from_stream.__func__.__code__
    interrupted = False

    def record_close(spool: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        real_close(spool)

    def interrupt_completed_spool(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        if (
            frame.f_code is target_code
            and event == "line"
            and frame.f_locals.get("completed") is not None
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("completed spool return interrupted")
        return interrupt_completed_spool

    monkeypatch.setattr(ReplaySpool, "close", record_close)
    sys.settrace(interrupt_completed_spool)
    try:
        with pytest.raises(MemoryError, match="completed spool return interrupted"):
            handle._ensure_spool()
    finally:
        sys.settrace(None)

    assert interrupted
    assert close_calls == 1
    assert source.tell() == 3
    handle.close()


def test_adopted_spool_survives_restore_and_repeated_process_cleanup_failures(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class RestoreOnceBytesIO(NamedBytesIO):
        enabled = False
        restore_calls = 0

        def seek(self, position: int, whence: int = 0) -> int:
            if self.enabled and whence == 0 and position == 3:
                self.restore_calls += 1
                if self.restore_calls == 2:
                    raise OSError("final restoration failed")
            return super().seek(position, whence)

    source = RestoreOnceBytesIO(b"complete", "source.bin")
    source.seek(3)
    handle = SourceHandle(source)
    source.restore_calls = 0
    source.enabled = True
    close_calls = 0
    real_close = ReplaySpool.close

    def fail_close_three_times(spool: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        if close_calls <= 3:
            raise MemoryError("spool cleanup interrupted")
        real_close(spool)

    monkeypatch.setattr(ReplaySpool, "close", fail_close_three_times)

    with pytest.raises(MemoryError, match="spool cleanup interrupted"):
        handle._ensure_spool()
    with pytest.raises(MemoryError, match="spool cleanup interrupted"):
        handle.close()
    with pytest.raises(MemoryError, match="spool cleanup interrupted"):
        handle.close()
    handle.close()

    assert close_calls == 4
    assert source.closed is False


def test_interrupted_adopter_does_not_unlink_owner_retained_spill(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"spilled", "source.bin")
    handle = SourceHandle.prepare(source)
    real_from_stream = ReplaySpool.from_stream.__func__
    real_adopt = handle._adopt_spool
    real_close = ReplaySpool.close

    def force_spill(
        cls: type[ReplaySpool],
        stream: object,
        _memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> ReplaySpool:
        return real_from_stream(
            cls,
            stream,
            memory_limit=1,
            _adopter=_adopter,
        )

    def adopt_then_interrupt(spool: ReplaySpool) -> None:
        real_adopt(spool)
        raise MemoryError("adopter interrupted after ownership transfer")

    def interrupt_cleanup(_spool: ReplaySpool) -> None:
        raise MemoryError("owner cleanup interrupted")

    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(force_spill))
    monkeypatch.setattr(handle, "_adopt_spool", adopt_then_interrupt)
    monkeypatch.setattr(ReplaySpool, "close", interrupt_cleanup)

    with pytest.raises(MemoryError, match="adopter interrupted"):
        handle._ensure_spool()

    retained = handle._spool
    assert retained is not None
    assert retained._path is not None
    assert retained._path.exists()

    monkeypatch.setattr(ReplaySpool, "close", real_close)
    handle.close()
    assert source.closed is False


def test_path_backend_stays_a_path(tmp_path: Path) -> None:
    path = tmp_path / "source.csv"
    path.write_bytes(b"Name,Value\nA,1\n")
    handle = SourceHandle(path)

    with handle.open_backend() as backend:
        assert backend == path
        assert isinstance(backend, Path)


def test_custom_handler_fallback_gets_fresh_raw_borrow() -> None:
    observations: list[tuple[str, object, bytes]] = []

    class ConsumingFailure(FormatHandler):
        def can_handle(self, format_type: str) -> bool:
            return format_type == "xlsx"

        def parse(self, file_source, sheet, options):
            observations.append(("primary", file_source, file_source.read(4)))
            raise ValueError("try fallback")

        def get_sheet_names(self, file_source):
            return ["Data"]

        def validate(self, file_source):
            return True, None

    class SuccessfulFallback(FormatHandler):
        def can_handle(self, format_type: str) -> bool:
            return format_type == "xlsx"

        def parse(self, file_source, sheet, options):
            observations.append(("fallback", file_source, file_source.read(4)))
            return pd.DataFrame({"recovered": [True]})

        def get_sheet_names(self, file_source):
            return ["Data"]

        def validate(self, file_source):
            return True, None

    source = io.BytesIO(_xlsx_content())
    source.seek(7)
    registry = HandlerRegistry(handlers=[ConsumingFailure(), SuccessfulFallback()])

    result = registry.parse(
        source,
        sheet="Data",
        options=ParseOptions(),
        format_type="xlsx",
    )

    assert result.to_dict(orient="list") == {"recovered": [True]}
    assert [(name, prefix) for name, _, prefix in observations] == [
        ("primary", b"PK\x03\x04"),
        ("fallback", b"PK\x03\x04"),
    ]
    assert all(not isinstance(raw, SourceHandle) for _, raw, _ in observations)
    assert source.tell() == 7


def test_custom_detector_gets_raw_source_and_position_is_restored() -> None:
    observed: list[object] = []

    class CustomDetector:
        def detect(self, file_source, filename=None):
            observed.append(file_source)
            assert file_source.read(4) == b"Name"
            return FormatInfo(format_type="csv")

    source = io.BytesIO(b"Name,Value\nA,1\n")
    source.seek(2)
    registry = HandlerRegistry(detector=CustomDetector())  # type: ignore[arg-type]

    detected = registry.detect_format(source, filename="source.csv")

    assert detected.format_type == "csv"
    assert observed == [source]
    assert source.tell() == 2


def test_one_argument_custom_detector_remains_compatible_across_registry_operations() -> None:
    calls: list[object] = []

    class OneArgumentDetector:
        def detect(self, file_source):
            calls.append(file_source)
            return FormatInfo(format_type="csv")

    source = io.BytesIO(b"Name,Value\nA,1\n")
    source.seek(2)
    registry = HandlerRegistry(detector=OneArgumentDetector())  # type: ignore[arg-type]

    names = registry.get_sheet_names(source)
    is_valid, error = registry.validate(source)
    frame = registry.parse(source, options=ParseOptions())

    assert names == ["Sheet1"]
    assert (is_valid, error) == (True, None)
    assert frame.to_dict(orient="list") == {"Name": ["A"], "Value": [1]}
    assert calls == [source, source, source]
    assert source.tell() == 2


def test_legacy_builtin_handler_subclass_gets_raw_source_by_default() -> None:
    observed: list[object] = []

    class LegacyCSVHandler(CSVHandler):
        def parse(self, file_source, sheet, options):
            observed.append(file_source)
            assert callable(getattr(file_source, "read", None))
            return super().parse(file_source, sheet, options)

    source = io.BytesIO(b"Name,Value\nA,1\n")
    source.seek(2)
    registry = HandlerRegistry(handlers=[LegacyCSVHandler()])

    frame = registry.parse(
        source,
        options=ParseOptions(),
        format_type="csv",
    )

    assert frame.to_dict(orient="list") == {"Name": ["A"], "Value": [1]}
    assert observed == [source]
    assert source.tell() == 2


def test_legacy_format_detector_subclass_gets_raw_source_by_default() -> None:
    observed: list[object] = []

    class LegacyDetector(FormatDetector):
        def detect(self, file_source, filename=None):
            observed.append(file_source)
            assert callable(getattr(file_source, "read", None))
            return FormatInfo(format_type="csv")

    source = io.BytesIO(b"Name,Value\nA,1\n")
    source.seek(2)
    registry = HandlerRegistry(detector=LegacyDetector())

    detected = registry.detect_format(source)

    assert detected.format_type == "csv"
    assert observed == [source]
    assert source.tell() == 2


def test_legacy_registry_subclass_gets_fresh_raw_source_for_each_workbook_operation() -> None:
    observations: list[tuple[str, object, int]] = []

    class LegacyRegistry(HandlerRegistry):
        def _record(self, operation: str, file_source: object) -> None:
            assert callable(getattr(file_source, "read", None))
            observations.append((operation, file_source, file_source.tell()))  # type: ignore[attr-defined]

        def detect_format(self, file_source, filename=None):
            self._record("detect", file_source)
            return super().detect_format(file_source, filename=filename)

        def get_sheet_names(self, file_source, format_type=None):
            self._record("sheets", file_source)
            return super().get_sheet_names(file_source, format_type=format_type)

        def validate(self, file_source, format_type=None):
            self._record("validate", file_source)
            return super().validate(file_source, format_type=format_type)

        def parse(self, file_source, sheet=None, options=None, format_type=None):
            self._record("parse", file_source)
            return super().parse(file_source, sheet, options, format_type)

    source = io.BytesIO(_xlsx_content())
    source.seek(3)
    registry = LegacyRegistry()

    with MessyWorkbook(source, filename="source.xlsx", registry=registry) as workbook:
        frame = workbook.to_dataframe("Data")

    assert frame.iloc[:, 0].tolist() == ["Alpha", "Beta"]
    assert [operation for operation, _, _ in observations] == [
        "detect",
        "sheets",
        "validate",
        "parse",
    ]
    assert all(raw is source for _, raw, _ in observations)
    assert all(position == 0 for _, _, position in observations)
    assert source.tell() == 3


def test_builtin_registry_reuses_one_handle_without_a_permanent_byte_cache() -> None:
    observations: list[tuple[str, SourceHandle]] = []

    class HandleDetector:
        _accepts_source_handle = True

        def detect(self, file_source, filename=None):
            assert isinstance(file_source, SourceHandle)
            observations.append(("detect", file_source))
            file_source.read_bytes()
            return FormatInfo(format_type="xlsx")

    class HandleHandler(FormatHandler):
        _accepts_source_handle = True

        def _record(self, operation: str, file_source: object) -> SourceHandle:
            assert isinstance(file_source, SourceHandle)
            observations.append((operation, file_source))
            file_source.read_bytes()
            return file_source

        def can_handle(self, format_type: str) -> bool:
            return format_type == "xlsx"

        def parse(self, file_source, sheet, options):
            self._record("parse", file_source)
            return pd.DataFrame({"value": [1]})

        def get_sheet_names(self, file_source):
            self._record("sheets", file_source)
            return ["Data"]

        def validate(self, file_source):
            self._record("validate", file_source)
            return True, None

    source = CountingBytesIO(b"opaque but repeatable backend payload")
    source.seek(4)
    registry = HandlerRegistry(
        handlers=[HandleHandler()],
        detector=HandleDetector(),  # type: ignore[arg-type]
    )
    config = SheetConfig(auto_detect=False)

    with MessyWorkbook(
        source,
        filename="source.xlsx",
        registry=registry,
        sheet_config=config,
    ) as workbook:
        frame = workbook.to_dataframe("Data")

    assert frame.to_dict(orient="list") == {"value": [1]}
    assert [operation for operation, _ in observations] == [
        "detect",
        "detect",
        "sheets",
        "validate",
        "parse",
    ]
    assert len({id(handle) for _, handle in observations}) == 1
    assert source.read_sizes == [-1, -1, -1, -1, -1]
    assert source.tell() == 4
    assert source.closed is False


def test_persistent_cell_views_use_owned_streams_and_restore_caller_position() -> None:
    source = io.BytesIO(_xlsx_content())
    source.seek(3)

    with MessyWorkbook(source, filename="source.xlsx") as workbook:
        assert workbook.get_cell("Data", 2, 1).value == "Alpha"
        assert workbook._get_cached_cell_value("Data", 2, 2) == 1
        primary_source = workbook._wb_source
        cached_source = workbook._cached_wb_source

        assert primary_source is not None
        assert cached_source is not None
        assert primary_source is not source
        assert cached_source is not source
        assert source.tell() == 3

    assert primary_source.closed is True
    assert cached_source.closed is True
    assert source.closed is False
    assert source.tell() == 3


def test_source_handle_retains_spool_after_ordinary_close_failure_for_retry(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = NonSeekableStream(b"x" * 64, "source.bin")
    handle = SourceHandle(source)
    spool = handle._spool
    assert spool is not None
    assert spool._path is None
    close_calls = 0
    real_close = ReplaySpool.close

    def fail_once(target: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        if close_calls == 1:
            raise OSError("unlink interrupted")
        real_close(target)

    monkeypatch.setattr(ReplaySpool, "close", fail_once)
    with pytest.raises(OSError, match="unlink interrupted"):
        handle.close()

    assert handle._spool is spool
    assert handle.closed is False
    handle.close()
    assert close_calls == 2
    assert handle._spool is None
    assert source.closed is False


def test_source_handle_retains_ordinary_cursor_restore_failure_for_retry() -> None:
    class RestoreOnce(io.BytesIO):
        enabled = False
        restore_calls = 0

        def seek(self, position: int, whence: int = 0) -> int:
            if self.enabled and position == 5 and whence == 0:
                self.restore_calls += 1
                if self.restore_calls == 1:
                    raise OSError("restore interrupted")
            return super().seek(position, whence)

    source = RestoreOnce(b"complete source")
    source.seek(5)
    handle = SourceHandle(source)
    source.enabled = True

    with (
        pytest.raises(OSError, match="restore interrupted"),
        handle.open_binary() as borrowed,
    ):
        borrowed.read()

    assert handle._pending_restore_position == 5
    handle.close()
    assert source.tell() == 5
    assert source.restore_calls == 2
    assert source.closed is False


@pytest.mark.parametrize(
    "restoration_error",
    [
        OSError("workbook cursor restore interrupted"),
        MemoryError("workbook cursor restore interrupted"),
    ],
    ids=["ordinary", "process"],
)
def test_workbook_retains_source_close_until_all_handle_obligations_finish(
    monkeypatch: pytest.MonkeyPatch,
    restoration_error: BaseException,
) -> None:
    class RestoreTwice(io.BytesIO):
        enabled = False
        restore_calls = 0

        def seek(self, position: int, whence: int = 0) -> int:
            if self.enabled and position == 5 and whence == 0:
                self.restore_calls += 1
                if self.restore_calls <= 2:
                    raise type(restoration_error)(*restoration_error.args)
            return super().seek(position, whence)

    source = RestoreTwice(b"complete source")
    source.seek(5)
    handle = SourceHandle(source)
    spool = ReplaySpool.from_stream(io.BytesIO(b"x" * 32), memory_limit=1)
    path = spool._path
    assert path is not None
    handle._spool = spool
    source.enabled = True
    close_calls = 0
    real_spool_close = ReplaySpool.close

    def track_spool_close(target: ReplaySpool) -> None:
        nonlocal close_calls
        close_calls += 1
        real_spool_close(target)

    monkeypatch.setattr(ReplaySpool, "close", track_spool_close)
    with (
        pytest.raises(type(restoration_error), match="cursor restore interrupted"),
        handle.open_binary() as borrowed,
    ):
        borrowed.read()

    assert source.tell() == len(source.getvalue())
    assert handle._pending_restore_position == 5
    assert handle._spool is spool
    assert source.restore_calls == 1

    workbook = object.__new__(MessyWorkbook)
    workbook._closed = False
    workbook._source_handle = handle
    workbook._source_handle_close_pending = True

    with pytest.raises(type(restoration_error), match="cursor restore interrupted"):
        workbook.close()

    assert source.tell() == len(source.getvalue())
    assert source.closed is False
    assert handle.closed is False
    assert handle._pending_restore_position == 5
    assert handle._spool is None
    assert workbook._source_handle_close_pending is True
    assert source.restore_calls == 2
    assert close_calls == 1
    assert not path.exists()

    workbook.close()
    workbook.close()

    assert source.tell() == 5
    assert source.closed is False
    assert handle.closed is True
    assert handle._pending_restore_position is None
    assert workbook._source_handle_close_pending is False
    assert source.restore_calls == 3
    assert close_calls == 1
    assert not path.exists()


def test_direct_source_handle_return_gap_finalizer_removes_spill(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    created: list[Path] = []
    real_mkstemp = spool_module.tempfile.mkstemp
    real_from_stream = ReplaySpool.from_stream.__func__
    target_code = SourceHandle.__init__.__code__
    interrupted = False

    def track_mkstemp(*args: object, **kwargs: object) -> tuple[int, str]:
        descriptor, raw_path = real_mkstemp(*args, **kwargs)
        created.append(Path(raw_path))
        return descriptor, raw_path

    def force_spill(
        cls: type[ReplaySpool],
        stream: object,
        _memory_limit: int = DEFAULT_MEMORY_LIMIT,
        *,
        _adopter: Any = None,
    ) -> ReplaySpool:
        return real_from_stream(
            cls,
            stream,
            memory_limit=1,
            _adopter=_adopter,
        )

    def interrupt_constructor_return(frame: object, event: str, _arg: object) -> object:
        nonlocal interrupted
        if getattr(frame, "f_code", None) is target_code and event == "return" and not interrupted:
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None  # type: ignore[attr-defined]
            raise MemoryError("source handle return interrupted")
        return interrupt_constructor_return

    monkeypatch.setattr(spool_module.tempfile, "mkstemp", track_mkstemp)
    monkeypatch.setattr(ReplaySpool, "from_stream", classmethod(force_spill))
    source = NonSeekableStream(b"x" * 64, "source.bin")
    sys.settrace(interrupt_constructor_return)
    try:
        with pytest.raises(MemoryError, match="source handle return interrupted"):
            SourceHandle(source)
    finally:
        sys.settrace(None)
    gc.collect()

    assert interrupted
    assert len(created) == 1
    assert not created[0].exists()
    assert source.closed is False
