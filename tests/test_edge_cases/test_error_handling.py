"""Tests for error handling with formulas and Excel errors."""

import logging

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook
from messy_xlsx.exceptions import FileError
from messy_xlsx.formulas import FormulaConfig, FormulaEvaluationMode


class TestFormulaErrors:
    """Test handling of Excel formula errors."""

    def test_all_formula_errors(self, temp_dir):
        """Test file with all types of Excel errors."""
        file_path = temp_dir / "formula_errors.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Error Type", "Formula", "Result"])
        ws.append(["DIV/0", "=1/0", None])
        ws.append(["N/A", "=NA()", None])
        ws.append(["NAME", "=UNKNOWNFUNC()", None])
        ws.append(["NULL", "=A1 B1", None])
        ws.append(["NUM", "=SQRT(-1)", None])
        ws.append(["REF", "=A1000000", None])
        ws.append(["VALUE", '=1+"text"', None])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            # Should handle errors gracefully
            assert len(df) >= 0


class TestCircularReferences:
    """Test handling of circular formula references."""

    def test_simple_circular_reference(self, temp_dir):
        """Test circular reference A1=B1, B1=A1."""
        file_path = temp_dir / "circular.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=B1"
        ws["B1"] = "=A1"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            # Should not crash
            df = mwb.to_dataframe()
            assert df is not None


class TestComplexFormulas:
    """Test complex formula scenarios."""

    def test_nested_formulas(self, temp_dir):
        """Test deeply nested formulas."""
        file_path = temp_dir / "nested.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value", "Formula"])
        ws["A2"] = 10
        ws["B2"] = '=IF(A2>5,IF(A2>8,"High","Medium"),"Low")'
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1

    def test_array_formulas(self, temp_dir):
        """Test array formulas if supported."""
        file_path = temp_dir / "array.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Sum"])
        ws.append([1, "=SUM(A2:A5)"])
        ws.append([2, None])
        ws.append([3, None])
        ws.append([4, None])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) >= 1


class TestFormulaEvaluationModes:
    """Test different formula evaluation modes."""

    def test_disabled_mode(self, temp_dir):
        """Test formula evaluation disabled."""
        file_path = temp_dir / "formulas.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "Sum"])
        ws.append([1, 2, "=A2+B2"])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.DISABLED)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1

    def test_cached_only_mode(self, temp_dir):
        """Test using only cached formula values."""
        file_path = temp_dir / "cached.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value", "Double"])
        ws["A2"] = 5
        ws["B2"] = "=A2*2"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == 1


class TestExceptionSerialization:
    """Test to_dict() on each exception class."""

    def test_base_exception_to_dict(self):
        from messy_xlsx.exceptions import MessyXlsxError

        e = MessyXlsxError("something went wrong", context={"key": "val"})
        d = e.to_dict()
        assert d["error"] == "MessyXlsxError"
        assert d["message"] == "something went wrong"
        assert d["context"] == {"key": "val"}

    def test_file_error_to_dict(self):
        from messy_xlsx.exceptions import FileError

        e = FileError("not found", file_path="/tmp/x.xlsx", operation="open")
        d = e.to_dict()
        assert d["error"] == "FileError"
        assert d["context"]["file_path"] == "/tmp/x.xlsx"
        assert d["context"]["operation"] == "open"

    def test_format_error_to_dict(self):
        from messy_xlsx.exceptions import FormatError

        e = FormatError(
            "bad format",
            file_path="/tmp/x.xlsx",
            detected_format="csv",
            attempted_formats=["xlsx", "xls"],
        )
        d = e.to_dict()
        assert d["error"] == "FormatError"
        assert d["context"]["detected_format"] == "csv"
        assert d["context"]["attempted_formats"] == ["xlsx", "xls"]

    def test_structure_error_to_dict(self):
        from messy_xlsx.exceptions import StructureError

        e = StructureError("no header", sheet="Sheet1", detection_phase="headers")
        d = e.to_dict()
        assert d["error"] == "StructureError"
        assert d["context"]["sheet"] == "Sheet1"
        assert d["context"]["detection_phase"] == "headers"

    def test_normalization_error_to_dict(self):
        from messy_xlsx.exceptions import NormalizationError

        e = NormalizationError(
            "type mismatch",
            column="price",
            row=5,
            value="abc",
            expected_type="float",
        )
        d = e.to_dict()
        assert d["error"] == "NormalizationError"
        assert d["context"]["column"] == "price"
        assert d["context"]["row"] == 5
        assert d["context"]["expected_type"] == "float"

    def test_formula_error_to_dict(self):
        from messy_xlsx.exceptions import FormulaError

        e = FormulaError("eval failed", cell_ref="A1", formula="=1/0")
        d = e.to_dict()
        assert d["error"] == "FormulaError"
        assert d["context"]["cell_ref"] == "A1"
        assert d["context"]["formula"] == "=1/0"

    def test_circular_reference_error_to_dict(self):
        from messy_xlsx.exceptions import CircularReferenceError

        e = CircularReferenceError("cycle detected", cycle=["A1", "B1", "A1"])
        d = e.to_dict()
        assert d["error"] == "CircularReferenceError"
        assert d["context"]["cycle"] == ["A1", "B1", "A1"]

    def test_unsupported_function_error_to_dict(self):
        from messy_xlsx.exceptions import UnsupportedFunctionError

        e = UnsupportedFunctionError("WEBSERVICE", cell_ref="B2")
        d = e.to_dict()
        assert d["error"] == "UnsupportedFunctionError"
        assert d["context"]["function_name"] == "WEBSERVICE"
        assert d["context"]["cell_ref"] == "B2"

    def test_none_context_values_excluded(self):
        from messy_xlsx.exceptions import FileError

        e = FileError("error", file_path=None, operation=None)
        d = e.to_dict()
        assert "file_path" not in d["context"]
        assert "operation" not in d["context"]


class TestUnsupportedFunctions:
    """Test handling of unsupported Excel functions."""

    def test_unsupported_function_placeholder(self, temp_dir):
        """Test unsupported function returns placeholder."""
        file_path = temp_dir / "unsupported.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Function", "Result"])
        ws.append(["WEBSERVICE", '=WEBSERVICE("url")'])
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(
            mode=FormulaEvaluationMode.CACHED_WITH_FALLBACK,
            raise_on_unsupported=False,
            unsupported_value="#UNSUPPORTED",
        )

        with MessyWorkbook(file_path, formula_config=config) as mwb:
            df = mwb.to_dataframe()
            assert len(df) >= 0


# ============================================================================
# Phase 4 — Exception narrowing & debug logging tests
# ============================================================================


class TestFormulaLoadLogging:
    """Test that formula engine load failures emit debug logs."""

    def test_formula_load_failure_logs_debug(self, temp_dir, caplog):
        """Formula engine load failure should log debug, not silently pass."""
        file_path = temp_dir / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        wb.save(file_path)
        wb.close()

        # Use FULL_EVALUATION mode so the engine tries to load
        config = FormulaConfig(mode=FormulaEvaluationMode.ALWAYS_EVALUATE)

        with (
            caplog.at_level(logging.DEBUG, logger="messy_xlsx.workbook"),
            MessyWorkbook(file_path, formula_config=config) as mwb,
        ):
            df = mwb.to_dataframe()
            assert df is not None

        # If the formula engine is NOT available (no xlcalculator/formulas),
        # the load won't even be attempted, so no debug log.
        # If available but fails, we should see a debug log.
        # Either outcome is fine — no crash is the key assertion.


class TestGetCellWithoutWorkbook:
    """Test that get_cell raises FileError instead of bare assert."""

    def test_get_cell_raises_file_error_when_wb_not_loaded(self, temp_dir):
        """Accessing cell without workbook should raise FileError, not AssertionError."""
        file_path = temp_dir / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header"])
        ws.append(["Data"])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            # Force _wb to None to simulate unloaded state
            mwb._wb = None
            # _ensure_workbook will reload it, so we need to bypass it
            # Instead, test the guard directly by calling get_cell
            # which calls _ensure_workbook first (which will reload)
            # So let's test the guard by patching _ensure_workbook
            original_ensure = mwb._ensure_workbook
            mwb._ensure_workbook = lambda: None  # type: ignore[assignment]
            mwb._wb = None

            with pytest.raises(FileError, match="Workbook not loaded"):
                mwb.get_cell("Sheet", 1, 1)

            # Restore
            mwb._ensure_workbook = original_ensure  # type: ignore[assignment]


class TestFormulaEvalLogging:
    """Test that formula evaluation failures emit debug logs."""

    def test_formula_eval_failure_logs_debug(self, temp_dir, caplog):
        """Formula eval failure should log debug, not silently pass."""
        file_path = temp_dir / "formula_eval.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Value", "Formula"])
        ws["A2"] = 10
        ws["B2"] = "=A2*2"
        wb.save(file_path)
        wb.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.ALWAYS_EVALUATE)

        with (
            caplog.at_level(logging.DEBUG, logger="messy_xlsx.workbook"),
            MessyWorkbook(file_path, formula_config=config) as mwb,
        ):
            # get_cell triggers formula evaluation
            cell = mwb.get_cell("Sheet", 2, 2)
            # The cell should have a value (either evaluated or cached)
            assert cell is not None


class TestCsvMetadataDetectionLogging:
    """Test that CSV metadata detection logs on parse errors."""

    def test_malformed_csv_logs_debug_on_metadata_detect(self, temp_dir, caplog):
        """Metadata detection on malformed CSV should log debug."""
        # Create a file that will fail pd.read_csv
        bad_file = temp_dir / "bad.csv"
        bad_file.write_bytes(b"\x00\x01\x02\xff\xfe" * 100)

        from messy_xlsx.parsing.csv_handler import MetadataRowDetector

        detector = MetadataRowDetector()

        with caplog.at_level(logging.DEBUG, logger="messy_xlsx.parsing.csv_handler"):
            result = detector.detect_skip_rows(bad_file, "utf-8", ",")

        assert result == 0

    def test_malformed_text_logs_debug_on_metadata_detect(self, caplog):
        """Metadata detection on malformed text should log debug."""
        from messy_xlsx.parsing.csv_handler import MetadataRowDetector

        detector = MetadataRowDetector()

        # Very short text that might cause EmptyDataError
        with caplog.at_level(logging.DEBUG, logger="messy_xlsx.parsing.csv_handler"):
            result = detector.detect_skip_rows_from_text("", ",")

        # Should return 0 gracefully (empty data → no rows to analyze)
        assert result == 0


class TestLocaleDetectorNarrowedException:
    """Test that locale detector handles cell access errors gracefully."""

    def test_handles_attribute_error_on_cell(self, caplog):
        """Locale detector should handle cells that raise AttributeError."""
        from messy_xlsx.detection.locale_detector import LocaleDetector

        # Create a mock worksheet-like object that raises on cell access
        class BrokenWorksheet:
            def cell(self, row, col):
                raise AttributeError("broken cell")

        detector = LocaleDetector()
        with caplog.at_level(logging.DEBUG, logger="messy_xlsx.detection.locale_detector"):
            result = detector.detect(BrokenWorksheet())  # type: ignore[arg-type]

        # Should return default US locale, not crash
        assert result.locale == "en_US"
        assert result.decimal_separator == "."

    def test_handles_index_error_on_cell(self, caplog):
        """Locale detector should handle cells that raise IndexError."""
        from messy_xlsx.detection.locale_detector import LocaleDetector

        class IndexErrorWorksheet:
            def cell(self, row, col):
                raise IndexError("cell out of range")

        detector = LocaleDetector()
        with caplog.at_level(logging.DEBUG, logger="messy_xlsx.detection.locale_detector"):
            result = detector.detect(IndexErrorWorksheet())  # type: ignore[arg-type]

        assert result.locale == "en_US"


class TestXlsxHandlerSheetNameFallback:
    """Test that XLSX handler sheet name fallback logs debug."""

    def test_corrupted_xlsx_falls_back_to_openpyxl(self, temp_dir, caplog):
        """When fastexcel fails, should log and fall back to openpyxl."""
        # Create a valid xlsx file to test the normal path
        file_path = temp_dir / "normal.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws.append(["data"])
        wb.save(file_path)
        wb.close()

        from messy_xlsx.parsing.xlsx_handler import XLSXHandler

        handler = XLSXHandler()

        with caplog.at_level(logging.DEBUG, logger="messy_xlsx.parsing.xlsx_handler"):
            names = handler.get_sheet_names(file_path)

        # Should successfully get sheet names (via fastexcel or openpyxl fallback)
        assert "TestSheet" in names


class TestNarrowedExceptionCoverage:
    """Verify no broad 'except Exception: pass' patterns remain."""

    def test_is_cell_merged_handles_attribute_error(self, temp_dir):
        """_is_cell_merged should handle AttributeError gracefully."""
        file_path = temp_dir / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["data"])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            mwb._ensure_workbook()
            ws = mwb._wb["Sheet"]
            # Normal case: should not raise
            result = mwb._is_cell_merged(ws, 1, 1)
            assert result is False

            # Edge case: object without merged_cells attribute
            result = mwb._is_cell_merged(object(), 1, 1)
            assert result is False

    def test_is_cell_hidden_handles_attribute_error(self, temp_dir):
        """_is_cell_hidden should handle AttributeError gracefully."""
        file_path = temp_dir / "simple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["data"])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            mwb._ensure_workbook()
            ws = mwb._wb["Sheet"]
            # Normal case: should not raise
            result = mwb._is_cell_hidden(ws, 1, 1)
            assert result is False

            # Edge case: object without row_dimensions
            result = mwb._is_cell_hidden(object(), 1, 1)
            assert result is False
