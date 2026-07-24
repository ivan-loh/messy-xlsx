"""Review regressions for the Task 13 multi-sheet planning architecture."""

from __future__ import annotations

import gc
import weakref
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pytest

import messy_xlsx.workbook as workbook_module
from messy_xlsx import (
    MessyWorkbook,
    MultiSheetOptions,
    MultiSheetParser,
    SheetConfig,
    SheetInfo,
    read_all_sheets,
)
from messy_xlsx.detection.structure_sampler import StructureEvidence, StructureSampler
from messy_xlsx.exceptions import FormatError
from messy_xlsx.ooxml.models import IntervalIndex, RowBitSet, SheetManifest
from messy_xlsx.parsing.base_handler import FormatHandler, ParseOptions
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.parsing.parse_plan import compile_parse_plan
from messy_xlsx.parsing.sheet_planner import PlannedSheetState, SheetPlanner


def _write_xlsx(path: Path, *sheet_names: str) -> Path:
    source = openpyxl.Workbook()
    for index, name in enumerate(sheet_names):
        sheet = source.active if index == 0 else source.create_sheet()
        sheet.title = name
        sheet.append(["name", "value"])
        sheet.append([name.lower(), index + 1])
    source.save(path)
    source.close()
    return path


def test_style_only_tail_cell_does_not_change_sheet_info_or_drop_output(
    tmp_path: Path,
) -> None:
    path = tmp_path / "styled-tail.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["name", "value"])
    sheet.append(["alpha", 1])
    sheet["Z100"].number_format = "0.00"
    source.save(path)
    source.close()

    info = MultiSheetParser(path).analyze_sheets()[0]
    frames = read_all_sheets(path)

    assert (info.row_count, info.col_count) == (2, 2)
    assert info.skip_reason is None
    assert list(frames) == ["Data"]
    assert frames["Data"].to_dict("records") == [{"name": "alpha", "value": 1}]


def test_offset_value_rectangle_controls_counts_and_legacy_minimum_selection(
    tmp_path: Path,
) -> None:
    path = tmp_path / "offset.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["C4"] = "name"
    sheet["D4"] = "value"
    sheet["C5"] = "alpha"
    sheet["D5"] = 1
    source.save(path)
    source.close()

    info = MultiSheetParser(path).analyze_sheets()[0]
    strict = MultiSheetParser(path, MultiSheetOptions(min_cols=3)).analyze_sheets()[0]
    frames = read_all_sheets(path, min_cols=3)

    assert (info.row_count, info.col_count) == (2, 2)
    assert strict.skip_reason == "Too small"
    assert frames == {}


class _VirtualHandler(FormatHandler):
    _accepts_source_handle = True

    def __init__(self, *, fail: bool = False) -> None:
        self.fail = fail
        self.parse_calls: list[str | None] = []

    def can_handle(self, format_type: str) -> bool:
        return format_type == "xlsx"

    def parse(
        self,
        _source: Any,
        sheet: str | None,
        _options: ParseOptions,
    ) -> pd.DataFrame:
        self.parse_calls.append(sheet)
        if self.fail:
            raise ValueError("primary rejected source")
        return pd.DataFrame({"custom": [7]})

    def get_sheet_names(self, _source: Any) -> list[str]:
        return ["Virtual"]

    def validate(self, _source: Any) -> tuple[bool, str | None]:
        return True, None


def test_default_iter_sheets_uses_authoritative_custom_registry_without_analysis(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_xlsx(tmp_path / "physical.xlsx", "Physical")
    handler = _VirtualHandler()
    registry = HandlerRegistry(handlers=[handler])
    monkeypatch.setattr(
        workbook_module,
        "ManifestReader",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("custom registry must not construct an OOXML manifest")
        ),
    )

    with (
        MessyWorkbook(path, registry=registry) as workbook,
        workbook.iter_sheets() as stream,
    ):
        results = list(stream)

    assert [result.name for result in results] == ["Virtual"]
    assert results[0].error is None
    assert results[0].dataframe is not None
    assert results[0].dataframe.to_dict("records") == [{"custom": 7}]
    assert handler.parse_calls == ["Virtual"]


def test_filter_mutated_sheet_name_controls_legacy_materialization_and_key(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_xlsx(tmp_path / "rename.xlsx", "First", "Last")
    materialized: list[str] = []
    original = MessyWorkbook._materialize_raw_frame

    def record(
        workbook: MessyWorkbook,
        sheet: str,
        format_type: str,
        plan: Any,
    ) -> pd.DataFrame:
        materialized.append(sheet)
        return original(workbook, sheet, format_type, plan)

    def rename(info: SheetInfo) -> bool:
        if info.name == "First":
            info.name = "Last"
        return True

    monkeypatch.setattr(MessyWorkbook, "_materialize_raw_frame", record)
    frames = read_all_sheets(path, sheet_filter=rename)

    assert list(frames) == ["Last"]
    assert materialized == ["Last", "Last"]


def test_parse_sheet_ooxml_has_only_one_full_materialization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_xlsx(tmp_path / "single.xlsx", "Data")
    materialized: list[str] = []
    original = MessyWorkbook._materialize_raw_frame

    def record(
        workbook: MessyWorkbook,
        sheet: str,
        format_type: str,
        plan: Any,
    ) -> pd.DataFrame:
        materialized.append(sheet)
        return original(workbook, sheet, format_type, plan)

    monkeypatch.setattr(MessyWorkbook, "_materialize_raw_frame", record)
    frame = MultiSheetParser(path).parse_sheet("Data")

    assert frame.to_dict("records") == [{"name": "data", "value": 1}]
    assert materialized == ["Data"]


def test_parse_sheet_missing_ooxml_name_preserves_legacy_format_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_xlsx(tmp_path / "missing.xlsx", "Data")
    parser = MultiSheetParser(path)

    with monkeypatch.context() as baseline_patch:
        baseline_patch.setattr(
            MessyWorkbook,
            "_uses_builtin_ooxml_planner",
            lambda _workbook: False,
        )
        with pytest.raises(FormatError) as baseline:
            parser.parse_sheet("missing")

    with pytest.raises(FormatError) as optimized:
        parser.parse_sheet("missing")

    assert type(optimized.value) is type(baseline.value)
    assert str(optimized.value) == str(baseline.value)
    assert optimized.value.context == baseline.value.context


def test_read_all_sheets_xls_materializes_only_selected_output_once(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    xlwt = pytest.importorskip("xlwt")
    pytest.importorskip("xlrd")
    path = tmp_path / "selected.xls"
    source = xlwt.Workbook()
    for index, name in enumerate(("First", "Second"), start=1):
        sheet = source.add_sheet(name)
        sheet.write(0, 0, "name")
        sheet.write(0, 1, "value")
        sheet.write(1, 0, name.lower())
        sheet.write(1, 1, index)
    source.save(str(path))

    materialized: list[str] = []
    original = MessyWorkbook._materialize_raw_frame

    def record(
        workbook: MessyWorkbook,
        sheet: str,
        format_type: str,
        plan: Any,
    ) -> pd.DataFrame:
        materialized.append(sheet)
        return original(workbook, sheet, format_type, plan)

    monkeypatch.setattr(MessyWorkbook, "_materialize_raw_frame", record)
    monkeypatch.setattr(
        workbook_module,
        "ManifestReader",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("XLS must not construct an OOXML manifest")
        ),
    )

    frames = read_all_sheets(path, sheets=["Second"])

    assert list(frames) == ["Second"]
    assert materialized == ["Second"]


def test_metrics_count_pivot_and_structure_bounded_reads(
    tmp_path: Path,
) -> None:
    path = _write_xlsx(tmp_path / "samples.xlsx", "Data")
    parser = MultiSheetParser(path)

    with MessyWorkbook(path) as workbook:
        planned = parser._plan_shared_ooxml(
            workbook,
            compile_outputs=False,
            select_all=True,
        )
        metrics = workbook.parse_metrics

    assert planned[0].state is PlannedSheetState.SKIPPED
    assert metrics.sample_reads == 2
    assert metrics.full_materializations == 0


def test_custom_primary_failure_and_fallback_success_are_counted(
    tmp_path: Path,
) -> None:
    path = _write_xlsx(tmp_path / "fallback.xlsx", "Physical")
    primary = _VirtualHandler(fail=True)
    fallback = _VirtualHandler()
    registry = HandlerRegistry(handlers=[primary, fallback])

    with (
        MessyWorkbook(path, registry=registry) as workbook,
        workbook.iter_sheets() as stream,
    ):
        results = list(stream)
        metrics = workbook.parse_metrics

    assert results[0].error is None
    assert results[0].dataframe is not None
    assert results[0].dataframe.to_dict("records") == [{"custom": 7}]
    assert primary.parse_calls == ["Virtual"]
    assert fallback.parse_calls == ["Virtual"]
    assert metrics.failed_attempts == 1
    assert metrics.full_materializations == 1


class _SparseManifestReader:
    def __init__(self, manifest: SheetManifest) -> None:
        self.manifest = manifest

    def sheet(self, _name: str) -> SheetManifest:
        return self.manifest


class _CellBudgetReader:
    def __init__(self) -> None:
        self.requests: list[tuple[int, int, int]] = []

    def sample_windows(
        self,
        _sheet: str,
        *,
        windows: Any,
        min_column: int,
        max_column: int,
    ) -> StructureEvidence:
        rows = sum(window.n_rows for window in windows)
        width = max_column - min_column + 1
        cells = rows * width
        self.requests.append((rows, width, cells))
        if cells > 1_000_000:
            raise AssertionError(f"dense sample requested {cells} cells")
        return StructureEvidence(
            row_numbers=(1,),
            values=pd.DataFrame([["head"]], index=(1,)),
        )


def test_sparse_structure_sampling_has_a_hard_cell_budget_and_keeps_bounds() -> None:
    manifest = SheetManifest(
        name="Sparse",
        target="xl/worksheets/sheet1.xml",
        declared_dimension=(1, 1, 10_000, 16_384),
        observed_max_row=10_000,
        observed_max_col=16_384,
        hidden_rows=IntervalIndex(()),
        hidden_columns=IntervalIndex(()),
        merged_ranges=(),
        has_formulas=False,
        formula_samples=(),
        observed_min_col=1,
        semantic_data_region=(1, 10_000, 1, 16_384),
        semantic_nonempty_rows=RowBitSet(b"\x01"),
    )
    reader = _CellBudgetReader()

    structure = StructureSampler(reader, _SparseManifestReader(manifest)).analyze("Sparse")

    assert len(reader.requests) == 1
    assert reader.requests[0][2] <= 1_000_000
    assert structure.data_start_row == 1
    assert structure.data_end_row == 10_000
    assert structure.data_start_col == 1
    assert structure.data_end_col == 16_384


class _CountingName(str):
    comparisons = 0

    def __eq__(self, other: object) -> bool:
        type(self).comparisons += 1
        return super().__eq__(other)

    __hash__ = str.__hash__


def test_explicit_sheet_selection_uses_near_linear_frozen_membership() -> None:
    count = 2_000
    names = [f"Sheet-{index}" for index in range(count)]
    options = MultiSheetOptions(sheets=[_CountingName(name) for name in names])
    parse_plan = compile_parse_plan(SheetConfig(auto_detect=False), None, "xlsx")
    planner = SheetPlanner(
        lambda name: SheetInfo(name=name, row_count=2, col_count=2, header_row=0),
        lambda _name, _info: parse_plan,
        should_propagate=lambda _error: False,
        analysis_failure_info=lambda name, error: SheetInfo(
            name=name,
            row_count=0,
            col_count=0,
            header_row=0,
            skip_reason=str(error),
        ),
    )
    _CountingName.comparisons = 0

    planned = planner.plan(names, options=options)

    assert len(planned) == count
    assert all(item.state is PlannedSheetState.READY for item in planned)
    assert _CountingName.comparisons < 10_000


def test_filter_mutation_of_explicit_list_keeps_legacy_timing() -> None:
    names = ["First", "Second"]
    options = MultiSheetOptions(sheets=["missing"])

    def clear_selection(_info: SheetInfo) -> bool:
        assert options.sheets is not None
        options.sheets.clear()
        return True

    options.sheet_filter = clear_selection
    parse_plan = compile_parse_plan(SheetConfig(auto_detect=False), None, "xlsx")
    planner = SheetPlanner(
        lambda name: SheetInfo(name=name, row_count=2, col_count=2, header_row=0),
        lambda _name, _info: parse_plan,
        should_propagate=lambda _error: False,
        analysis_failure_info=lambda name, error: SheetInfo(
            name=name,
            row_count=0,
            col_count=0,
            header_row=0,
            skip_reason=str(error),
        ),
    )

    planned = planner.plan(names, options=options)

    assert [item.state for item in planned] == [
        PlannedSheetState.READY,
        PlannedSheetState.READY,
    ]


def test_planning_failure_does_not_retain_analyzer_traceback_locals() -> None:
    class Payload:
        pass

    retained: weakref.ReferenceType[Payload] | None = None
    parse_plan = compile_parse_plan(SheetConfig(auto_detect=False), None, "xlsx")

    def fail(_name: str) -> SheetInfo:
        nonlocal retained
        payload = Payload()
        retained = weakref.ref(payload)
        raise ValueError("planning failed")

    planner = SheetPlanner(
        fail,
        lambda _name, _info: parse_plan,
        should_propagate=lambda _error: False,
        analysis_failure_info=lambda name, error: SheetInfo(
            name=name,
            row_count=0,
            col_count=0,
            header_row=0,
            skip_reason=str(error),
        ),
    )

    planned = planner.plan(["Data"])
    gc.collect()

    assert planned[0].state is PlannedSheetState.ERROR
    assert isinstance(planned[0].error, ValueError)
    assert str(planned[0].error) == "planning failed"
    assert retained is not None
    assert retained() is None
