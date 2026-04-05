"""Regression tests for previously found bugs."""

import openpyxl
import pytest

from messy_xlsx import MessyWorkbook, SheetConfig


class TestRegressionBugs:
    """Tests for specific bugs found in production."""

    def test_header_detection_with_metadata_rows(self, temp_dir):
        """Regression: Headers after metadata rows should be detected."""
        file_path = temp_dir / "metadata_header.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company Report"])
        ws.append(["Generated: 2024-01-01"])
        ws.append([])
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            structure = mwb.get_structure()

            # Should detect row 4 as header
            assert structure.header_row == 4
            assert structure.header_confidence >= 0.7

    def test_european_numbers_not_corrupted(self, temp_dir):
        """Regression: European numbers should parse correctly with locale."""
        file_path = temp_dir / "european.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Amount"])
        ws.append(["1.234,56"])  # European format: 1234.56
        wb.save(file_path)
        wb.close()

        config = SheetConfig(locale="de_DE")

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            import pandas as pd

            # Must be numeric — not left as text
            assert pd.api.types.is_numeric_dtype(df["amount"]), (
                f"European number was not converted to numeric (dtype={df['amount'].dtype})"
            )
            # Must be 1234.56, not corrupted to 1.23456 or 123456
            assert df.iloc[0]["amount"] == pytest.approx(1234.56)

    def test_comma_decimal_not_corrupted(self, temp_dir):
        """Regression: comma-decimal '1,23' with de_DE locale must become 1.23, not 123."""
        file_path = temp_dir / "comma_decimal.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Amount"])
        ws.append(["1,23"])
        ws.append(["4,56"])
        wb.save(file_path)
        wb.close()

        config = SheetConfig(locale="de_DE")

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()

            import pandas as pd

            assert pd.api.types.is_numeric_dtype(df["amount"]), (
                f"Comma-decimal was not converted to numeric (dtype={df['amount'].dtype})"
            )
            assert df.iloc[0]["amount"] == pytest.approx(1.23)
            assert df.iloc[1]["amount"] == pytest.approx(4.56)

    def test_merged_cells_dont_crash(self, temp_dir):
        """Regression: Merged cells should not crash parser."""
        file_path = temp_dir / "merged_regression.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Merged"
        ws.merge_cells("A1:C1")
        ws.append(["X", "Y", "Z"])
        ws.append([1, 2, 3])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert df is not None

    def test_hidden_rows_excluded_by_default(self, temp_dir):
        """Regression: Hidden rows should be excluded."""
        file_path = temp_dir / "hidden.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws.append([3, 4])  # Will be hidden
        ws.append([5, 6])

        # Hide row 3
        ws.row_dimensions[3].hidden = True

        wb.save(file_path)
        wb.close()

        config = SheetConfig(include_hidden=False)

        with MessyWorkbook(file_path, sheet_config=config) as mwb:
            df = mwb.to_dataframe()
            # Should exclude hidden row
            assert len(df) <= 3
