"""Contracts for shared multi-sheet planning and sequential sheet results."""

from __future__ import annotations

import gc
import inspect
import warnings
import weakref
from dataclasses import FrozenInstanceError
from pathlib import Path
from typing import Any, get_args

import openpyxl
import pandas as pd
import pytest

import messy_xlsx as api
import messy_xlsx.models as models_module
import messy_xlsx.multi_sheet as multi_sheet_module
import messy_xlsx.workbook as workbook_module
from messy_xlsx import (
    LegacyAPIWarning,
    MessyWorkbook,
    MultiSheetOptions,
    MultiSheetParser,
    SheetConfig,
    SheetInfo,
    SheetResult,
    read_all_sheets,
)
from messy_xlsx.parsing.contracts import ParseMetrics
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.parsing.streams import SheetStream


@pytest.fixture
def streaming_multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Create ordered data, empty, pivot-like, and data sheets."""
    path = tmp_path / "multi.xlsx"
    source = openpyxl.Workbook()

    first = source.active
    first.title = "First"
    first.append(["name", "value"])
    first.append(["a", 1])

    source.create_sheet("Empty")

    pivot = source.create_sheet("Pivot")
    pivot.append(["Row Labels", "Sum of value"])
    pivot.append(["a", 1])
    pivot.append(["Grand Total", 1])

    last = source.create_sheet("Last")
    last.append(["name", "value"])
    last.append(["b", 2])

    source.save(path)
    source.close()
    return path


def _replace_raw_materializer(
    monkeypatch: pytest.MonkeyPatch,
    workbook: MessyWorkbook,
    replacement: Any,
) -> Any:
    original = workbook._materialize_raw_frame
    monkeypatch.setattr(workbook, "_materialize_raw_frame", replacement)
    return original


def test_iter_sheets_preserves_order_every_sheet_and_result_invariant(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        expected = workbook.sheet_names
        with workbook.iter_sheets() as stream:
            assert iter(stream) is stream
            results = list(stream)

    assert [result.name for result in results] == expected
    assert all(isinstance(result, SheetResult) for result in results)
    assert all((result.dataframe is None) != (result.error is None) for result in results)
    empty = next(result for result in results if result.name == "Empty")
    assert empty.error is None
    assert empty.dataframe is not None
    assert empty.dataframe.empty


def test_sheet_result_remains_frozen_and_rejects_non_xor_payloads() -> None:
    frame = pd.DataFrame({"value": [1]})
    result = SheetResult(name="Data", dataframe=frame)

    with pytest.raises(FrozenInstanceError):
        result.name = "Changed"  # type: ignore[misc]
    with pytest.raises(ValueError, match="exactly one"):
        SheetResult(name="Data")
    with pytest.raises(ValueError, match="exactly one"):
        SheetResult(
            name="Data",
            dataframe=frame,
            error=workbook_module.SheetError(
                sheet_name="Data",
                error_type="ValueError",
                message="bad",
            ),
        )


def test_iter_sheets_emits_no_legacy_warning(streaming_multi_sheet_xlsx: Path) -> None:
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        with (
            MessyWorkbook(streaming_multi_sheet_xlsx) as workbook,
            workbook.iter_sheets() as stream,
        ):
            list(stream)

    assert not any(isinstance(item.message, LegacyAPIWarning) for item in caught)


def test_iter_sheets_context_and_early_close_release_the_active_operation(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        stream = workbook.iter_sheets()
        first = next(stream)
        assert first.name == "First"

        with pytest.raises(
            RuntimeError,
            match="MessyWorkbook already has an active parse or stream",
        ):
            workbook.to_arrow("Last")

        stream.close()
        stream.close()
        with pytest.raises(StopIteration):
            next(stream)

        table = workbook.to_arrow("Last", SheetConfig(auto_detect=False))
        assert table.num_rows == 1

        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as second:
            assert next(second).name == "First"
        with pytest.raises(StopIteration):
            next(second)


def test_parent_close_invalidates_an_active_sheet_stream(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    workbook = MessyWorkbook(streaming_multi_sheet_xlsx)
    stream = workbook.iter_sheets(SheetConfig(auto_detect=False))

    workbook.close()

    with pytest.raises(RuntimeError, match="MessyWorkbook is closed"):
        next(stream)
    stream.close()


def test_iter_sheets_converts_ordinary_error_and_continues(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        original = workbook._materialize_raw_frame

        def fail_first(sheet: str, format_type: str, plan: Any) -> pd.DataFrame:
            if sheet == "First":
                raise ValueError("ordinary sheet failure")
            return original(sheet, format_type, plan)

        _replace_raw_materializer(monkeypatch, workbook, fail_first)
        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream:
            results = list(stream)

        assert [result.name for result in results] == workbook.sheet_names
        failure = results[0]
        assert failure.dataframe is None
        assert failure.error is not None
        assert failure.error.sheet_name == "First"
        assert failure.error.error_type == "ValueError"
        assert failure.error.message == "ordinary sheet failure"
        assert results[-1].error is None

        monkeypatch.setattr(workbook, "_materialize_raw_frame", original)
        assert workbook.to_arrow("Last", SheetConfig(auto_detect=False)).num_rows == 1


@pytest.mark.parametrize("failure_type", [MemoryError, KeyboardInterrupt, SystemExit])
def test_iter_sheets_propagates_direct_process_failures_after_cleanup(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    failure_type: type[BaseException],
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        original = workbook._materialize_raw_frame

        def fail(_sheet: str, _format_type: str, _plan: Any) -> pd.DataFrame:
            raise failure_type("process failure")

        _replace_raw_materializer(monkeypatch, workbook, fail)
        stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
        with pytest.raises(failure_type, match="process failure"):
            next(stream)

        monkeypatch.setattr(workbook, "_materialize_raw_frame", original)
        assert workbook.to_arrow("Last", SheetConfig(auto_detect=False)).num_rows == 1


def test_iter_sheets_propagates_nested_process_failure_tree_after_cleanup(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        original = workbook._materialize_raw_frame

        def fail(_sheet: str, _format_type: str, _plan: Any) -> pd.DataFrame:
            raise BaseExceptionGroup(
                "nested",
                [RuntimeError("ordinary"), MemoryError("nested process failure")],
            )

        _replace_raw_materializer(monkeypatch, workbook, fail)
        stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
        with pytest.raises(BaseExceptionGroup, match="nested"):
            next(stream)

        monkeypatch.setattr(workbook, "_materialize_raw_frame", original)
        assert workbook.to_arrow("Last", SheetConfig(auto_detect=False)).num_rows == 1


@pytest.mark.parametrize("wrapped", [False, True])
def test_iter_sheets_propagates_reentrant_failure_instead_of_returning_error(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    wrapped: bool,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        original = workbook._materialize_raw_frame

        def reenter(_sheet: str, _format_type: str, _plan: Any) -> pd.DataFrame:
            if not wrapped:
                return workbook.to_arrow("Last").to_pandas()
            try:
                workbook.to_arrow("Last")
            except BaseException as cause:
                raise RuntimeError("wrapped reentrant failure") from cause
            raise AssertionError("reentrant operation unexpectedly succeeded")

        _replace_raw_materializer(monkeypatch, workbook, reenter)
        stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
        message = (
            "wrapped reentrant failure"
            if wrapped
            else "MessyWorkbook already has an active parse or stream"
        )
        with pytest.raises(RuntimeError, match=message):
            next(stream)

        monkeypatch.setattr(workbook, "_materialize_raw_frame", original)
        assert workbook.to_arrow("Last", SheetConfig(auto_detect=False)).num_rows == 1


def test_iter_sheets_freezes_config_before_return(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    config = SheetConfig(auto_detect=False, skip_rows=0)
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        stream = workbook.iter_sheets(config)
        config.skip_rows = 1
        with stream:
            first = next(stream)

    assert first.dataframe is not None
    assert list(first.dataframe.columns) == ["name", "value"]
    assert first.dataframe.to_dict("records") == [{"name": "a", "value": 1}]


def test_multi_sheet_metrics_are_cumulative_and_manifest_is_constructed_once(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    real_manifest_reader = workbook_module.ManifestReader
    constructions = 0

    def count_manifest(*args: Any, **kwargs: Any) -> Any:
        nonlocal constructions
        constructions += 1
        return real_manifest_reader(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "ManifestReader", count_manifest)

    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        metrics = workbook.parse_metrics
        assert metrics == ParseMetrics()
        with pytest.raises(AttributeError):
            workbook.parse_metrics = ParseMetrics()  # type: ignore[misc]

        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as first:
            first_results = list(first)
        assert all(result.error is None for result in first_results)
        assert constructions == 1
        assert metrics.manifest_builds == 1
        assert metrics.full_materializations == len(workbook.sheet_names)

        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as second:
            second_results = list(second)
        assert all(result.error is None for result in second_results)
        assert constructions == 1
        assert metrics.manifest_builds == 1
        assert metrics.full_materializations == 2 * len(workbook.sheet_names)


def test_failed_materialization_is_not_counted_as_success(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        original = workbook._materialize_raw_frame

        def fail_first(sheet: str, format_type: str, plan: Any) -> pd.DataFrame:
            if sheet == "First":
                raise ValueError("failed materialization")
            return original(sheet, format_type, plan)

        _replace_raw_materializer(monkeypatch, workbook, fail_first)
        with workbook.iter_sheets(
            SheetConfig(
                auto_detect=False,
                include_hidden=True,
                merge_strategy="skip",
            )
        ) as stream:
            results = list(stream)

        assert results[0].error is not None
        assert workbook.parse_metrics.full_materializations == len(workbook.sheet_names) - 1
        assert workbook.parse_metrics.failed_attempts == 1


def test_legacy_filtering_happens_before_any_full_materialization(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    materialized: list[str] = []
    filtered: list[str] = []
    original = MessyWorkbook._materialize_raw_frame

    def record(
        workbook: MessyWorkbook,
        sheet: str,
        format_type: str,
        plan: Any,
    ) -> pd.DataFrame:
        materialized.append(sheet)
        return original(workbook, sheet, format_type, plan)

    def eligible(info: Any) -> bool:
        filtered.append(info.name)
        return True

    monkeypatch.setattr(MessyWorkbook, "_materialize_raw_frame", record)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", LegacyAPIWarning)
        results = read_all_sheets(
            streaming_multi_sheet_xlsx,
            sheets=["Last"],
            sheet_filter=eligible,
        )

    assert list(results) == ["Last"]
    assert filtered == ["First", "Last"]
    assert materialized == ["Last"]


def test_sheet_info_identity_signature_and_mutability_remain_legacy_visible() -> None:
    assert api.SheetInfo is multi_sheet_module.SheetInfo is models_module.SheetInfo
    assert SheetInfo.__module__ == "messy_xlsx.multi_sheet"
    assert str(inspect.signature(SheetInfo)) == (
        "(name: str, row_count: int, col_count: int, header_row: int, "
        "is_empty: bool = False, is_pivot: bool = False, "
        "skip_reason: str | None = None) -> None"
    )
    assert "messy_xlsx.multi_sheet.SheetInfo" in str(inspect.signature(MultiSheetOptions))

    info = SheetInfo("Data", 2, 3, 0)
    info.col_count = 4
    assert info.column_count == 4


@pytest.mark.parametrize(
    ("requested", "expected"),
    [
        (["Last", "First"], ["First", "Last"]),
        ([], ["First", "Last"]),
        (["Unknown", "Last"], ["Last"]),
    ],
)
def test_legacy_explicit_selection_preserves_workbook_order_and_truthiness(
    streaming_multi_sheet_xlsx: Path,
    requested: list[str],
    expected: list[str],
) -> None:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", LegacyAPIWarning)
        result = read_all_sheets(streaming_multi_sheet_xlsx, sheets=requested)
    assert list(result) == expected


def test_legacy_filter_error_propagates_after_skip_checks(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    called: list[str] = []

    def fail(info: SheetInfo) -> bool:
        called.append(info.name)
        raise LookupError("filter failure")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", LegacyAPIWarning)
        with pytest.raises(LookupError, match="filter failure"):
            read_all_sheets(streaming_multi_sheet_xlsx, sheet_filter=fail)
    assert called == ["First"]


def test_parse_sheet_and_analysis_bypass_legacy_selection_filter(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    called: list[str] = []

    def reject(info: SheetInfo) -> bool:
        called.append(info.name)
        raise AssertionError("selection filter must not run")

    parser = MultiSheetParser(
        streaming_multi_sheet_xlsx,
        MultiSheetOptions(sheets=["Last"], sheet_filter=reject),
    )
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", LegacyAPIWarning)
        frame = parser.parse_sheet("First")
    infos = parser.analyze_sheets()

    assert frame.to_dict("records") == [{"name": "a", "value": 1}]
    assert [info.name for info in infos] == ["First", "Empty", "Pivot", "Last"]
    assert called == []


def test_planning_failure_policy_differs_by_adapter(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original = MessyWorkbook._get_sheet_manifest

    def fail_first(workbook: MessyWorkbook, name: str) -> Any:
        if name == "First":
            raise ValueError("bounded planning failure")
        return original(workbook, name)

    monkeypatch.setattr(MessyWorkbook, "_get_sheet_manifest", fail_first)

    with (
        MessyWorkbook(streaming_multi_sheet_xlsx) as workbook,
        workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream,
    ):
        streamed = list(stream)
    assert streamed[0].error is not None
    assert streamed[0].error.message == "bounded planning failure"
    assert streamed[-1].error is None

    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        frames, errors = workbook.to_dataframes(
            SheetConfig(auto_detect=False),
            include_errors=True,
        )
    assert "First" not in frames
    assert [error.sheet_name for error in errors] == ["First"]

    parser = MultiSheetParser(streaming_multi_sheet_xlsx)
    infos = parser.analyze_sheets()
    first = infos[0]
    assert first.skip_reason == "Parse error: bounded planning failure"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", LegacyAPIWarning)
        parsed = parser.parse_all()
    assert "First" not in parsed
    assert "Last" in parsed


def test_planning_process_failure_propagates_before_stream_return(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    original = MessyWorkbook._get_sheet_manifest

    def fail_first(workbook: MessyWorkbook, name: str) -> Any:
        if name == "First":
            raise MemoryError("planning process failure")
        return original(workbook, name)

    monkeypatch.setattr(MessyWorkbook, "_get_sheet_manifest", fail_first)
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        with pytest.raises(MemoryError, match="planning process failure"):
            workbook.iter_sheets(SheetConfig(auto_detect=False))
        monkeypatch.setattr(MessyWorkbook, "_get_sheet_manifest", original)
        assert (
            workbook.to_arrow(
                "Last",
                SheetConfig(auto_detect=False, include_hidden=True, merge_strategy="skip"),
            ).num_rows
            == 1
        )


def test_stream_creation_reserves_operation_and_deep_freezes_config(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    config = SheetConfig(
        auto_detect=False,
        column_renames={"name": "frozen_name"},
        drop_conditions=[{"column": "name", "value": "never"}],
    )
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        stream = workbook.iter_sheets(config)
        with pytest.raises(
            RuntimeError,
            match="MessyWorkbook already has an active parse or stream",
        ):
            workbook.to_arrow("Last")

        config.column_renames["name"] = "mutated_name"
        config.drop_conditions[0]["value"] = "a"
        with stream:
            first = next(stream)

    assert first.dataframe is not None
    assert list(first.dataframe.columns) == ["frozen_name", "value"]
    assert first.dataframe.to_dict("records") == [{"frozen_name": "a", "value": 1}]


def test_stream_creation_deep_freezes_workbook_default_config(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    default = SheetConfig(
        auto_detect=False,
        column_renames={"name": "default_name"},
        drop_conditions=[{"column": "name", "value": "never"}],
    )
    with MessyWorkbook(streaming_multi_sheet_xlsx, sheet_config=default) as workbook:
        stream = workbook.iter_sheets()
        default.column_renames["name"] = "mutated_name"
        default.drop_conditions[0]["value"] = "a"
        with stream:
            first = next(stream)

    assert first.dataframe is not None
    assert list(first.dataframe.columns) == ["default_name", "value"]
    assert first.dataframe.to_dict("records") == [{"default_name": "a", "value": 1}]


def test_sheet_stream_runtime_generic_is_sheet_result() -> None:
    assert get_args(SheetStream.__orig_bases__[0]) == (SheetResult,)


def test_early_close_materializes_no_sheet(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
        stream.close()
        assert workbook.parse_metrics.manifest_builds == 1
        assert workbook.parse_metrics.full_materializations == 0
        assert workbook.parse_metrics.failed_attempts == 0


def test_failed_manifest_construction_is_retryable_and_only_success_is_counted(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    real_manifest_reader = workbook_module.ManifestReader
    constructions = 0

    def fail_once(*args: Any, **kwargs: Any) -> Any:
        nonlocal constructions
        constructions += 1
        if constructions == 1:
            raise ValueError("manifest build failed")
        return real_manifest_reader(*args, **kwargs)

    monkeypatch.setattr(workbook_module, "ManifestReader", fail_once)
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream:
            results = list(stream)

        assert constructions == 2
        assert results[0].error is not None
        assert all(result.error is None for result in results[1:])
        assert workbook.parse_metrics.manifest_builds == 1
        assert workbook.parse_metrics.full_materializations == 3


def test_custom_registry_controls_order_and_results_without_manifest(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class SentinelRegistry(HandlerRegistry):
        def get_sheet_names(self, *_args: Any, **_kwargs: Any) -> list[str]:
            return ["Last", "First"]

        def parse(
            self,
            _source: Any,
            sheet: str | None,
            **_kwargs: Any,
        ) -> pd.DataFrame:
            return pd.DataFrame({"sentinel": [sheet]})

    monkeypatch.setattr(
        workbook_module,
        "ManifestReader",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("custom registries must not build OOXML manifests")
        ),
    )
    with MessyWorkbook(
        streaming_multi_sheet_xlsx,
        registry=SentinelRegistry(),
    ) as workbook:
        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream:
            results = list(stream)
        metrics = workbook.parse_metrics

    assert [result.name for result in results] == ["Last", "First"]
    assert [
        result.dataframe["sentinel"].iat[0]  # type: ignore[index]
        for result in results
    ] == ["Last", "First"]
    assert metrics.manifest_builds == 0
    assert metrics.full_materializations == 2


def test_xls_iter_sheets_never_constructs_ooxml_manifest(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    xlwt = pytest.importorskip("xlwt")
    path = tmp_path / "multi.xls"
    source = xlwt.Workbook()
    for name, value in (("Second", 2), ("First", 1)):
        sheet = source.add_sheet(name)
        sheet.write(0, 0, "name")
        sheet.write(0, 1, "value")
        sheet.write(1, 0, name.lower())
        sheet.write(1, 1, value)
    source.save(str(path))

    monkeypatch.setattr(
        workbook_module,
        "ManifestReader",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("XLS must not build an OOXML manifest")
        ),
    )
    with MessyWorkbook(path) as workbook:
        with workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream:
            results = list(stream)
        metrics = workbook.parse_metrics

    assert [result.name for result in results] == ["Second", "First"]
    assert all(result.error is None for result in results)
    assert metrics.manifest_builds == 0
    assert metrics.full_materializations == 2


def test_generator_releases_prior_frame_before_materializing_next_sheet(
    streaming_multi_sheet_xlsx: Path,
) -> None:
    previous: list[weakref.ReferenceType[pd.DataFrame]] = []

    class ReleasingRegistry(HandlerRegistry):
        def parse(
            self,
            _source: Any,
            sheet: str | None,
            **_kwargs: Any,
        ) -> pd.DataFrame:
            if sheet == "Empty":
                gc.collect()
                assert previous[0]() is None
            return pd.DataFrame({"sheet": [sheet]})

    with (
        MessyWorkbook(
            streaming_multi_sheet_xlsx,
            registry=ReleasingRegistry(),
        ) as workbook,
        workbook.iter_sheets(SheetConfig(auto_detect=False)) as stream,
    ):
        first = next(stream)
        assert first.dataframe is not None
        previous.append(weakref.ref(first.dataframe))
        del first
        second = next(stream)

    assert second.name == "Empty"


def test_sheet_stream_process_cleanup_wins_over_reentrant_materialization_error(
    streaming_multi_sheet_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with MessyWorkbook(streaming_multi_sheet_xlsx) as workbook:
        stream = workbook.iter_sheets(SheetConfig(auto_detect=False))
        original_release = stream._close_callback

        def fail_materialization(_name: str, _plan: Any) -> pd.DataFrame:
            return workbook.to_arrow("Last").to_pandas()

        def fail_release() -> None:
            raise MemoryError("process cleanup failure")

        monkeypatch.setattr(workbook, "_materialize_compiled_plan", fail_materialization)
        stream._close_callback = fail_release
        try:
            with pytest.raises(MemoryError, match="process cleanup failure") as raised:
                next(stream)
        finally:
            stream._close_callback = original_release
            stream.close()
        assert raised.value.__dict__["backend_context"]["operation_failure"] == {
            "type": "_ActiveOperationError"
        }
