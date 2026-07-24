"""Public Arrow, batch, and pandas-chunk API contracts."""

from __future__ import annotations

import gc
import inspect
import io
import sys
import warnings
import weakref
from dataclasses import FrozenInstanceError
from datetime import UTC, date, datetime, time, timedelta, timezone, tzinfo
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from typing import Any, BinaryIO
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import pyarrow as pa
import pytest
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

import messy_xlsx
import messy_xlsx.parsing.materialized_streaming as materialized_streaming_module
import messy_xlsx.parsing.physical_values as physical_values_module
import messy_xlsx.parsing.xlsx_streaming as xlsx_streaming_module
import messy_xlsx.workbook as workbook_module
from messy_xlsx.models import FormatInfo, SheetConfig, SheetError
from messy_xlsx.parsing.coordinates import CoordinateOperation
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.parsing.parse_plan import compile_parse_plan
from messy_xlsx.warnings import LegacyAPIWarning


def _write_book(
    path: Path,
    rows: list[list[object | None]],
    *,
    title: str = "Data",
) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _public(name: str) -> Any:
    return getattr(messy_xlsx, name)


def _assert_parameters(
    callable_object: Any,
    expected: list[tuple[str, object]],
) -> inspect.Signature:
    signature = inspect.signature(callable_object)
    assert [
        (name, parameter.default) for name, parameter in signature.parameters.items()
    ] == expected
    assert all(
        parameter.kind is inspect.Parameter.POSITIONAL_OR_KEYWORD
        for parameter in signature.parameters.values()
    )
    return signature


class _InjectedReader:
    def __init__(
        self,
        *,
        batches: list[pa.RecordBatch] | None = None,
        read_error: BaseException | None = None,
        close_error: BaseException | None = None,
    ) -> None:
        self.schema = batches[0].schema if batches else pa.schema([pa.field("value", pa.int64())])
        self._batches = iter(batches or [])
        self._read_error = read_error
        self._close_error = close_error
        self.close_calls = 0

    def read_next_batch(self) -> pa.RecordBatch | None:
        if self._read_error is not None:
            error = self._read_error
            self._read_error = None
            raise error
        return next(self._batches, None)

    def close(self) -> None:
        self.close_calls += 1
        if self._close_error is not None:
            raise self._close_error


def test_new_public_exports_and_exact_signatures_are_complete() -> None:
    expected = {
        "BatchStream",
        "DataFrameChunkStream",
        "SheetStream",
        "SheetResult",
        "LegacyAPIWarning",
        "StreamingTypeError",
        "read_excel_arrow",
        "read_excel_batches",
    }
    assert expected <= set(messy_xlsx.__all__)

    workbook_type = messy_xlsx.MessyWorkbook
    to_arrow = _assert_parameters(
        workbook_type.to_arrow,
        [("self", inspect.Parameter.empty), ("sheet", None), ("config", None)],
    )
    assert to_arrow.return_annotation is pa.Table

    iter_batches = _assert_parameters(
        workbook_type.iter_batches,
        [
            ("self", inspect.Parameter.empty),
            ("sheet", None),
            ("batch_size", 65_536),
            ("config", None),
        ],
    )
    assert iter_batches.return_annotation is _public("BatchStream")

    dataframe_chunks = _assert_parameters(
        workbook_type.iter_dataframe_chunks,
        [
            ("self", inspect.Parameter.empty),
            ("sheet", None),
            ("batch_size", 65_536),
            ("config", None),
        ],
    )
    assert dataframe_chunks.return_annotation is _public("DataFrameChunkStream")

    read_arrow = _assert_parameters(
        _public("read_excel_arrow"),
        [
            ("file_path_or_buffer", inspect.Parameter.empty),
            ("sheet", None),
            ("config", None),
            ("filename", None),
        ],
    )
    assert read_arrow.return_annotation is pa.Table

    read_batches = _assert_parameters(
        _public("read_excel_batches"),
        [
            ("file_path_or_buffer", inspect.Parameter.empty),
            ("sheet", None),
            ("batch_size", 65_536),
            ("config", None),
            ("filename", None),
        ],
    )
    assert read_batches.return_annotation is _public("BatchStream")


def test_top_level_source_annotations_cover_paths_and_binary_buffers() -> None:
    read_arrow = inspect.signature(_public("read_excel_arrow"))
    read_batches = inspect.signature(_public("read_excel_batches"))
    expected = str | Path | BinaryIO
    assert read_arrow.parameters["file_path_or_buffer"].annotation == expected
    assert read_batches.parameters["file_path_or_buffer"].annotation == expected


def test_sheet_result_is_frozen_and_enforces_xor() -> None:
    result_type = _public("SheetResult")
    frame = pd.DataFrame({"value": [1]})
    success = result_type(name="Data", dataframe=frame)
    failure = result_type(
        name="Data",
        error=SheetError("Data", "ValueError", "bad"),
    )

    assert success.dataframe is frame and success.error is None
    assert failure.dataframe is None and failure.error is not None
    with pytest.raises(FrozenInstanceError):
        success.name = "Other"
    with pytest.raises(ValueError, match="exactly one"):
        result_type(name="Data")
    with pytest.raises(ValueError, match="exactly one"):
        result_type(name="Data", dataframe=frame, error=failure.error)


def test_every_new_api_emits_no_legacy_warning(sample_xlsx: Path) -> None:
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
            workbook.to_arrow()
            workbook.iter_batches().close()
            workbook.iter_dataframe_chunks().close()
        _public("read_excel_arrow")(sample_xlsx)
        _public("read_excel_batches")(sample_xlsx).close()

    assert not [item for item in captured if issubclass(item.category, LegacyAPIWarning)]


def test_to_arrow_uses_global_materialized_contract_and_drops_null_columns(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "null-column.xlsx",
        [["Name", "Always Null"], [" Alice ", None], ["Bob", None]],
    )
    with messy_xlsx.MessyWorkbook(path) as workbook:
        table = workbook.to_arrow("Data", SheetConfig(auto_detect=False))

    assert isinstance(table, pa.Table)
    assert table.column_names == ["name"]
    assert table.column(0).to_pylist() == ["Alice", "Bob"]


def test_to_arrow_matches_materialized_dataframe_values_labels_and_order(
    sample_xlsx: Path,
) -> None:
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", LegacyAPIWarning)
            expected = workbook.to_dataframe()
        table = workbook.to_arrow()

    pd.testing.assert_frame_equal(table.to_pandas(), expected)


def test_to_arrow_empty_sheet_returns_empty_zero_column_table(tmp_path: Path) -> None:
    path = _write_book(tmp_path / "empty.xlsx", [])
    with messy_xlsx.MessyWorkbook(path) as workbook:
        table = workbook.to_arrow()
    assert table.num_rows == 0
    assert table.num_columns == 0


def test_read_excel_arrow_restores_caller_cursor_and_leaves_path_owned_by_caller(
    sample_xlsx: Path,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(9)
    table = _public("read_excel_arrow")(
        source,
        filename=sample_xlsx.name,
    )

    assert table.num_rows == 3
    assert source.tell() == 9
    assert not source.closed
    assert sample_xlsx.exists()
    assert sample_xlsx.read_bytes()[:4] == b"PK\x03\x04"


def test_iter_batches_has_stable_public_schema_and_post_filter_batch_bound(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "filtered.xlsx",
        [
            ["Name", "Amount", "Declared Null"],
            ["drop", "1", None],
            ["a", "2", None],
            ["drop", "3", None],
            ["b", "4", None],
            ["c", "5", None],
        ],
    )
    config = SheetConfig(drop_regex="^drop$", sanitize_column_names=False)
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches("Data", batch_size=2, config=config) as stream,
    ):
        schema = stream.schema
        batches = list(stream)

    assert schema.names == ["Name", "Amount", "Declared Null"]
    assert batches
    assert all(batch.schema == schema for batch in batches)
    assert all(0 < batch.num_rows <= 2 for batch in batches)
    assert sum(batch.num_rows for batch in batches) == 3
    assert pa.types.is_null(schema.field(2).type)


def test_iter_batches_normalize_false_preserves_native_scalar_types(
    tmp_path: Path,
) -> None:
    first_timestamp = datetime(2025, 1, 2, 3, 4)
    second_timestamp = datetime(2025, 1, 3, 4, 5)
    path = _write_book(
        tmp_path / "native-types.xlsx",
        [
            ["Count", "Flag", "When"],
            [1, True, first_timestamp],
            [2, False, second_timestamp],
        ],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.int64(), pa.bool_(), pa.timestamp("us")]
    assert table.to_pydict() == {
        "Count": [1, 2],
        "Flag": [True, False],
        "When": [first_timestamp, second_timestamp],
    }


def test_empty_batch_stream_exposes_schema_and_is_one_shot(tmp_path: Path) -> None:
    path = _write_book(tmp_path / "headers-only.xlsx", [["Name", "Value"]])
    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(batch_size=2)
        assert iter(stream) is stream
        assert stream.schema.names == ["name", "value"]
        assert list(stream) == []
        assert list(stream) == []


def test_late_incompatible_value_raises_before_failing_batch_and_releases(
    tmp_path: Path,
) -> None:
    rows: list[list[object | None]] = [["Amount"]]
    rows.extend([str(index)] for index in range(1_000))
    rows.append(["not-a-number"])
    path = _write_book(tmp_path / "late.xlsx", rows)

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(batch_size=500)
        assert next(stream).num_rows == 500
        assert next(stream).num_rows == 500
        with pytest.raises(_public("StreamingTypeError")):
            next(stream)
        with pytest.raises(StopIteration):
            next(stream)
        assert workbook._active_operation_token is None


@pytest.mark.parametrize("batch_size", [True, 0, -1, 1.5, "2", None])
def test_iter_batches_rejects_invalid_batch_size_before_return(
    sample_xlsx: Path,
    batch_size: object,
) -> None:
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        with pytest.raises((TypeError, ValueError), match="batch_size"):
            workbook.iter_batches(batch_size=batch_size)  # type: ignore[arg-type]
        assert workbook._active_operation_token is None


def test_stream_preflight_rejects_missing_sheet_and_invalid_config(
    sample_xlsx: Path,
) -> None:
    invalid = SheetConfig()
    invalid.skip_rows = True  # type: ignore[assignment]
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        with pytest.raises(messy_xlsx.FormatError, match="not found"):
            workbook.iter_batches("Missing")
        with pytest.raises(TypeError, match="skip_rows"):
            workbook.iter_batches(config=invalid)
        assert workbook._active_operation_token is None


def test_stream_preflight_sampling_and_schema_failures_release_reservation(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        failure = ValueError("sample compilation failed")

        def fail(*_args: object, **_kwargs: object) -> object:
            raise failure

        monkeypatch.setattr(workbook, "_prepare_streaming_operation", fail, raising=False)
        with pytest.raises(ValueError) as captured:
            workbook.iter_batches()
        assert captured.value is failure
        assert workbook._active_operation_token is None


def test_distinct_sample_failure_happens_before_full_reader_and_restores_cursor(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(11)
    failure = ValueError("bounded sample failed")
    with messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name) as workbook:
        monkeypatch.setattr(
            workbook,
            "_produce_normalization_sample",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(failure),
        )
        monkeypatch.setattr(
            workbook_module,
            "OpenpyxlStreamingReader",
            lambda *_args, **_kwargs: pytest.fail("full reader opened after sample failure"),
        )
        with pytest.raises(ValueError) as captured:
            workbook.iter_batches()

        assert captured.value is failure
        assert source.tell() == 11
        assert workbook._active_operation_token is None


def test_schema_compilation_failure_happens_before_full_reader_and_releases(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    failure = ValueError("stable schema failed")
    compile_calls = 0
    real_compile = workbook_module.compile_normalization_plan

    def fail_schema(*args: object, **kwargs: object) -> object:
        nonlocal compile_calls
        compile_calls += 1
        if compile_calls == 1:
            raise failure
        return real_compile(*args, **kwargs)

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(workbook_module, "compile_normalization_plan", fail_schema)
        monkeypatch.setattr(
            workbook_module,
            "OpenpyxlStreamingReader",
            lambda *_args, **_kwargs: pytest.fail("full reader opened after schema failure"),
        )
        with pytest.raises(ValueError) as captured:
            workbook.iter_batches()

        assert captured.value is failure
        assert compile_calls == 1
        assert workbook._active_operation_token is None


@pytest.mark.parametrize(
    ("close_error", "expected_type"),
    [(OSError("ordinary cleanup"), ValueError), (MemoryError("process cleanup"), MemoryError)],
)
def test_iteration_failure_cleanup_precedence_and_reservation_release(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    close_error: BaseException,
    expected_type: type[BaseException],
) -> None:
    primary = ValueError("iteration")
    reader = _InjectedReader(read_error=primary, close_error=close_error)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )
        stream = workbook.iter_batches()
        token = workbook._active_operation_token
        with pytest.raises(expected_type) as captured:
            next(stream)

        if expected_type is ValueError:
            assert captured.value is primary
        else:
            assert captured.value is close_error
        assert reader.close_calls == 1
        if expected_type is MemoryError:
            assert workbook._active_operation_token is token
            assert workbook._active_stream is stream
            reader._close_error = None
            stream.close()
            assert reader.close_calls == 2
        assert workbook._active_operation_token is None


def test_explicit_and_context_close_attempt_cleanup_and_release_reservation(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    close_error = OSError("ordinary cleanup")
    reader = _InjectedReader(close_error=close_error)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )
        stream = workbook.iter_batches()
        with pytest.raises(OSError) as captured:
            stream.close()
        assert captured.value is close_error
        assert workbook._active_operation_token is None

    body_error = ValueError("context body")
    reader = _InjectedReader(close_error=OSError("context cleanup"))
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )
        stream = workbook.iter_batches()
        with pytest.raises(ValueError) as captured, stream:
            raise body_error
        assert captured.value is body_error
        assert workbook._active_operation_token is None


@pytest.mark.parametrize("termination", ["exhaustion", "early", "context-error"])
def test_workbook_batch_stream_restores_caller_cursor_on_terminal_paths(
    sample_xlsx: Path,
    termination: str,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(7)
    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    stream = workbook.iter_batches(batch_size=2)

    if termination == "exhaustion":
        list(stream)
    elif termination == "early":
        next(stream)
        stream.close()
    else:
        body_error = ValueError("body")
        with pytest.raises(ValueError) as captured, stream:
            next(stream)
            raise body_error
        assert captured.value is body_error

    assert source.tell() == 7
    assert not source.closed
    assert workbook._active_operation_token is None
    workbook.close()


def test_parent_close_invalidates_batch_stream_and_restores_buffer(
    sample_xlsx: Path,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(5)
    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    stream = workbook.iter_batches(batch_size=2)
    workbook.close()

    assert source.tell() == 5
    assert not source.closed
    with pytest.raises(RuntimeError, match="MessyWorkbook is closed"):
        next(stream)


def test_active_operation_and_reentrancy_reject_then_release(sample_xlsx: Path) -> None:
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        first = workbook.iter_batches(batch_size=2)
        with pytest.raises(RuntimeError, match="active parse or stream"):
            workbook.iter_batches()
        with pytest.raises(RuntimeError, match="active parse or stream"):
            workbook.to_arrow()
        first.close()

        second = workbook.iter_batches(batch_size=2)
        second.close()
        assert workbook._active_operation_token is None


def test_top_level_batch_stream_owns_workbook_without_closing_caller_buffer(
    sample_xlsx: Path,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(7)
    with _public("read_excel_batches")(
        source,
        filename=sample_xlsx.name,
        batch_size=2,
    ) as batches:
        next(batches)
        assert source.tell() != 7
    assert source.tell() == 7
    assert not source.closed


def test_top_level_batch_construction_failure_closes_owned_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    failure = ValueError("construction")
    instances: list[Any] = []

    class FailingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            self.closed = False
            instances.append(self)

        def iter_batches(self, *_args: object, **_kwargs: object) -> object:
            raise failure

        def close(self) -> None:
            self.closed = True

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", FailingWorkbook)
    with pytest.raises(ValueError) as captured:
        _public("read_excel_batches")("book.xlsx")
    assert captured.value is failure
    assert len(instances) == 1 and instances[0].closed


def test_top_level_batch_cleanup_process_failure_wins(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    child_error = OSError("child cleanup")
    workbook_error = MemoryError("workbook cleanup")

    class Child:
        schema = pa.schema([])

        def __iter__(self) -> Child:
            return self

        def __next__(self) -> pa.RecordBatch:
            raise StopIteration

        def close(self) -> None:
            raise child_error

    class FailingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            return None

        def iter_batches(self, *_args: object, **_kwargs: object) -> Child:
            return Child()

        def close(self) -> None:
            raise workbook_error

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", FailingWorkbook)
    stream = _public("read_excel_batches")(sample_xlsx)
    with pytest.raises(MemoryError) as captured:
        stream.close()
    assert captured.value is workbook_error
    assert workbook_error.__dict__["backend_context"]["operation_failure"] == {"type": "OSError"}


@pytest.mark.parametrize("termination", ["exhaustion", "iteration-error"])
def test_top_level_batch_stream_closes_owned_workbook_on_all_terminal_paths(
    monkeypatch: pytest.MonkeyPatch,
    termination: str,
) -> None:
    instances: list[Any] = []
    iteration_error = ValueError("top-level iteration")

    class Child:
        schema = pa.schema([pa.field("value", pa.int64())])

        def __init__(self) -> None:
            self.closed = False

        def __iter__(self) -> Child:
            return self

        def __next__(self) -> pa.RecordBatch:
            if termination == "iteration-error":
                raise iteration_error
            raise StopIteration

        def close(self) -> None:
            self.closed = True

    class TrackingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            self.child = Child()
            self.closed = False
            instances.append(self)

        def iter_batches(self, *_args: object, **_kwargs: object) -> Child:
            return self.child

        def close(self) -> None:
            self.closed = True

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", TrackingWorkbook)
    stream = _public("read_excel_batches")("book.xlsx")
    if termination == "iteration-error":
        with pytest.raises(ValueError) as captured:
            next(stream)
        assert captured.value is iteration_error
    else:
        assert list(stream) == []

    assert len(instances) == 1
    assert instances[0].child.closed
    assert instances[0].closed


def test_dataframe_chunks_have_global_range_index_labels_and_arrow_dtypes(
    sample_xlsx: Path,
) -> None:
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx) as workbook,
        workbook.iter_dataframe_chunks("Data", batch_size=2) as chunks,
    ):
        frames = list(chunks)

    combined = pd.concat(frames)
    assert combined.index.tolist() == list(range(len(combined)))
    assert list(combined.columns) == ["name", "age", "city"]
    assert str(combined.dtypes["age"]).endswith("[pyarrow]")


def test_dataframe_chunks_restore_exact_duplicate_labels_and_close_early(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "duplicates.xlsx",
        [["Value", "Value"], ["a", "b"], ["c", "d"]],
    )
    config = SheetConfig(sanitize_column_names=False)
    with messy_xlsx.MessyWorkbook(path) as workbook:
        chunks = workbook.iter_dataframe_chunks(batch_size=1, config=config)
        frame = next(chunks)
        assert list(frame.columns) == ["Value", "Value"]
        chunks.close()
        assert workbook._active_operation_token is None


def test_empty_dataframe_chunk_stream_yields_no_frames(tmp_path: Path) -> None:
    path = _write_book(tmp_path / "headers.xlsx", [["Value"]])
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_dataframe_chunks() as chunks,
    ):
        assert list(chunks) == []


def test_dataframe_chunk_conversion_failure_closes_child_and_releases(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    conversion_error = ValueError("pandas conversion")

    def fail_types_mapper(_data_type: object) -> object:
        raise conversion_error

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        chunks = workbook.iter_dataframe_chunks(batch_size=2)
        monkeypatch.setattr(workbook_module.pd, "ArrowDtype", fail_types_mapper)
        with pytest.raises(ValueError) as captured:
            next(chunks)

        assert captured.value is conversion_error
        assert workbook._active_operation_token is None


def test_sheet_config_evaluate_formulas_selects_expression_or_cached_value(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "formula.xlsx",
        [["ID", "Formula"], [1, "=1+1"]],
    )
    expression_config = SheetConfig(evaluate_formulas=False, sanitize_column_names=False)
    cached_config = SheetConfig(evaluate_formulas=True, sanitize_column_names=False)
    with messy_xlsx.MessyWorkbook(path) as workbook:
        with workbook.iter_batches(config=expression_config) as expression_stream:
            expression = pa.Table.from_batches(list(expression_stream))
        with workbook.iter_batches(config=cached_config) as cached_stream:
            cached = pa.Table.from_batches(list(cached_stream))

    assert expression.column("Formula").to_pylist() == ["=1+1"]
    assert cached.column("Formula").to_pylist() == [None]


def test_custom_registry_parse_remains_authoritative_for_batch_api(
    sample_xlsx: Path,
) -> None:
    class CustomRegistry(HandlerRegistry):
        calls = 0

        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            self.calls += 1
            return pd.DataFrame([["custom", 9]], columns=["label", "value"])

    registry = CustomRegistry()
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=registry) as workbook,
        workbook.iter_batches(batch_size=1) as stream,
    ):
        table = pa.Table.from_batches(list(stream))

    assert registry.calls == 1
    assert table.to_pydict() == {"label": ["custom"], "value": [9]}


def test_custom_registry_detection_validation_sheet_and_parse_overrides_are_not_bypassed(
    sample_xlsx: Path,
) -> None:
    class FullyCustomRegistry(HandlerRegistry):
        def __init__(self) -> None:
            super().__init__()
            self.events: list[str] = []

        def detect_format(self, *args: object, **kwargs: object) -> object:
            self.events.append("detect")
            return super().detect_format(*args, **kwargs)  # type: ignore[arg-type]

        def get_sheet_names(self, *args: object, **kwargs: object) -> list[str]:
            self.events.append("sheets")
            return super().get_sheet_names(*args, **kwargs)  # type: ignore[arg-type]

        def validate(self, *args: object, **kwargs: object) -> tuple[bool, str | None]:
            self.events.append("validate")
            return super().validate(*args, **kwargs)  # type: ignore[arg-type]

        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            self.events.append("parse")
            return pd.DataFrame({"source": ["custom"]})

    registry = FullyCustomRegistry()
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=registry) as workbook,
        workbook.iter_batches(batch_size=1) as stream,
    ):
        table = pa.Table.from_batches(list(stream))

    assert registry.events.count("detect") == 1
    assert registry.events.count("sheets") == 1
    assert registry.events.count("validate") == 1
    assert registry.events.count("parse") == 1
    assert table.to_pydict() == {"source": ["custom"]}


def test_csv_batch_api_uses_supported_format_route_without_warning() -> None:
    source = io.BytesIO(b"name,value\na,1\nb,2\n")
    source.seek(3)
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        with _public("read_excel_batches")(
            source,
            filename="data.csv",
            batch_size=1,
        ) as stream:
            table = pa.Table.from_batches(list(stream))

    assert table.column_names == ["name", "value"]
    assert table.num_rows == 2
    assert source.tell() == 3
    assert not source.closed
    assert not [item for item in captured if issubclass(item.category, LegacyAPIWarning)]


# Task 12 review remediation A: bounded retained-output sampling.


def test_large_projection_sample_never_terminalizes_the_truncated_coordinate_operation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_book(
        tmp_path / "large-projection.xlsx",
        [["Value"], [1]],
    )
    config = SheetConfig(
        auto_detect=False,
        cell_range="A1:A100000",
        sanitize_column_names=False,
    )
    synthesized_spans: list[int] = []
    captured_samples: list[Any] = []
    real_null_range_batch = CoordinateOperation._null_range_batch
    real_compile = workbook_module.compile_normalization_plan

    def reject_synthesized_remainder(
        operation: CoordinateOperation,
        start_row: int,
        end_row: int,
    ) -> object:
        synthesized_spans.append(end_row - start_row + 1)
        return real_null_range_batch(operation, start_row, end_row)

    def capture_sample(sample: object, plan: object) -> object:
        captured_samples.append(sample)
        return real_compile(sample, plan)  # type: ignore[arg-type]

    monkeypatch.setattr(
        CoordinateOperation,
        "_null_range_batch",
        reject_synthesized_remainder,
    )
    monkeypatch.setattr(workbook_module, "compile_normalization_plan", capture_sample)

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        stream.close()

    assert synthesized_spans == []
    assert len(captured_samples) == 1
    sample = captured_samples[0]
    assert sample.row_count <= 1_000
    assert sample.row_count * len(sample.columns) <= 1_000_000


def test_sample_locates_retained_evidence_after_large_skip_rows(tmp_path: Path) -> None:
    rows: list[list[object | None]] = [["ignored"] for _ in range(2_500)]
    rows.extend([["Amount"], [1], [2]])
    path = _write_book(tmp_path / "large-skip-rows.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        skip_rows=2_500,
        header_rows=1,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=128, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column_names == ["Amount"]
    assert table.schema.types == [pa.int64()]
    assert table.column(0).to_pylist() == [1, 2]


def test_sample_locates_retained_evidence_before_large_skip_footer(tmp_path: Path) -> None:
    rows: list[list[object | None]] = [["Amount"], [1], [2]]
    rows.extend([["footer"] for _ in range(2_501)])
    path = _write_book(tmp_path / "large-skip-footer.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        header_rows=1,
        skip_footer=2_501,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=128, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column_names == ["Amount"]
    assert table.schema.types == [pa.int64()]
    assert table.column(0).to_pylist() == [1, 2]


# Task 12 review remediation B: nonempty zero-column row counts.


def test_to_arrow_preserves_nonempty_zero_column_row_count(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    frame = pd.DataFrame(index=pd.RangeIndex(5))
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.num_columns == 0
    assert table.num_rows == 5


def test_zero_column_batches_and_dataframe_chunks_preserve_rows_and_global_index(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "hidden-only-column.xlsx",
        [[1], [2], [3], [4], [5]],
    )
    physical = openpyxl.load_workbook(path)
    physical["Data"].column_dimensions["A"].hidden = True
    physical.save(path)
    physical.close()
    config = SheetConfig(
        auto_detect=False,
        header_rows=0,
        normalize=False,
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        with workbook.iter_batches(batch_size=2, config=config) as stream:
            batches = list(stream)
        with workbook.iter_dataframe_chunks(batch_size=2, config=config) as chunks:
            frames = list(chunks)

    assert [batch.num_rows for batch in batches] == [2, 2, 1]
    assert all(batch.num_columns == 0 for batch in batches)
    assert [frame.shape for frame in frames] == [(2, 0), (2, 0), (1, 0)]
    assert [index for frame in frames for index in frame.index] == list(range(5))


# Task 12 review remediation C: deterministic mixed materialized columns.


def test_to_arrow_uses_dense_union_for_mixed_values_and_keeps_positional_labels(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    frame = pd.DataFrame(
        [
            [1, "left"],
            ["two", "right"],
        ],
    )
    frame.columns = [7, 7]

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column_names == ["7", "7"]
    assert pa.types.is_union(table.column(0).type)
    assert table.column(0).type.mode == "dense"
    assert table.column(0).to_pylist() == [1, "two"]
    assert table.column(1).to_pylist() == ["left", "right"]


# Task 12 review remediation D: sparse hints and raw OOXML provenance.


def test_explicit_hint_accepts_compatible_value_after_all_null_sample(
    tmp_path: Path,
) -> None:
    rows: list[list[object | None]] = [["Amount"]]
    rows.extend([[None] for _ in range(2_100)])
    rows.append([5])
    path = _write_book(tmp_path / "sparse-hinted.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        type_hints={"Amount": "INTEGER"},
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=256, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.int64()]
    assert table.column(0).to_pylist() == [5]


def test_normalize_false_never_lexically_casts_late_original_text(
    tmp_path: Path,
) -> None:
    rows: list[list[object | None]] = [["Amount"]]
    rows.extend([[index] for index in range(2_100)])
    rows.append(["1005"])
    path = _write_book(tmp_path / "late-original-text.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(batch_size=500, config=config)
        with pytest.raises(messy_xlsx.StreamingTypeError) as captured:
            list(stream)

    assert captured.value.context["row_offset"] == 2_100
    assert captured.value.context["value_description"] == "str(length=4)"


# Task 12 review remediation E: staged compatibility streams retain null columns.


def test_csv_compatibility_stream_retains_declared_all_null_column() -> None:
    source = io.BytesIO(b"a,null\n1,\n2,\n")
    with _public("read_excel_batches")(
        source,
        filename="all-null.csv",
        batch_size=1,
        config=SheetConfig(sanitize_column_names=False),
    ) as stream:
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column_names == ["a", "null"]
    assert table.to_pydict() == {"a": [1, 2], "null": [None, None]}


def test_custom_compatibility_stream_retains_null_column_order_and_filters(
    sample_xlsx: Path,
) -> None:
    class CustomRegistry(HandlerRegistry):
        calls = 0

        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            self.calls += 1
            return pd.DataFrame(
                {
                    "status": ["keep", "drop"],
                    "declared_null": [None, None],
                    "value": [1, 2],
                }
            )

    registry = CustomRegistry()
    config = SheetConfig(
        auto_detect=False,
        drop_regex="^drop$",
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=registry) as workbook,
        workbook.iter_batches(batch_size=1, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert registry.calls == 1
    assert table.column_names == ["status", "declared_null", "value"]
    assert table.to_pydict() == {
        "status": ["keep"],
        "declared_null": [None],
        "value": [1],
    }


# Task 12 review remediation F: transactional construction and sticky ownership.


def test_iter_batches_owner_allocation_failure_precedes_reader_construction(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    prepare_calls = 0

    def prepare(*_args: object, **_kwargs: object) -> SimpleNamespace:
        nonlocal prepare_calls
        prepare_calls += 1
        return SimpleNamespace(
            reader=reader,
            display_names=("value",),
        )

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(workbook, "_prepare_streaming_operation", prepare)

        def fail_owner() -> object:
            raise MemoryError("owner allocation")

        monkeypatch.setattr(workbook_module, "_CloseOnceOwner", fail_owner)
        with pytest.raises(MemoryError, match="owner allocation"):
            workbook.iter_batches()

        assert prepare_calls == 0
        assert reader.close_calls == 0
        assert workbook._active_operation_token is None


@pytest.mark.parametrize(
    "failure",
    [ValueError("dataframe wrapper"), MemoryError("dataframe wrapper process")],
)
def test_dataframe_wrapper_construction_failure_rolls_back_exact_child_token(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    failure: BaseException,
) -> None:
    reader = _InjectedReader()
    released: list[object] = []
    created: list[object] = []
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        real_end = workbook._end_operation

        def record_release(token: object) -> None:
            released.append(token)
            real_end(token)

        def fail_wrapper(*_args: object, **_kwargs: object) -> object:
            token = workbook._active_operation_token
            assert token is not None
            created.append(token)
            raise failure

        monkeypatch.setattr(workbook, "_end_operation", record_release)
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )
        monkeypatch.setattr(workbook_module, "DataFrameChunkStream", fail_wrapper)

        with pytest.raises(type(failure)) as captured:
            workbook.iter_dataframe_chunks()

        assert captured.value is failure
        assert len(created) == 1
        assert len(released) == 1 and released[0] is created[0]
        assert reader.close_calls == 1
        assert workbook._active_operation_token is None


@pytest.mark.parametrize(
    "failure",
    [ValueError("outer wrapper"), MemoryError("outer wrapper process")],
)
def test_top_level_wrapper_construction_failure_closes_child_and_workbook_once(
    monkeypatch: pytest.MonkeyPatch,
    failure: BaseException,
) -> None:
    instances: list[Any] = []

    class Child:
        schema = pa.schema([pa.field("value", pa.int64())])
        _display_names = ("value",)

        def __init__(self) -> None:
            self.close_calls = 0

        def __iter__(self) -> Child:
            return self

        def __next__(self) -> pa.RecordBatch:
            raise StopIteration

        def close(self) -> None:
            self.close_calls += 1

    class TrackingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            self.child = Child()
            self.close_calls = 0
            instances.append(self)

        def iter_batches(self, *_args: object, **_kwargs: object) -> Child:
            return self.child

        def close(self) -> None:
            self.close_calls += 1

    def fail_wrapper(*_args: object, **_kwargs: object) -> object:
        raise failure

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", TrackingWorkbook)
    monkeypatch.setattr(messy_xlsx, "BatchStream", fail_wrapper)

    with pytest.raises(type(failure)) as captured:
        _public("read_excel_batches")("book.xlsx")

    assert captured.value is failure
    assert len(instances) == 1
    assert instances[0].child.close_calls == 1
    assert instances[0].close_calls == 1


def test_dataframe_owner_invalidation_is_sticky_on_every_next(
    sample_xlsx: Path,
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    chunks = workbook.iter_dataframe_chunks(batch_size=2)
    workbook.close()

    for _attempt in range(2):
        with pytest.raises(RuntimeError, match="MessyWorkbook is closed"):
            next(chunks)


def test_top_level_construction_cleanup_bypasses_hostile_traceback_accessor(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    traceback_reads: list[bool] = []

    class HostileError(RuntimeError):
        def __getattribute__(self, name: str) -> object:
            if name == "__traceback__":
                traceback_reads.append(True)
            return BaseException.__getattribute__(self, name)

    failure = HostileError("construction")
    instances: list[Any] = []

    class FailingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            self.close_calls = 0
            instances.append(self)

        def iter_batches(self, *_args: object, **_kwargs: object) -> object:
            raise failure

        def close(self) -> None:
            self.close_calls += 1

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", FailingWorkbook)
    captured: BaseException | None = None
    try:
        _public("read_excel_batches")("book.xlsx")
    except HostileError as error:
        captured = error

    assert captured is failure
    assert len(instances) == 1 and instances[0].close_calls == 1
    assert traceback_reads == []


# Task 12 review remediation G: public validation and exact late context.


@pytest.mark.parametrize("invalid_config", [0, False, "", object()])
def test_materialized_arrow_apis_reject_non_sheet_config_before_sheet_parse(
    sample_xlsx: Path,
    invalid_config: object,
) -> None:
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx) as workbook,
        pytest.raises(TypeError, match="config must be a SheetConfig or None"),
    ):
        workbook.to_arrow(config=invalid_config)  # type: ignore[arg-type]

    with pytest.raises(TypeError, match="config must be a SheetConfig or None"):
        _public("read_excel_arrow")(
            sample_xlsx,
            config=invalid_config,  # type: ignore[arg-type]
        )


def test_late_physical_type_error_reports_exact_scalar_after_nonzero_offset(
    tmp_path: Path,
) -> None:
    rows: list[list[object | None]] = [["ID", "Amount"]]
    rows.extend([[index, index] for index in range(2_100)])
    rows.append([2_100, "not-a-number"])
    path = _write_book(tmp_path / "late-exact-context.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(batch_size=500, config=config)
        with pytest.raises(messy_xlsx.StreamingTypeError) as captured:
            list(stream)

    assert captured.value.context == {
        "expected_type": "int64",
        "ordinal": 1,
        "display_label": "str label(length=6)",
        "row_offset": 2_100,
        "value_description": "str(length=12)",
    }


# Task 12 final-review remediation 1: pre-allocation sample budgets and jumps.


def test_sample_rejects_overwide_raw_window_before_arrow_allocation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_book(
        tmp_path / "overwide-sample.xlsx",
        [["A", "B", "C"], [1, 2, 3]],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    raw_string_allocations: list[int] = []
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        if kwargs.get("type") == pa.string() and isinstance(values, list):
            raw_string_allocations.append(len(values))
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(workbook_module, "_NORMALIZATION_SAMPLE_CELLS", 2)
    monkeypatch.setattr(pa, "array", track_array)
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        pytest.raises(ValueError, match=r"sample.*cells"),
    ):
        workbook.iter_batches(config=config)

    assert raw_string_allocations == []


def test_sample_rejects_large_scalar_before_arrow_value_allocation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    oversized = "x" * 512
    path = _write_book(
        tmp_path / "oversized-sample-scalar.xlsx",
        [["Value"], [oversized]],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    oversized_arrow_allocations: list[int] = []
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        if isinstance(values, list):
            for value in values:
                if isinstance(value, str) and len(value) >= len(oversized):
                    oversized_arrow_allocations.append(len(value))
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(
        workbook_module,
        "_NORMALIZATION_SAMPLE_BYTES",
        128,
        raising=False,
    )
    monkeypatch.setattr(pa, "array", track_array)
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        pytest.raises(ValueError, match=r"sample.*bytes"),
    ):
        workbook.iter_batches(config=config)

    assert oversized_arrow_allocations == []


@pytest.mark.parametrize("framing", ["skip_rows", "skip_footer"])
def test_sample_jumps_over_large_explicit_framing_regions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    framing: str,
) -> None:
    if framing == "skip_rows":
        rows: list[list[object | None]] = [["ignored"] for _ in range(2_500)]
        rows.extend([["Amount"], [1], [2]])
        config = SheetConfig(
            auto_detect=False,
            skip_rows=2_500,
            header_rows=1,
            normalize=False,
            sanitize_column_names=False,
        )
    else:
        rows = [["Amount"], [1], [2]]
        rows.extend([["footer"] for _ in range(2_501)])
        config = SheetConfig(
            auto_detect=False,
            header_rows=1,
            skip_footer=2_501,
            normalize=False,
            sanitize_column_names=False,
        )
    path = _write_book(tmp_path / f"bounded-{framing}.xlsx", rows)
    iter_rows_calls: list[tuple[int | None, int | None]] = []
    real_iter_rows = ReadOnlyWorksheet.iter_rows

    def counted_iter_rows(
        worksheet: ReadOnlyWorksheet,
        *args: object,
        **kwargs: object,
    ) -> Any:
        iter_rows_calls.append(
            (
                kwargs.get("min_row"),  # type: ignore[arg-type]
                kwargs.get("max_row"),  # type: ignore[arg-type]
            )
        )
        return real_iter_rows(worksheet, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", counted_iter_rows)
    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        # openpyxl's read-only iterator is forward-only: a high ``min_row``
        # still parses the worksheet prefix. The bounded contract is one
        # monotonically advancing sample pass, not a false yielded-row bound.
        assert len(iter_rows_calls) == 2
        stream.close()


def test_sample_skip_jump_preserves_merge_anchor_crossing_retained_start(
    tmp_path: Path,
) -> None:
    path = tmp_path / "merge-crosses-skip-boundary.xlsx"
    source = openpyxl.Workbook()
    sheet = source.active
    sheet["A1"] = "Merged"
    sheet.merge_cells("A1:A3")
    sheet["A4"] = "tail"
    source.save(path)
    source.close()
    config = SheetConfig(
        auto_detect=False,
        skip_rows=1,
        header_rows=0,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.string()]
    assert table.column(0).to_pylist() == ["Merged", "Merged", "tail"]


def test_sample_accumulates_without_intermediate_detached_arrow_batches(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    rows: list[list[object | None]] = [["Value"]]
    rows.extend([[f"{index:04d}-{'x' * 64}"] for index in range(2_000)])
    path = _write_book(tmp_path / "detached-sample.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    detached_calls = 0
    real_detached = workbook_module.CoordinateBatch.detached

    def track_detached(batch: object) -> object:
        nonlocal detached_calls
        if batch.batch.num_rows:  # type: ignore[attr-defined]
            detached_calls += 1
        return real_detached(batch)  # type: ignore[arg-type]

    monkeypatch.setattr(workbook_module.CoordinateBatch, "detached", track_detached)
    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        stream.close()

    # The accumulator owns decoded evidence, not a detached copy of every
    # transformed Arrow window. This avoids the former raw + detached + final
    # three-buffer peak.
    assert detached_calls == 0


# Task 12 final-review remediation 2: normalized zero-field rows.


def test_normalized_zero_column_streams_preserve_nonzero_row_counts(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "normalized-hidden-only-column.xlsx",
        [[1], [2], [3], [4], [5]],
    )
    physical = openpyxl.load_workbook(path)
    physical["Data"].column_dimensions["A"].hidden = True
    physical.save(path)
    physical.close()
    config = SheetConfig(
        auto_detect=False,
        header_rows=0,
        normalize=True,
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        with workbook.iter_batches(batch_size=2, config=config) as stream:
            batches = list(stream)
        with workbook.iter_dataframe_chunks(batch_size=2, config=config) as chunks:
            frames = list(chunks)

    assert [batch.num_rows for batch in batches] == [2, 2, 1]
    assert all(batch.num_columns == 0 for batch in batches)
    assert [frame.shape for frame in frames] == [(2, 0), (2, 0), (1, 0)]
    assert [index for frame in frames for index in frame.index] == list(range(5))


# Task 12 final-review remediation 3: silent homogeneous Arrow coercions.


@pytest.mark.parametrize(
    "values",
    [
        [date(2024, 1, 2), datetime(2024, 1, 3, 4, 5, 6)],
        [b"binary", "text"],
    ],
)
def test_materialized_mixed_scalars_use_lossless_dense_union(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    values: list[object],
) -> None:
    frame = pd.DataFrame({"mixed": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert pa.types.is_union(table.column(0).type)
    assert table.column(0).type.mode == "dense"
    assert table.column(0).to_pylist() == values


# Task 12 final-review remediation 4: bounded nonterminal header identity.


def test_header_only_truncated_projection_keeps_known_header_identity(
    tmp_path: Path,
) -> None:
    path = _write_book(tmp_path / "header-only-projection.xlsx", [["Known Header"]])
    config = SheetConfig(
        auto_detect=False,
        cell_range="A1:A100000",
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        assert stream.schema.names == ["Known Header"]
        stream.close()


# Task 12 final-review remediation 5: exact normalize=False provenance.


def test_normalize_false_rejects_sampled_native_text_heterogeneity(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "sampled-native-text-mix.xlsx",
        [["Value"], [1], ["two"]],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )

    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        with pytest.raises(messy_xlsx.StreamingTypeError):
            next(stream)


def test_normalize_false_staged_route_preserves_binary_time_and_duration(
    sample_xlsx: Path,
) -> None:
    expected_time = time(3, 4, 5, 6)
    expected_delta = timedelta(days=2, seconds=3, microseconds=4)

    class NativeRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                {
                    "blob": pd.Series([b"\x00\xff"], dtype=object),
                    "clock": pd.Series([expected_time], dtype=object),
                    "elapsed": pd.Series([expected_delta], dtype=object),
                }
            )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=NativeRegistry()) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.binary(), pa.time64("us"), pa.duration("us")]
    assert table.to_pydict() == {
        "blob": [b"\x00\xff"],
        "clock": [expected_time],
        "elapsed": [expected_delta],
    }


def test_normalize_false_unsupported_custom_value_fails_without_stringification(
    sample_xlsx: Path,
) -> None:
    callbacks: list[str] = []

    class HostileValue:
        def __str__(self) -> str:
            callbacks.append("str")
            raise AssertionError("unsupported values must not be stringified")

        def __repr__(self) -> str:
            callbacks.append("repr")
            raise AssertionError("unsupported values must not be represented")

        def __format__(self, _spec: str) -> str:
            callbacks.append("format")
            raise AssertionError("unsupported values must not be formatted")

    class HostileRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"value": pd.Series([HostileValue()], dtype=object)})

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=HostileRegistry()) as workbook,
        pytest.raises(messy_xlsx.StreamingTypeError),
    ):
        stream = workbook.iter_batches(config=config)
        list(stream)

    assert callbacks == []


# Task 12 final-review remediation 6: custom registry manifest independence.


def test_custom_xlsx_registry_does_not_open_builtin_ooxml_manifest() -> None:
    class ProprietaryRegistry(HandlerRegistry):
        def detect_format(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> FormatInfo:
            return FormatInfo("xlsx")

        def get_sheet_names(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> list[str]:
            return ["Data"]

        def validate(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> tuple[bool, str | None]:
            return True, None

        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"value": [7]})

    source = io.BytesIO(b"proprietary workbook bytes")
    config = SheetConfig(auto_detect=False, sanitize_column_names=False)
    with (
        messy_xlsx.MessyWorkbook(
            source,
            filename="custom.xlsx",
            registry=ProprietaryRegistry(),
        ) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.to_pydict() == {"value": [7]}
    assert not source.closed


# Task 12 final-review remediation 7: sample borrow acquisition transaction.


def test_sample_process_failure_after_borrow_entry_runs_borrow_cleanup(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    exit_calls = 0

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        workbook._get_sheet_manifest("Data")
        target_code = workbook._produce_normalization_sample.__func__.__code__

        class EnteredBorrow:
            def __enter__(self) -> Path:
                target_frame = sys._getframe(1)
                while target_frame is not None and target_frame.f_code is not target_code:
                    target_frame = target_frame.f_back
                assert target_frame is not None

                def fail_after_entry(
                    frame: Any,
                    event: str,
                    _arg: object,
                ) -> Any:
                    if frame.f_code is target_code and event == "line":
                        sys.settrace(None)
                        frame.f_trace = None
                        raise MemoryError("after sample borrow entry")
                    return fail_after_entry

                target_frame.f_trace = fail_after_entry
                sys.settrace(fail_after_entry)
                return sample_xlsx

            def __exit__(
                self,
                _exc_type: object,
                _exc_value: object,
                _traceback: object,
            ) -> None:
                nonlocal exit_calls
                exit_calls += 1

        monkeypatch.setattr(
            workbook._source_handle,
            "open_backend",
            lambda: EnteredBorrow(),
        )
        try:
            with pytest.raises(MemoryError, match="after sample borrow entry"):
                workbook.iter_batches(
                    config=SheetConfig(auto_detect=False),
                )
        finally:
            sys.settrace(None)

    assert exit_calls == 1


# Task 12 final-review remediation 8: reader factory/adoption transaction.


def test_prepare_to_attach_failure_closes_unadopted_reader_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    failure = MemoryError("reader adoption")

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )

        def fail_attach(
            _owner: object,
            _reader: object,
        ) -> object:
            raise failure

        monkeypatch.setattr(workbook_module._CloseOnceOwner, "attach", fail_attach)
        with pytest.raises(MemoryError) as captured:
            workbook.iter_batches()

    assert captured.value is failure
    assert reader.close_calls == 1


def test_reader_factory_return_process_gap_closes_produced_reader_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    target_code = workbook_module._StreamOperationLease.adopt.__code__

    def prepare(*_args: object, **_kwargs: object) -> SimpleNamespace:
        target_frame = sys._getframe(1)
        while target_frame is not None and target_frame.f_code is not target_code:
            target_frame = target_frame.f_back
        assert target_frame is not None

        def fail_after_factory_return(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame is target_frame and event == "line":
                sys.settrace(None)
                frame.f_trace = None
                raise MemoryError("after reader factory return")
            return fail_after_factory_return

        target_frame.f_trace = fail_after_factory_return
        sys.settrace(fail_after_factory_return)
        return SimpleNamespace(
            reader=reader,
            display_names=("value",),
        )

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(workbook, "_prepare_streaming_operation", prepare)
        try:
            with pytest.raises(MemoryError, match="after reader factory return"):
                workbook.iter_batches()
        finally:
            sys.settrace(None)

    assert reader.close_calls == 1


# Task 12 final-review remediation 9: internal wrapper/sidecar transactions.


@pytest.mark.parametrize("stage", ["normalized", "sidecar"])
@pytest.mark.parametrize(
    "failure",
    [ValueError("internal wrapper"), MemoryError("internal wrapper process")],
)
def test_internal_wrapper_failure_closes_raw_reader_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    stage: str,
    failure: BaseException,
) -> None:
    close_calls = 0
    real_close = materialized_streaming_module._EncodedDataFrameReader.close

    def record_close(reader: object) -> None:
        nonlocal close_calls
        close_calls += 1
        real_close(reader)  # type: ignore[arg-type]

    class CustomRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"value": [1]})

    def fail_constructor(*_args: object, **_kwargs: object) -> object:
        raise failure

    monkeypatch.setattr(
        materialized_streaming_module._EncodedDataFrameReader,
        "close",
        record_close,
    )
    if stage == "normalized":
        monkeypatch.setattr(
            materialized_streaming_module,
            "NormalizedStreamingReader",
            fail_constructor,
        )
    else:
        monkeypatch.setattr(
            materialized_streaming_module,
            "PreparedStreamingReader",
            fail_constructor,
        )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=CustomRegistry()) as workbook,
        pytest.raises(type(failure)) as captured,
    ):
        workbook.iter_batches(config=config)

    assert captured.value is failure
    assert close_calls == 1


def test_internal_constructor_return_process_gap_closes_raw_reader_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    close_calls = 0
    real_close = materialized_streaming_module._EncodedDataFrameReader.close
    real_physical_reader = materialized_streaming_module.PhysicalTypeReader
    target_code = materialized_streaming_module.wrap_normalized_streaming_reader.__code__

    def record_close(reader: object) -> None:
        nonlocal close_calls
        close_calls += 1
        real_close(reader)  # type: ignore[arg-type]

    class CustomRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"value": [1]})

    def traced_physical_reader(*args: object, **kwargs: object) -> object:
        physical_reader = real_physical_reader(*args, **kwargs)
        target_frame = sys._getframe(1)
        while target_frame is not None and target_frame.f_code is not target_code:
            target_frame = target_frame.f_back
        assert target_frame is not None

        def fail_after_constructor_return(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame is target_frame and event == "line":
                sys.settrace(None)
                frame.f_trace = None
                raise MemoryError("after internal constructor return")
            return fail_after_constructor_return

        target_frame.f_trace = fail_after_constructor_return
        sys.settrace(fail_after_constructor_return)
        return physical_reader

    monkeypatch.setattr(
        materialized_streaming_module._EncodedDataFrameReader,
        "close",
        record_close,
    )
    monkeypatch.setattr(
        materialized_streaming_module,
        "PhysicalTypeReader",
        traced_physical_reader,
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=CustomRegistry()) as workbook:
        try:
            with pytest.raises(
                MemoryError,
                match="after internal constructor return",
            ):
                workbook.iter_batches(config=config)
        finally:
            sys.settrace(None)

    assert close_calls == 1


# Task 12 final-review remediation 10: top-level workbook rollback begins first.


def test_top_level_process_failure_immediately_after_workbook_creation_closes_it(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    instances: list[Any] = []
    target_code = _public("read_excel_batches").__code__

    class TrackingWorkbook:
        def __init__(self, *_args: object, **_kwargs: object) -> None:
            self.close_calls = 0
            instances.append(self)
            caller = sys._getframe(1)

            def fail_after_creation(
                frame: Any,
                event: str,
                _arg: object,
            ) -> Any:
                if frame.f_code is target_code and event == "line":
                    sys.settrace(None)
                    frame.f_trace = None
                    raise MemoryError("after workbook creation")
                return fail_after_creation

            caller.f_trace = fail_after_creation
            sys.settrace(fail_after_creation)

        def iter_batches(self, *_args: object, **_kwargs: object) -> object:
            raise AssertionError("failure must precede child construction")

        def close(self) -> None:
            self.close_calls += 1

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", TrackingWorkbook)
    try:
        with pytest.raises(MemoryError, match="after workbook creation"):
            _public("read_excel_batches")("book.xlsx")
    finally:
        sys.settrace(None)

    assert len(instances) == 1
    assert instances[0].close_calls == 1


# Task 12 final-review remediation 11: top-level validation before parsing.


@pytest.mark.parametrize(
    ("api_name", "kwargs", "error_type", "message"),
    [
        (
            "read_excel_arrow",
            {"config": 0},
            TypeError,
            "config must be a SheetConfig or None",
        ),
        (
            "read_excel_batches",
            {"config": False},
            TypeError,
            "config must be a SheetConfig or None",
        ),
        (
            "read_excel_batches",
            {"batch_size": 0},
            ValueError,
            "batch_size",
        ),
    ],
)
def test_top_level_validation_precedes_missing_source_parsing(
    tmp_path: Path,
    api_name: str,
    kwargs: dict[str, object],
    error_type: type[BaseException],
    message: str,
) -> None:
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(error_type, match=message):
        _public(api_name)(missing, **kwargs)


# Task 12 final-review remediation 12: exact structural label kind.


def test_late_physical_type_error_preserves_integer_label_kind(
    sample_xlsx: Path,
) -> None:
    class IntegerLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            values: list[object] = list(range(2_100))
            values.append("late")
            frame = pd.DataFrame({7: pd.Series(values, dtype=object)})
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=IntegerLabelRegistry(),
    ) as workbook:
        stream = workbook.iter_batches(batch_size=500, config=config)
        with pytest.raises(messy_xlsx.StreamingTypeError) as captured:
            list(stream)

    assert captured.value.context["display_label"] == "int label"


# Task 12 second final-review remediation 1: raw rows are consumed directly.


def test_sample_and_full_reader_do_not_pass_row_major_lists_to_raw_builder(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _write_book(
        tmp_path / "direct-column-builder.xlsx",
        [["Value"], [1], [2], [3], [4]],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    row_inputs: list[tuple[str, object]] = []
    real_sample_builder = workbook_module._raw_coordinate_batch
    real_full_builder = xlsx_streaming_module._raw_coordinate_batch

    def sample_builder(rows: object, *args: object, **kwargs: object) -> object:
        row_inputs.append(("sample", rows))
        return real_sample_builder(rows, *args, **kwargs)  # type: ignore[arg-type]

    def full_builder(rows: object, *args: object, **kwargs: object) -> object:
        row_inputs.append(("full", rows))
        return real_full_builder(rows, *args, **kwargs)  # type: ignore[arg-type]

    monkeypatch.setattr(workbook_module, "_raw_coordinate_batch", sample_builder)
    monkeypatch.setattr(xlsx_streaming_module, "_raw_coordinate_batch", full_builder)

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=2, config=config) as stream,
    ):
        assert pa.Table.from_batches(list(stream), schema=stream.schema).column(0).to_pylist() == [
            1,
            2,
            3,
            4,
        ]

    assert {stage for stage, _rows in row_inputs} == {"sample", "full"}
    assert all(not isinstance(rows, list) for _stage, rows in row_inputs)


# Task 12 second final-review remediation 2: cumulative bounded framing windows.


@pytest.mark.parametrize("framing", ["skip_rows", "skip_footer"])
def test_sample_jumps_large_framing_with_unrelated_hidden_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    framing: str,
) -> None:
    if framing == "skip_rows":
        rows: list[list[object | None]] = [["ignored"] for _ in range(2_500)]
        rows.extend([["Amount"], [1], [2]])
        hidden_row = 2_600
        config = SheetConfig(
            auto_detect=False,
            skip_rows=2_500,
            header_rows=1,
            normalize=False,
            sanitize_column_names=False,
        )
    else:
        rows = [["Amount"], [1], [2]]
        rows.extend([["footer"] for _ in range(2_501)])
        hidden_row = len(rows)
        config = SheetConfig(
            auto_detect=False,
            header_rows=1,
            skip_footer=2_500,
            normalize=False,
            sanitize_column_names=False,
        )
    path = _write_book(tmp_path / f"hidden-{framing}.xlsx", rows)
    physical = openpyxl.load_workbook(path)
    physical["Data"].row_dimensions[hidden_row].hidden = True
    physical.save(path)
    physical.close()

    iter_rows_calls: list[tuple[int | None, int | None]] = []
    real_iter_rows = ReadOnlyWorksheet.iter_rows

    def counted_iter_rows(
        worksheet: ReadOnlyWorksheet,
        *args: object,
        **kwargs: object,
    ) -> Any:
        iter_rows_calls.append(
            (
                kwargs.get("min_row"),  # type: ignore[arg-type]
                kwargs.get("max_row"),  # type: ignore[arg-type]
            )
        )
        return real_iter_rows(worksheet, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", counted_iter_rows)
    with messy_xlsx.MessyWorkbook(path) as workbook:
        stream = workbook.iter_batches(config=config)
        assert stream.schema.types == [pa.int64()]
        assert len(iter_rows_calls) == 2
        stream.close()


def test_sample_jumps_disjoint_merge_anchor_and_projection_windows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "disjoint-merge-projection.xlsx"
    physical = openpyxl.Workbook()
    sheet = physical.active
    sheet.title = "Data"
    sheet["A1"] = "Merged"
    sheet.merge_cells("A1:A5000")
    sheet["A5001"] = "tail"
    physical.save(path)
    physical.close()
    config = SheetConfig(
        auto_detect=False,
        cell_range="A5000:A5001",
        header_rows=0,
        normalize=False,
        sanitize_column_names=False,
    )

    active_iterators = 0
    max_active_iterators = 0
    iter_rows_calls: list[tuple[int | None, int | None]] = []
    real_iter_rows = ReadOnlyWorksheet.iter_rows

    def counted_iter_rows(
        worksheet: ReadOnlyWorksheet,
        *args: object,
        **kwargs: object,
    ) -> Any:
        nonlocal active_iterators, max_active_iterators
        source = real_iter_rows(worksheet, *args, **kwargs)
        iter_rows_calls.append(
            (
                kwargs.get("min_row"),  # type: ignore[arg-type]
                kwargs.get("max_row"),  # type: ignore[arg-type]
            )
        )
        active_iterators += 1
        max_active_iterators = max(max_active_iterators, active_iterators)

        class TrackedIterator:
            def __init__(self) -> None:
                self._closed = False

            def __iter__(self) -> TrackedIterator:
                return self

            def __next__(self) -> tuple[Any, ...]:
                return next(source)

            def close(self) -> None:
                nonlocal active_iterators
                if self._closed:
                    return
                self._closed = True
                active_iterators -= 1
                close = getattr(source, "close", None)
                if callable(close):
                    close()

        return TrackedIterator()

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", counted_iter_rows)
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column(0).to_pylist() == ["Merged", "tail"]
    # Exactly one sample pass and one full-reader pass are created. The sample
    # keeps only one active row source while monotonically discarding gaps.
    assert len(iter_rows_calls) == 2
    assert max_active_iterators == 1
    assert active_iterators == 0


@pytest.mark.parametrize("budget_kind", ["cells", "bytes"])
def test_sample_rejects_required_headers_that_exceed_cumulative_budget(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    budget_kind: str,
) -> None:
    if budget_kind == "cells":
        rows = [[f"header-{index}"] for index in range(11)]
        rows.append([1])
        monkeypatch.setattr(workbook_module, "_NORMALIZATION_SAMPLE_CELLS", 10)
        expected = "cells"
    else:
        rows = [[f"{index}-{'x' * 64}"] for index in range(3)]
        rows.append([1])
        monkeypatch.setattr(workbook_module, "_NORMALIZATION_SAMPLE_BYTES", 180)
        expected = "bytes"
    path = _write_book(tmp_path / f"header-budget-{budget_kind}.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        header_rows=len(rows) - 1,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        pytest.raises(ValueError, match=rf"sample.*{expected}"),
    ):
        workbook.iter_batches(config=config)


# Task 12 second final-review remediation 3: projected padding owns footer semantics.


def test_projection_footer_is_applied_to_padded_endpoint_not_observed_endpoint(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "padded-projection-footer.xlsx",
        [["Amount"], [7]],
    )
    config = SheetConfig(
        auto_detect=False,
        cell_range="A1:A100",
        header_rows=1,
        skip_footer=98,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.int64()]
    assert table.column(0).to_pylist() == [7]


# Task 12 second final-review remediation 4: homogeneous materialized fast path.


def test_to_arrow_non_object_dtype_skips_python_scalar_scan(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    frame = pd.DataFrame({"value": pd.Series(range(10_000), dtype="int64")})
    direct_series_calls: list[str] = []
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        if isinstance(values, pd.Series):
            direct_series_calls.append(str(values.dtype))
        return real_array(values, *args, **kwargs)

    def reject_scalar_scan(*_args: object, **_kwargs: object) -> object:
        raise AssertionError("safe non-object dtypes must not be scanned per scalar")

    monkeypatch.setattr(pa, "array", track_array)
    monkeypatch.setattr(workbook_module, "_materialized_scalar_types", reject_scalar_scan)
    monkeypatch.setattr(
        workbook_module,
        "ensure_supported_physical_value",
        reject_scalar_scan,
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column(0).to_pylist() == list(range(10_000))
    assert direct_series_calls == ["int64"]


# Task 12 second final-review remediation 5: value-sensitive datetime types.


def test_materialized_naive_and_timezone_aware_datetimes_use_lossless_union(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        datetime(2024, 1, 1, 2, 3),
        datetime(2024, 1, 2, 3, 4, tzinfo=timezone(timedelta(hours=8))),
    ]
    frame = pd.DataFrame({"when": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert pa.types.is_union(table.column(0).type)
    assert table.column(0).to_pylist() == values


# Task 12 second final-review remediation 6: batch-level physical conversion.


def test_normalize_false_avoids_per_scalar_arrow_type_probes(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    integers = list(range(128))
    decimals = [Decimal("12.34")] * 128

    class NativeRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                {
                    "integer": pd.Series(integers, dtype=object),
                    "decimal": pd.Series(decimals, dtype=object),
                }
            )

    singleton_calls: list[type[object]] = []
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        if isinstance(values, list) and len(values) == 1:
            singleton_calls.append(type(values[0]))
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(pa, "array", track_array)
    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=NativeRegistry()) as workbook,
        workbook.iter_batches(batch_size=128, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.to_pydict() == {
        "integer": integers,
        "decimal": decimals,
    }
    assert singleton_calls == []


# Task 12 second final-review remediation 7: custom auto-detection authority.


def test_custom_xlsx_registry_default_auto_detect_never_uses_builtin_structure() -> None:
    parse_calls = 0

    class ProprietaryRegistry(HandlerRegistry):
        def detect_format(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> FormatInfo:
            return FormatInfo("xlsx")

        def get_sheet_names(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> list[str]:
            return ["Data"]

        def validate(
            self,
            *_args: object,
            **_kwargs: object,
        ) -> tuple[bool, str | None]:
            return True, None

        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            nonlocal parse_calls
            parse_calls += 1
            return pd.DataFrame({"value": [7]})

    source = io.BytesIO(b"proprietary workbook bytes")
    with (
        messy_xlsx.MessyWorkbook(
            source,
            filename="custom.xlsx",
            registry=ProprietaryRegistry(),
        ) as workbook,
        workbook.iter_batches() as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.to_pydict() == {"value": [7]}
    assert parse_calls == 1
    assert not source.closed


# Task 12 second final-review remediation 8: raw-reader return transactions.


def test_ooxml_raw_reader_return_gap_closes_reader_and_restores_cursor(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(13)
    entry = source.tell()
    target_code = workbook_module.MessyWorkbook._prepare_ooxml_stream.__code__
    real_reader_type = workbook_module.OpenpyxlStreamingReader
    created: list[Any] = []
    close_calls = 0

    def traced_reader(*args: object, **kwargs: object) -> object:
        nonlocal close_calls
        reader = real_reader_type(*args, **kwargs)
        created.append(reader)
        real_close = reader.close

        def record_close() -> None:
            nonlocal close_calls
            close_calls += 1
            real_close()

        reader.close = record_close
        target_frame = sys._getframe(1)
        while target_frame is not None and target_frame.f_code is not target_code:
            target_frame = target_frame.f_back
        assert target_frame is not None

        def fail_after_reader_return(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame is target_frame and event == "line":
                sys.settrace(None)
                frame.f_trace = None
                raise MemoryError("after OOXML raw reader return")
            return fail_after_reader_return

        target_frame.f_trace = fail_after_reader_return
        sys.settrace(fail_after_reader_return)
        return reader

    monkeypatch.setattr(workbook_module, "OpenpyxlStreamingReader", traced_reader)
    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    try:
        with pytest.raises(MemoryError, match="after OOXML raw reader return"):
            workbook.iter_batches(config=SheetConfig(auto_detect=False))
        assert close_calls == 1
        assert source.tell() == entry
    finally:
        sys.settrace(None)
        for reader in created:
            reader.close()
        workbook.close()


def test_materialized_raw_reader_return_gap_closes_and_releases_frame(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    target_code = materialized_streaming_module.prepare_materialized_streaming_reader.__code__
    real_reader_type = materialized_streaming_module._EncodedDataFrameReader
    created: list[Any] = []
    close_calls = 0

    class CustomRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"value": [1, 2]})

    def traced_reader(*args: object, **kwargs: object) -> object:
        nonlocal close_calls
        reader = real_reader_type(*args, **kwargs)
        created.append(reader)
        real_close = reader.close

        def record_close() -> None:
            nonlocal close_calls
            close_calls += 1
            real_close()

        reader.close = record_close
        target_frame = sys._getframe(1)
        while target_frame is not None and target_frame.f_code is not target_code:
            target_frame = target_frame.f_back
        assert target_frame is not None

        def fail_after_reader_return(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame is target_frame and event == "line":
                sys.settrace(None)
                frame.f_trace = None
                raise MemoryError("after materialized raw reader return")
            return fail_after_reader_return

        target_frame.f_trace = fail_after_reader_return
        sys.settrace(fail_after_reader_return)
        return reader

    monkeypatch.setattr(
        materialized_streaming_module,
        "_EncodedDataFrameReader",
        traced_reader,
    )
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx, registry=CustomRegistry())
    try:
        with pytest.raises(MemoryError, match="after materialized raw reader return"):
            workbook.iter_batches(
                config=SheetConfig(
                    auto_detect=False,
                    normalize=False,
                    sanitize_column_names=False,
                )
            )
        assert close_calls == 1
        assert created and created[0]._frame is None
    finally:
        sys.settrace(None)
        for reader in created:
            reader.close()
        workbook.close()


# Task 12 second final-review remediation 9: source-context entry transaction.


def test_openpyxl_reader_process_gap_after_source_entry_exits_borrow_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(17)
    entry = source.tell()
    target_code = xlsx_streaming_module.OpenpyxlStreamingReader._open.__code__
    entered_contexts: list[Any] = []
    target_exit_calls = 0

    with messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name) as workbook:
        real_open_backend = workbook._source_handle.open_backend

        class TrackingBorrow:
            def __init__(self) -> None:
                self._context = real_open_backend()
                self._target = False
                self._exited = False
                entered_contexts.append(self)

            def __enter__(self) -> object:
                backend = self._context.__enter__()
                target_frame = sys._getframe(1)
                while target_frame is not None and target_frame.f_code is not target_code:
                    target_frame = target_frame.f_back
                if target_frame is None:
                    return backend
                self._target = True

                def fail_after_source_entry(
                    frame: Any,
                    event: str,
                    _arg: object,
                ) -> Any:
                    if frame is target_frame and event == "line":
                        sys.settrace(None)
                        frame.f_trace = None
                        raise MemoryError("after full-reader source entry")
                    return fail_after_source_entry

                target_frame.f_trace = fail_after_source_entry
                sys.settrace(fail_after_source_entry)
                return backend

            def __exit__(
                self,
                exc_type: object,
                exc_value: object,
                traceback: object,
            ) -> object:
                nonlocal target_exit_calls
                self._exited = True
                if self._target:
                    target_exit_calls += 1
                return self._context.__exit__(exc_type, exc_value, traceback)

        monkeypatch.setattr(
            workbook._source_handle,
            "open_backend",
            lambda: TrackingBorrow(),
        )
        try:
            with pytest.raises(MemoryError, match="after full-reader source entry"):
                workbook.iter_batches(config=SheetConfig(auto_detect=False))
            assert target_exit_calls == 1
            assert source.tell() == entry
        finally:
            sys.settrace(None)
            for context in entered_contexts:
                if not context._exited:
                    context.__exit__(None, None, None)


# Task 12 second final-review remediation 10: partial owner attachment.


def test_partial_owner_attachment_process_gap_closes_reader_exactly_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    target_code = workbook_module._CloseOnceOwner.attach.__code__

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )

        def fail_after_resource_store(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame.f_code is target_code and event == "line":
                owner = frame.f_locals["self"]
                if owner._resource is reader and not owner._attached:
                    sys.settrace(None)
                    frame.f_trace = None
                    raise MemoryError("during owner attachment")
            return fail_after_resource_store

        sys.settrace(fail_after_resource_store)
        try:
            with pytest.raises(MemoryError, match="during owner attachment"):
                workbook.iter_batches()
        finally:
            sys.settrace(None)

    assert reader.close_calls == 1


def test_committed_owner_attachment_return_gap_closes_reader_exactly_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    target_code = workbook_module._CloseOnceOwner.attach.__code__

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_prepare_streaming_operation",
            lambda *_args, **_kwargs: SimpleNamespace(
                reader=reader,
                display_names=("value",),
            ),
        )

        def fail_after_attachment_commit(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame.f_code is target_code and event == "line":
                owner = frame.f_locals["self"]
                if owner._resource is reader and owner._attached:
                    sys.settrace(None)
                    frame.f_trace = None
                    raise MemoryError("after owner attachment commit")
            return fail_after_attachment_commit

        sys.settrace(fail_after_attachment_commit)
        try:
            with pytest.raises(MemoryError, match="after owner attachment commit"):
                workbook.iter_batches()
        finally:
            sys.settrace(None)

    assert reader.close_calls == 1


# Task 12 second final-review remediation 11: retryable lease release.


def test_lease_release_process_interruption_remains_retryable(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    reader = _InjectedReader()
    real_run_cleanups = workbook_module._run_cleanups
    interrupted = False

    def interrupt_once(*args: object, **kwargs: object) -> object:
        nonlocal interrupted
        if not interrupted:
            interrupted = True
            raise MemoryError("before lease cleanup")
        return real_run_cleanups(*args, **kwargs)

    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    lease = workbook._stream_operation()
    lease.own(reader)
    monkeypatch.setattr(workbook_module, "_run_cleanups", interrupt_once)
    try:
        with pytest.raises(MemoryError, match="before lease cleanup"):
            lease.release()
        assert reader.close_calls == 0
        assert workbook._active_operation_token is lease._token

        lease.release()
        assert reader.close_calls == 1
        assert workbook._active_operation_token is None
    finally:
        monkeypatch.setattr(workbook_module, "_run_cleanups", real_run_cleanups)
        lease.release()
        workbook.close()


# Task 12 acceptance review A: transport tags are not normalization lexemes.


def test_default_normalization_decodes_native_xlsx_float_before_normalizing(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "native-float.xlsx",
        [["value"], [1.5], [2.25]],
    )
    config = SheetConfig(
        auto_detect=False,
        normalize=True,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.float64()]
    assert table.column(0).to_pylist() == [1.5, 2.25]


def test_default_normalization_decodes_native_staged_csv_float() -> None:
    source = io.BytesIO(b"value\n1.5\n2.25\n")
    config = SheetConfig(
        auto_detect=False,
        normalize=True,
        sanitize_column_names=False,
    )

    with _public("read_excel_batches")(
        source,
        filename="native-float.csv",
        config=config,
    ) as stream:
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.float64()]
    assert table.column(0).to_pylist() == [1.5, 2.25]


def test_default_normalization_preserves_custom_binary_scalars(
    sample_xlsx: Path,
) -> None:
    expected = [b"\x00\xff", b"abc"]

    class BinaryRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"payload": expected})

    config = SheetConfig(
        auto_detect=False,
        normalize=True,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=BinaryRegistry()) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.schema.types == [pa.binary()]
    assert table.column(0).to_pylist() == expected


# Task 12 acceptance review B: deterministic footer exclusion is a read plan.


@pytest.mark.parametrize("projected", [False, True])
def test_full_reader_excludes_large_footer_before_coordinate_buffering(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    projected: bool,
) -> None:
    if projected:
        rows: list[list[object | None]] = [["value"], [7]]
        cell_range = "A1:A10000"
    else:
        rows = [["value"], [7], *([["footer"]] * 9_998)]
        cell_range = None
    path = _write_book(tmp_path / f"bounded-full-footer-{projected}.xlsx", rows)
    config = SheetConfig(
        auto_detect=False,
        cell_range=cell_range,
        header_rows=1,
        skip_footer=9_998,
        normalize=False,
        sanitize_column_names=False,
    )
    buffered_peaks: list[int] = []
    real_push = CoordinateOperation.push

    def track_buffered(
        operation: CoordinateOperation,
        batch: object,
    ) -> tuple[object, ...]:
        emitted = real_push(operation, batch)  # type: ignore[arg-type]
        buffered_peaks.append(operation._buffered_rows)
        return emitted

    monkeypatch.setattr(CoordinateOperation, "push", track_buffered)
    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=64, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column_names == ["value"]
    assert table.column(0).to_pylist() == [7]
    assert max(buffered_peaks, default=0) <= 65


@pytest.mark.parametrize(
    ("case", "expected"),
    [
        ("hidden_tail", [1, 2]),
        ("skip_header_footer", [10, 20]),
        ("merge_crosses_cutoff", [7, 7]),
        ("all_dropped", []),
    ],
)
def test_footer_execution_window_preserves_coordinate_framing_matrix(
    tmp_path: Path,
    case: str,
    expected: list[int],
) -> None:
    path = tmp_path / f"footer-matrix-{case}.xlsx"
    physical = openpyxl.Workbook()
    sheet = physical.active
    sheet.title = "Data"
    if case == "hidden_tail":
        for row in [["value"], [1], [2], [3], [4], [5]]:
            sheet.append(row)
        sheet.row_dimensions[6].hidden = True
        config = SheetConfig(
            auto_detect=False,
            header_rows=1,
            skip_footer=2,
            normalize=False,
            sanitize_column_names=False,
        )
    elif case == "skip_header_footer":
        for row in [["ignored"], ["value"], [10], [20], [30], [40]]:
            sheet.append(row)
        config = SheetConfig(
            auto_detect=False,
            skip_rows=1,
            header_rows=1,
            skip_footer=2,
            normalize=False,
            sanitize_column_names=False,
        )
    elif case == "merge_crosses_cutoff":
        sheet["A1"] = "value"
        sheet["A2"] = 7
        sheet.merge_cells("A2:A5")
        config = SheetConfig(
            auto_detect=False,
            header_rows=1,
            skip_footer=2,
            normalize=False,
            sanitize_column_names=False,
        )
    else:
        for row in [["value"], [1], [2]]:
            sheet.append(row)
        config = SheetConfig(
            auto_detect=False,
            header_rows=1,
            skip_footer=2,
            normalize=False,
            sanitize_column_names=False,
        )
    physical.save(path)
    physical.close()

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(batch_size=2, config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column_names == ["value"]
    assert table.column(0).to_pylist() == expected


def test_projected_footer_keeps_range_precedence_over_hidden_rows(
    tmp_path: Path,
) -> None:
    path = _write_book(
        tmp_path / "projected-hidden-footer.xlsx",
        [["value"], [1], [2], [3]],
    )
    physical = openpyxl.load_workbook(path)
    physical["Data"].row_dimensions[3].hidden = True
    physical.save(path)
    physical.close()
    config = SheetConfig(
        auto_detect=False,
        cell_range="A1:A4",
        header_rows=1,
        skip_footer=1,
        normalize=False,
        sanitize_column_names=False,
    )

    with (
        messy_xlsx.MessyWorkbook(path) as workbook,
        workbook.iter_batches(config=config) as stream,
    ):
        table = pa.Table.from_batches(list(stream), schema=stream.schema)

    assert table.column(0).to_pylist() == [1, 2]


# Task 12 acceptance review E: cleanup ownership survives process failure.


def test_lease_retries_owned_reader_cleanup_until_success_then_is_idempotent(
    sample_xlsx: Path,
) -> None:
    class RetryableReader(_InjectedReader):
        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise MemoryError("reader close interrupted")

    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    lease = workbook._stream_operation()
    reader = RetryableReader()
    lease.own(reader)
    try:
        with pytest.raises(MemoryError, match="reader close interrupted"):
            lease.release()
        assert reader.close_calls == 1

        lease.release()
        lease.release()
        assert reader.close_calls == 2
        assert workbook._active_operation_token is None
    finally:
        lease.release()
        workbook.close()


def test_openpyxl_reader_cleanup_retries_until_success_then_is_idempotent() -> None:
    class RetryableWorkbook:
        def __init__(self) -> None:
            self.close_calls = 0

        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise MemoryError("workbook close interrupted")

    reader = object.__new__(xlsx_streaming_module.OpenpyxlStreamingReader)
    workbook = RetryableWorkbook()
    reader._closed = False
    reader._terminal = False
    reader._rows = iter(())
    reader._workbook = workbook
    reader._backend_stack = None
    reader._operation = object()
    reader._rechunker = object()

    with pytest.raises(MemoryError, match="workbook close interrupted"):
        reader.close()
    assert workbook.close_calls == 1

    reader.close()
    reader.close()
    assert workbook.close_calls == 2


# Task 12 acceptance review F: context ownership is armed before __enter__.


@pytest.mark.parametrize("stage", ["sample", "full"])
def test_source_context_enter_failure_is_prearmed_for_sample_and_full_reader(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    stage: str,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(19)
    entry = source.tell()
    exit_calls = 0

    with messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name) as workbook:
        workbook._get_sheet_manifest("Data")
        real_open_backend = workbook._source_handle.open_backend
        open_calls = 0

        class EnterThenFail:
            def __init__(self, context: object) -> None:
                self._context = context

            def __enter__(self) -> object:
                backend = self._context.__enter__()  # type: ignore[attr-defined]
                del backend
                raise MemoryError(f"{stage} source enter interrupted")

            def __exit__(
                self,
                exc_type: object,
                exc_value: object,
                traceback: object,
            ) -> object:
                nonlocal exit_calls
                exit_calls += 1
                return self._context.__exit__(  # type: ignore[attr-defined]
                    exc_type,
                    exc_value,
                    traceback,
                )

        def open_backend() -> object:
            nonlocal open_calls
            open_calls += 1
            context = real_open_backend()
            target_call = 1 if stage == "sample" else 2
            return EnterThenFail(context) if open_calls == target_call else context

        monkeypatch.setattr(workbook._source_handle, "open_backend", open_backend)
        with pytest.raises(MemoryError, match=rf"{stage} source enter interrupted"):
            workbook.iter_batches(config=SheetConfig(auto_detect=False))

        assert source.tell() == entry

    assert exit_calls == 1


# Task 12 acceptance review G: constructor return is rollback-owned immediately.


def test_openpyxl_workbook_return_gap_closes_workbook_and_source(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(23)
    entry = source.tell()
    target_code = xlsx_streaming_module.OpenpyxlStreamingReader._open.__code__
    real_load_workbook = xlsx_streaming_module.openpyxl.load_workbook
    target_workbooks: list[Any] = []
    close_calls = 0

    def traced_load_workbook(*args: object, **kwargs: object) -> object:
        nonlocal close_calls
        loaded = real_load_workbook(*args, **kwargs)
        caller = sys._getframe(1)
        if caller.f_code is not target_code:
            return loaded
        target_workbooks.append(loaded)
        real_close = loaded.close

        def record_close() -> None:
            nonlocal close_calls
            close_calls += 1
            real_close()

        loaded.close = record_close

        def fail_after_constructor_return(
            frame: Any,
            event: str,
            _arg: object,
        ) -> Any:
            if frame is caller and event == "line":
                sys.settrace(None)
                frame.f_trace = None
                raise MemoryError("after openpyxl workbook return")
            return fail_after_constructor_return

        caller.f_trace = fail_after_constructor_return
        sys.settrace(fail_after_constructor_return)
        return loaded

    monkeypatch.setattr(
        xlsx_streaming_module.openpyxl,
        "load_workbook",
        traced_load_workbook,
    )
    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    try:
        with pytest.raises(MemoryError, match="after openpyxl workbook return"):
            workbook.iter_batches(config=SheetConfig(auto_detect=False))
        assert close_calls == 1
        assert source.tell() == entry
    finally:
        sys.settrace(None)
        for loaded in target_workbooks:
            loaded.close()
        workbook.close()


# Task 12 fourth acceptance remediation 1: pandas restoration is positional.


def test_dataframe_chunks_restore_colliding_labels_positionally(
    sample_xlsx: Path,
) -> None:
    class CollidingLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame(
                {
                    "integer": pd.Series([1, 2, 3], dtype="int64"),
                    "text": pd.Series(["a", "b", "c"], dtype="string"),
                    "flag": pd.Series([True, False, True], dtype="bool"),
                }
            )
            frame.columns = [7, "7", 7]
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=CollidingLabelRegistry(),
        ) as workbook,
        workbook.iter_dataframe_chunks(batch_size=2, config=config) as chunks,
    ):
        frames = list(chunks)

    assert [list(frame.columns) for frame in frames] == [[7, "7", 7], [7, "7", 7]]
    assert [str(dtype) for dtype in frames[0].dtypes] == [
        "int64[pyarrow]",
        "string[pyarrow]",
        "bool[pyarrow]",
    ]
    combined = pd.concat(frames)
    assert combined.index.tolist() == [0, 1, 2]
    assert combined.iloc[:, 0].tolist() == [1, 2, 3]
    assert combined.iloc[:, 1].tolist() == ["a", "b", "c"]
    assert combined.iloc[:, 2].tolist() == [True, False, True]


def test_dataframe_chunks_restore_typed_and_all_null_duplicates_positionally(
    sample_xlsx: Path,
) -> None:
    class NullDuplicateRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame(
                {
                    "typed": pd.Series([1, 2], dtype="int64"),
                    "null": pd.Series([None, None], dtype=object),
                }
            )
            frame.columns = ["duplicate", "duplicate"]
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=NullDuplicateRegistry(),
        ) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        frame = next(chunks)

    assert list(frame.columns) == ["duplicate", "duplicate"]
    assert [str(dtype) for dtype in frame.dtypes] == [
        "int64[pyarrow]",
        "null[pyarrow]",
    ]
    assert frame.iloc[:, 0].tolist() == [1, 2]
    assert frame.iloc[:, 1].isna().all()


# Task 12 fourth acceptance remediation 2/3: retryable workbook and stream owners.


def test_top_level_batch_close_retries_process_failed_owned_workbook(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    real_workbook = messy_xlsx.MessyWorkbook
    instances: list[Any] = []

    class RetryOnceResource:
        def __init__(self) -> None:
            self.close_calls = 0

        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise MemoryError("owned workbook resource interrupted")

    class TrackingWorkbook(real_workbook):
        def __init__(self, *args: object, **kwargs: object) -> None:
            super().__init__(*args, **kwargs)  # type: ignore[arg-type]
            instances.append(self)

    monkeypatch.setattr(messy_xlsx, "MessyWorkbook", TrackingWorkbook)
    stream = _public("read_excel_batches")(sample_xlsx)
    retry_resource = RetryOnceResource()
    instances[0]._fastexcel_session = retry_resource
    workbook_ref = weakref.ref(instances[0])

    with pytest.raises(MemoryError, match="owned workbook resource interrupted"):
        stream.close()
    assert retry_resource.close_calls == 1
    assert workbook_ref() is instances[0]

    stream.close()
    stream.close()
    assert retry_resource.close_calls == 2


def test_iteration_failure_reader_process_cleanup_retries_from_public_close(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    iteration_error = ValueError("iteration failed")

    class RetryOnceReader(_InjectedReader):
        def close(self) -> None:
            self.close_calls += 1
            if self.close_calls == 1:
                raise MemoryError("reader close interrupted")

    reader = RetryOnceReader(read_error=iteration_error)
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    monkeypatch.setattr(
        workbook,
        "_prepare_streaming_operation",
        lambda *_args, **_kwargs: SimpleNamespace(
            reader=reader,
            display_names=("value",),
        ),
    )
    stream = workbook.iter_batches()
    token = workbook._active_operation_token
    try:
        with pytest.raises(MemoryError, match="reader close interrupted"):
            next(stream)

        assert reader.close_calls == 1
        assert workbook._active_operation_token is token
        assert workbook._active_stream is stream

        stream.close()
        stream.close()
        assert reader.close_calls == 2
        assert workbook._active_operation_token is None
        assert workbook._active_stream is None
    finally:
        stream.close()
        workbook.close()


# Task 12 fourth acceptance remediation 4: prearmed source exit remains retryable.


@pytest.mark.parametrize("stage", ["sample", "full"])
def test_source_cursor_restoration_process_failure_retries_exact_context(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    stage: str,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(29)
    entry = source.tell()
    contexts: list[Any] = []

    class RetryExit:
        def __init__(self, context: object) -> None:
            self._context = context
            self.exit_calls = 0
            self.exited = False
            contexts.append(self)

        def __enter__(self) -> object:
            return self._context.__enter__()  # type: ignore[attr-defined]

        def __exit__(
            self,
            exc_type: object,
            exc_value: object,
            traceback: object,
        ) -> object:
            self.exit_calls += 1
            if self.exit_calls == 1:
                raise MemoryError(f"{stage} cursor restoration interrupted")
            self.exited = True
            return self._context.__exit__(  # type: ignore[attr-defined]
                exc_type,
                exc_value,
                traceback,
            )

    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    workbook._get_sheet_manifest("Data")
    real_open_backend = workbook._source_handle.open_backend
    open_calls = 0

    def open_backend() -> object:
        nonlocal open_calls
        open_calls += 1
        context = real_open_backend()
        target_call = 1 if stage == "sample" else 2
        return RetryExit(context) if open_calls == target_call else context

    monkeypatch.setattr(workbook._source_handle, "open_backend", open_backend)
    stream: Any | None = None
    try:
        if stage == "sample":
            with pytest.raises(
                MemoryError,
                match="sample cursor restoration interrupted",
            ):
                workbook.iter_batches(config=SheetConfig(auto_detect=False))
            assert len(contexts) == 1 and contexts[0].exit_calls == 1
            assert source.tell() != entry

            workbook.close()
            workbook.close()
        else:
            stream = workbook.iter_batches(config=SheetConfig(auto_detect=False))
            with pytest.raises(
                MemoryError,
                match="full cursor restoration interrupted",
            ):
                stream.close()
            assert len(contexts) == 1 and contexts[0].exit_calls == 1
            assert source.tell() != entry

            stream.close()
            stream.close()
            workbook.close()

        assert contexts[0].exit_calls == 2
        assert contexts[0].exited
        assert source.tell() == entry
        assert workbook._source_handle._active_borrow is False
    finally:
        if stream is not None:
            try:
                stream.close()
            except BaseException:
                pass
        for context in contexts:
            if not context.exited:
                context._context.__exit__(None, None, None)
        try:
            workbook.close()
        except BaseException:
            pass


@pytest.mark.parametrize("stage", ["sample", "full"])
def test_source_cursor_seek_process_failure_retries_saved_entry_position(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    stage: str,
) -> None:
    class RetrySeekBuffer(io.BytesIO):
        def __init__(self, content: bytes) -> None:
            super().__init__(content)
            self.restoration_position = 0
            self.fail_next_restoration = False
            self.restoration_failures = 0

        def seek(self, position: int, whence: int = 0) -> int:
            if self.fail_next_restoration and whence == 0 and position == self.restoration_position:
                self.fail_next_restoration = False
                self.restoration_failures += 1
                raise MemoryError(f"{stage} cursor seek interrupted")
            return super().seek(position, whence)

    source = RetrySeekBuffer(sample_xlsx.read_bytes())
    source.seek(37)
    source.restoration_position = source.tell()
    entry = source.tell()
    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    workbook._get_sheet_manifest("Data")
    real_open_backend = workbook._source_handle.open_backend
    open_calls = 0

    class FailRestoreOnExit:
        def __init__(self, context: object) -> None:
            self._context = context
            self._armed = False

        def __enter__(self) -> object:
            return self._context.__enter__()  # type: ignore[attr-defined]

        def __exit__(
            self,
            exc_type: object,
            exc_value: object,
            traceback: object,
        ) -> object:
            if not self._armed:
                self._armed = True
                source.fail_next_restoration = True
            return self._context.__exit__(  # type: ignore[attr-defined]
                exc_type,
                exc_value,
                traceback,
            )

    def open_backend() -> object:
        nonlocal open_calls
        open_calls += 1
        context = real_open_backend()
        target_call = 1 if stage == "sample" else 2
        return FailRestoreOnExit(context) if open_calls == target_call else context

    monkeypatch.setattr(workbook._source_handle, "open_backend", open_backend)
    stream: Any | None = None
    try:
        if stage == "sample":
            with pytest.raises(MemoryError, match=rf"{stage} cursor seek interrupted"):
                workbook.iter_batches(config=SheetConfig(auto_detect=False))
            assert source.tell() != entry
            workbook.close()
            workbook.close()
        else:
            stream = workbook.iter_batches(config=SheetConfig(auto_detect=False))
            with pytest.raises(MemoryError, match=rf"{stage} cursor seek interrupted"):
                stream.close()
            assert source.tell() != entry
            stream.close()
            stream.close()
            workbook.close()

        assert source.restoration_failures == 1
        assert source.tell() == entry
        assert workbook._source_handle._active_borrow is False
    finally:
        if stream is not None:
            try:
                stream.close()
            except BaseException:
                pass
        try:
            workbook.close()
        except BaseException:
            pass


# Task 12 fourth acceptance remediation 7: wide merge anchors fail pre-return.


def test_wide_merge_anchor_sample_never_publishes_late_null_schema(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "wide-merge-anchor-sample.xlsx"
    physical = openpyxl.Workbook()
    sheet = physical.active
    sheet.title = "Data"
    for column in range(1, 726):
        sheet.cell(column, column).value = column
        sheet.merge_cells(
            start_row=column,
            start_column=column,
            end_row=726,
            end_column=column,
        )
    physical.save(path)
    physical.close()
    config = SheetConfig(
        auto_detect=False,
        cell_range="A726:AAW726",
        header_rows=0,
        normalize=False,
        sanitize_column_names=False,
    )
    full_reader_calls = 0

    def reject_full_reader(*_args: object, **_kwargs: object) -> object:
        nonlocal full_reader_calls
        full_reader_calls += 1
        raise AssertionError("invalid sampled schema reached the full reader")

    monkeypatch.setattr(
        workbook_module,
        "OpenpyxlStreamingReader",
        reject_full_reader,
    )
    with messy_xlsx.MessyWorkbook(path) as workbook:
        with pytest.raises(ValueError, match=r"sample.*(rows|cells|bytes)"):
            workbook.iter_batches(config=config)
        assert workbook._active_operation_token is None

    assert full_reader_calls == 0


# Task 12 fifth acceptance remediation A/B: sample resources remain parent-owned.


@pytest.mark.parametrize("failed_resource", ["row_iterator", "workbook"])
def test_sample_process_failed_resource_is_retried_from_parent_close(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    failed_resource: str,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(41)
    entry = source.tell()
    events: list[str] = []
    row_close_calls = 0
    workbook_close_calls = 0
    real_load_workbook = workbook_module.openpyxl.load_workbook

    class RetryableRows:
        def __init__(self, rows: Any) -> None:
            self._rows = iter(rows)

        def __iter__(self) -> Any:
            return self

        def __next__(self) -> Any:
            return next(self._rows)

        def close(self) -> None:
            nonlocal row_close_calls
            row_close_calls += 1
            events.append(f"rows:{row_close_calls}")
            if failed_resource == "row_iterator" and row_close_calls == 1:
                raise MemoryError("sample row cleanup interrupted")
            close = getattr(self._rows, "close", None)
            if callable(close):
                close()

    class WorksheetProxy:
        def __init__(self, worksheet: Any) -> None:
            self._worksheet = worksheet

        def __getattr__(self, name: str) -> Any:
            return getattr(self._worksheet, name)

        def iter_rows(self, *args: object, **kwargs: object) -> RetryableRows:
            return RetryableRows(self._worksheet.iter_rows(*args, **kwargs))

    class WorkbookProxy:
        def __init__(self, workbook: Any) -> None:
            self._workbook = workbook

        def __getitem__(self, name: str) -> WorksheetProxy:
            return WorksheetProxy(self._workbook[name])

        def close(self) -> None:
            nonlocal workbook_close_calls
            workbook_close_calls += 1
            events.append(f"workbook:{workbook_close_calls}")
            if failed_resource == "workbook" and workbook_close_calls == 1:
                raise MemoryError("sample workbook cleanup interrupted")
            self._workbook.close()

    def tracked_load_workbook(*args: object, **kwargs: object) -> WorkbookProxy:
        return WorkbookProxy(real_load_workbook(*args, **kwargs))

    workbook = messy_xlsx.MessyWorkbook(source, filename=sample_xlsx.name)
    workbook._get_sheet_manifest("Data")
    real_open_backend = workbook._source_handle.open_backend

    class SourceContextProxy:
        def __init__(self, context: Any) -> None:
            self._context = context

        def __enter__(self) -> Any:
            return self._context.__enter__()

        def __exit__(
            self,
            exc_type: object,
            exc_value: object,
            traceback: object,
        ) -> object:
            events.append("source")
            return self._context.__exit__(exc_type, exc_value, traceback)

    monkeypatch.setattr(workbook_module.openpyxl, "load_workbook", tracked_load_workbook)
    monkeypatch.setattr(
        workbook._source_handle,
        "open_backend",
        lambda: SourceContextProxy(real_open_backend()),
    )

    expected_message = (
        "sample row cleanup interrupted"
        if failed_resource == "row_iterator"
        else "sample workbook cleanup interrupted"
    )
    try:
        with pytest.raises(MemoryError, match=expected_message):
            workbook.iter_batches(config=SheetConfig(auto_detect=False))

        assert events[:3] == ["rows:1", "workbook:1", "source"]
        assert source.tell() == entry
        assert workbook._source_handle._active_borrow is False

        workbook.close()
        workbook.close()
        assert row_close_calls == (2 if failed_resource == "row_iterator" else 1)
        assert workbook_close_calls == (2 if failed_resource == "workbook" else 1)
    finally:
        try:
            workbook.close()
        except BaseException:
            pass


# Task 12 fifth acceptance remediation D: exact safe pandas labels are a sidecar.


def test_dataframe_chunks_restore_exact_decimal_and_timestamp_labels_positionally(
    sample_xlsx: Path,
) -> None:
    decimal_label = Decimal("1.25")
    timestamp_label = pd.Timestamp("2024-01-02 03:04:05.123456789")
    labels = pd.Index(
        [decimal_label, timestamp_label, decimal_label, 7],
        dtype=object,
    )

    class LabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [[1, "two", 3, True], [4, "five", 6, False]],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=LabelRegistry()) as workbook:
        with workbook.iter_batches(batch_size=1, config=config) as batches:
            assert all(type(name) is str for name in batches.schema.names)
            list(batches)
        with workbook.iter_dataframe_chunks(batch_size=1, config=config) as chunks:
            frames = list(chunks)

    assert [frame.index.tolist() for frame in frames] == [[0], [1]]
    for frame in frames:
        restored = frame.columns.tolist()
        assert type(restored[0]) is Decimal
        assert restored[0] == decimal_label
        assert type(restored[1]) is pd.Timestamp
        assert restored[1] == timestamp_label
        assert restored[1].nanosecond == 789
        assert type(restored[2]) is Decimal
        assert restored[2] == decimal_label
        assert type(restored[3]) is int
        assert restored[3] == 7


@pytest.mark.parametrize(
    "config,expected",
    [
        (
            SheetConfig(
                auto_detect=False,
                normalize=False,
                sanitize_column_names=False,
                column_renames={Decimal("1.25"): "renamed"},
            ),
            "renamed",
        ),
        (
            SheetConfig(
                auto_detect=False,
                normalize=False,
                sanitize_column_names=True,
            ),
            "unsafe_label",
        ),
    ],
)
def test_dataframe_label_sidecar_preserves_renames_and_sanitization(
    sample_xlsx: Path,
    config: SheetConfig,
    expected: str,
) -> None:
    class LabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[1]], columns=pd.Index([Decimal("1.25")], dtype=object))

    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=LabelRegistry()) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        frame = next(chunks)

    assert frame.columns.tolist() == [expected]


def test_dataframe_label_sidecar_never_executes_or_retains_hostile_labels(
    sample_xlsx: Path,
) -> None:
    callbacks: list[str] = []
    label_refs: list[weakref.ReferenceType[object]] = []

    class HostileLabel:
        def __str__(self) -> str:
            callbacks.append("str")
            raise AssertionError("hostile label text executed")

        def __repr__(self) -> str:
            callbacks.append("repr")
            raise AssertionError("hostile label representation executed")

        def __hash__(self) -> int:
            callbacks.append("hash")
            raise AssertionError("hostile label hash executed")

        def __eq__(self, other: object) -> bool:
            callbacks.append("eq")
            raise AssertionError("hostile label equality executed")

    class HostileRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            label = HostileLabel()
            label_refs.append(weakref.ref(label))
            frame = pd.DataFrame([[1]])
            frame.columns = pd.Index([label], dtype=object)
            callbacks.clear()
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx, registry=HostileRegistry())
    stream = workbook.iter_dataframe_chunks(config=config)
    frame = next(stream)
    assert callbacks == []
    del frame
    stream.close()
    workbook.close()
    gc.collect()

    assert callbacks == []
    assert label_refs and label_refs[0]() is None


# Task 12 sixth acceptance remediation: exact safe label projection remains
# per-column, selective, and hostile-hook free.


def test_dataframe_label_sidecar_budget_is_per_top_level_label(
    sample_xlsx: Path,
) -> None:
    labels = tuple(Decimal(ordinal) for ordinal in range(300))

    class WideLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [range(len(labels))],
                columns=pd.Index(labels, dtype=object),
            )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=WideLabelRegistry()) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        frame = next(chunks)

    restored = frame.columns.tolist()
    assert len(restored) == len(labels)
    assert all(type(label) is Decimal for label in restored)
    assert restored == list(labels)


def test_dataframe_chunks_restore_exact_pandas_timedelta_label(
    sample_xlsx: Path,
) -> None:
    label = pd.Timedelta(days=2, nanoseconds=123)

    class TimedeltaLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[1]], columns=pd.Index([label], dtype=object))

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=TimedeltaLabelRegistry()) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        frame = next(chunks)

    restored = frame.columns[0]
    assert type(restored) is pd.Timedelta
    assert restored == label
    assert restored.nanoseconds == 123


def test_dataframe_chunks_keep_untouched_safe_labels_with_selective_rename(
    sample_xlsx: Path,
) -> None:
    renamed = Decimal("1.25")
    untouched = Decimal("2.50")
    timestamp = pd.Timestamp("2024-01-02 03:04:05.123456789")
    duration = pd.Timedelta(days=1, nanoseconds=7)
    labels = pd.Index([renamed, untouched, timestamp, duration], dtype=object)

    class SelectiveRenameRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[" 1 ", " 2 ", " 3 ", " 4 "]], columns=labels)

    config = SheetConfig(
        auto_detect=False,
        normalize=True,
        sanitize_column_names=False,
        column_renames={renamed: "renamed"},
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=SelectiveRenameRegistry()) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        frame = next(chunks)

    restored = frame.columns.tolist()
    assert restored[0] == "renamed"
    assert type(restored[1]) is Decimal and restored[1] == untouched
    assert type(restored[2]) is pd.Timestamp and restored[2] == timestamp
    assert type(restored[3]) is pd.Timedelta and restored[3] == duration


def test_to_arrow_never_executes_hostile_label_text_hooks(
    sample_xlsx: Path,
) -> None:
    callbacks: list[str] = []

    class HostileLabel:
        def __str__(self) -> str:
            callbacks.append("str")
            raise AssertionError("hostile label text executed")

        def __repr__(self) -> str:
            callbacks.append("repr")
            raise AssertionError("hostile label representation executed")

        def __hash__(self) -> int:
            callbacks.append("hash")
            raise AssertionError("hostile label hash executed")

        def __eq__(self, _other: object) -> bool:
            callbacks.append("eq")
            raise AssertionError("hostile label equality executed")

    class HostileLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame([[1]])
            frame.columns = pd.Index([HostileLabel()], dtype=object)
            callbacks.clear()
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=HostileLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)

    assert table.num_columns == 1
    assert all(type(name) is str for name in table.column_names)
    assert callbacks == []


def test_to_arrow_preserves_safe_non_string_label_text(
    sample_xlsx: Path,
) -> None:
    labels = (
        Decimal("1.25"),
        pd.Timestamp("2024-01-02 03:04:05.123456789"),
        pd.Timedelta(days=2, nanoseconds=123),
    )

    class SafeLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame([[1, 2, 3]])
            frame.columns = pd.Index(labels, dtype=object)
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=SafeLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)

    assert table.column_names == [
        "1.25",
        "2024-01-02 03:04:05.123456789",
        "2 days 00:00:00.000000123",
    ]


# Task 12 ninth acceptance remediation: trusted label projection, constructor
# and operation handoffs, and vector-first materialized object conversion.


def test_to_arrow_timestamp_label_never_executes_untrusted_timezone_hooks(
    sample_xlsx: Path,
) -> None:
    callbacks: list[str] = []

    class ArmedTimezone(tzinfo):
        armed = False

        def _record(self, name: str) -> None:
            if self.armed:
                callbacks.append(name)
                raise AssertionError("untrusted timezone hook executed")

        def utcoffset(self, _value: datetime | None) -> timedelta:
            self._record("utcoffset")
            return timedelta(hours=8)

        def dst(self, _value: datetime | None) -> timedelta:
            self._record("dst")
            return timedelta(0)

        def tzname(self, _value: datetime | None) -> str:
            self._record("tzname")
            return "HOSTILE"

    hostile_timezone = ArmedTimezone()
    label = pd.Timestamp(datetime(2024, 1, 2, 3, 4, tzinfo=hostile_timezone))

    class HostileTimezoneLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame([[1]])
            frame.columns = pd.Index([label], dtype=object)
            callbacks.clear()
            hostile_timezone.armed = True
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=HostileTimezoneLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)

    assert table.column_names == ["<pandas.Timestamp>"]
    assert callbacks == []


def test_to_arrow_timestamp_label_preserves_trusted_timezone_text(
    sample_xlsx: Path,
) -> None:
    labels = (
        pd.Timestamp("2024-01-02 03:04:05.123456789"),
        pd.Timestamp(datetime(2024, 1, 2, 3, 4, tzinfo=timezone(timedelta(hours=8)))),
        pd.Timestamp(datetime(2024, 1, 2, 3, 4, tzinfo=ZoneInfo("UTC"))),
    )

    class TrustedTimezoneLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[1, 2, 3]], columns=pd.Index(labels, dtype=object))

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=TrustedTimezoneLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)

    assert table.column_names == [pd.Timestamp.__str__(label) for label in labels]


def test_workbook_source_handle_handoff_process_failure_closes_acquired_handle(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(b"caller-owned")
    acquired: list[Any] = []

    class AcquiredHandle:
        def __init__(self, original: object, filename: str | None = None) -> None:
            self.original = original
            self.filename = filename
            self.close_calls = 0
            acquired.append(self)

        def close(self) -> None:
            self.close_calls += 1

    monkeypatch.setattr(workbook_module, "SourceHandle", AcquiredHandle)
    target_code = messy_xlsx.MessyWorkbook.__init__.__code__
    interrupted = False

    def interrupt_after_handle_assignment(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        workbook = frame.f_locals.get("self")
        if (
            frame.f_code is target_code
            and event == "line"
            and acquired
            and getattr(workbook, "_source_handle", None) is acquired[0]
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("source handle handoff interrupted")
        return interrupt_after_handle_assignment

    sys.settrace(interrupt_after_handle_assignment)
    try:
        with pytest.raises(MemoryError, match="source handle handoff interrupted"):
            messy_xlsx.MessyWorkbook(source, filename="input.xlsx")
    finally:
        sys.settrace(None)

    assert interrupted
    assert len(acquired) == 1
    assert acquired[0].close_calls == 1
    assert not source.closed


def _interrupt_after_operation_token_commit(workbook: messy_xlsx.MessyWorkbook) -> None:
    target_code = workbook._begin_operation.__func__.__code__
    interrupted = False

    def interrupt(
        frame: Any,
        event: str,
        _arg: object,
    ) -> Any:
        nonlocal interrupted
        if (
            frame.f_code is target_code
            and event == "line"
            and workbook._active_operation_token is not None
            and not interrupted
        ):
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("operation token commit interrupted")
        return interrupt

    sys.settrace(interrupt)


@pytest.mark.parametrize("entry", ["to_arrow", "_to_dataframes_compat", "_parse_sheet"])
def test_to_arrow_operation_token_commit_process_failure_is_rolled_back(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    entry: str,
) -> None:
    frame = pd.DataFrame({"value": [1]})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        _interrupt_after_operation_token_commit(workbook)
        try:
            with pytest.raises(MemoryError, match="operation token commit interrupted"):
                if entry == "to_arrow":
                    workbook.to_arrow()
                elif entry == "_to_dataframes_compat":
                    workbook._to_dataframes_compat()
                else:
                    workbook._parse_sheet("Data")
        finally:
            sys.settrace(None)

        assert workbook._active_operation_token is None
        assert workbook.to_arrow().to_pydict() == {"value": [1]}


@pytest.mark.parametrize("entry", ["_stream_operation", "iter_batches"])
def test_stream_operation_token_commit_process_failure_is_rolled_back(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    entry: str,
) -> None:
    frame = pd.DataFrame({"value": [1]})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        _interrupt_after_operation_token_commit(workbook)
        try:
            with pytest.raises(MemoryError, match="operation token commit interrupted"):
                if entry == "_stream_operation":
                    workbook._stream_operation()
                else:
                    workbook.iter_batches(config=SheetConfig(auto_detect=False))
        finally:
            sys.settrace(None)

        assert workbook._active_operation_token is None
        assert workbook.to_arrow().to_pydict() == {"value": [1]}


def test_to_arrow_object_decimals_infer_one_common_decimal_type(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [Decimal("1.25"), Decimal("12.25"), Decimal("123.25")]
    frame = pd.DataFrame({"amount": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column(0).type == pa.decimal128(5, 2)
    assert table.column(0).to_pylist() == values


def test_to_arrow_homogeneous_aware_object_datetimes_stay_timestamp(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_timezone = timezone(timedelta(hours=8))
    values = [
        datetime(2024, 1, 1, 2, 3, tzinfo=shared_timezone),
        datetime(2024, 1, 2, 3, 4, tzinfo=shared_timezone),
    ]
    frame = pd.DataFrame({"when": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column(0).type == pa.timestamp("us", tz="+08:00")
    assert table.column(0).to_pylist() == values


def test_to_arrow_large_homogeneous_object_columns_do_not_probe_singletons(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    row_count = 20_000
    shared_timezone = UTC
    decimals = [Decimal("12.34")] * row_count
    datetimes = [
        datetime(2024, 1, 1, tzinfo=shared_timezone) + timedelta(minutes=ordinal)
        for ordinal in range(row_count)
    ]
    frame = pd.DataFrame(
        {
            "amount": pd.Series(decimals, dtype=object),
            "when": pd.Series(datetimes, dtype=object),
        }
    )
    singleton_calls = 0
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        nonlocal singleton_calls
        if isinstance(values, list) and len(values) == 1:
            singleton_calls += 1
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(pa, "array", track_array)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.num_rows == row_count
    assert table.column(0).type == pa.decimal128(4, 2)
    assert pa.types.is_timestamp(table.column(1).type)
    assert singleton_calls == 0


def test_to_arrow_object_pandas_temporals_preserve_nanoseconds(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    timestamps = [
        pd.Timestamp("2024-01-01 00:00:00.000000001"),
        pd.Timestamp("2024-01-02 00:00:00.000000002"),
    ]
    timedeltas = [
        pd.Timedelta(nanoseconds=1),
        pd.Timedelta(days=1, nanoseconds=2),
    ]
    frame = pd.DataFrame(
        {
            "when": pd.Series(timestamps, dtype=object),
            "elapsed": pd.Series(timedeltas, dtype=object),
        }
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column(0).type == pa.timestamp("ns")
    assert table.column(1).type == pa.duration("ns")
    assert table.column(0).to_pylist() == timestamps
    assert table.column(1).to_pylist() == timedeltas


# Task 12 eleventh acceptance remediation: canonical temporal labels, lossless
# temporal values, family-bounded mixed conversion, and materialized leases.


def test_pandas_temporal_labels_match_equivalent_stdlib_rename_keys(
    sample_xlsx: Path,
) -> None:
    timestamp = pd.Timestamp("2024-01-02 03:04:05.123456")
    duration = pd.Timedelta(days=2, microseconds=7)
    untouched = pd.Timedelta(days=3, nanoseconds=11)
    labels = pd.Index([timestamp, duration, untouched], dtype=object)

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[1, 2, 3]], columns=labels)

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
        column_renames={
            datetime(2024, 1, 2, 3, 4, 5, 123456): "when",
            timedelta(days=2, microseconds=7): "elapsed",
        },
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=TemporalRegistry()) as workbook:
        materialized = workbook.to_arrow(config=config)
        with workbook.iter_batches(config=config) as batches:
            batch_table = pa.Table.from_batches(list(batches), schema=batches.schema)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            chunk = next(chunks)

    assert materialized.column_names == ["when", "elapsed", str(untouched)]
    assert batch_table.column_names == ["when", "elapsed", str(untouched)]
    restored = chunk.columns.tolist()
    assert restored[:2] == ["when", "elapsed"]
    assert type(restored[2]) is pd.Timedelta
    assert restored[2] == untouched
    assert restored[2].nanoseconds == 11


def test_pandas_temporal_labels_match_unique_stdlib_hint_and_condition_keys(
    sample_xlsx: Path,
) -> None:
    timestamp = pd.Timestamp("2024-01-02 03:04:05.123456")
    first_duration = pd.Timedelta(days=1)
    second_duration = pd.Timedelta(days=2)
    labels = pd.Index([timestamp, first_duration, second_duration], dtype=object)

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [["1", "drop", "2"], ["3", "keep", "4"]],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={
            datetime(2024, 1, 2, 3, 4, 5, 123456): "INTEGER",
            timedelta(days=1): "TEXT",
            timedelta(days=2): "INTEGER",
        },
        drop_conditions=[{"column": timedelta(days=1), "value": "drop"}],
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=TemporalRegistry()) as workbook:
        materialized = workbook.to_arrow(config=config)
        with workbook.iter_batches(config=config) as batches:
            batch_table = pa.Table.from_batches(list(batches), schema=batches.schema)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            chunk = next(chunks)

    for table in (materialized, batch_table):
        assert table.num_rows == 1
        assert table.schema.types[0] == pa.int64()
        assert pa.types.is_string(table.schema.types[1]) or pa.types.is_large_string(
            table.schema.types[1]
        )
        assert table.schema.types[2] == pa.int64()
        assert table.to_pylist() == [
            {
                str(timestamp): 3,
                str(first_duration): "keep",
                str(second_duration): 4,
            }
        ]
    assert chunk.iloc[0].tolist() == [3, "keep", 4]
    restored = chunk.columns.tolist()
    assert [type(label) for label in restored] == [pd.Timestamp, pd.Timedelta, pd.Timedelta]
    assert restored == [timestamp, first_duration, second_duration]


@pytest.mark.parametrize("payload_kind", ["offset", "name"])
def test_exact_timezone_with_untrusted_payload_degrades_temporal_labels_without_hooks(
    sample_xlsx: Path,
    payload_kind: str,
) -> None:
    callbacks: list[str] = []

    class ArmedDelta(timedelta):
        armed = False

        def __str__(self) -> str:
            if self.armed:
                callbacks.append("offset_str")
                raise AssertionError("hostile offset text hook executed")
            return timedelta.__str__(self)

        def total_seconds(self) -> float:
            if self.armed:
                callbacks.append("total_seconds")
                raise AssertionError("hostile offset arithmetic hook executed")
            return timedelta.total_seconds(self)

    class ArmedName(str):
        armed = False

        def __str__(self) -> str:
            if self.armed:
                callbacks.append("name_str")
                raise AssertionError("hostile timezone name hook executed")
            return str.__str__(self)

    offset = ArmedDelta(hours=8) if payload_kind == "offset" else timedelta(hours=8)
    name = ArmedName("CUSTOM") if payload_kind == "name" else "CUSTOM"
    label = pd.Timestamp(
        datetime(2024, 1, 2, 3, 4, tzinfo=timezone(offset, name)),
    )

    class HostileTimezoneRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            frame = pd.DataFrame([[1]], columns=pd.Index([label], dtype=object))
            callbacks.clear()
            ArmedDelta.armed = True
            ArmedName.armed = True
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=HostileTimezoneRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            frame = next(chunks)

    assert table.column_names == ["<pandas.Timestamp>"]
    assert frame.columns.tolist() == ["<pandas.Timestamp>"]
    assert callbacks == []


@pytest.mark.parametrize("normalize", [False, True])
@pytest.mark.parametrize("entry", ["to_arrow", "iter_batches", "chunks"])
def test_timezone_aware_time_data_is_rejected_before_arrow_can_strip_timezone(
    sample_xlsx: Path,
    normalize: bool,
    entry: str,
) -> None:
    values = [time(1, 2, 3) for _ in range(1_000)]
    values.append(time(4, 5, 6, tzinfo=ZoneInfo("UTC")))
    frame = pd.DataFrame({"clock": pd.Series(values, dtype=object)})

    class AwareTimeRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=normalize,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=AwareTimeRegistry()) as workbook:
        if entry == "to_arrow":
            with pytest.raises(
                messy_xlsx.StreamingTypeError,
                match="incompatible with the fixed schema",
            ) as captured:
                workbook.to_arrow(config=config)
        else:
            stream = (
                workbook.iter_batches(batch_size=256, config=config)
                if entry == "iter_batches"
                else workbook.iter_dataframe_chunks(batch_size=256, config=config)
            )
            with (
                stream,
                pytest.raises(
                    messy_xlsx.StreamingTypeError,
                    match="incompatible with the fixed schema",
                ) as captured,
            ):
                list(stream)

    assert captured.value.context["ordinal"] == 0
    assert captured.value.context["row_offset"] == 1_000
    assert captured.value.context["display_label"] == "str label(length=5)"
    assert captured.value.context["value_description"] == "time"


def test_timezone_aware_time_rejection_never_invokes_custom_timezone_hooks(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    callbacks: list[str] = []

    class ArmedTimezone(tzinfo):
        armed = False

        def _record(self, name: str) -> None:
            if self.armed:
                callbacks.append(name)
                raise AssertionError("custom time timezone hook executed")

        def utcoffset(self, _value: datetime | None) -> timedelta | None:
            self._record("utcoffset")
            return None

        def dst(self, _value: datetime | None) -> timedelta | None:
            self._record("dst")
            return None

        def tzname(self, _value: datetime | None) -> str | None:
            self._record("tzname")
            return None

    zone = ArmedTimezone()
    value = time(1, 2, 3, tzinfo=zone)
    frame = pd.DataFrame({"clock": pd.Series([value], dtype=object)})
    zone.armed = True

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        with pytest.raises(messy_xlsx.StreamingTypeError):
            workbook.to_arrow()

    assert callbacks == []


@pytest.mark.parametrize(
    "entry",
    ["to_arrow", "iter_batches", "iter_dataframe_chunks"],
)
def test_hostile_stdlib_datetime_timezone_is_rejected_without_callbacks_with_context(
    sample_xlsx: Path,
    entry: str,
) -> None:
    callbacks: list[str] = []

    class ArmedTimezone(tzinfo):
        armed = False

        def _record(self, name: str) -> None:
            if self.armed:
                callbacks.append(name)
                raise AssertionError("custom datetime timezone hook executed")

        def utcoffset(self, _value: datetime | None) -> timedelta | None:
            self._record("utcoffset")
            return timedelta(hours=8)

        def dst(self, _value: datetime | None) -> timedelta | None:
            self._record("dst")
            return timedelta(0)

        def tzname(self, _value: datetime | None) -> str | None:
            self._record("tzname")
            return "HOSTILE"

    zone = ArmedTimezone()
    hostile = datetime(2024, 1, 2, 3, 4, 5, 678901, tzinfo=zone, fold=1)
    frame = pd.DataFrame(
        {
            "safe": pd.Series([1, 2], dtype=object),
            "hostile moment": pd.Series(
                [datetime(2024, 1, 1, tzinfo=UTC), hostile],
                dtype=object,
            ),
        }
    )

    class HostileDatetimeRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    zone.armed = True
    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=HostileDatetimeRegistry(),
        ) as workbook,
        pytest.raises(messy_xlsx.StreamingTypeError) as captured,
    ):
        if entry == "to_arrow":
            workbook.to_arrow(config=config)
        else:
            stream = (
                workbook.iter_batches(batch_size=1, config=config)
                if entry == "iter_batches"
                else workbook.iter_dataframe_chunks(batch_size=1, config=config)
            )
            with stream:
                list(stream)

    assert captured.value.context == {
        "expected_type": "supported Arrow scalar",
        "ordinal": 1,
        "display_label": "str label(length=14)",
        "row_offset": 1,
        "value_description": "datetime",
    }
    assert callbacks == []


def test_naive_time_data_remains_lossless_for_materialized_and_streaming(
    sample_xlsx: Path,
) -> None:
    values = [time(1, 2, 3, 4), time(5, 6, 7, 8)]

    class NaiveTimeRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"clock": pd.Series(values, dtype=object)})

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(sample_xlsx, registry=NaiveTimeRegistry()) as workbook:
        materialized = workbook.to_arrow(config=config)
        with workbook.iter_batches(config=config) as batches:
            streamed = pa.Table.from_batches(list(batches), schema=batches.schema)

    assert materialized.column(0).to_pylist() == values
    assert streamed.column(0).to_pylist() == values


def test_equivalent_timezone_instances_and_temporal_classes_share_vector_bucket(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixed_values = [
        datetime(2024, 1, 1, tzinfo=timezone(timedelta(hours=8))),
        pd.Timestamp(
            datetime(2024, 1, 2, tzinfo=timezone(timedelta(hours=8))),
        )
        + pd.Timedelta(nanoseconds=7),
    ]
    zone_values = [
        datetime(2024, 1, 1, tzinfo=ZoneInfo.no_cache("UTC")),
        pd.Timestamp(datetime(2024, 1, 2, tzinfo=ZoneInfo.no_cache("UTC"))),
    ]
    frame = pd.DataFrame(
        {
            "fixed": pd.Series(fixed_values, dtype=object),
            "zone": pd.Series(zone_values, dtype=object),
        }
    )
    singleton_calls = 0
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        nonlocal singleton_calls
        if hasattr(values, "__len__") and len(values) == 1:  # type: ignore[arg-type]
            singleton_calls += 1
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(pa, "array", track_array)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.column(0).type == pa.timestamp("ns", tz="+08:00")
    assert table.column(1).type == pa.timestamp("us", tz="UTC")
    assert not pa.types.is_union(table.column(0).type)
    assert not pa.types.is_union(table.column(1).type)
    assert singleton_calls == 0


def test_distinct_timezone_semantics_remain_lossless_union_variants(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        datetime(2024, 1, 1, tzinfo=timezone(timedelta(hours=8))),
        datetime(2024, 1, 2, tzinfo=ZoneInfo.no_cache("UTC")),
    ]
    frame = pd.DataFrame({"when": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    column_type = table.column(0).type
    assert pa.types.is_union(column_type)
    assert {field.type.tz for field in column_type if pa.types.is_timestamp(field.type)} == {
        "+08:00",
        "UTC",
    }
    assert table.column(0).to_pylist() == values


def test_large_true_mixed_columns_use_arrow_calls_per_family_not_per_row(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    row_count = 4_000
    decimals = [
        Decimal("1E-30") if ordinal % 2 == 0 else Decimal("9" * 40)
        for ordinal in range(row_count // 2)
    ]
    moments = [
        datetime(2024, 1, 1, tzinfo=timezone(timedelta(hours=8))) + timedelta(minutes=ordinal)
        for ordinal in range(row_count // 2)
    ]
    decimal_values: list[object] = [
        value
        for pair in zip(decimals, (f"text-{index}" for index in range(len(decimals))), strict=True)
        for value in pair
    ]
    datetime_values: list[object] = [
        value
        for pair in zip(moments, (f"text-{index}" for index in range(len(moments))), strict=True)
        for value in pair
    ]
    frame = pd.DataFrame(
        {
            "decimal_mixed": pd.Series(decimal_values, dtype=object),
            "datetime_mixed": pd.Series(datetime_values, dtype=object),
        }
    )
    array_calls = 0
    singleton_calls = 0
    real_array = pa.array

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        nonlocal array_calls, singleton_calls
        array_calls += 1
        if hasattr(values, "__len__") and len(values) == 1:  # type: ignore[arg-type]
            singleton_calls += 1
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(pa, "array", track_array)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        table = workbook.to_arrow()

    assert table.num_rows == row_count
    assert all(pa.types.is_union(column.type) for column in table.columns)
    assert table.column(0).to_pylist() == decimal_values
    assert table.column(1).to_pylist() == datetime_values
    decimal_types = [
        field.type for field in table.column(0).type if pa.types.is_decimal(field.type)
    ]
    assert len(decimal_types) == 1
    assert decimal_types[0].precision >= 70
    assert decimal_types[0].scale == 30
    assert singleton_calls == 0
    assert array_calls <= 12


@pytest.mark.parametrize("entry", ["to_arrow", "_to_dataframes_compat", "_parse_sheet"])
def test_materialized_operation_release_retries_process_failure_and_unpoisons_workbook(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    entry: str,
) -> None:
    frame = pd.DataFrame({"value": [1]})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        real_end = workbook._end_operation
        end_calls = 0

        def fail_once(token: object) -> None:
            nonlocal end_calls
            end_calls += 1
            if end_calls == 1:
                raise MemoryError("materialized release interrupted")
            real_end(token)

        monkeypatch.setattr(workbook, "_end_operation", fail_once)
        with pytest.raises(MemoryError, match="materialized release interrupted"):
            if entry == "to_arrow":
                workbook.to_arrow()
            elif entry == "_to_dataframes_compat":
                workbook._to_dataframes_compat()
            else:
                workbook._parse_sheet("Data")

        assert end_calls == 2
        assert workbook._active_operation_token is None
        monkeypatch.setattr(workbook, "_end_operation", real_end)
        assert workbook.to_arrow().to_pydict() == {"value": [1]}


def test_materialized_operation_normal_success_releases_exact_token_once(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    frame = pd.DataFrame({"value": [1]})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        real_end = workbook._end_operation
        released: list[object] = []

        def record(token: object) -> None:
            released.append(token)
            real_end(token)

        monkeypatch.setattr(workbook, "_end_operation", record)
        workbook.to_arrow()

    assert len(released) == 1


def test_materialized_operation_foreign_token_is_a_noop(
    sample_xlsx: Path,
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    lease = workbook._materialized_operation()
    lease.__enter__()
    assert workbook._active_materialized_lease is lease
    foreign = object()
    workbook._active_operation_token = foreign
    lease.__exit__(None, None, None)

    assert workbook._active_operation_token is foreign
    assert workbook._active_materialized_lease is None
    workbook._end_operation(foreign)
    workbook.close()


def test_reentrant_materialized_operation_preserves_outer_lease(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        outer_lease: object | None = None

        def reenter(*_args: object, **_kwargs: object) -> pd.DataFrame:
            nonlocal outer_lease
            outer_lease = workbook._active_materialized_lease
            with pytest.raises(
                RuntimeError,
                match="already has an active parse or stream",
            ):
                workbook.to_arrow()
            assert workbook._active_materialized_lease is outer_lease
            raise ValueError("stop outer operation")

        monkeypatch.setattr(workbook, "_parse_sheet_unreserved", reenter)
        with pytest.raises(ValueError, match="stop outer operation"):
            workbook.to_arrow()

        assert outer_lease is not None
        assert workbook._active_operation_token is None
        assert workbook._active_materialized_lease is None


# Task 12 temporal follow-up: pandas chunks keep inert exact temporal labels and
# honor semantic aware-instant rename keys.


def test_dataframe_chunks_preserve_exact_stdlib_temporal_labels(
    sample_xlsx: Path,
) -> None:
    zone = ZoneInfo.no_cache("America/New_York")
    moment = datetime(2024, 11, 3, 1, 30, 0, 123456, tzinfo=zone, fold=1)
    elapsed = timedelta(days=2, microseconds=7)

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [[1, 2]],
                columns=pd.Index([moment, elapsed], dtype=object),
            )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=TemporalLabelRegistry(),
        ) as workbook,
        workbook.iter_dataframe_chunks(config=config) as chunks,
    ):
        chunk = next(chunks)

    restored = chunk.columns.tolist()
    assert type(restored[0]) is datetime
    assert restored[0].replace(tzinfo=None) == moment.replace(tzinfo=None)
    assert restored[0].fold == 1
    assert type(restored[0].tzinfo) is ZoneInfo
    assert type(restored[1]) is timedelta
    assert restored[1] == elapsed


def test_public_apis_apply_equivalent_aware_timestamp_rename_key(
    sample_xlsx: Path,
) -> None:
    label = pd.Timestamp(
        datetime(
            2024,
            1,
            2,
            8,
            tzinfo=timezone(timedelta(hours=8), "MALAYSIA"),
        )
    )

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([[1]], columns=pd.Index([label], dtype=object))

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
        column_renames={datetime(2024, 1, 2, tzinfo=ZoneInfo("UTC")): "when"},
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=TemporalLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            chunk = next(chunks)

    assert table.column_names == ["when"]
    assert chunk.columns.tolist() == ["when"]


def test_dataframe_label_sidecar_keeps_ordinary_rename_lookup_linear(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    width = 80
    source_names = tuple(f"source-{ordinal}" for ordinal in range(width))
    renamed_names = tuple(f"renamed-{ordinal}" for ordinal in range(width))
    plan = compile_parse_plan(
        SheetConfig(
            auto_detect=False,
            normalize=False,
            sanitize_column_names=False,
            column_renames=dict(zip(source_names, renamed_names, strict=True)),
        ),
        None,
        "xlsx",
    )
    match_calls = 0
    real_match = materialized_streaming_module._label_tokens_match

    def counted_match(left: object, right: object) -> bool:
        nonlocal match_calls
        match_calls += 1
        return real_match(left, right)  # type: ignore[arg-type]

    monkeypatch.setattr(
        materialized_streaming_module,
        "_label_tokens_match",
        counted_match,
    )

    restored = materialized_streaming_module._public_dataframe_display_names(
        source_names,
        renamed_names,
        plan,
    )

    assert restored == renamed_names
    assert match_calls <= width * 2


# Task 12 final acceptance: temporal sentinel parity, exact timezone values,
# Decimal bounds, and durable materialized-operation ownership.


def test_public_apis_apply_zoneinfo_time_pseudo_naive_config_semantics(
    sample_xlsx: Path,
) -> None:
    zone = ZoneInfo.no_cache("America/New_York")
    clock = time(9, 30, tzinfo=zone)
    nested = (time(10, 0, tzinfo=zone), "nested")

    class TimeLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [["1", "drop"], ["2", "keep"]],
                columns=pd.Index([clock, nested], dtype=object, tupleize_cols=False),
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={time(9, 30): "INTEGER"},
        column_renames={time(9, 30): "clock"},
        drop_conditions=[
            {
                "column": (time(10, 0), "nested"),
                "value": "drop",
            }
        ],
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=TimeLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)
        with workbook.iter_batches(config=config) as batches:
            streamed = pa.Table.from_batches(list(batches), schema=batches.schema)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            chunk = next(chunks)

    assert table.column_names[0] == streamed.column_names[0] == "clock"
    assert table.num_rows == streamed.num_rows == 1
    assert table.column(0).to_pylist() == streamed.column(0).to_pylist() == [2]
    assert chunk.iloc[:, 0].tolist() == [2]
    assert chunk.columns[0] == "clock"
    assert chunk.columns[1] == nested


def test_public_apis_preserve_nat_label_tuple_config_and_chunk_sidecar(
    sample_xlsx: Path,
) -> None:
    nested = (pd.NaT, "nested")

    class NatLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [["drop", "1"], ["keep", "2"]],
                columns=pd.Index([pd.NaT, nested], dtype=object, tupleize_cols=False),
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={pd.NaT: "TEXT", nested: "INTEGER"},
        column_renames={nested: "nested_nat"},
        drop_conditions=[{"column": pd.NaT, "value": "drop"}],
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=NatLabelRegistry(),
    ) as workbook:
        table = workbook.to_arrow(config=config)
        with workbook.iter_batches(config=config) as batches:
            streamed = pa.Table.from_batches(list(batches), schema=batches.schema)
        with workbook.iter_dataframe_chunks(config=config) as chunks:
            chunk = next(chunks)

    assert table.column_names == streamed.column_names == ["NaT", "nested_nat"]
    assert table.to_pylist() == streamed.to_pylist() == [{"NaT": "keep", "nested_nat": 2}]
    assert chunk.columns[0] is pd.NaT
    assert chunk.columns[1] == "nested_nat"
    assert chunk.iloc[0].tolist() == ["keep", 2]


@pytest.mark.parametrize(
    ("value_factory", "boundary", "bad_offset"),
    [
        (
            lambda offset: datetime(2024, 1, 2, tzinfo=timezone(offset)),
            "sample",
            timedelta(seconds=30),
        ),
        (
            lambda offset: pd.Timestamp(datetime(2024, 1, 2, tzinfo=timezone(offset))),
            "sample",
            timedelta(seconds=30, microseconds=1),
        ),
        (
            lambda offset: datetime(2024, 1, 2, tzinfo=timezone(offset)),
            "late",
            timedelta(seconds=30),
        ),
        (
            lambda offset: pd.Timestamp(datetime(2024, 1, 2, tzinfo=timezone(offset))),
            "late",
            timedelta(seconds=30, microseconds=1),
        ),
    ],
)
def test_exact_timezone_subminute_timestamp_is_rejected_with_context(
    sample_xlsx: Path,
    value_factory: Any,
    boundary: str,
    bad_offset: timedelta,
) -> None:
    bad_position = 0 if boundary == "sample" else 1_000
    values = [datetime(2024, 1, 1, tzinfo=UTC) for _ in range(bad_position)]
    values.append(value_factory(bad_offset))
    frame = pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    class TimestampRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with messy_xlsx.MessyWorkbook(
        sample_xlsx,
        registry=TimestampRegistry(),
    ) as workbook:
        if boundary == "sample":
            with pytest.raises(messy_xlsx.StreamingTypeError) as captured:
                workbook.iter_batches(batch_size=256, config=config)
        else:
            stream = workbook.iter_batches(batch_size=256, config=config)
            with stream, pytest.raises(messy_xlsx.StreamingTypeError) as captured:
                list(stream)

    assert captured.value.context["ordinal"] == 0
    assert captured.value.context["row_offset"] == bad_position
    assert captured.value.context["display_label"] == "str label(length=6)"
    assert captured.value.context["value_description"] == "datetime"


def test_materialized_exact_timezone_subminute_timestamp_is_rejected_with_context(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        datetime(2024, 1, 1, tzinfo=UTC),
        pd.Timestamp(
            datetime(
                2024,
                1,
                2,
                tzinfo=timezone(timedelta(seconds=30, microseconds=1)),
            )
        ),
    ]
    frame = pd.DataFrame({"moment": pd.Series(values, dtype=object)})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        with pytest.raises(messy_xlsx.StreamingTypeError) as captured:
            workbook.to_arrow()

    assert captured.value.context["ordinal"] == 0
    assert captured.value.context["row_offset"] == 1
    assert captured.value.context["value_description"] == "datetime"


def test_decimal_precision_limit_is_validated_before_arrow_group_conversion(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    valid = Decimal("9" * 76)
    invalid = Decimal("9" * 77)
    invalid_frame = pd.DataFrame({"amount": pd.Series([valid, invalid], dtype=object)})
    real_array = pa.array
    oversized_arrow_calls = 0

    def track_array(values: object, *args: object, **kwargs: object) -> pa.Array:
        nonlocal oversized_arrow_calls
        if isinstance(values, (list, pd.Series)) and any(
            type(value) is Decimal and len(value.as_tuple().digits) > 76 for value in values
        ):
            oversized_arrow_calls += 1
        return real_array(values, *args, **kwargs)

    monkeypatch.setattr(pa, "array", track_array)
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: invalid_frame,
        )
        with pytest.raises(ValueError, match="decimal256 precision"):
            workbook.to_arrow()

        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: pd.DataFrame({"amount": pd.Series([valid], dtype=object)}),
        )
        valid_table = workbook.to_arrow()

    assert oversized_arrow_calls == 0
    assert valid_table.column(0).type == pa.decimal256(76, 0)
    assert valid_table.column(0).to_pylist() == [valid]


def test_materialized_lease_is_retained_until_next_operation_releases_exact_token(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    frame = pd.DataFrame({"value": [1]})
    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        real_end = workbook._end_operation
        calls = 0

        def fail_twice(token: object) -> None:
            nonlocal calls
            calls += 1
            if calls <= 2:
                raise MemoryError("materialized release interrupted")
            real_end(token)

        monkeypatch.setattr(workbook, "_end_operation", fail_twice)
        with pytest.raises(MemoryError, match="materialized release interrupted"):
            workbook.to_arrow()

        assert workbook._active_operation_token is not None
        assert workbook._active_materialized_lease is not None
        assert workbook.to_arrow().to_pydict() == {"value": [1]}
        assert workbook._active_operation_token is None
        assert workbook._active_materialized_lease is None


def test_workbook_close_retries_materialized_lease_after_arbitrary_process_failures(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    monkeypatch.setattr(
        workbook,
        "_parse_sheet_unreserved",
        lambda *_args, **_kwargs: pd.DataFrame({"value": [1]}),
    )
    real_end = workbook._end_operation
    calls = 0

    def fail_four_times(token: object) -> None:
        nonlocal calls
        calls += 1
        if calls <= 4:
            raise MemoryError("materialized release interrupted")
        real_end(token)

    monkeypatch.setattr(workbook, "_end_operation", fail_four_times)
    with pytest.raises(MemoryError, match="materialized release interrupted"):
        workbook.to_arrow()
    with pytest.raises(MemoryError, match="materialized release interrupted"):
        workbook.close()
    workbook.close()

    assert calls == 5
    assert workbook._active_operation_token is None
    assert workbook._active_materialized_lease is None


@pytest.mark.parametrize(
    ("body_error", "cleanup_error", "expected_type"),
    [
        (ValueError("body failed"), MemoryError("cleanup interrupted"), MemoryError),
        (MemoryError("body interrupted"), OSError("cleanup failed"), MemoryError),
    ],
)
def test_materialized_lease_preserves_body_cleanup_precedence(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    body_error: BaseException,
    cleanup_error: BaseException,
    expected_type: type[BaseException],
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    real_end = workbook._end_operation

    def fail_body(*_args: object, **_kwargs: object) -> pd.DataFrame:
        raise body_error

    def fail_cleanup(_token: object) -> None:
        raise cleanup_error

    monkeypatch.setattr(workbook, "_parse_sheet_unreserved", fail_body)
    monkeypatch.setattr(workbook, "_end_operation", fail_cleanup)
    with pytest.raises(expected_type):
        workbook.to_arrow()

    monkeypatch.setattr(workbook, "_end_operation", real_end)
    workbook.close()


# Task 12 final semantic/performance remediation: exact pandas physical
# transport and bounded mixed-materialized conversion planning.


def test_streaming_custom_frame_preserves_pandas_nanoseconds_in_sample_and_late_rows(
    sample_xlsx: Path,
) -> None:
    row_count = 1_001
    timestamps = [pd.Timestamp("2024-01-01 00:00:00.000000001", tz="UTC") for _ in range(row_count)]
    durations = [pd.Timedelta(nanoseconds=1) for _ in range(row_count)]
    timestamps[-1] = pd.Timestamp("2024-01-02 00:00:00.000000002", tz="UTC")
    durations[-1] = pd.Timedelta(days=1, nanoseconds=2)

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                {
                    "moment": pd.Series(timestamps, dtype=object),
                    "elapsed": pd.Series(durations, dtype=object),
                }
            )

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with (
            messy_xlsx.MessyWorkbook(
                sample_xlsx,
                registry=TemporalRegistry(),
            ) as workbook,
            workbook.iter_batches(batch_size=257, config=config) as batches,
        ):
            table = pa.Table.from_batches(list(batches), schema=batches.schema)
            with workbook.iter_dataframe_chunks(
                batch_size=257,
                config=config,
            ) as chunks:
                frame = pd.concat(list(chunks))

    assert table.schema.types == [pa.timestamp("ns", tz="UTC"), pa.duration("ns")]
    assert table.column(0)[0].as_py() == timestamps[0]
    assert table.column(0)[row_count - 1].as_py() == timestamps[-1]
    assert table.column(1)[0].as_py() == durations[0]
    assert table.column(1)[row_count - 1].as_py() == durations[-1]
    assert frame.iloc[0, 0] == timestamps[0]
    assert frame.iloc[-1, 0] == timestamps[-1]
    assert frame.iloc[0, 1] == durations[0]
    assert frame.iloc[-1, 1] == durations[-1]


def test_physical_codec_round_trips_out_of_nanosecond_range_pandas_timestamp() -> None:
    value = pd.Timestamp("2500-01-02 03:04:05.123456")

    encoded = physical_values_module.encode_physical_value(value)
    decoded = physical_values_module.decode_physical_value(encoded)

    assert type(decoded) is pd.Timestamp
    assert decoded == value
    assert decoded.unit == value.unit


def test_materialized_union_limit_is_checked_before_any_arrow_conversion(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        datetime(
            2024,
            1,
            1,
            tzinfo=timezone(timedelta(minutes=minute)),
        )
        for minute in range(-64, 64)
    ]
    calls = 0
    real_array = pa.array

    def count_array(*args: object, **kwargs: object) -> pa.Array:
        nonlocal calls
        calls += 1
        return real_array(*args, **kwargs)

    monkeypatch.setattr(pa, "array", count_array)
    with pytest.raises(ValueError, match="union type limit"):
        workbook_module._materialized_arrow_array(
            pd.Series(values, dtype=object),
            ordinal=0,
            display_label="moment",
        )

    assert calls == 0


def test_materialized_union_coalesces_supported_scalar_subclasses_without_hooks(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    callbacks: list[str] = []

    class ArmedStr(str):
        armed = False

        def __str__(self) -> str:
            if self.armed:
                callbacks.append("str")
                raise AssertionError("string subclass hook executed")
            return str.__str__(self)

    subclasses = tuple(type(f"StringVariant{index}", (ArmedStr,), {}) for index in range(128))
    values = [subclass(f"value-{index}") for index, subclass in enumerate(subclasses)]
    ArmedStr.armed = True
    real_array = pa.array
    calls = 0

    def count_array(*args: object, **kwargs: object) -> pa.Array:
        nonlocal calls
        calls += 1
        return real_array(*args, **kwargs)

    monkeypatch.setattr(pa, "array", count_array)
    result = workbook_module._materialized_arrow_array(
        pd.Series(values, dtype=object),
        ordinal=0,
        display_label="text",
    )

    assert result.type == pa.string()
    assert result.to_pylist() == [f"value-{index}" for index in range(128)]
    assert calls <= 2
    assert callbacks == []


# Task 12 final lifecycle remediation: public construction/return and release
# gaps always have an independent retry owner.


def test_top_level_batches_preowns_real_workbook_before_constructor_return(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = io.BytesIO(sample_xlsx.read_bytes())
    source.seek(37)
    entry = source.tell()
    target_code = workbook_module.MessyWorkbook.__init__.__code__
    real_close = workbook_module.MessyWorkbook.close
    close_calls = 0
    interrupted = False

    def track_close(workbook: Any) -> None:
        nonlocal close_calls
        close_calls += 1
        real_close(workbook)

    def interrupt_constructor_return(frame: Any, event: str, _arg: object) -> Any:
        nonlocal interrupted
        if frame.f_code is target_code and event == "return" and not interrupted:
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("workbook constructor return interrupted")
        return interrupt_constructor_return

    monkeypatch.setattr(workbook_module.MessyWorkbook, "close", track_close)
    sys.settrace(interrupt_constructor_return)
    try:
        with pytest.raises(MemoryError, match="constructor return interrupted"):
            messy_xlsx.read_excel_batches(
                source,
                filename=sample_xlsx.name,
            )
    finally:
        sys.settrace(None)

    assert interrupted
    assert close_calls == 1
    assert source.tell() == entry
    assert source.closed is False


def test_top_level_batches_return_gap_finalizer_closes_child_before_workbook(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    events: list[str] = []
    target_code = messy_xlsx.read_excel_batches.__code__
    real_stream_close = messy_xlsx.BatchStream.close
    real_workbook_close = workbook_module.MessyWorkbook.close
    interrupted = False

    def track_stream_close(stream: Any) -> None:
        events.append("child")
        real_stream_close(stream)

    def track_workbook_close(workbook: Any) -> None:
        events.append("workbook")
        real_workbook_close(workbook)

    def interrupt_public_return(frame: Any, event: str, arg: object) -> Any:
        nonlocal interrupted
        if frame.f_code is target_code and event == "return" and not interrupted:
            interrupted = True
            del arg
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("batch convenience return interrupted")
        return interrupt_public_return

    monkeypatch.setattr(messy_xlsx.BatchStream, "close", track_stream_close)
    monkeypatch.setattr(workbook_module.MessyWorkbook, "close", track_workbook_close)
    sys.settrace(interrupt_public_return)
    try:
        with pytest.raises(MemoryError, match="convenience return interrupted"):
            messy_xlsx.read_excel_batches(sample_xlsx)
    finally:
        sys.settrace(None)
    gc.collect()

    assert interrupted
    assert events[:2] == ["child", "workbook"]


def test_materialized_release_is_armed_before_exit_first_line(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    monkeypatch.setattr(
        workbook,
        "_parse_sheet_unreserved",
        lambda *_args, **_kwargs: pd.DataFrame({"value": [1]}),
    )
    target_code = workbook_module._MaterializedOperationLease.__exit__.__code__
    interrupted = False

    def interrupt_exit(frame: Any, event: str, _arg: object) -> Any:
        nonlocal interrupted
        if frame.f_code is target_code and event == "line" and not interrupted:
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("materialized exit interrupted")
        return interrupt_exit

    sys.settrace(interrupt_exit)
    try:
        with pytest.raises(MemoryError, match="materialized exit interrupted"):
            workbook.to_arrow()
    finally:
        sys.settrace(None)

    try:
        assert workbook.to_arrow().to_pydict() == {"value": [1]}
        assert workbook._active_operation_token is None
    finally:
        workbook.close()


@pytest.mark.parametrize("entry", ["to_arrow", "_to_dataframes_compat", "_parse_sheet"])
def test_materialized_enter_return_interruption_is_recovered_by_next_operation(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    entry: str,
) -> None:
    workbook = messy_xlsx.MessyWorkbook(sample_xlsx)
    monkeypatch.setattr(
        workbook,
        "_parse_sheet_unreserved",
        lambda *_args, **_kwargs: pd.DataFrame({"value": [1]}),
    )
    target_code = workbook_module._MaterializedOperationLease.__enter__.__code__
    interrupted = False

    def interrupt_enter_return(frame: Any, event: str, _arg: object) -> Any:
        nonlocal interrupted
        if frame.f_code is target_code and event == "return" and not interrupted:
            interrupted = True
            sys.settrace(None)
            frame.f_trace = None
            raise MemoryError("materialized enter return interrupted")
        return interrupt_enter_return

    sys.settrace(interrupt_enter_return)
    try:
        with pytest.raises(MemoryError, match="enter return interrupted"):
            if entry == "to_arrow":
                workbook.to_arrow()
            elif entry == "_to_dataframes_compat":
                workbook._to_dataframes_compat()
            else:
                workbook._parse_sheet("Data")
    finally:
        sys.settrace(None)

    try:
        assert interrupted
        assert workbook.to_arrow().to_pydict() == {"value": [1]}
    finally:
        workbook.close()


def test_to_arrow_preserves_out_of_stdlib_range_pandas_timestamp_storage(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        pd.Timestamp(-62_198_755_200, unit="s"),
        pd.Timestamp(253_402_300_800, unit="s"),
    ]
    frame = pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
            monkeypatch.setattr(
                workbook,
                "_parse_sheet_unreserved",
                lambda *_args, **_kwargs: frame,
            )
            table = workbook.to_arrow()

    assert table.column(0).type == pa.timestamp("us")
    assert table.column(0).cast(pa.int64()).to_pylist() == [
        -62_198_755_200_000_000,
        253_402_300_800_000_000,
    ]


def test_to_arrow_splits_lossless_pandas_timestamp_units_into_bounded_union(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = [
        pd.Timestamp(1, unit="ns"),
        pd.Timestamp(253_402_300_800, unit="s"),
    ]
    frame = pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
            monkeypatch.setattr(
                workbook,
                "_parse_sheet_unreserved",
                lambda *_args, **_kwargs: frame,
            )
            table = workbook.to_arrow()

    union = table.column(0).chunk(0)
    assert pa.types.is_union(union.type)
    temporal_children = {
        child.type.unit: child.cast(pa.int64()).to_pylist()
        for child in (union.field(index) for index in range(union.type.num_fields))
        if pa.types.is_timestamp(child.type)
    }
    assert temporal_children == {"ns": [1], "us": [253_402_300_800_000_000]}
    assert union.type.num_fields <= 4


@pytest.mark.parametrize(
    ("zone", "expected_timezone"),
    [
        (timezone(timedelta(hours=8)), "+08:00"),
        (ZoneInfo("UTC"), "UTC"),
    ],
)
def test_out_of_range_pandas_timestamp_keeps_timezone_metadata_and_raw_utc(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
    zone: timezone | ZoneInfo,
    expected_timezone: str,
) -> None:
    value = pd.Timestamp(253_402_300_800, unit="s", tz=zone)
    frame = pd.DataFrame({"moment": pd.Series([value], dtype=object)})

    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
            monkeypatch.setattr(
                workbook,
                "_parse_sheet_unreserved",
                lambda *_args, **_kwargs: frame,
            )
            table = workbook.to_arrow()

    assert table.column(0).type == pa.timestamp("us", tz=expected_timezone)
    assert table.column(0).cast(pa.int64()).to_pylist() == [253_402_300_800_000_000]


def test_streaming_preserves_out_of_range_pandas_timestamp_sample_and_late_rows(
    sample_xlsx: Path,
) -> None:
    values = [pd.Timestamp(-62_198_755_200, unit="s") for _ in range(1_001)]
    values[-1] = pd.Timestamp(253_402_300_800, unit="s")

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(  # noqa: SIM117
            sample_xlsx,
            registry=TemporalRegistry(),
        ) as workbook:
            with workbook.iter_batches(batch_size=257, config=config) as batches:
                table = pa.Table.from_batches(list(batches), schema=batches.schema)

    assert table.column(0).type == pa.timestamp("us")
    assert table.column(0).cast(pa.int64())[0].as_py() == -62_198_755_200_000_000
    assert table.column(0).cast(pa.int64())[-1].as_py() == 253_402_300_800_000_000


def test_hostile_pandas_timedelta_subclass_keeps_one_nanosecond_without_hooks(
    sample_xlsx: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    callbacks: list[str] = []

    class HostileTimedelta(pd.Timedelta):
        armed = False

        @property
        def asm8(self) -> object:
            if self.armed:
                callbacks.append("asm8")
                raise AssertionError("hostile asm8 callback executed")
            return super().asm8

        @property
        def unit(self) -> str:
            if self.armed:
                callbacks.append("unit")
                raise AssertionError("hostile unit callback executed")
            return super().unit

        @property
        def value(self) -> int:
            if self.armed:
                callbacks.append("value")
                raise AssertionError("hostile value callback executed")
            return super().value

    value = HostileTimedelta(1, unit="ns")
    frame = pd.DataFrame({"elapsed": pd.Series([value], dtype=object)})
    HostileTimedelta.armed = True

    with messy_xlsx.MessyWorkbook(sample_xlsx) as workbook:
        monkeypatch.setattr(
            workbook,
            "_parse_sheet_unreserved",
            lambda *_args, **_kwargs: frame,
        )
        materialized = workbook.to_arrow()

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=TemporalRegistry()) as workbook,
        workbook.iter_batches(config=config) as batches,
    ):
        streamed = pa.Table.from_batches(list(batches), schema=batches.schema)

    assert materialized.column(0).type == pa.duration("ns")
    assert materialized.column(0).cast(pa.int64()).to_pylist() == [1]
    assert streamed.column(0).type == pa.duration("ns")
    assert streamed.column(0).cast(pa.int64()).to_pylist() == [1]
    assert callbacks == []


def test_negative_year_pandas_label_drives_hint_rename_condition_and_chunk(
    sample_xlsx: Path,
) -> None:
    negative = pd.Timestamp(-62_198_755_200, unit="s")
    future = pd.Timestamp(253_402_300_800, unit="s")
    labels = pd.Index([negative, future], dtype=object)

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame([["1", "drop"], ["2", "keep"]], columns=labels)

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={negative: "INTEGER"},
        column_renames={future: "future"},
        drop_conditions=[{"column": negative, "value": 1}],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with (
            messy_xlsx.MessyWorkbook(
                sample_xlsx,
                registry=TemporalLabelRegistry(),
            ) as workbook,
            workbook.iter_dataframe_chunks(config=config) as chunks,
        ):
            chunk = next(chunks)

    assert chunk.iloc[:, 0].tolist() == [2]
    assert chunk.iloc[:, 1].tolist() == ["keep"]
    assert type(chunk.columns[0]) is pd.Timestamp
    assert chunk.columns[0].unit == "s"
    assert int(chunk.columns[0].asm8.view("i8")) == -62_198_755_200
    assert chunk.columns[1] == "future"


def test_streaming_rejects_late_pandas_timestamp_that_cannot_fit_sampled_unit(
    sample_xlsx: Path,
) -> None:
    values = [pd.Timestamp(1, unit="ns") for _ in range(1_001)]
    values[-1] = pd.Timestamp(253_402_300_800, unit="s")

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=TemporalRegistry()) as workbook,
        workbook.iter_batches(batch_size=257, config=config) as batches,
        pytest.raises(messy_xlsx.StreamingTypeError) as captured,
    ):
        list(batches)

    assert captured.value.context["ordinal"] == 0
    assert captured.value.context["row_offset"] == 1_000
    assert captured.value.context["expected_type"] == "timestamp[ns]"


def test_streaming_mixed_stdlib_and_pandas_timestamps_use_lossless_raw_buffers(
    sample_xlsx: Path,
) -> None:
    values = [
        datetime(2024, 1, 1),
        pd.Timestamp("2024-01-02 00:00:00.000000001"),
    ]

    class TemporalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame({"moment": pd.Series(values, dtype=object)})

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    with (
        messy_xlsx.MessyWorkbook(sample_xlsx, registry=TemporalRegistry()) as workbook,
        workbook.iter_batches(config=config) as batches,
    ):
        table = pa.Table.from_batches(list(batches), schema=batches.schema)

    assert table.column(0).type == pa.timestamp("ns")
    assert table.column(0).cast(pa.int64()).to_pylist() == [
        1_704_067_200_000_000_000,
        1_704_153_600_000_000_001,
    ]


# Task 12 final semantic acceptance: public label text must not route trusted
# extended-range pandas timestamps through pandas or stdlib datetime formatters.


def test_extended_timestamp_payload_has_one_hook_free_label_formatter(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    callbacks: list[str] = []

    class HostileTimestamp(pd.Timestamp):
        armed = False

        def __str__(self) -> str:
            if self.armed:
                callbacks.append("str")
                raise AssertionError("timestamp text hook executed")
            return pd.Timestamp.__str__(self)

        def isoformat(self, *_args: object, **_kwargs: object) -> str:
            if self.armed:
                callbacks.append("isoformat")
                raise AssertionError("timestamp ISO hook executed")
            return pd.Timestamp.isoformat(self, *_args, **_kwargs)

        def to_pydatetime(self, *_args: object, **_kwargs: object) -> datetime:
            if self.armed:
                callbacks.append("to_pydatetime")
                raise AssertionError("timestamp conversion hook executed")
            return pd.Timestamp.to_pydatetime(self, *_args, **_kwargs)

    label = HostileTimestamp(
        -62_198_755_200,
        unit="s",
        tz=ZoneInfo.no_cache("UTC"),
    )
    HostileTimestamp.armed = True
    payload = physical_values_module.pandas_temporal_payload(label)

    assert payload is not None
    assert physical_values_module.pandas_timestamp_label_text(label) == "-001-01-01 00:00:00+00:00"

    class ForbiddenDatetime:
        def __new__(cls, *_args: object, **_kwargs: object) -> object:
            raise AssertionError("stdlib datetime conversion executed")

    monkeypatch.setattr(physical_values_module, "datetime", ForbiddenDatetime)
    assert (
        physical_values_module.pandas_timestamp_payload_label_text(payload)
        == "-001-01-01 00:00:00+00:00"
    )
    assert callbacks == []


def test_public_apis_format_zoneinfo_extended_timestamp_labels_positionally(
    sample_xlsx: Path,
) -> None:
    zone = ZoneInfo.no_cache("UTC")
    negative = pd.Timestamp(-62_198_755_200, unit="s", tz=zone)
    future = pd.Timestamp(253_402_300_800, unit="s", tz=zone)
    condition_label = pd.Timestamp(253_402_300_801, unit="s", tz=zone)
    labels = pd.Index(
        [negative, negative, future, condition_label],
        dtype=object,
    )

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [
                    ["1", "10", "before", "drop"],
                    ["2", "20", "after", "keep"],
                ],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={negative: "INTEGER"},
        column_renames={future: "future"},
        drop_conditions=[{"column": condition_label, "value": "drop"}],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=TemporalLabelRegistry(),
        ) as workbook:
            materialized = workbook.to_arrow(config=config)
            with workbook.iter_batches(config=config) as batches:
                streamed = pa.Table.from_batches(
                    list(batches),
                    schema=batches.schema,
                )
            with workbook.iter_dataframe_chunks(config=config) as chunks:
                chunk = next(chunks)

    expected_names = [
        "-001-01-01 00:00:00+00:00",
        "-001-01-01 00:00:00+00:00",
        "future",
        "10000-01-01 00:00:01+00:00",
    ]
    assert materialized.column_names == streamed.column_names == expected_names
    for table in (materialized, streamed):
        assert table.column(0).to_pylist() == [2]
        assert table.column(1).to_pylist() == [20]
        assert table.column(2).to_pylist() == ["after"]
        assert table.column(3).to_pylist() == ["keep"]

    restored = chunk.columns.tolist()
    assert restored[0] is negative
    assert restored[1] is negative
    assert restored[2] == "future"
    assert restored[3] is condition_label
    assert all(type(restored[ordinal]) is pd.Timestamp for ordinal in (0, 1, 3))
    assert all(type(restored[ordinal].tzinfo) is ZoneInfo for ordinal in (0, 1, 3))
    assert [chunk.iloc[0, ordinal] for ordinal in range(4)] == [
        2,
        20,
        "after",
        "keep",
    ]


def test_public_apis_resolve_exact_complex_labels_by_python_numeric_semantics(
    sample_xlsx: Path,
) -> None:
    labels = pd.Index(
        [1 + 2j, 3 + 4j, 1 + 0j],
        dtype=object,
    )

    class ComplexLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [
                    ["1", "left", "drop"],
                    ["2", "right", "keep"],
                ],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={1 + 2j: "INTEGER", 1: "TEXT"},
        column_renames={1 + 2j: "selected"},
        drop_conditions=[{"column": 1, "value": "drop"}],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=ComplexLabelRegistry(),
        ) as workbook:
            materialized = workbook.to_arrow(config=config)
            with workbook.iter_batches(config=config) as batches:
                streamed = pa.Table.from_batches(list(batches), schema=batches.schema)
            with workbook.iter_dataframe_chunks(config=config) as chunks:
                chunk = next(chunks)

    expected_names = ["selected", "(3+4j)", "(1+0j)"]
    assert materialized.column_names == streamed.column_names == expected_names
    for table in (materialized, streamed):
        assert table.schema.types == [pa.int64(), pa.string(), pa.string()]
        assert table.to_pylist() == [
            {
                "selected": 2,
                "(3+4j)": "right",
                "(1+0j)": "keep",
            }
        ]
    assert chunk.columns.tolist() == ["selected", 3 + 4j, 1 + 0j]
    assert chunk.iloc[0].tolist() == [2, "right", "keep"]


def test_public_apis_preserve_nested_non_reflexive_complex_label_identity(
    sample_xlsx: Path,
) -> None:
    first_complex = complex(float("nan"), -0.0)
    second_complex = complex(float("nan"), -0.0)
    first_label = (first_complex, "nested")
    second_label = (second_complex, "nested")
    labels = pd.Index(
        [first_label, second_label],
        dtype=object,
        tupleize_cols=False,
    )

    class ComplexLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [["1", "drop"], ["2", "keep"]],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={first_label: "INTEGER"},
        column_renames={first_label: "selected"},
        drop_conditions=[{"column": second_label, "value": "drop"}],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=ComplexLabelRegistry(),
        ) as workbook:
            materialized = workbook.to_arrow(config=config)
            with workbook.iter_batches(config=config) as batches:
                streamed = pa.Table.from_batches(list(batches), schema=batches.schema)
            with workbook.iter_dataframe_chunks(config=config) as chunks:
                chunk = next(chunks)

    second_name = "((nan-0j), 'nested')"
    assert (
        materialized.column_names
        == streamed.column_names
        == [
            "selected",
            second_name,
        ]
    )
    for table in (materialized, streamed):
        assert table.schema.types == [pa.int64(), pa.string()]
        assert table.to_pylist() == [{"selected": 2, second_name: "keep"}]
    restored = chunk.columns.tolist()
    assert restored[0] == "selected"
    assert restored[1] is second_label
    assert restored[1][0] is second_complex
    assert chunk.iloc[0].tolist() == [2, "keep"]


def test_public_apis_format_nested_extended_timestamp_labels_recursively(
    sample_xlsx: Path,
) -> None:
    zone = ZoneInfo.no_cache("UTC")
    negative = pd.Timestamp(-62_198_755_200, unit="s", tz=zone)
    future = pd.Timestamp(253_402_300_800, unit="s", tz=zone)
    condition_timestamp = pd.Timestamp(253_402_300_801, unit="s", tz=zone)
    negative_label = (negative, "negative")
    future_label = (future, "future")
    condition_label = (condition_timestamp, "condition")
    labels = pd.Index(
        [negative_label, future_label, condition_label],
        dtype=object,
        tupleize_cols=False,
    )

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [
                    ["1", "before", "drop"],
                    ["2", "after", "keep"],
                ],
                columns=labels,
            )

    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        type_hints={negative_label: "INTEGER"},
        column_renames={future_label: "future"},
        drop_conditions=[{"column": condition_label, "value": "drop"}],
    )
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        with messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=TemporalLabelRegistry(),
        ) as workbook:
            materialized = workbook.to_arrow(config=config)
            with workbook.iter_batches(config=config) as batches:
                streamed = pa.Table.from_batches(list(batches), schema=batches.schema)
            with workbook.iter_dataframe_chunks(config=config) as chunks:
                chunk = next(chunks)

    expected_names = [
        "(Timestamp('-001-01-01 00:00:00+00:00'), 'negative')",
        "future",
        "(Timestamp('10000-01-01 00:00:01+00:00'), 'condition')",
    ]
    assert materialized.column_names == streamed.column_names == expected_names
    for table in (materialized, streamed):
        assert table.schema.types == [pa.int64(), pa.string(), pa.string()]
        assert table.to_pylist() == [
            {
                expected_names[0]: 2,
                "future": "after",
                expected_names[2]: "keep",
            }
        ]
    restored = chunk.columns.tolist()
    assert restored[0] is negative_label
    assert restored[1] == "future"
    assert restored[2] is condition_label
    assert chunk.iloc[0].tolist() == [2, "after", "keep"]


@pytest.mark.parametrize("config_role", ["type_hint", "rename", "condition"])
@pytest.mark.parametrize(
    "api_name",
    ["to_arrow", "iter_batches", "iter_dataframe_chunks"],
)
def test_materialized_and_streaming_apis_share_dst_fold_ambiguity_policy(
    sample_xlsx: Path,
    config_role: str,
    api_name: str,
) -> None:
    zone = ZoneInfo.no_cache("America/New_York")
    fold_zero = pd.Timestamp(datetime(2024, 11, 3, 1, 30, tzinfo=zone, fold=0))
    fold_one = pd.Timestamp(datetime(2024, 11, 3, 1, 30, tzinfo=zone, fold=1))
    stdlib_key = datetime(2024, 11, 3, 1, 30, tzinfo=zone, fold=0)
    labels = pd.Index([fold_zero, fold_one], dtype=object)

    class TemporalLabelRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return pd.DataFrame(
                [["drop", "drop"], ["keep", "keep"]],
                columns=labels,
            )

    role_config: dict[str, object]
    if config_role == "type_hint":
        role_config = {"type_hints": {stdlib_key: "TEXT"}}
    elif config_role == "rename":
        role_config = {"column_renames": {stdlib_key: "ambiguous"}}
    else:
        role_config = {
            "drop_conditions": [
                {
                    "column": stdlib_key,
                    "value": "drop",
                }
            ]
        }
    config = SheetConfig(
        auto_detect=False,
        sanitize_column_names=False,
        **role_config,
    )

    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=TemporalLabelRegistry(),
        ) as workbook,
        pytest.raises(ValueError, match="ambiguous temporal label configuration"),
    ):
        if api_name == "to_arrow":
            workbook.to_arrow(config=config)
        elif api_name == "iter_batches":
            with workbook.iter_batches(config=config) as batches:
                list(batches)
        else:
            with workbook.iter_dataframe_chunks(config=config) as chunks:
                list(chunks)


@pytest.mark.parametrize(
    "entry",
    ["to_arrow", "iter_batches", "iter_dataframe_chunks"],
)
@pytest.mark.parametrize(
    ("boundary", "bad_position"),
    [("sample", 1), ("late", 1_000)],
)
def test_public_arrow_apis_reject_unsupported_hostile_physical_values_without_hooks(
    sample_xlsx: Path,
    entry: str,
    boundary: str,
    bad_position: int,
) -> None:
    callbacks: list[str] = []

    class ArmedUnsupported:
        armed = False

        def _record(self, name: str) -> None:
            if self.armed:
                callbacks.append(name)

        def __repr__(self) -> str:
            self._record("repr")
            return "<unsupported>"

        def __str__(self) -> str:
            self._record("str")
            return "<unsupported>"

        def __hash__(self) -> int:
            self._record("hash")
            return object.__hash__(self)

        def __eq__(self, other: object) -> bool:
            self._record("eq")
            return self is other

    hostile = ArmedUnsupported()
    hostile_values: list[object] = [0] * bad_position
    hostile_values.append(hostile)
    frame = pd.DataFrame(
        {
            "safe": pd.Series(range(len(hostile_values)), dtype=object),
            "hostile value": pd.Series(hostile_values, dtype=object),
        }
    )

    class HostilePhysicalRegistry(HandlerRegistry):
        def parse(self, *_args: object, **_kwargs: object) -> pd.DataFrame:
            return frame

    config = SheetConfig(
        auto_detect=False,
        normalize=False,
        sanitize_column_names=False,
    )
    ArmedUnsupported.armed = True
    callbacks.clear()
    with (
        messy_xlsx.MessyWorkbook(
            sample_xlsx,
            registry=HostilePhysicalRegistry(),
        ) as workbook,
        pytest.raises(messy_xlsx.StreamingTypeError) as captured,
    ):
        if entry == "to_arrow":
            workbook.to_arrow(config=config)
        else:
            stream = (
                workbook.iter_batches(batch_size=257, config=config)
                if entry == "iter_batches"
                else workbook.iter_dataframe_chunks(batch_size=257, config=config)
            )
            with stream:
                list(stream)

    assert captured.value.context == {
        "expected_type": "supported Arrow scalar",
        "ordinal": 1,
        "display_label": "str label(length=13)",
        "row_offset": bad_position,
        "value_description": "unsupported value",
    }
    assert callbacks == []


@pytest.mark.parametrize(
    "value",
    [
        None,
        "text",
        True,
        7,
        1.25,
        datetime(2024, 1, 2, 3, 4, 5),
        date(2024, 1, 2),
        time(3, 4, 5),
        timedelta(days=1, microseconds=2),
        pd.Timestamp("2024-01-02 03:04:05.000000001"),
        pd.Timedelta(days=1, nanoseconds=2),
        b"bytes",
        bytearray(b"bytes"),
        memoryview(b"bytes"),
        Decimal("1.25"),
    ],
    ids=[
        "null",
        "string",
        "boolean",
        "integer",
        "floating",
        "datetime",
        "date",
        "time",
        "timedelta",
        "pandas-timestamp",
        "pandas-timedelta",
        "bytes",
        "bytearray",
        "memoryview",
        "decimal",
    ],
)
def test_physical_value_gate_accepts_every_exact_supported_base_scalar(
    value: object,
) -> None:
    physical_values_module.ensure_supported_physical_value(value)
