"""Characterization tests protecting public behavior during architectural refactors."""

import io
import xml.etree.ElementTree as ET
import zipfile
from types import MethodType

import openpyxl
import pandas as pd
from pandas.testing import assert_frame_equal

from messy_xlsx import MessyWorkbook, SheetConfig, read_all_sheets, read_excel
from messy_xlsx.formulas import FormulaConfig, FormulaEvaluationMode
from messy_xlsx.ooxml.manifest import ManifestReader
from messy_xlsx.parsing import HandlerRegistry, XLSXHandler
from messy_xlsx.parsing.xlsx_materialized import FastexcelMaterializedReader


def _set_formula_cache(path, cell_ref, formula, value):
    """Inject an Excel-style cached value into an openpyxl-created workbook."""
    replacement_path = path.with_name(f"{path.stem}-cached{path.suffix}")
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    qualified = f"{{{spreadsheet_namespace}}}"
    ET.register_namespace("", spreadsheet_namespace)
    replacement_made = False

    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(replacement_path, "w") as target,
    ):
        for entry in source.infolist():
            content = source.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                root = ET.fromstring(content)
                cell = root.find(f".//{qualified}c[@r='{cell_ref}']")
                formula_node = None if cell is None else cell.find(f"{qualified}f")
                if formula_node is not None and formula_node.text == formula:
                    value_node = cell.find(f"{qualified}v")
                    if value_node is None:
                        value_node = ET.SubElement(cell, f"{qualified}v")
                    value_node.text = str(value)
                    content = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                    replacement_made = True
            target.writestr(entry, content)

    assert replacement_made, "formula cache placeholder was not found"
    replacement_path.replace(path)


class TestPublicApiParity:
    """Equivalent public entry points should produce equivalent data."""

    def test_quick_read_matches_workbook_and_sheet_apis(self, sample_xlsx):
        quick = read_excel(sample_xlsx)

        with MessyWorkbook(sample_xlsx) as workbook:
            workbook_df = workbook.to_dataframe()
            sheet_df = workbook.get_sheet("Data").to_dataframe()

        assert_frame_equal(quick, workbook_df)
        assert_frame_equal(quick, sheet_df)

    def test_path_and_buffer_inputs_match(self, sample_xlsx):
        with MessyWorkbook(sample_xlsx) as workbook:
            from_path = workbook.to_dataframe()

        content = sample_xlsx.read_bytes()
        with MessyWorkbook(io.BytesIO(content), filename=sample_xlsx.name) as workbook:
            from_buffer = workbook.to_dataframe()

        assert_frame_equal(from_path, from_buffer)

    def test_multi_sheet_helper_matches_workbook_for_data_sheets(self, temp_dir):
        path = temp_dir / "two_sheets.xlsx"
        source = openpyxl.Workbook()
        first = source.active
        first.title = "First"
        first.append(["Name", "Value"])
        first.append(["A", 1])
        second = source.create_sheet("Second")
        second.append(["Name", "Value"])
        second.append(["B", 2])
        source.save(path)
        source.close()

        helper_results = read_all_sheets(path)
        with MessyWorkbook(path) as workbook:
            workbook_results = workbook.to_dataframes()

        assert helper_results.keys() == workbook_results.keys()
        for name in helper_results:
            assert_frame_equal(helper_results[name], workbook_results[name])


class TestStructureConfigurationContracts:
    """Detection configuration must not leak through structure caching."""

    def test_get_structure_respects_workbook_header_patterns(self, temp_dir):
        path = temp_dir / "pattern_cache.xlsx"
        source = openpyxl.Workbook()
        sheet = source.active
        sheet.append(["First field", "Second field", "Third field"])
        sheet.append([1, 2, 3])
        source.save(path)
        source.close()

        plain_config = SheetConfig(header_patterns=None)
        pattern_config = SheetConfig(
            header_patterns=[r"first field", r"second field", r"third field"]
        )

        with MessyWorkbook(path, sheet_config=plain_config) as workbook:
            plain = workbook.get_structure()
        with MessyWorkbook(path, sheet_config=pattern_config) as workbook:
            patterned = workbook.get_structure()

        assert patterned.header_confidence > plain.header_confidence


class TestFormulaBoundaryContracts:
    """Formula expressions and cached results are separate concepts."""

    def test_formula_expression_is_not_passed_as_cached_value(self, temp_dir):
        path = temp_dir / "formula.xlsx"
        source = openpyxl.Workbook()
        sheet = source.active
        sheet.append(["A", "B", "Sum"])
        sheet.append([10, 20, "=A2+B2"])
        source.save(path)
        source.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_WITH_FALLBACK)
        with MessyWorkbook(path, formula_config=config) as workbook:
            calls = []

            def evaluate(sheet_name, row, col, cached_value=None):
                calls.append((sheet_name, row, col, cached_value))
                return 30

            workbook._formula_engine.evaluate = evaluate
            cell = workbook.get_cell("Sheet", 2, 3)

        assert calls == [("Sheet", 2, 3, None)]
        assert cell.formula == "=A2+B2"
        assert cell.value == 30

    def test_cached_only_reads_excel_cached_result(self, temp_dir):
        path = temp_dir / "cached_formula.xlsx"
        source = openpyxl.Workbook()
        sheet = source.active
        sheet.append(["A", "B", "Sum"])
        sheet.append([10, 20, "=A2+B2"])
        source.save(path)
        source.close()
        _set_formula_cache(path, "C2", "A2+B2", 30)

        config = FormulaConfig(mode=FormulaEvaluationMode.CACHED_ONLY)
        with MessyWorkbook(path, formula_config=config) as workbook:
            cell = workbook.get_cell("Sheet", 2, 3)

        assert cell.formula == "=A2+B2"
        assert cell.value == 30

    def test_disabled_formula_mode_preserves_expression_from_buffer(self, temp_dir):
        path = temp_dir / "buffer_formula.xlsx"
        source = openpyxl.Workbook()
        sheet = source.active
        sheet["A1"] = "Formula"
        sheet["A2"] = "=1+1"
        source.save(path)
        source.close()

        config = FormulaConfig(mode=FormulaEvaluationMode.DISABLED)
        with MessyWorkbook(io.BytesIO(path.read_bytes()), formula_config=config) as workbook:
            cell = workbook.get_cell("Sheet", 2, 1)

        assert cell.formula == "=1+1"
        assert cell.value == "=1+1"


class TestRegistryContracts:
    """The registry supplied by callers must drive workbook parsing."""

    def test_workbook_uses_injected_registry(self, sample_xlsx):
        class RecordingRegistry(HandlerRegistry):
            def __init__(self):
                super().__init__()
                self.parsed = False

            def parse(self, *args, **kwargs):
                self.parsed = True
                return super().parse(*args, **kwargs)

        registry = RecordingRegistry()
        with MessyWorkbook(sample_xlsx, registry=registry) as workbook:
            result = workbook.to_dataframe()

        assert registry.parsed is True
        assert isinstance(result, pd.DataFrame)

    def test_registry_subclass_sentinel_is_never_bypassed(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"registry_sentinel": [101]})

        class SentinelRegistry(HandlerRegistry):
            def parse(self, *args, **kwargs):
                return expected.copy()

        registry = SentinelRegistry()
        with MessyWorkbook(sample_xlsx, registry=registry) as workbook:
            monkeypatch.setattr(
                FastexcelMaterializedReader,
                "read_table",
                lambda _self: (_ for _ in ()).throw(
                    AssertionError("registry subclass must stay on its DataFrame SPI")
                ),
            )
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)

    def test_custom_registry_coordinate_options_preserve_sentinel_output(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"coordinate_registry_sentinel": [404]})

        class SentinelRegistry(HandlerRegistry):
            def parse(self, *args, **kwargs):
                return expected.copy()

        monkeypatch.setattr(
            ManifestReader,
            "__init__",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(
                AssertionError("custom registry must bypass coordinate metadata")
            ),
        )
        registry = SentinelRegistry()
        config = SheetConfig(
            auto_detect=False,
            cell_range="A1:B2",
            include_hidden=False,
            merge_strategy="fill",
        )

        with MessyWorkbook(
            sample_xlsx,
            registry=registry,
            sheet_config=config,
        ) as workbook:
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)

    def test_registry_subclass_cannot_claim_builtin_eligibility(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"lying_registry_sentinel": [111]})

        class LyingRegistry(HandlerRegistry):
            def _uses_builtin_components(self):
                return True

            def parse(self, *args, **kwargs):
                return expected.copy()

        registry = LyingRegistry()
        with MessyWorkbook(sample_xlsx, registry=registry) as workbook:
            monkeypatch.setattr(
                FastexcelMaterializedReader,
                "read_table",
                lambda _self: (_ for _ in ()).throw(
                    AssertionError("registry subclasses cannot self-authorize the fast path")
                ),
            )
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)

    def test_custom_handler_sentinel_is_never_bypassed(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"handler_sentinel": [202]})

        class SentinelHandler(XLSXHandler):
            def parse(self, *args, **kwargs):
                return expected.copy()

        registry = HandlerRegistry(handlers=[SentinelHandler()])
        with MessyWorkbook(sample_xlsx, registry=registry) as workbook:
            monkeypatch.setattr(
                FastexcelMaterializedReader,
                "read_table",
                lambda _self: (_ for _ in ()).throw(
                    AssertionError("custom handler must stay on its DataFrame SPI")
                ),
            )
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)

    def test_custom_handler_coordinate_options_preserve_sentinel_output(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"coordinate_handler_sentinel": [505]})

        class SentinelHandler(XLSXHandler):
            def parse(self, *args, **kwargs):
                return expected.copy()

        monkeypatch.setattr(
            ManifestReader,
            "__init__",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(
                AssertionError("custom handler must bypass coordinate metadata")
            ),
        )
        config = SheetConfig(
            auto_detect=False,
            cell_range="A1:B2",
            include_hidden=False,
            merge_strategy="fill",
        )
        registry = HandlerRegistry(handlers=[SentinelHandler()])

        with MessyWorkbook(
            sample_xlsx,
            registry=registry,
            sheet_config=config,
        ) as workbook:
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)

    def test_builtin_handler_instance_override_sentinel_is_never_bypassed(
        self,
        sample_xlsx,
        monkeypatch,
    ):
        expected = pd.DataFrame({"instance_sentinel": [303]})
        registry = HandlerRegistry()
        handler = registry.get_handler("xlsx")
        assert isinstance(handler, XLSXHandler)

        def sentinel_parse(self, *args, **kwargs):
            return expected.copy()

        handler.parse = MethodType(sentinel_parse, handler)  # type: ignore[method-assign]
        with MessyWorkbook(sample_xlsx, registry=registry) as workbook:
            monkeypatch.setattr(
                FastexcelMaterializedReader,
                "read_table",
                lambda _self: (_ for _ in ()).throw(
                    AssertionError("instance override must stay on its DataFrame SPI")
                ),
            )
            actual = workbook.to_dataframe()

        assert_frame_equal(actual, expected)
