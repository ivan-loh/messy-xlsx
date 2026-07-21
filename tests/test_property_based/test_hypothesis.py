"""Property-based tests using hypothesis."""

import openpyxl
import pandas as pd
import pytest

pytest.importorskip("hypothesis")

from hypothesis import HealthCheck, assume, given, settings
from hypothesis import strategies as st
from openpyxl.utils.exceptions import IllegalCharacterError

from messy_xlsx import MessyWorkbook
from messy_xlsx.normalization import MissingValueHandler


class TestPropertyBased:
    """Property-based tests for robustness."""

    @given(st.integers(min_value=1, max_value=100))
    @settings(
        max_examples=10, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture]
    )
    def test_any_number_of_rows(self, temp_dir, num_rows):
        """Test parser handles any reasonable number of rows."""
        file_path = temp_dir / f"rows_{num_rows}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Col1", "Col2"])

        for i in range(num_rows):
            ws.append([i, i * 2])

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df) == num_rows

    @given(st.integers(min_value=1, max_value=50))
    @settings(
        max_examples=10, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture]
    )
    def test_any_number_of_columns(self, temp_dir, num_cols):
        """Test parser handles any reasonable number of columns."""
        file_path = temp_dir / f"cols_{num_cols}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        headers = [f"Col{i}" for i in range(num_cols)]
        ws.append(headers)
        ws.append(list(range(num_cols)))

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            assert len(df.columns) == num_cols

    @given(st.text(min_size=0, max_size=100))
    @settings(
        max_examples=20, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture]
    )
    def test_any_text_value(self, temp_dir, text_value):
        """Test parser handles any text value without crashing."""
        file_path = temp_dir / "text_test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Text"])

        # Skip illegal characters that openpyxl can't handle
        try:
            cell = ws.cell(row=2, column=1, value=text_value)
            # A leading "=" makes openpyxl serialize the value as a formula.
            # This property is specifically about arbitrary text, so preserve
            # the generated value's intended Excel cell type.
            cell.data_type = "s"
        except IllegalCharacterError:
            assume(False)  # Skip this example
            return

        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            # Empty/blank text values may be cleaned by missing-value normalization,
            # resulting in 0 rows. Non-empty values must produce exactly 1 row.
            # Some text values can legitimately produce 0 rows:
            # - Missing value patterns (na, null, none, etc.) get cleaned
            # - Footer-like words (sum, total, subtotal) get skipped by structure detection
            # - Empty/whitespace-only values get dropped
            # For all other values, we should get exactly 1 row.
            footer_words = {"sum", "total", "subtotal", "grand total"}
            missing_words = {
                str(value).strip().lower() for value in MissingValueHandler().missing_values
            }
            stripped = text_value.strip().lower()
            if stripped and stripped not in missing_words and stripped not in footer_words:
                assert len(df) == 1
            else:
                assert len(df) <= 1

    @given(st.floats(allow_nan=False, allow_infinity=False))
    @settings(
        max_examples=20, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture]
    )
    def test_any_numeric_value(self, temp_dir, number):
        """Test parser handles any valid number."""
        file_path = temp_dir / "number_test.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Number"])
        ws.append([number])
        wb.save(file_path)
        wb.close()

        with MessyWorkbook(file_path) as mwb:
            df = mwb.to_dataframe()
            # Must produce exactly 1 row
            assert len(df) == 1


class TestFuzzTesting:
    """Fuzz testing for robustness."""

    @given(
        st.lists(
            st.lists(
                st.one_of(st.none(), st.integers(), st.floats(), st.text(max_size=50)),
                min_size=1,
                max_size=10,
            ),
            min_size=2,
            max_size=20,
        )
    )
    @settings(
        max_examples=10,
        deadline=None,
        suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.filter_too_much],
    )
    def test_random_data_structure(self, temp_dir, data):
        """Test parser handles randomly structured data."""
        file_path = temp_dir / "fuzz.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # Skip illegal characters that openpyxl can't handle
        try:
            for row in data:
                ws.append(row)
        except IllegalCharacterError:
            assume(False)  # Skip this example
            return

        wb.save(file_path)
        wb.close()

        from messy_xlsx.exceptions import MessyXlsxError

        try:
            with MessyWorkbook(file_path) as mwb:
                df = mwb.to_dataframe()
                # Must produce a valid DataFrame with consistent shape
                assert isinstance(df, pd.DataFrame)
                assert all(isinstance(c, str) for c in df.columns)
        except MessyXlsxError:
            # Known library errors on unparseable data are acceptable
            pass
