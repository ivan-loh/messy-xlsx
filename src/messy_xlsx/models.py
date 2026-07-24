"""Data models for messy-xlsx."""

# ============================================================================
# Imports
# ============================================================================

from dataclasses import dataclass, field
from typing import Any

import pandas as pd

from messy_xlsx.enums import (
    DataType,
    FormatType,
    HeaderDetectionMode,
    HeaderFallback,
    MergeStrategy,
)

# ============================================================================
# Format Models
# ============================================================================


@dataclass
class FormatInfo:
    """Information about detected file format."""

    format_type: FormatType | str
    confidence: float = 1.0
    version: str | None = None
    encoding: str | None = None
    has_macros: bool = False
    is_encrypted: bool = False
    is_compressed: bool = False

    def __post_init__(self) -> None:
        if isinstance(self.format_type, str) and not isinstance(self.format_type, FormatType):
            self.format_type = FormatType(self.format_type)


# ============================================================================
# Structure Models
# ============================================================================


@dataclass
class StructureInfo:
    """Results from structure analysis of a sheet."""

    data_start_row: int
    data_end_row: int
    data_start_col: int
    data_end_col: int
    header_row: int | None
    header_rows_count: int
    header_confidence: float
    metadata_rows: list[int] = field(default_factory=list)
    metadata_type: str = "unknown"
    merged_ranges: list[tuple[int, int, int, int]] = field(default_factory=list)
    merged_in_headers: bool = False
    merged_in_data: bool = False
    hidden_rows: list[int] = field(default_factory=list)
    hidden_columns: list[int] = field(default_factory=list)
    detected_locale: str = "en_US"
    decimal_separator: str = "."
    thousands_separator: str = ","
    num_tables: int = 1
    table_ranges: list[dict[str, Any]] = field(default_factory=list)
    blank_rows: list[int] = field(default_factory=list)
    inconsistent_columns: bool = False
    has_formulas: bool = False
    sparse_columns: list[int] = field(default_factory=list)
    suggested_skip_rows: int = 0
    suggested_skip_footer: int = 0
    suggested_overrides: dict[str, Any] = field(default_factory=dict)


@dataclass
class TableInfo:
    """Information about a detected table within a sheet."""

    start_row: int
    end_row: int
    start_col: int
    end_col: int
    has_header: bool = True
    header_row: int | None = None
    confidence: float = 1.0
    title_rows: list[int] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        """Number of data rows (excluding header)."""
        header_offset = 1 if self.has_header else 0
        return self.end_row - self.start_row + 1 - header_offset

    @property
    def column_count(self) -> int:
        """Number of columns."""
        return self.end_col - self.start_col + 1

    def to_range_string(self) -> str:
        """Convert to Excel range notation (e.g., 'A1:F100')."""
        from openpyxl.utils import get_column_letter

        start_col_letter = get_column_letter(self.start_col)
        end_col_letter = get_column_letter(self.end_col)
        return f"{start_col_letter}{self.start_row}:{end_col_letter}{self.end_row}"


# ============================================================================
# Configuration Models
# ============================================================================


@dataclass
class SheetError:
    """Error information for a sheet that failed to parse."""

    sheet_name: str
    error_type: str
    message: str
    context: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "sheet_name": self.sheet_name,
            "error_type": self.error_type,
            "message": self.message,
            "context": self.context,
        }


@dataclass
class SheetInfo:
    """Information about a sheet's structure."""

    name: str
    row_count: int
    col_count: int
    header_row: int
    is_empty: bool = False
    is_pivot: bool = False
    skip_reason: str | None = None

    @property
    def column_count(self) -> int:
        """Number of columns (descriptive alias for ``col_count``)."""
        return self.col_count


# ``SheetInfo`` historically lived in ``multi_sheet``.  Keep its public type
# identity and rendered annotations stable while allowing parser-layer code to
# import the model without depending on the legacy adapter module.
SheetInfo.__module__ = "messy_xlsx.multi_sheet"


@dataclass(frozen=True)
class SheetResult:
    """One successful or failed per-sheet materialized result."""

    name: str
    dataframe: pd.DataFrame | None = None
    error: SheetError | None = None

    def __post_init__(self) -> None:
        if (self.dataframe is None) == (self.error is None):
            raise ValueError("exactly one of dataframe and error must be set")


@dataclass
class SheetConfig:
    """Configuration options for parsing a sheet."""

    skip_rows: int = 0
    header_rows: int = 1
    skip_footer: int = 0
    cell_range: str | None = None
    column_renames: dict[str, str] = field(default_factory=dict)
    type_hints: dict[str, str] = field(default_factory=dict)
    auto_detect: bool = True
    include_hidden: bool = False
    merge_strategy: MergeStrategy | str = MergeStrategy.FILL
    locale: str | None = None
    evaluate_formulas: bool = True
    drop_regex: str | None = None
    drop_conditions: list[dict[str, Any]] = field(default_factory=list)

    # Header detection configuration
    header_detection_mode: HeaderDetectionMode | str = HeaderDetectionMode.SMART
    header_confidence_threshold: float = 0.7
    header_fallback: HeaderFallback | str = HeaderFallback.FIRST_ROW
    multi_row_headers: bool = False
    header_patterns: list[str] | None = None

    # Normalization configuration
    normalize: bool = True  # Master switch for all normalization
    normalize_dates: bool = True  # Convert date-like columns to datetime
    normalize_numbers: bool = True  # Convert number-like strings to numeric
    normalize_whitespace: bool = True  # Clean whitespace in text columns
    use_extended_missing_list: bool = False
    preserve_types: bool = True
    ensure_type_consistency: bool = True
    decimal_separator: str | None = None
    thousands_separator: str | None = None

    # Column name sanitization (ON by default for BigQuery compatibility)
    sanitize_column_names: bool = True

    def __post_init__(self) -> None:
        # Enum coercion — backward compatible with raw strings
        if isinstance(self.merge_strategy, str) and not isinstance(
            self.merge_strategy, MergeStrategy
        ):
            self.merge_strategy = MergeStrategy(self.merge_strategy)
        if isinstance(self.header_detection_mode, str) and not isinstance(
            self.header_detection_mode, HeaderDetectionMode
        ):
            self.header_detection_mode = HeaderDetectionMode(self.header_detection_mode)
        if isinstance(self.header_fallback, str) and not isinstance(
            self.header_fallback, HeaderFallback
        ):
            self.header_fallback = HeaderFallback(self.header_fallback)

        # Bounds validation
        if self.skip_rows < 0:
            raise ValueError(f"skip_rows must be >= 0, got {self.skip_rows}")
        if self.header_rows < 0:
            raise ValueError(f"header_rows must be >= 0, got {self.header_rows}")
        if self.skip_footer < 0:
            raise ValueError(f"skip_footer must be >= 0, got {self.skip_footer}")
        if not (0.0 <= self.header_confidence_threshold <= 1.0):
            raise ValueError(
                f"header_confidence_threshold must be between 0.0 and 1.0, "
                f"got {self.header_confidence_threshold}"
            )


# ============================================================================
# Cell Models
# ============================================================================


@dataclass
class CellValue:
    """Represents a single cell's value with metadata."""

    value: Any
    formula: str | None = None
    is_merged: bool = False
    is_hidden: bool = False
    data_type: DataType | str = DataType.EMPTY
    original_format: str | None = None

    def __post_init__(self) -> None:
        if isinstance(self.data_type, str) and not isinstance(self.data_type, DataType):
            self.data_type = DataType(self.data_type)

    @property
    def is_formula(self) -> bool:
        """True if cell contains a formula."""
        return self.formula is not None

    @property
    def is_error(self) -> bool:
        """True if cell contains an Excel error value."""
        if isinstance(self.value, str):
            return self.value.startswith("#") and self.value.endswith("!")
        return False

    def __repr__(self) -> str:
        if self.formula:
            return f"CellValue({self.value!r}, formula={self.formula!r})"
        return f"CellValue({self.value!r})"
