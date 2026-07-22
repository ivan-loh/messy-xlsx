"""Bounded-row OOXML reader that emits schema-stable Arrow batches."""

from __future__ import annotations

from collections import deque
from collections.abc import Iterator
from contextlib import AbstractContextManager
from dataclasses import dataclass
from types import TracebackType
from typing import Any, Self

import openpyxl
import pyarrow as pa

from messy_xlsx._fallback_signals import _exception_traceback
from messy_xlsx._source import BackendSource, SourceHandle
from messy_xlsx.enums import MergeStrategy
from messy_xlsx.ooxml.models import SheetManifest
from messy_xlsx.parsing.contracts import OutputMode, ParseMetrics, StreamingBatchReader
from messy_xlsx.parsing.coordinates import (
    CoordinateBatch,
    CoordinateCompatibilityError,
    CoordinateOperation,
    CoordinateTransform,
    PreparedCoordinateSchema,
)
from messy_xlsx.parsing.parse_plan import ParsePlan
from messy_xlsx.parsing.streams import _run_cleanups

_MAX_EXCEL_ROW = 1_048_576
_MAX_EXCEL_COLUMN = 16_384


@dataclass(frozen=True, slots=True)
class StreamingWorksheetLayout:
    """Validated absolute worksheet bounds and immutable Arrow schemas."""

    sheet_name: str
    raw_schema: pa.Schema
    raw_column_numbers: tuple[int, ...]
    output_schema: pa.Schema
    output_column_numbers: tuple[int, ...]
    min_row: int | None
    max_row: int | None
    min_col: int | None
    max_col: int | None

    def __post_init__(self) -> None:
        _validate_layout_schemas(self)
        _validate_layout_bounds(self)

    @property
    def read_min_row(self) -> int | None:
        return self.min_row

    @property
    def read_max_row(self) -> int | None:
        return self.max_row

    @property
    def read_min_col(self) -> int | None:
        return self.min_col

    @property
    def read_max_col(self) -> int | None:
        return self.max_col

    @property
    def prepared_schema(self) -> PreparedCoordinateSchema:
        return PreparedCoordinateSchema(
            raw_schema=self.raw_schema,
            raw_column_numbers=self.raw_column_numbers,
            output_schema=self.output_schema,
            output_column_numbers=self.output_column_numbers,
        )

    @classmethod
    def compile(
        cls,
        manifest: SheetManifest,
        plan: ParsePlan,
        raw_schema: pa.Schema,
        transform: CoordinateTransform,
        *,
        raw_column_numbers: tuple[int, ...] | None = None,
    ) -> StreamingWorksheetLayout:
        """Compile exact read bounds and the physical output schema without I/O."""
        _validate_streaming_inputs(manifest, plan, raw_schema, transform)
        bounds = _read_bounds(manifest, plan)
        if raw_column_numbers is None:
            if bounds is None:
                raw_column_numbers = tuple(range(1, len(raw_schema) + 1))
            else:
                raw_column_numbers = tuple(range(bounds[2], bounds[3] + 1))
        prepared = transform.prepare_schema(plan, raw_schema, raw_column_numbers)
        if bounds is None:
            return cls(
                sheet_name=manifest.name,
                raw_schema=prepared.raw_schema,
                raw_column_numbers=prepared.raw_column_numbers,
                output_schema=prepared.output_schema,
                output_column_numbers=prepared.output_column_numbers,
                min_row=None,
                max_row=None,
                min_col=None,
                max_col=None,
            )
        layout = cls(
            sheet_name=manifest.name,
            raw_schema=prepared.raw_schema,
            raw_column_numbers=prepared.raw_column_numbers,
            output_schema=prepared.output_schema,
            output_column_numbers=prepared.output_column_numbers,
            min_row=bounds[0],
            max_row=bounds[1],
            min_col=bounds[2],
            max_col=bounds[3],
        )
        return layout


def _validate_layout_schemas(layout: StreamingWorksheetLayout) -> None:
    if not layout.sheet_name:
        raise ValueError("streaming worksheet name cannot be empty")
    if not isinstance(layout.raw_schema, pa.Schema) or not isinstance(
        layout.output_schema, pa.Schema
    ):
        raise TypeError("streaming layout schemas must be pyarrow.Schema values")
    if not isinstance(layout.raw_column_numbers, tuple) or not isinstance(
        layout.output_column_numbers, tuple
    ):
        raise TypeError("streaming layout coordinates must be immutable tuples")
    if len(layout.raw_schema) != len(layout.raw_column_numbers):
        raise ValueError("raw schema width does not match absolute columns")
    if len(layout.output_schema) != len(layout.output_column_numbers):
        raise ValueError("output schema width does not match absolute columns")
    if tuple(layout.output_schema.names) != tuple(
        str(index) for index in range(len(layout.output_schema))
    ):
        raise ValueError("streaming output schema fields must use ordinal names")
    if any(not field.nullable or field.metadata for field in layout.output_schema):
        raise ValueError("streaming output schema fields must be nullable without metadata")
    if layout.output_schema.metadata:
        raise ValueError("streaming output schema must not contain metadata")


def _validate_layout_bounds(layout: StreamingWorksheetLayout) -> None:
    bounds = (layout.min_row, layout.max_row, layout.min_col, layout.max_col)
    if all(value is None for value in bounds):
        if layout.raw_column_numbers and any(
            column < 1 or column > _MAX_EXCEL_COLUMN for column in layout.raw_column_numbers
        ):
            raise ValueError("raw column coordinates exceed Excel worksheet bounds")
        return
    if any(value is None for value in bounds):
        raise ValueError("streaming worksheet bounds must be all present or all absent")
    assert layout.min_row is not None
    assert layout.max_row is not None
    assert layout.min_col is not None
    assert layout.max_col is not None
    if (
        layout.min_row < 1
        or layout.max_row > _MAX_EXCEL_ROW
        or layout.min_col < 1
        or layout.max_col > _MAX_EXCEL_COLUMN
        or layout.min_row > layout.max_row
        or layout.min_col > layout.max_col
    ):
        raise ValueError("streaming worksheet bounds are invalid")
    expected_columns = tuple(range(layout.min_col, layout.max_col + 1))
    if layout.raw_column_numbers != expected_columns:
        raise ValueError("raw absolute columns must be contiguous and match read bounds")


def _validate_streaming_inputs(
    manifest: SheetManifest,
    plan: ParsePlan,
    raw_schema: pa.Schema,
    transform: CoordinateTransform,
) -> None:
    if not isinstance(manifest, SheetManifest):
        raise TypeError("manifest must be a SheetManifest")
    if not manifest.name:
        raise ValueError("manifest sheet name cannot be empty")
    if plan.output_mode is not OutputMode.STREAMING:
        raise ValueError("OpenpyxlStreamingReader requires streaming output")
    if plan.batch_size is None or plan.batch_size < 1:
        raise ValueError("batch_size must be >= 1")
    if not isinstance(raw_schema, pa.Schema):
        raise TypeError("raw_schema must be a pyarrow.Schema")
    if (
        transform.hidden_rows != manifest.hidden_rows
        or transform.hidden_columns != manifest.hidden_columns
        or transform.merged_ranges != manifest.merged_ranges
    ):
        raise ValueError("coordinate transform does not match sheet manifest")


def _read_bounds(
    manifest: SheetManifest,
    plan: ParsePlan,
) -> tuple[int, int, int, int] | None:
    projection = CoordinateOperation._parse_projection(plan.cell_range)
    if projection is not None:
        min_row = projection.min_row
        max_row = projection.max_row
        min_col = projection.min_col
        max_col = projection.max_col
        if MergeStrategy(plan.merge_strategy) is MergeStrategy.FILL:
            for merged_range in manifest.merged_ranges:
                if CoordinateOperation._merge_intersects_projection(
                    merged_range,
                    projection,
                ):
                    min_row = min(min_row, merged_range.min_row)
                    min_col = min(min_col, merged_range.min_col)
        return min_row, max_row, min_col, max_col

    relevant_merges = manifest.merged_ranges
    max_row = max(
        (manifest.observed_max_row, *(merged.max_row for merged in relevant_merges)),
    )
    max_col = max(
        (manifest.observed_max_col, *(merged.max_col for merged in relevant_merges)),
    )
    if max_row == 0 or max_col == 0:
        return None
    return 1, max_row, 1, max_col


class _BatchRechunker:
    """Bounded carry that emits exact-size physical Arrow batches."""

    def __init__(self, schema: pa.Schema, batch_size: int) -> None:
        self._schema = schema
        self._batch_size = batch_size
        self._batches: deque[pa.RecordBatch] = deque()
        self._rows = 0

    def push(self, batch: CoordinateBatch) -> None:
        physical = batch.batch
        if physical.num_rows == 0:
            return
        if not physical.schema.equals(self._schema, check_metadata=True):
            raise CoordinateCompatibilityError(
                "coordinate output schema changed across streaming batches"
            )
        self._batches.append(physical)
        self._rows += physical.num_rows

    def pop(self, *, terminal: bool = False) -> pa.RecordBatch | None:
        if self._rows < self._batch_size and not terminal:
            return None
        if self._rows == 0:
            return None
        count = min(self._batch_size, self._rows)
        pieces: list[pa.RecordBatch] = []
        remaining = count
        while remaining:
            batch = self._batches[0]
            if batch.num_rows <= remaining:
                pieces.append(self._batches.popleft())
                remaining -= batch.num_rows
                continue
            pieces.append(batch.slice(0, remaining))
            self._batches[0] = _detach_record_batch(
                batch.slice(remaining, batch.num_rows - remaining)
            )
            remaining = 0
        self._rows -= count
        if len(pieces) == 1 and pieces[0].num_rows == count:
            result = pieces[0]
        else:
            arrays = [
                pa.concat_arrays([piece.column(column) for piece in pieces])
                for column in range(len(self._schema))
            ]
            result = _record_batch_with_row_count(arrays, self._schema, count)
        if not result.schema.equals(self._schema, check_metadata=True):
            raise CoordinateCompatibilityError("streaming output schema is unstable")
        return result


def _detach_record_batch(batch: pa.RecordBatch) -> pa.RecordBatch:
    arrays = [pa.concat_arrays([array]) for array in batch.columns]
    return _record_batch_with_row_count(arrays, batch.schema, batch.num_rows)


def _record_batch_with_row_count(
    arrays: list[pa.Array],
    schema: pa.Schema,
    row_count: int,
) -> pa.RecordBatch:
    if arrays:
        return pa.record_batch(arrays, schema=schema)
    return pa.record_batch([pa.nulls(row_count)], names=["_row_count"]).select([])


class OpenpyxlStreamingReader:
    """Read one OOXML worksheet pass into bounded Arrow row windows."""

    def __init__(
        self,
        source: SourceHandle,
        manifest: SheetManifest,
        plan: ParsePlan,
        raw_schema_or_layout: pa.Schema | StreamingWorksheetLayout,
        transform: CoordinateTransform,
        metrics: ParseMetrics | None = None,
        *,
        raw_column_numbers: tuple[int, ...] | None = None,
    ) -> None:
        if not isinstance(source, SourceHandle):
            raise TypeError("source must be a SourceHandle")
        if isinstance(raw_schema_or_layout, StreamingWorksheetLayout):
            layout = raw_schema_or_layout
            _validate_streaming_inputs(manifest, plan, layout.raw_schema, transform)
            expected = StreamingWorksheetLayout.compile(
                manifest,
                plan,
                layout.raw_schema,
                transform,
                raw_column_numbers=layout.raw_column_numbers,
            )
            if expected != layout:
                raise ValueError("streaming worksheet layout does not match parser inputs")
        else:
            layout = StreamingWorksheetLayout.compile(
                manifest,
                plan,
                raw_schema_or_layout,
                transform,
                raw_column_numbers=raw_column_numbers,
            )
        del metrics
        batch_size = plan.batch_size
        assert batch_size is not None

        self._source = source
        self._manifest = manifest
        self._plan = plan
        self._layout = layout
        self._schema = layout.output_schema
        self._operation = transform.open(plan, layout.prepared_schema)
        self._batch_size = batch_size
        self._rechunker = _BatchRechunker(self._schema, batch_size)
        self._backend_context: AbstractContextManager[BackendSource] | None = None
        self._workbook: Any | None = None
        self._rows: Iterator[tuple[Any, ...]] = iter(())
        self._next_input_row = layout.min_row
        self._row_source_exhausted = layout.min_row is None
        self._coordinate_finished = False
        self._terminal = False
        self._closed = False

        try:
            self._open()
        except BaseException as error:
            self._close_resources(
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise

    @property
    def schema(self) -> pa.Schema:
        """Return the persistent physical output schema."""
        return self._schema

    def _open(self) -> None:
        backend_context: AbstractContextManager[BackendSource]
        if self._source.is_path:
            backend_context = self._source.open_binary()
        else:
            backend_context = self._source.open_backend()
        backend = backend_context.__enter__()
        self._backend_context = backend_context
        workbook = openpyxl.load_workbook(
            backend,
            read_only=True,
            data_only=self._plan.data_only,
            keep_links=False,
            keep_vba=False,
        )
        self._workbook = workbook
        worksheet = workbook[self._manifest.name]
        reset_dimensions = getattr(worksheet, "reset_dimensions", None)
        if callable(reset_dimensions):
            reset_dimensions()
        if self._layout.min_row is None:
            return
        assert self._layout.max_row is not None
        assert self._layout.min_col is not None
        assert self._layout.max_col is not None
        self._rows = iter(
            worksheet.iter_rows(
                values_only=True,
                min_row=self._layout.min_row,
                max_row=self._layout.max_row,
                min_col=self._layout.min_col,
                max_col=self._layout.max_col,
            )
        )

    def read_next_batch(self) -> pa.RecordBatch | None:
        """Return one non-empty bounded batch, with sticky terminal EOF."""
        if self._closed or self._terminal:
            return None
        while True:
            ready = self._rechunker.pop()
            if ready is not None:
                return ready

            raw = self._read_raw_window()
            if raw is not None:
                for transformed in self._operation.push(raw):
                    self._rechunker.push(transformed)
                continue

            if not self._coordinate_finished:
                self._coordinate_finished = True
                for transformed in self._operation.finish():
                    self._rechunker.push(transformed)
                continue

            ready = self._rechunker.pop(terminal=True)
            if ready is not None:
                return ready
            self._terminal = True
            return None

    def _read_raw_window(self) -> CoordinateBatch | None:
        next_row = self._next_input_row
        max_row = self._layout.max_row
        if next_row is None or max_row is None or next_row > max_row:
            return None
        count = min(self._batch_size, max_row - next_row + 1)
        assert count > 0
        columns: list[list[object]] = [[] for _column in self._layout.raw_column_numbers]
        for _offset in range(count):
            row: tuple[Any, ...]
            if self._row_source_exhausted:
                row = ()
            else:
                try:
                    row = tuple(next(self._rows))
                except StopIteration:
                    self._row_source_exhausted = True
                    row = ()
            for position, values in enumerate(columns):
                raw_value = row[position] if position < len(row) else None
                values.append(
                    _coerce_raw_value(
                        raw_value,
                        self._layout.raw_schema.field(position).type,
                    )
                )
        arrays = [
            pa.array(values, type=field.type)
            for values, field in zip(columns, self._layout.raw_schema, strict=True)
        ]
        batch = pa.record_batch(arrays, schema=self._layout.raw_schema)
        row_numbers = pa.array(range(next_row, next_row + count), type=pa.int64())
        self._next_input_row = next_row + count
        return CoordinateBatch(
            batch=batch,
            row_numbers=row_numbers,
            column_numbers=self._layout.raw_column_numbers,
        )

    def close(self) -> None:
        """Close the workbook and active source borrow exactly once."""
        if self._closed:
            return
        self._closed = True
        self._close_resources()

    def _close_resources(
        self,
        *,
        primary_error: BaseException | None = None,
        primary_traceback: TracebackType | None = None,
    ) -> None:
        workbook = self._workbook
        backend_context = self._backend_context
        self._workbook = None
        self._backend_context = None
        cleanups: list[tuple[str, Any]] = []
        if workbook is not None:
            cleanups.append(("openpyxl workbook cleanup", workbook.close))
        if backend_context is not None:
            cleanups.append(
                (
                    "source borrow cleanup",
                    lambda: backend_context.__exit__(None, None, None),
                )
            )
        _run_cleanups(
            cleanups,
            primary_error=primary_error,
            primary_traceback=primary_traceback,
        )

    def __enter__(self) -> Self:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        del exc_type
        if exc_value is None:
            self.close()
            return
        if self._closed:
            return
        self._closed = True
        self._close_resources(primary_error=exc_value, primary_traceback=traceback)


class _ReaderBatchIterator(Iterator[pa.RecordBatch]):
    """One-shot closable adapter for a low-level streaming reader."""

    def __init__(self, reader: StreamingBatchReader) -> None:
        self._reader: StreamingBatchReader | None = reader
        self._closed = False

    def __iter__(self) -> Self:
        return self

    def __next__(self) -> pa.RecordBatch:
        reader = self._reader
        if self._closed or reader is None:
            raise StopIteration
        try:
            batch = reader.read_next_batch()
        except BaseException as error:
            self._close(
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise
        if batch is None:
            self.close()
            raise StopIteration
        return batch

    def close(self) -> None:
        self._close()

    def _close(
        self,
        *,
        primary_error: BaseException | None = None,
        primary_traceback: TracebackType | None = None,
    ) -> None:
        if self._closed:
            return
        self._closed = True
        reader = self._reader
        self._reader = None
        cleanups = [] if reader is None else [("streaming reader cleanup", reader.close)]
        _run_cleanups(
            cleanups,
            primary_error=primary_error,
            primary_traceback=primary_traceback,
        )

    def __enter__(self) -> Self:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        del exc_type
        if exc_value is None:
            self.close()
            return
        self._close(primary_error=exc_value, primary_traceback=traceback)


def reader_batches(reader: StreamingBatchReader) -> _ReaderBatchIterator:
    """Return a closable one-shot iterator that owns reader cleanup."""
    return _ReaderBatchIterator(reader)


def _coerce_raw_value(value: object, data_type: pa.DataType) -> object:
    if value is None:
        return None
    if pa.types.is_string(data_type) or pa.types.is_large_string(data_type):
        return value if isinstance(value, str) else str(value)
    return value


__all__ = ["OpenpyxlStreamingReader", "StreamingWorksheetLayout", "reader_batches"]
