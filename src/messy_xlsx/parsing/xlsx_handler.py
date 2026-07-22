"""XLSX/XLSM file handler using fastexcel with openpyxl fallback."""

from __future__ import annotations

# ============================================================================
# Imports
# ============================================================================
import logging
from collections.abc import Callable
from contextlib import ExitStack, closing
from typing import Any

import fastexcel
import numpy as np
import openpyxl
import pandas as pd
import pyarrow as pa

from messy_xlsx._source import SourceHandle
from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.parsing.base_handler import (
    FileSource,
    FormatHandler,
    ParseOptions,
)
from messy_xlsx.parsing.contracts import OutputMode, ParseMetrics
from messy_xlsx.parsing.coordinates import (
    CoordinateCompatibilityError,
    CoordinateTransform,
)
from messy_xlsx.parsing.fallback import FallbackCoordinator
from messy_xlsx.parsing.fastexcel_session import FastexcelSession
from messy_xlsx.parsing.legacy_adapter import LegacyDataFrameAdapter
from messy_xlsx.parsing.parse_plan import ParsePlan
from messy_xlsx.parsing.xlsx_materialized import (
    FastexcelMaterializedReader,
    _coerce_materialized_table,
)

logger = logging.getLogger(__name__)

# ============================================================================
# Constants
# ============================================================================

EXCEL_ERRORS = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
    "#GETTING_DATA",
}

_FASTEXCEL_COMPATIBILITY_ERRORS = (
    fastexcel.UnsupportedColumnTypeCombinationError,
    fastexcel.CannotRetrieveCellDataError,
    fastexcel.CalamineCellError,
    CoordinateCompatibilityError,
)


def is_fastexcel_compatibility_error(error: Exception) -> bool:
    """Return whether openpyxl can safely retry this typed cell limitation."""
    return isinstance(error, _FASTEXCEL_COMPATIBILITY_ERRORS)


def _is_fastexcel_materialized_plan(plan: ParsePlan) -> bool:
    """Gate features whose shared coordinate transforms arrive in Task 8."""
    return (
        plan.output_mode is OutputMode.MATERIALIZED
        and (plan.skip_rows == 0 or bool(plan.cell_range))
        and plan.data_only
    )


class _FastexcelDataFrameReader:
    """Own a session while adapting one Arrow result to the legacy frame SPI."""

    def __init__(
        self,
        handler: XLSXHandler,
        source: SourceHandle,
        sheet: str | None,
        options: ParseOptions,
        plan: ParsePlan | None,
        transform: CoordinateTransform | None = None,
        session_factory: Callable[[], FastexcelSession] | None = None,
    ) -> None:
        self._handler = handler
        self._options = options
        self._plan = plan
        self._skip_rows_already_applied = plan is None
        self._owns_session = session_factory is None
        if session_factory is None:

            def make_owned_session() -> FastexcelSession:
                return FastexcelSession(source)

            get_session = make_owned_session
        else:
            get_session = session_factory
        try:
            session = get_session()
        except _FASTEXCEL_COMPATIBILITY_ERRORS:
            raise
        except PermissionError:
            raise
        except MemoryError:
            raise
        except Exception as error:
            raise FormatError(
                f"Cannot open Excel file: {error}",
                file_path=source.description,
            ) from error
        self._session: FastexcelSession | None = session
        self._reader: FastexcelMaterializedReader | _LegacyFastexcelArrowReader | None = None
        try:
            resolved_sheet = handler._resolve_session_sheet(
                session,
                sheet,
                source.description,
            )
            if plan is None:
                self._reader = _LegacyFastexcelArrowReader(
                    session,
                    resolved_sheet,
                    options.skip_rows,
                )
            else:
                self._reader = FastexcelMaterializedReader(
                    session,
                    resolved_sheet,
                    plan,
                    transform,
                )
        except BaseException:
            if self._owns_session:
                session.close()
            self._session = None
            raise

    def read_table(self) -> pd.DataFrame:
        """Read Arrow once, bridge once, then apply handler-owned framing."""
        reader = self._reader
        if reader is None:
            raise RuntimeError("Fastexcel materialized operation is closed")
        table = reader.read_table()
        frame = LegacyDataFrameAdapter().to_dataframe(table, self._plan)
        if (
            self._plan is not None
            and isinstance(reader, FastexcelMaterializedReader)
            and reader._transform is not None
        ):
            return self._handler._clean_excel_data(frame, self._options)
        return self._handler._frame_fastexcel_data(
            frame,
            self._options,
            skip_rows_already_applied=self._skip_rows_already_applied,
        )

    def close(self) -> None:
        """Release only the session; the SourceHandle retains spill ownership."""
        session = self._session
        self._reader = None
        self._session = None
        if self._owns_session and session is not None:
            session.close()


class _LegacyFastexcelArrowReader:
    """Raw operation for the ParseOptions-only legacy handler SPI."""

    def __init__(
        self,
        session: FastexcelSession,
        sheet: str,
        skip_rows: int,
    ) -> None:
        self._session = session
        self._sheet = sheet
        self._skip_rows = skip_rows

    def read_table(self) -> pa.Table:
        materialized = self._session.materialize(
            self._sheet,
            skip_rows=self._skip_rows,
        )
        return _coerce_materialized_table(materialized)


class _OpenpyxlDataFrameReader:
    """Coordinator adapter for the existing compatibility DataFrame reader."""

    def __init__(
        self,
        handler: XLSXHandler,
        source: SourceHandle,
        sheet: str | None,
        options: ParseOptions,
    ) -> None:
        self._handler = handler
        self._source = source
        self._sheet = sheet
        self._options = options

    def read_table(self) -> pd.DataFrame:
        """Return the existing openpyxl compatibility frame unchanged."""
        return self._handler._parse_openpyxl(
            self._source,
            self._sheet,
            self._options,
        )


# ============================================================================
# Core
# ============================================================================


class XLSXHandler(FormatHandler):
    """Handler for XLSX and XLSM files."""

    _accepts_source_handle = True

    def can_handle(self, format_type: str) -> bool:
        return format_type in ("xlsx", "xlsm")

    def parse(
        self,
        file_source: FileSource | SourceHandle,
        sheet: str | None,
        options: ParseOptions,
    ) -> pd.DataFrame:
        """Parse XLSX/XLSM file to DataFrame."""
        source = SourceHandle.coerce(file_source)
        try:
            needs_openpyxl = (
                options.merge_strategy != "skip"
                or options.ignore_hidden
                or options.cell_range
                or not options.data_only
            )

            if needs_openpyxl:
                return self._parse_openpyxl(source, sheet, options)
            return self._parse_fastexcel(source, sheet, options, plan=None)
        finally:
            if source is not file_source:
                source.close()

    def _parse_materialized_plan(
        self,
        file_source: FileSource | SourceHandle,
        sheet: str | None,
        plan: ParsePlan,
        session_factory: Callable[[], FastexcelSession],
        *,
        metrics: ParseMetrics | None = None,
        transform: CoordinateTransform | None = None,
    ) -> pd.DataFrame:
        """Parse an eligible built-in plan without reconstructing it."""
        if not _is_fastexcel_materialized_plan(plan):
            return self.parse(file_source, sheet, plan.to_parse_options())

        source = SourceHandle.coerce(file_source)
        try:
            return self._parse_fastexcel(
                source,
                sheet,
                plan.to_parse_options(),
                plan=plan,
                metrics=metrics,
                session_factory=session_factory,
                transform=transform,
            )
        finally:
            if source is not file_source:
                source.close()

    # -------------------------------------------------------------------------
    # Fastexcel (fast path)
    # -------------------------------------------------------------------------

    def _parse_fastexcel(
        self,
        source: SourceHandle,
        sheet: str | None,
        options: ParseOptions,
        *,
        plan: ParsePlan | None,
        metrics: ParseMetrics | None = None,
        session_factory: Callable[[], FastexcelSession] | None = None,
        transform: CoordinateTransform | None = None,
    ) -> pd.DataFrame:
        """Fast Arrow path with one narrowly classified transactional fallback."""
        file_desc = source.description
        try:
            return FallbackCoordinator(
                is_fastexcel_compatibility_error,
                metrics=metrics,
            ).materialize(
                lambda: _FastexcelDataFrameReader(
                    self,
                    source,
                    sheet,
                    options,
                    plan,
                    transform,
                    session_factory,
                ),
                lambda: _OpenpyxlDataFrameReader(self, source, sheet, options),
            )
        except PermissionError as e:
            raise FileError(
                f"Permission denied: {file_desc}", file_path=file_desc, operation="open"
            ) from e
        except FormatError:
            raise
        except MemoryError:
            raise
        except Exception as e:
            raise FormatError(f"Error reading Excel file: {e}", file_path=file_desc) from e

    def _resolve_session_sheet(
        self,
        session: FastexcelSession,
        sheet: str | None,
        file_desc: str,
    ) -> str:
        if not sheet:
            return session.sheet_names[0]
        if sheet not in session.sheet_names:
            raise FormatError(
                f"Sheet '{sheet}' not found",
                file_path=file_desc,
                detected_format="xlsx",
            )
        return sheet

    def _frame_fastexcel_data(
        self,
        frame: pd.DataFrame,
        options: ParseOptions,
        *,
        skip_rows_already_applied: bool,
    ) -> pd.DataFrame:
        """Apply the legacy handler sequence after raw coordinate materialization."""
        if frame.empty:
            return pd.DataFrame()
        if not skip_rows_already_applied:
            frame = frame.iloc[options.skip_rows :].reset_index(drop=True)
            if frame.empty:
                return pd.DataFrame()
        frame = self._apply_options(frame, options)
        return self._clean_excel_data(frame, options)

    # -------------------------------------------------------------------------
    # Openpyxl (fallback for advanced features)
    # -------------------------------------------------------------------------

    def _parse_openpyxl(
        self,
        source: SourceHandle,
        sheet: str | None,
        options: ParseOptions,
    ) -> pd.DataFrame:
        """Fallback path using openpyxl for advanced features."""
        read_only = options.merge_strategy == "skip" and not options.ignore_hidden
        file_desc = source.description

        with ExitStack() as stack:
            try:
                backend_source = stack.enter_context(source.open_backend())
                wb = openpyxl.load_workbook(
                    backend_source,
                    read_only=read_only,
                    data_only=options.data_only,
                    keep_links=False,
                )
            except PermissionError as e:
                raise FileError(
                    f"Permission denied: {file_desc}", file_path=file_desc, operation="open"
                ) from e
            except Exception as e:
                raise FormatError(f"Cannot open Excel file: {e}", file_path=file_desc) from e

            # Close the workbook before the borrowed source view is restored or
            # released. This matters for read-only openpyxl archives.
            stack.callback(wb.close)
            ws = self._resolve_worksheet(wb, sheet, file_desc)

            if options.merge_strategy != "skip" and not read_only:
                self._handle_merged_cells(ws, options.merge_strategy)

            data = self._read_worksheet(ws, options)
            if not data:
                return pd.DataFrame()

            df = pd.DataFrame(data)

            if options.skip_rows > 0 and not options.cell_range:
                df = df.iloc[options.skip_rows :]

            df = self._apply_options(df, options)
            return self._clean_excel_data(df, options)

    def _resolve_worksheet(self, wb: Any, sheet: str | None, file_desc: str) -> Any:
        if not sheet:
            return wb.active
        if sheet not in wb.sheetnames:
            raise FormatError(
                f"Sheet '{sheet}' not found", file_path=file_desc, detected_format="xlsx"
            )
        return wb[sheet]

    def _read_worksheet(self, ws: Any, options: ParseOptions) -> list[list[Any]]:
        """Read worksheet data using openpyxl."""
        if options.cell_range:
            return self._read_cell_range(ws, options.cell_range)
        return self._read_full_sheet(ws, options.ignore_hidden)

    def _read_cell_range(self, ws: Any, cell_range: str) -> list[list[Any]]:
        try:
            rows_iter = ws[cell_range]
            if not isinstance(rows_iter[0], tuple):
                rows_iter = [rows_iter]

            data = []
            for row in rows_iter:
                if not isinstance(row, tuple):
                    row = [row]
                data.append([cell.value for cell in row])
            return data

        except Exception as e:
            raise FormatError(f"Invalid cell range: {cell_range}", detected_format="xlsx") from e

    def _read_full_sheet(self, ws: Any, ignore_hidden: bool) -> list[list[Any]]:
        hidden_rows: set[int] = set()
        hidden_cols: set[str] = set()

        if ignore_hidden:
            hidden_rows = {r for r, d in ws.row_dimensions.items() if d.hidden}
            hidden_cols = {c for c, d in ws.column_dimensions.items() if d.hidden}

        data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx in hidden_rows:
                continue

            if hidden_cols:
                from openpyxl.utils import get_column_letter

                row_values = [
                    cell.value for cell in row if get_column_letter(cell.column) not in hidden_cols
                ]
            else:
                row_values = [cell.value for cell in row]
            data.append(row_values)

        return data

    def _handle_merged_cells(self, ws: Any, strategy: str) -> None:
        """Handle merged cells according to strategy."""
        for merged_range in list(ws.merged_cells.ranges):
            top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
            ws.unmerge_cells(str(merged_range))

            if strategy == "fill":
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        ws.cell(r, c).value = top_left

            elif strategy == "first_only":
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        if r != merged_range.min_row or c != merged_range.min_col:
                            ws.cell(r, c).value = None

    # -------------------------------------------------------------------------
    # Common
    # -------------------------------------------------------------------------

    def _apply_options(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        """Apply skip_footer and header options."""
        if options.skip_footer > 0:
            df = df.iloc[: -options.skip_footer]

        if options.header_rows > 0 and len(df) >= options.header_rows:
            df, columns = self._generate_column_names(df, options.header_rows)
            df.columns = columns
            df = df.reset_index(drop=True)
        else:
            df.columns = [f"col_{i}" for i in range(len(df.columns))]

        return df

    def _clean_excel_data(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        """Replace Excel errors and custom NA values with NaN."""
        na_values = EXCEL_ERRORS | set(options.na_values or [])

        for idx in range(len(df.columns)):
            series = df.iloc[:, idx]
            if series.dtype == object:
                df.iloc[:, idx] = series.map(lambda x: np.nan if x in na_values else x)

        return df

    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    def get_sheet_names(self, file_source: FileSource | SourceHandle) -> list[str]:
        """Get list of sheet names."""
        source = SourceHandle.coerce(file_source)
        try:
            file_desc = source.description
            try:
                with source.open_path_or_bytes() as content:
                    excel_file = fastexcel.read_excel(content)
                    return list(excel_file.sheet_names)
            except (OSError, ValueError, RuntimeError) as e:
                logger.debug("fastexcel sheet name read failed, falling back to openpyxl: %s", e)

            # Keep the openpyxl workbook inside the borrowed source lifetime.
            try:
                with (
                    source.open_backend() as backend_source,
                    closing(openpyxl.load_workbook(backend_source, read_only=True)) as wb,
                ):
                    return list(wb.sheetnames)
            except PermissionError as e:
                raise FileError(
                    f"Permission denied: {file_desc}",
                    file_path=file_desc,
                    operation="get_sheets",
                ) from e
            except Exception as e:
                raise FormatError(f"Cannot read sheet names: {e}", file_path=file_desc) from e
        finally:
            if source is not file_source:
                source.close()

    def validate(self, file_source: FileSource | SourceHandle) -> tuple[bool, str | None]:
        """Validate that file can be parsed."""
        source = SourceHandle.coerce(file_source)
        try:
            with source.open_path_or_bytes() as content:
                fastexcel.read_excel(content)
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            if source is not file_source:
                source.close()
