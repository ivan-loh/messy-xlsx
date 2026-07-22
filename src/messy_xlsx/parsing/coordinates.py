"""Arrow transforms that retain original worksheet coordinates."""

from __future__ import annotations

import math
from bisect import bisect_right
from collections import deque
from dataclasses import dataclass, field
from datetime import date
from itertools import pairwise

import numpy as np
import pyarrow as pa
import pyarrow.compute as pc
from openpyxl.utils.cell import range_boundaries

from messy_xlsx.enums import MergeStrategy
from messy_xlsx.ooxml.models import (
    IntervalIndex,
    MergeRange,
    SheetManifest,
)
from messy_xlsx.parsing.parse_plan import ParsePlan

_MAX_EXCEL_ROW = 1_048_576
_MAX_EXCEL_COLUMN = 16_384


class CoordinateCompatibilityError(Exception):
    """Signal an exact coordinate result that Arrow cannot represent."""


@dataclass(frozen=True, slots=True)
class ColumnIdentity:
    """Final positional identity for one output column."""

    ordinal: int
    display_name: object

    def __post_init__(self) -> None:
        if self.ordinal < 0:
            raise ValueError("column identity ordinal must be non-negative")


@dataclass(frozen=True, slots=True)
class CoordinateBatch:
    """An Arrow batch paired with original worksheet coordinates."""

    batch: pa.RecordBatch
    row_numbers: pa.Int64Array
    column_numbers: tuple[int, ...]
    column_identities: tuple[ColumnIdentity, ...] = ()

    def __post_init__(self) -> None:
        if not pa.types.is_int64(self.row_numbers.type):
            raise TypeError("row_numbers must be an int64 Arrow array")
        if len(self.row_numbers) != self.batch.num_rows:
            raise ValueError("row coordinate count does not match batch rows")
        if len(self.column_numbers) != self.batch.num_columns:
            raise ValueError("column coordinate count does not match batch columns")
        if any(number < 1 for number in self.column_numbers):
            raise ValueError("column coordinates must be positive")
        if any(left >= right for left, right in pairwise(self.column_numbers)):
            raise ValueError("column coordinates must be strictly increasing")
        if self.column_identities and len(self.column_identities) != self.batch.num_columns:
            raise ValueError("column identity count does not match batch columns")
        if self.column_identities and tuple(
            identity.ordinal for identity in self.column_identities
        ) != tuple(range(self.batch.num_columns)):
            raise ValueError("column identity ordinals must match output positions")

    def slice_rows(self, offset: int, length: int) -> CoordinateBatch:
        """Slice Arrow rows and their coordinate sidecar together."""
        return CoordinateBatch(
            batch=self.batch.slice(offset, length),
            row_numbers=self.row_numbers.slice(offset, length),
            column_numbers=self.column_numbers,
            column_identities=self.column_identities,
        )


@dataclass(frozen=True, slots=True)
class CoordinateTransform:
    """Immutable coordinate metadata shared by independent operations."""

    hidden_rows: IntervalIndex
    hidden_columns: IntervalIndex
    merged_ranges: tuple[MergeRange, ...]
    _observed_max_row: int = field(default=0, init=False, repr=False)
    _observed_max_col: int = field(default=0, init=False, repr=False)

    @classmethod
    def from_manifest(cls, manifest: SheetManifest) -> CoordinateTransform:
        """Create a transform from one selected sheet manifest."""
        transform = cls(
            hidden_rows=manifest.hidden_rows,
            hidden_columns=manifest.hidden_columns,
            merged_ranges=manifest.merged_ranges,
        )
        object.__setattr__(
            transform,
            "_observed_max_row",
            max(
                (
                    manifest.observed_max_row,
                    *(merged.max_row for merged in manifest.merged_ranges),
                )
            ),
        )
        object.__setattr__(
            transform,
            "_observed_max_col",
            max(
                (
                    manifest.observed_max_col,
                    *(merged.max_col for merged in manifest.merged_ranges),
                )
            ),
        )
        return transform

    def open(self, plan: ParsePlan) -> CoordinateOperation:
        """Open fresh mutable state for one parse operation."""
        return CoordinateOperation(self, plan)


@dataclass(frozen=True, slots=True)
class _RangeProjection:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def column_numbers(self) -> tuple[int, ...]:
        return tuple(range(self.min_col, self.max_col + 1))


class CoordinateOperation:
    """State for one ordered coordinate transformation."""

    __slots__ = (
        "_active_anchors",
        "_active_merge_ranges",
        "_buffer",
        "_buffered_rows",
        "_emitted_any",
        "_finished",
        "_hidden_interval_cursor",
        "_identities",
        "_last_row",
        "_merge_cursor",
        "_merge_ranges",
        "_merge_strategy",
        "_next_range_row",
        "_plan",
        "_projection",
        "_range_types",
        "_remaining_skip",
        "_template",
        "_transform",
    )

    def __init__(self, transform: CoordinateTransform, plan: ParsePlan) -> None:
        self._transform = transform
        self._plan = plan
        self._merge_strategy = MergeStrategy(plan.merge_strategy)
        self._active_anchors: dict[MergeRange, pa.Scalar] = {}
        self._projection = self._parse_projection(plan.cell_range)
        self._merge_ranges = tuple(
            sorted(
                (
                    merged_range
                    for merged_range in transform.merged_ranges
                    if self._projection is None
                    or self._merge_intersects_projection(
                        merged_range,
                        self._projection,
                    )
                ),
                key=lambda merged_range: (
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                ),
            )
        )
        self._merge_cursor = 0
        self._active_merge_ranges: list[MergeRange] = []
        self._next_range_row = self._projection.min_row if self._projection is not None else None
        self._range_types: tuple[pa.DataType, ...] | None = None
        self._remaining_skip = 0 if self._projection is not None else plan.skip_rows
        self._buffer: deque[CoordinateBatch] = deque()
        self._buffered_rows = 0
        self._identities: tuple[ColumnIdentity, ...] | None = None
        self._template: CoordinateBatch | None = None
        self._emitted_any = False
        self._hidden_interval_cursor = 0
        self._last_row: int | None = None
        self._finished = False

    @staticmethod
    def _parse_projection(cell_range: str | None) -> _RangeProjection | None:
        if not cell_range:
            return None
        if ":" not in cell_range:
            raise ValueError(f"{cell_range} is not a valid coordinate or range")
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if None in (min_col, min_row, max_col, max_row):
            raise ValueError(f"{cell_range} is not a valid coordinate or range")
        if min_row < 1 or min_col < 1:
            raise ValueError("cell range coordinates must be one-based")
        if min_row > max_row or min_col > max_col:
            raise ValueError("cell range coordinates must be ordered")
        if max_row > _MAX_EXCEL_ROW or max_col > _MAX_EXCEL_COLUMN:
            raise ValueError("cell range exceeds Excel worksheet bounds")
        return _RangeProjection(
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
        )

    def push(self, batch: CoordinateBatch) -> tuple[CoordinateBatch, ...]:
        """Validate and accept the next original-coordinate batch."""
        self._validate_batch(batch)
        return self._accept_projected(self._project_batch(batch))

    def _validate_batch(self, batch: CoordinateBatch) -> None:
        if self._finished:
            raise RuntimeError("coordinate operation is finished")
        if len(batch.row_numbers) != batch.batch.num_rows:
            raise ValueError("row coordinate count does not match batch rows")
        if len(batch.column_numbers) != batch.batch.num_columns:
            raise ValueError("column coordinate count does not match batch columns")

        rows = batch.row_numbers
        if rows.null_count:
            raise ValueError("row coordinates cannot contain nulls")
        if len(rows) and pc.min(rows).as_py() < 1:
            raise ValueError("row coordinates must be positive")

        first_row = rows[0].as_py() if len(rows) else None
        last_row = rows[-1].as_py() if len(rows) else None
        is_increasing = len(rows) < 2 or bool(
            pc.all(pc.greater(rows.slice(1), rows.slice(0, len(rows) - 1))).as_py()
        )
        if not is_increasing or (
            first_row is not None and self._last_row is not None and first_row <= self._last_row
        ):
            raise ValueError("coordinate batches are out of order")
        if last_row is not None:
            self._last_row = last_row

    def _project_batch(self, batch: CoordinateBatch) -> tuple[CoordinateBatch, ...]:
        if self._projection is not None:
            candidates = self._merge_candidates(batch)
            self._capture_merge_anchors(batch, candidates)
            projected = self._push_range(batch, candidates)
        else:
            projected = self._push_without_range(self._apply_merges(batch))
        return projected

    @staticmethod
    def _merge_intersects_projection(
        merged_range: MergeRange,
        projection: _RangeProjection,
    ) -> bool:
        return not (
            merged_range.max_row < projection.min_row
            or merged_range.min_row > projection.max_row
            or merged_range.max_col < projection.min_col
            or merged_range.min_col > projection.max_col
        )

    def _materialize_complete(
        self,
        batches: CoordinateBatch | tuple[CoordinateBatch, ...],
    ) -> tuple[CoordinateBatch, ...]:
        """Transform one complete materialized input before global framing."""
        if self._last_row is not None or self._template is not None or self._buffer:
            raise RuntimeError("materialized coordinate operation must be fresh")
        raw_batches = (batches,) if isinstance(batches, CoordinateBatch) else batches
        projected_list: list[CoordinateBatch] = []
        for batch in raw_batches:
            self._validate_batch(batch)
            projected_list.extend(self._project_batch(batch))
        projected = (*projected_list, *self._finish_projection())
        self._active_anchors.clear()
        emitted: list[CoordinateBatch] = []
        for batch in self._infer_materialized_batches(projected):
            emitted.extend(self._accept_projected((batch,)))
        emitted.extend(self._finish_framing())
        self._finished = True
        return tuple(emitted)

    def finish(self) -> tuple[CoordinateBatch, ...]:
        """Finish once; subsequent calls are idempotent."""
        if self._finished:
            return ()
        emitted = list(self._accept_projected(self._finish_projection()))
        self._active_anchors.clear()
        emitted.extend(self._finish_framing())
        self._finished = True
        return tuple(emitted)

    def _finish_projection(self) -> tuple[CoordinateBatch, ...]:
        if self._projection is None or self._next_range_row is None:
            return ()
        if self._next_range_row > self._projection.max_row:
            return ()
        result = self._null_range_batch(
            self._next_range_row,
            self._projection.max_row,
        )
        self._next_range_row = self._projection.max_row + 1
        return (result,)

    @classmethod
    def _infer_materialized_batches(
        cls,
        batches: tuple[CoordinateBatch, ...],
    ) -> tuple[CoordinateBatch, ...]:
        if not batches:
            return ()
        first = batches[0]
        targets = tuple(
            cls._materialized_target_type(tuple(batch.batch.column(index) for batch in batches))
            for index in range(first.batch.num_columns)
        )
        inferred: list[CoordinateBatch] = []
        for batch in batches:
            arrays = [
                cls._cast_materialized_array(array, targets[index])
                for index, array in enumerate(batch.batch.columns)
            ]
            inferred.append(
                CoordinateBatch(
                    batch=pa.record_batch(arrays, names=batch.batch.schema.names),
                    row_numbers=batch.row_numbers,
                    column_numbers=batch.column_numbers,
                )
            )
        return tuple(inferred)

    @staticmethod
    def _materialized_target_type(arrays: tuple[pa.Array, ...]) -> pa.DataType:
        types = {array.type for array in arrays if not pa.types.is_null(array.type)}
        if len(types) > 1:
            raise CoordinateCompatibilityError(
                "coordinate batches do not share one representable Arrow schema"
            )
        source_type = next(iter(types), pa.null())
        length = sum(len(array) for array in arrays)
        null_count = sum(array.null_count for array in arrays)
        if length and null_count == length:
            return pa.null()
        if pa.types.is_floating(source_type):
            return CoordinateOperation._floating_materialized_type(
                arrays,
                null_count,
            )
        if pa.types.is_integer(source_type):
            return CoordinateOperation._integer_materialized_type(arrays, null_count)
        if pa.types.is_timestamp(source_type):
            return CoordinateOperation._timestamp_materialized_type(
                arrays,
                source_type,
            )
        if pa.types.is_duration(source_type):
            return pa.duration("us")
        if (
            pa.types.is_string(source_type)
            or pa.types.is_large_string(source_type)
            or pa.types.is_boolean(source_type)
            or pa.types.is_null(source_type)
        ):
            return source_type
        raise CoordinateCompatibilityError(f"Arrow type {source_type} requires legacy inference")

    @staticmethod
    def _floating_materialized_type(
        arrays: tuple[pa.Array, ...],
        null_count: int,
    ) -> pa.DataType:
        if null_count:
            return pa.float64()
        all_integral = True
        all_exact = True
        for array in arrays:
            if pa.types.is_null(array.type):
                continue
            values = array.to_numpy(zero_copy_only=True)
            if not np.all(np.isfinite(values)):
                raise CoordinateCompatibilityError(
                    "non-finite numeric coordinates require legacy inference"
                )
            all_integral = all_integral and bool(np.all(np.equal(values, np.trunc(values))))
            all_exact = all_exact and bool(np.all(np.less(np.abs(values), 2**53)))
        if not all_integral:
            return pa.float64()
        if not all_exact:
            raise CoordinateCompatibilityError(
                "integral numeric coordinates exceed exact IEEE range"
            )
        return pa.int64()

    @staticmethod
    def _integer_materialized_type(
        arrays: tuple[pa.Array, ...],
        null_count: int,
    ) -> pa.DataType:
        if not null_count:
            return pa.int64()
        for array in arrays:
            if pa.types.is_null(array.type) or len(array) == array.null_count:
                continue
            bounds = pc.min_max(array)
            minimum = bounds["min"].as_py()
            maximum = bounds["max"].as_py()
            if minimum < -(2**53) or maximum > 2**53:
                raise CoordinateCompatibilityError(
                    "nullable integer coordinates exceed exact IEEE range"
                )
        return pa.float64()

    @staticmethod
    def _timestamp_materialized_type(
        arrays: tuple[pa.Array, ...],
        source_type: pa.TimestampType,
    ) -> pa.DataType:
        if source_type.tz is not None:
            raise CoordinateCompatibilityError("timezone-aware timestamps require legacy inference")
        for array in arrays:
            if pa.types.is_null(array.type):
                continue
            normalized = CoordinateOperation._safe_cast(
                array,
                pa.timestamp("us"),
                "timestamp coordinates cannot normalize losslessly",
            )
            dates = CoordinateOperation._safe_cast(
                normalized,
                pa.date32(),
                "timestamp coordinates cannot normalize losslessly",
            )
            for epoch in (date(1899, 12, 31), date(1904, 1, 1)):
                if bool(pc.any(pc.equal(dates, pa.scalar(epoch))).as_py()):
                    raise CoordinateCompatibilityError(
                        "time-only timestamp coordinates require legacy inference"
                    )
        return pa.timestamp("us")

    @staticmethod
    def _cast_materialized_array(array: pa.Array, target: pa.DataType) -> pa.Array:
        if array.type == target:
            return array
        if pa.types.is_null(target):
            return pa.nulls(len(array))
        return CoordinateOperation._safe_cast(
            array,
            target,
            f"coordinate values cannot cast losslessly to {target}",
        )

    @staticmethod
    def _safe_cast(array: pa.Array, target: pa.DataType, message: str) -> pa.Array:
        try:
            return pc.cast(array, target, safe=True)
        except (pa.ArrowInvalid, pa.ArrowTypeError) as error:
            raise CoordinateCompatibilityError(message) from error

    def _merge_candidates(
        self,
        batch: CoordinateBatch,
    ) -> tuple[MergeRange, ...]:
        if not self._merge_ranges or batch.batch.num_rows == 0:
            return ()
        first_row = batch.row_numbers[0].as_py()
        last_row = batch.row_numbers[-1].as_py()
        self._active_merge_ranges = [
            merged_range
            for merged_range in self._active_merge_ranges
            if merged_range.max_row >= first_row
        ]
        while (
            self._merge_cursor < len(self._merge_ranges)
            and self._merge_ranges[self._merge_cursor].min_row <= last_row
        ):
            merged_range = self._merge_ranges[self._merge_cursor]
            if merged_range.max_row >= first_row:
                self._active_merge_ranges.append(merged_range)
            self._merge_cursor += 1
        return tuple(self._active_merge_ranges)

    def _capture_merge_anchors(
        self,
        batch: CoordinateBatch,
        merged_ranges: tuple[MergeRange, ...],
    ) -> None:
        if self._merge_strategy is not MergeStrategy.FILL or not merged_ranges:
            return
        arrays = list(batch.batch.columns)
        position_by_column = {
            coordinate: position for position, coordinate in enumerate(batch.column_numbers)
        }
        first_row = batch.row_numbers[0].as_py()
        last_row = batch.row_numbers[-1].as_py()
        for merged_range in merged_ranges:
            self._resolve_anchor(
                batch,
                arrays,
                position_by_column,
                merged_range,
                first_row,
                last_row,
            )

    def _apply_merges(
        self,
        batch: CoordinateBatch,
        merged_ranges: tuple[MergeRange, ...] | None = None,
    ) -> CoordinateBatch:
        if (
            self._merge_strategy is MergeStrategy.SKIP
            or not self._merge_ranges
            or batch.batch.num_rows == 0
        ):
            return batch

        first_row = batch.row_numbers[0].as_py()
        last_row = batch.row_numbers[-1].as_py()
        candidates = merged_ranges if merged_ranges is not None else self._merge_candidates(batch)
        if not candidates:
            return batch
        for merged_range in tuple(self._active_anchors):
            if merged_range.max_row < first_row:
                del self._active_anchors[merged_range]

        arrays = list(batch.batch.columns)
        position_by_column = {
            coordinate: position for position, coordinate in enumerate(batch.column_numbers)
        }
        replacements = self._collect_merge_replacements(
            batch,
            arrays,
            position_by_column,
            candidates,
            first_row,
            last_row,
        )

        for position, segments in replacements.items():
            arrays[position] = self._rebuild_merged_array(arrays[position], segments)

        return CoordinateBatch(
            batch=pa.record_batch(arrays, names=batch.batch.schema.names),
            row_numbers=batch.row_numbers,
            column_numbers=batch.column_numbers,
            column_identities=batch.column_identities,
        )

    def _collect_merge_replacements(
        self,
        batch: CoordinateBatch,
        arrays: list[pa.Array],
        position_by_column: dict[int, int],
        candidates: tuple[MergeRange, ...],
        first_row: int,
        last_row: int,
    ) -> dict[int, list[tuple[int, int, pa.Scalar]]]:
        rows = batch.row_numbers.to_numpy(zero_copy_only=True)
        replacements: dict[int, list[tuple[int, int, pa.Scalar]]] = {}
        for merged_range in candidates:
            if merged_range.max_row < first_row or merged_range.min_row > last_row:
                continue
            self._collect_one_merge_replacements(
                batch,
                arrays,
                position_by_column,
                replacements,
                rows,
                merged_range,
                first_row,
                last_row,
            )
            if last_row >= merged_range.max_row:
                self._active_anchors.pop(merged_range, None)
        return replacements

    def _collect_one_merge_replacements(
        self,
        batch: CoordinateBatch,
        arrays: list[pa.Array],
        position_by_column: dict[int, int],
        replacements: dict[int, list[tuple[int, int, pa.Scalar]]],
        rows: np.ndarray[tuple[int], np.dtype[np.int64]],
        merged_range: MergeRange,
        first_row: int,
        last_row: int,
    ) -> None:
        anchor = self._resolve_anchor(
            batch,
            arrays,
            position_by_column,
            merged_range,
            first_row,
            last_row,
        )
        start = int(np.searchsorted(rows, merged_range.min_row, side="left"))
        end = int(np.searchsorted(rows, merged_range.max_row, side="right"))
        for column in range(merged_range.min_col, merged_range.max_col + 1):
            position = position_by_column.get(column)
            if position is None:
                continue
            segment_start = self._merge_segment_start(
                rows,
                merged_range,
                column,
                start,
                end,
            )
            if segment_start >= end:
                continue
            arrays[position], replacement = self._merge_replacement(
                arrays[position],
                anchor,
            )
            replacements.setdefault(position, []).append((segment_start, end, replacement))

    def _merge_segment_start(
        self,
        rows: np.ndarray[tuple[int], np.dtype[np.int64]],
        merged_range: MergeRange,
        column: int,
        start: int,
        end: int,
    ) -> int:
        is_anchor = (
            self._merge_strategy is MergeStrategy.FIRST_ONLY
            and column == merged_range.min_col
            and start < end
            and rows[start] == merged_range.min_row
        )
        return start + int(is_anchor)

    def _merge_replacement(
        self,
        array: pa.Array,
        anchor: pa.Scalar | None,
    ) -> tuple[pa.Array, pa.Scalar]:
        if self._merge_strategy is not MergeStrategy.FILL:
            return array, pa.scalar(None, type=array.type)
        if anchor is None:
            raise CoordinateCompatibilityError(
                "merged-cell anchor is unavailable in the coordinate stream"
            )
        return self._fill_replacement(array, anchor)

    @staticmethod
    def _rebuild_merged_array(
        array: pa.Array,
        segments: list[tuple[int, int, pa.Scalar]],
    ) -> pa.Array:
        pieces: list[pa.Array] = []
        cursor = 0
        for start, end, replacement in segments:
            if start < cursor:
                raise CoordinateCompatibilityError(
                    "overlapping merged cells require legacy inference"
                )
            if start > cursor:
                pieces.append(array.slice(cursor, start - cursor))
            pieces.append(pa.repeat(replacement, end - start))
            cursor = end
        if cursor < len(array):
            pieces.append(array.slice(cursor))
        if len(pieces) == 1:
            return pieces[0]
        return pa.concat_arrays(pieces)

    def _resolve_anchor(
        self,
        batch: CoordinateBatch,
        arrays: list[pa.Array],
        position_by_column: dict[int, int],
        merged_range: MergeRange,
        first_row: int,
        last_row: int,
    ) -> pa.Scalar | None:
        anchor = self._active_anchors.get(merged_range)
        anchor_position = position_by_column.get(merged_range.min_col)
        if (
            anchor is not None
            or anchor_position is None
            or not first_row <= merged_range.min_row <= last_row
        ):
            return anchor
        index = pc.index(
            batch.row_numbers,
            pa.scalar(merged_range.min_row, type=pa.int64()),
        ).as_py()
        if index < 0:
            return None
        anchor = arrays[anchor_position][index]
        if self._merge_strategy is MergeStrategy.FILL:
            self._active_anchors[merged_range] = anchor
        return anchor

    @staticmethod
    def _fill_replacement(
        array: pa.Array,
        anchor: pa.Scalar,
    ) -> tuple[pa.Array, pa.Scalar]:
        if not anchor.is_valid:
            return array, pa.scalar(None, type=array.type)
        if pa.types.is_null(array.type):
            return pc.cast(array, anchor.type), anchor
        if anchor.type != array.type and not (
            pa.types.is_integer(anchor.type) and pa.types.is_floating(array.type)
        ):
            raise CoordinateCompatibilityError("merged-cell fill requires lossy Arrow coercion")
        try:
            replacement = pa.scalar(anchor.as_py(), type=array.type)
        except (
            pa.ArrowInvalid,
            pa.ArrowNotImplementedError,
            pa.ArrowTypeError,
            OverflowError,
            TypeError,
            ValueError,
        ) as error:
            raise CoordinateCompatibilityError(
                "merged-cell fill requires lossy Arrow coercion"
            ) from error
        if (
            pa.types.is_integer(anchor.type)
            and pa.types.is_floating(array.type)
            and int(replacement.as_py()) != anchor.as_py()
        ):
            raise CoordinateCompatibilityError("merged-cell fill requires lossy Arrow coercion")
        return array, replacement

    def _push_without_range(
        self,
        batch: CoordinateBatch,
    ) -> tuple[CoordinateBatch, ...]:
        row_mask: pa.BooleanArray | pa.ChunkedArray | None = None
        if self._plan.ignore_hidden and self._transform.hidden_rows.intervals:
            row_mask = self._visible_row_mask(batch.row_numbers)

        selected_positions = tuple(
            position
            for position, coordinate in enumerate(batch.column_numbers)
            if not (
                self._plan.ignore_hidden and self._transform.hidden_columns.contains(coordinate)
            )
        )
        selected_rows = (
            batch.row_numbers
            if row_mask is None
            else pc.filter(
                batch.row_numbers,
                row_mask,
            )
        )
        arrays = [batch.batch.column(position) for position in selected_positions]
        if row_mask is not None:
            arrays = [pc.filter(array, row_mask) for array in arrays]
        if len(selected_rows) == 0 or not arrays:
            return ()
        columns = tuple(batch.column_numbers[position] for position in selected_positions)
        return (self._coordinate_batch(arrays, selected_rows, columns),)

    def _visible_row_mask(self, row_numbers: pa.Int64Array) -> pa.BooleanArray:
        coordinates = row_numbers.to_numpy(zero_copy_only=True)
        intervals = self._transform.hidden_rows.intervals
        first_row = int(coordinates[0])
        last_row = int(coordinates[-1])
        cursor = self._hidden_interval_cursor
        while cursor < len(intervals) and intervals[cursor].end < first_row:
            cursor += 1
        self._hidden_interval_cursor = cursor
        stop = bisect_right(
            self._transform.hidden_rows.starts,
            last_row,
            lo=cursor,
        )
        overlapping = intervals[cursor:stop]
        if not overlapping:
            return pc.equal(row_numbers, row_numbers)
        starts = np.fromiter(
            (interval.start for interval in overlapping),
            dtype=np.int64,
            count=len(overlapping),
        )
        ends = np.fromiter(
            (interval.end for interval in overlapping),
            dtype=np.int64,
            count=len(overlapping),
        )
        positions = np.searchsorted(starts, coordinates, side="right") - 1
        covered = np.zeros(len(coordinates), dtype=np.bool_)
        eligible = positions >= 0
        covered[eligible] = coordinates[eligible] <= ends[positions[eligible]]
        return pa.array(~covered, type=pa.bool_())

    def _push_range(
        self,
        batch: CoordinateBatch,
        merged_ranges: tuple[MergeRange, ...],
    ) -> tuple[CoordinateBatch, ...]:
        projection = self._projection
        next_row = self._next_range_row
        assert projection is not None
        assert next_row is not None

        requested_columns = projection.column_numbers
        position_by_column = {
            coordinate: position for position, coordinate in enumerate(batch.column_numbers)
        }
        if self._range_types is None:
            self._range_types = tuple(
                batch.batch.column(position_by_column[column]).type
                if column in position_by_column
                else pa.null()
                for column in requested_columns
            )

        rows = batch.row_numbers.to_numpy(zero_copy_only=True)
        start = int(
            np.searchsorted(
                rows,
                max(projection.min_row, next_row),
                side="left",
            )
        )
        stop = int(np.searchsorted(rows, projection.max_row, side="right"))
        if start >= stop:
            return ()

        last_selected = int(rows[stop - 1])
        target_rows = pa.array(
            range(next_row, last_selected + 1),
            type=pa.int64(),
        )
        indices = pc.index_in(target_rows, value_set=batch.row_numbers)
        aligned = [
            pc.take(batch.batch.column(position_by_column[column]), indices)
            if column in position_by_column
            else pa.nulls(len(target_rows), type=self._range_types[index])
            for index, column in enumerate(requested_columns)
        ]
        self._next_range_row = last_selected + 1
        projected = self._coordinate_batch(aligned, target_rows, requested_columns)
        return (self._apply_merges(projected, merged_ranges),)

    def _null_range_batch(self, start_row: int, end_row: int) -> CoordinateBatch:
        projection = self._projection
        assert projection is not None
        row_numbers = pa.array(range(start_row, end_row + 1), type=pa.int64())
        column_types = self._range_types or tuple(pa.null() for _ in projection.column_numbers)
        projected = self._coordinate_batch(
            [pa.nulls(len(row_numbers), type=column_type) for column_type in column_types],
            row_numbers,
            projection.column_numbers,
        )
        return self._apply_merges(projected)

    @staticmethod
    def _coordinate_batch(
        arrays: list[pa.Array | pa.ChunkedArray],
        row_numbers: pa.Array | pa.ChunkedArray,
        column_numbers: tuple[int, ...],
    ) -> CoordinateBatch:
        return CoordinateBatch(
            batch=pa.record_batch(
                arrays,
                names=[str(ordinal) for ordinal in range(len(arrays))],
            ),
            row_numbers=pa.array(row_numbers, type=pa.int64()),
            column_numbers=column_numbers,
        )

    def _accept_projected(
        self,
        batches: tuple[CoordinateBatch, ...],
    ) -> tuple[CoordinateBatch, ...]:
        emitted: list[CoordinateBatch] = []
        for batch in batches:
            if self._template is None:
                self._template = self._detach_batch(batch.slice_rows(0, 0))
            if self._remaining_skip:
                if batch.batch.num_rows <= self._remaining_skip:
                    self._remaining_skip -= batch.batch.num_rows
                    continue
                batch = batch.slice_rows(
                    self._remaining_skip,
                    batch.batch.num_rows - self._remaining_skip,
                )
                self._remaining_skip = 0
            if batch.batch.num_rows:
                self._buffer.append(batch)
                self._buffered_rows += batch.batch.num_rows
            emitted.extend(self._drain_ready())
        if emitted:
            self._emitted_any = True
        return tuple(emitted)

    def _drain_ready(self) -> tuple[CoordinateBatch, ...]:
        if self._identities is None:
            if self._plan.header_rows == 0:
                self._identities = self._generic_identities()
            elif self._buffered_rows > self._plan.header_rows + self._plan.skip_footer:
                header = self._take_prefix(self._plan.header_rows)
                self._identities = self._header_identities(header)
            else:
                return ()
        ready = self._buffered_rows - self._plan.skip_footer
        if ready <= 0:
            return ()
        return tuple(self._attach_identities(batch) for batch in self._take_prefix(ready))

    def _finish_framing(self) -> tuple[CoordinateBatch, ...]:
        self._drop_tail(min(self._plan.skip_footer, self._buffered_rows))
        if self._identities is None:
            if self._plan.header_rows > 0 and self._buffered_rows >= self._plan.header_rows:
                header = self._take_prefix(self._plan.header_rows)
                self._identities = self._header_identities(header)
            else:
                self._identities = self._generic_identities()

        emitted = tuple(
            self._attach_identities(batch) for batch in self._take_prefix(self._buffered_rows)
        )
        if emitted:
            self._emitted_any = True
            return emitted
        if not self._emitted_any and self._template is not None:
            return (self._attach_identities(self._template),)
        return ()

    def _take_prefix(self, count: int) -> tuple[CoordinateBatch, ...]:
        taken: list[CoordinateBatch] = []
        remaining = count
        while remaining:
            batch = self._buffer[0]
            if batch.batch.num_rows <= remaining:
                taken.append(self._buffer.popleft())
                remaining -= batch.batch.num_rows
                continue
            taken.append(batch.slice_rows(0, remaining))
            retained = batch.slice_rows(
                remaining,
                batch.batch.num_rows - remaining,
            )
            if self._buffered_rows - count <= self._plan.header_rows + self._plan.skip_footer:
                retained = self._detach_batch(retained)
            self._buffer[0] = retained
            remaining = 0
        self._buffered_rows -= count
        return tuple(taken)

    @staticmethod
    def _detach_batch(batch: CoordinateBatch) -> CoordinateBatch:
        arrays = [pa.concat_arrays([array]) for array in batch.batch.columns]
        row_numbers = pa.concat_arrays([batch.row_numbers])
        return CoordinateBatch(
            batch=pa.record_batch(arrays, names=batch.batch.schema.names),
            row_numbers=row_numbers,
            column_numbers=batch.column_numbers,
            column_identities=batch.column_identities,
        )

    def _drop_tail(self, count: int) -> None:
        remaining = count
        while remaining:
            batch = self._buffer[-1]
            if batch.batch.num_rows <= remaining:
                self._buffer.pop()
                remaining -= batch.batch.num_rows
                continue
            self._buffer[-1] = batch.slice_rows(0, batch.batch.num_rows - remaining)
            remaining = 0
        self._buffered_rows -= count

    def _generic_identities(self) -> tuple[ColumnIdentity, ...]:
        template = self._template
        if template is None:
            return ()
        return tuple(
            ColumnIdentity(ordinal, f"col_{ordinal}")
            for ordinal in range(template.batch.num_columns)
        )

    def _header_identities(
        self,
        header: tuple[CoordinateBatch, ...],
    ) -> tuple[ColumnIdentity, ...]:
        template = self._template
        assert template is not None
        values_by_column = tuple(
            tuple(
                batch.batch.column(column)[row].as_py()
                for batch in header
                for row in range(batch.batch.num_rows)
            )
            for column in range(template.batch.num_columns)
        )
        display_names: list[str] = []
        for ordinal, values in enumerate(values_by_column):
            if self._plan.header_rows == 1:
                value = values[0]
                display_names.append(
                    f"col_{ordinal}"
                    if value is None or (isinstance(value, float) and math.isnan(value))
                    else str(value)
                )
                continue
            parts = []
            for value in values:
                if value is None or (isinstance(value, float) and math.isnan(value)):
                    continue
                part = str(value).strip()
                if part and part.lower() != "nan":
                    parts.append(part)
            display_names.append("__".join(parts) if parts else f"col_{ordinal}")
        return tuple(
            ColumnIdentity(ordinal, display_name)
            for ordinal, display_name in enumerate(display_names)
        )

    def _attach_identities(self, batch: CoordinateBatch) -> CoordinateBatch:
        identities = self._identities
        assert identities is not None
        return CoordinateBatch(
            batch=batch.batch,
            row_numbers=batch.row_numbers,
            column_numbers=batch.column_numbers,
            column_identities=identities,
        )
