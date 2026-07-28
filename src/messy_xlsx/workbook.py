"""MessyWorkbook - Main entry point for parsing Excel files."""

# ============================================================================
# Imports
# ============================================================================

import logging
import re
import weakref
from collections import deque
from collections.abc import Callable, Iterator
from contextlib import contextmanager
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from itertools import islice
from pathlib import Path
from types import TracebackType
from typing import Any, BinaryIO, Generic, NoReturn, TypeVar, cast

import numpy as np
import openpyxl
import pandas as pd
import pyarrow as pa

import messy_xlsx.parsing.csv_native as csv_native
from messy_xlsx._fallback_signals import (
    _blocks_backend_retry,
    _bounded_exception_graph,
    _contains_process_failure,
    _exception_traceback,
    _FallbackBlockReason,
    _mark_fallback_blocked,
)
from messy_xlsx._source import BackendSource, SourceHandle, describe_source
from messy_xlsx.cache import get_structure_cache
from messy_xlsx.detection.structure_analyzer import StructureAnalyzer
from messy_xlsx.detection.structure_sampler import StructureSampler
from messy_xlsx.enums import MergeStrategy
from messy_xlsx.exceptions import FileError, FormatError, StreamingTypeError
from messy_xlsx.formulas.config import FormulaConfig, FormulaEvaluationMode
from messy_xlsx.formulas.engine import FormulaEngine
from messy_xlsx.models import (
    CellValue,
    SheetConfig,
    SheetError,
    SheetInfo,
    SheetResult,
    StructureInfo,
)
from messy_xlsx.normalization import (
    NormalizationSample,
    compile_normalization_plan,
)
from messy_xlsx.normalization.pipeline import NormalizationPipeline
from messy_xlsx.normalization.plan import (
    MAX_SAMPLE_BYTES,
    MAX_SAMPLE_CELLS,
    MAX_SAMPLE_VALUES,
    _display_label_token,
    _LabelResolutionIndex,
    _reject_ambiguous_temporal_resolution,
    _safe_name_text,
    _snapshot_display_name,
    _timezone_label_projection,
)
from messy_xlsx.ooxml.manifest import ManifestReader
from messy_xlsx.ooxml.models import IntervalIndex, SheetManifest
from messy_xlsx.parsing.contracts import BackendKind, OutputMode, ParseMetrics
from messy_xlsx.parsing.coordinates import (
    CoordinateBatch,
    CoordinateOperation,
    CoordinateTransform,
)
from messy_xlsx.parsing.csv_contracts import CSVExecutionKind, CSVExecutionReason
from messy_xlsx.parsing.csv_handler import CSVHandler
from messy_xlsx.parsing.csv_streaming import prepare_csv_streaming_reader
from messy_xlsx.parsing.fastexcel_session import FastexcelSession
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.parsing.materialized_streaming import (
    PreparedStreamingReader as _PreparedStreamingOperation,
)
from messy_xlsx.parsing.materialized_streaming import (
    _CloseOnceReader,
    _NormalizationSampleAccumulator,
    prepare_materialized_streaming_reader,
    wrap_normalized_streaming_reader,
)
from messy_xlsx.parsing.parse_plan import (
    ParsePlan,
    compile_parse_plan,
    requires_structure_analysis,
)
from messy_xlsx.parsing.physical_values import (
    PandasTemporalPayload,
    UnsupportedPhysicalValueError,
    arrow_temporal_array,
    common_temporal_arrow_type,
    ensure_supported_physical_value,
    pandas_temporal_payload,
    physical_label_description,
    physical_value_description,
    temporal_payload,
)
from messy_xlsx.parsing.router import BackendRouter, WorkbookContext
from messy_xlsx.parsing.sheet_planner import (
    PlannedSheet,
    PlannedSheetState,
    SheetPlanner,
)
from messy_xlsx.parsing.streams import (
    BatchStream,
    DataFrameChunkStream,
    SheetStream,
    _close_if_present,
    _run_cleanups,
)
from messy_xlsx.parsing.xls_handler import XLSHandler
from messy_xlsx.parsing.xlsx_handler import (
    XLSXHandler,
    _is_fastexcel_materialized_plan,
)
from messy_xlsx.parsing.xlsx_streaming import (
    OpenpyxlStreamingReader,
    StreamingWorksheetLayout,
    _enter_prearmed_context,
    _raw_coordinate_batch,
    _RawBatchBudget,
    _read_bounds,
    _record_batch_with_row_count,
    _RetryableSourceContext,
    reader_batches,
)
from messy_xlsx.sheet import MessySheet
from messy_xlsx.warnings import warn_legacy

# ============================================================================
# Core
# ============================================================================

logger = logging.getLogger(__name__)
_NO_BUILTIN_MATERIALIZATION = object()
_NORMALIZATION_SAMPLE_ROWS = MAX_SAMPLE_VALUES
_NORMALIZATION_SAMPLE_CELLS = MAX_SAMPLE_CELLS
_NORMALIZATION_SAMPLE_BYTES = MAX_SAMPLE_BYTES
_ARROW_ARRAY_ERRORS = (
    pa.ArrowInvalid,
    pa.ArrowNotImplementedError,
    pa.ArrowTypeError,
    OverflowError,
    TypeError,
    ValueError,
)


def _visible_row_intervals(
    start_row: int,
    end_row: int,
    hidden_rows: IntervalIndex,
    *,
    ignore_hidden: bool,
) -> tuple[tuple[int, int], ...]:
    """Return compact visible spans without visiting framed-out rows."""
    if start_row > end_row:
        return ()
    if not ignore_hidden or not hidden_rows.intervals:
        return ((start_row, end_row),)
    visible: list[tuple[int, int]] = []
    cursor = start_row
    for hidden in hidden_rows.intervals:
        if hidden.end < cursor:
            continue
        if hidden.start > end_row:
            break
        if cursor < hidden.start:
            visible.append((cursor, min(end_row, hidden.start - 1)))
        cursor = max(cursor, hidden.end + 1)
        if cursor > end_row:
            break
    if cursor <= end_row:
        visible.append((cursor, end_row))
    return tuple(visible)


def _slice_row_intervals(
    intervals: tuple[tuple[int, int], ...],
    *,
    skip: int,
    count: int,
) -> tuple[tuple[int, int], ...]:
    """Select a bounded ordinal slice from compact inclusive intervals."""
    if count <= 0:
        return ()
    selected: list[tuple[int, int]] = []
    remaining_skip = skip
    remaining_count = count
    for start, end in intervals:
        length = end - start + 1
        if remaining_skip >= length:
            remaining_skip -= length
            continue
        selected_start = start + remaining_skip
        remaining_skip = 0
        selected_end = min(end, selected_start + remaining_count - 1)
        selected.append((selected_start, selected_end))
        remaining_count -= selected_end - selected_start + 1
        if remaining_count == 0:
            break
    return tuple(selected)


def _interval_row_count(intervals: tuple[tuple[int, int], ...]) -> int:
    return sum(end - start + 1 for start, end in intervals)


def _row_is_in_intervals(
    row: int,
    intervals: tuple[tuple[int, int], ...],
) -> bool:
    return any(start <= row <= end for start, end in intervals)


def _row_range_intersects_intervals(
    start_row: int,
    end_row: int,
    intervals: tuple[tuple[int, int], ...],
) -> bool:
    return any(start <= end_row and end >= start_row for start, end in intervals)


def _coalesce_row_windows(
    intervals: tuple[tuple[int, int], ...],
    rows: set[int],
) -> tuple[tuple[int, int], ...]:
    """Merge selected spans and singleton state rows without filling gaps."""
    candidates = [*intervals, *((row, row) for row in rows)]
    if not candidates:
        return ()
    merged: list[tuple[int, int]] = []
    for start, end in sorted(candidates):
        if merged and start <= merged[-1][1] + 1:
            previous_start, previous_end = merged[-1]
            merged[-1] = (previous_start, max(previous_end, end))
        else:
            merged.append((start, end))
    return tuple(merged)


def _sample_capacity(width: int, header_rows: int) -> tuple[int, str]:
    """Bound required header state plus retained post-coordinate evidence."""
    capacities = [(_NORMALIZATION_SAMPLE_ROWS + header_rows, "rows")]
    if width:
        capacities.extend(
            (
                (_NORMALIZATION_SAMPLE_CELLS // width, "cells"),
                (_NORMALIZATION_SAMPLE_BYTES // (16 * width), "bytes"),
            )
        )
    return min(capacities, key=lambda item: item[0])


def _sample_budget_error(budget: _RawBatchBudget, reason: str | None = None) -> ValueError:
    exhausted = reason or budget.exhausted_reason or "rows"
    limit = {
        "rows": budget.max_rows,
        "cells": budget.max_cells,
        "bytes": budget.max_bytes,
    }[exhausted]
    return ValueError(f"sample raw window may retain at most {limit} {exhausted}")


def _arrow_table_from_dataframe(frame: pd.DataFrame) -> pa.Table:
    """Convert columns positionally so duplicate labels remain valid."""
    if len(frame.columns) == 0:
        schema = pa.schema([])
        batch = _record_batch_with_row_count([], schema, len(frame))
        return pa.Table.from_batches([batch], schema=schema)
    arrays = [
        _materialized_arrow_array(
            frame.iloc[:, ordinal],
            ordinal=ordinal,
            display_label=frame.columns[ordinal],
        )
        for ordinal in range(len(frame.columns))
    ]
    return pa.Table.from_arrays(
        arrays,
        names=[_safe_arrow_field_name(label) for label in frame.columns],
    )


def _safe_arrow_field_name(label: object) -> str:
    """Project one pandas label without invoking caller-controlled text hooks."""
    snapshot = _snapshot_display_name(label)
    if type(snapshot) is str:
        return snapshot
    return cast(str, _safe_name_text(snapshot))


def _rename_materialized_columns(
    frame: pd.DataFrame,
    items: list[tuple[object, object]],
) -> pd.DataFrame:
    """Apply legacy renames positionally without pandas temporal coercion."""
    _validate_materialized_label_targets(frame.columns, items)
    candidate_tokens = tuple(_display_label_token(label) for label, _value in items)
    candidate_index = _LabelResolutionIndex(candidate_tokens)
    renamed: list[object] = []
    for label in frame.columns:
        matches = candidate_index.matching(_display_label_token(label))
        renamed.append(items[matches[0][0]][1] if matches else label)
    result = frame.copy()
    result.columns = pd.Index(renamed, dtype=object, tupleize_cols=False)
    return result


def _materialized_condition_ordinals(
    columns: pd.Index,
    label: object,
) -> tuple[int, ...]:
    """Resolve legacy condition labels without DatetimeIndex inference."""
    tokens = tuple(_display_label_token(column) for column in columns)
    label_token = _display_label_token(label)
    matches = _LabelResolutionIndex(tokens).matching(label_token)
    _reject_ambiguous_temporal_resolution(
        label_token,
        tuple(token for _ordinal, token in matches),
    )
    return tuple(ordinal for ordinal, _token in matches)


def _validate_materialized_label_targets(
    columns: pd.Index,
    items: list[tuple[object, object]],
) -> None:
    """Apply the streaming ambiguity policy before legacy label lookup."""
    target_tokens = tuple(_display_label_token(column) for column in columns)
    target_index = _LabelResolutionIndex(target_tokens)
    for label, _value in items:
        label_token = _display_label_token(label)
        matches = target_index.matching(label_token)
        _reject_ambiguous_temporal_resolution(
            label_token,
            tuple(token for _ordinal, token in matches),
        )


def _dataframe_from_record_batch(
    batch: pa.RecordBatch,
    display_names: tuple[object, ...],
) -> pd.DataFrame:
    """Restore Arrow-backed columns by ordinal before assigning public labels."""
    if batch.num_columns == 0:
        frame = pd.DataFrame(index=pd.RangeIndex(batch.num_rows))
    else:
        columns = [
            batch.column(ordinal).to_pandas(types_mapper=pd.ArrowDtype)
            for ordinal in range(batch.num_columns)
        ]
        frame = pd.concat(columns, axis=1)
    frame.columns = pd.Index(display_names, dtype=object, tupleize_cols=False)
    return frame


def _materialized_arrow_array(
    series: pd.Series,
    *,
    ordinal: int,
    display_label: object,
) -> pa.Array:
    """Keep the homogeneous fast path and preserve supported mixed scalars."""
    if pd.api.types.is_object_dtype(series.dtype):
        for row_offset, value in enumerate(series):
            if _is_materialized_null(value):
                continue
            try:
                ensure_supported_physical_value(value)
            except UnsupportedPhysicalValueError:
                _raise_materialized_unsupported(
                    value,
                    ordinal=ordinal,
                    display_label=display_label,
                    row_offset=row_offset,
                )
        return _materialized_union_array(series)
    if isinstance(series.dtype, pd.DatetimeTZDtype):
        for row_offset, value in enumerate(series):
            if _is_materialized_null(value):
                continue
            try:
                ensure_supported_physical_value(value)
            except UnsupportedPhysicalValueError:
                _raise_materialized_unsupported(
                    value,
                    ordinal=ordinal,
                    display_label=display_label,
                    row_offset=row_offset,
                )
            break
    if _supports_direct_materialized_arrow_conversion(series.dtype):
        try:
            return pa.array(series, from_pandas=True)
        except _ARROW_ARRAY_ERRORS:
            pass
    return _materialized_union_array(series)


def _raise_materialized_unsupported(
    value: object,
    *,
    ordinal: int,
    display_label: object,
    row_offset: int,
) -> NoReturn:
    raise StreamingTypeError(
        "streamed value is incompatible with the fixed schema",
        ordinal=ordinal,
        display_label=physical_label_description(display_label),
        row_offset=row_offset,
        value_description=physical_value_description(value),
        expected_type="supported Arrow scalar",
    ) from None


def _supports_direct_materialized_arrow_conversion(dtype: object) -> bool:
    """Select pandas dtypes Arrow can consume without Python scalar discovery."""
    return bool(
        pd.api.types.is_bool_dtype(dtype)
        or pd.api.types.is_integer_dtype(dtype)
        or pd.api.types.is_float_dtype(dtype)
        or pd.api.types.is_datetime64_any_dtype(dtype)
        or pd.api.types.is_timedelta64_dtype(dtype)
    )


def _homogeneous_materialized_object_type(
    series: pd.Series,
    inferred_type: pa.DataType,
) -> tuple[bool, pa.DataType | None]:
    """Validate one vector conversion without retaining a duplicate value list."""
    family: tuple[object, ...] | None = None
    requires_timestamp_ns = False
    requires_duration_ns = False

    for value in series:
        if _is_materialized_null(value):
            continue
        current_family = _materialized_bucket_key(value)
        if family is None:
            family = current_family
        elif current_family != family:
            return False, None

        value_type = type(value)
        if value_type is pd.Timestamp:
            timestamp = cast(pd.Timestamp, value)
            requires_timestamp_ns = requires_timestamp_ns or timestamp.nanosecond != 0

        if value_type is pd.Timedelta:
            duration = cast(pd.Timedelta, value)
            requires_duration_ns = requires_duration_ns or duration.nanoseconds != 0

    if requires_timestamp_ns:
        if not pa.types.is_timestamp(inferred_type):
            return False, None
        return True, pa.timestamp("ns", tz=inferred_type.tz)
    if requires_duration_ns:
        if not pa.types.is_duration(inferred_type):
            return False, None
        return True, pa.duration("ns")
    return True, None


@dataclass(slots=True)
class _MaterializedBucket:
    """One hook-free logical family collected before vector conversion."""

    key: tuple[object, ...]
    values: list[object]
    positions: list[int]
    requires_ns: bool = False


@dataclass(slots=True)
class _MaterializedConversionPlan:
    """Complete scalar discovery before the first Arrow conversion."""

    buckets: tuple[_MaterializedBucket, ...]
    null_positions: tuple[int, ...]
    canonical_values: tuple[object | None, ...]
    child_count: int


def _materialized_union_array(series: pd.Series) -> pa.Array:
    """Build a dense union with Arrow work bounded by logical families."""
    plan = _plan_materialized_conversion(series)
    if not plan.buckets:
        return pa.nulls(len(series))

    converted: list[tuple[list[int], pa.Array, int]] = []
    encounter = 0
    for bucket in plan.buckets:
        for positions, child in _convert_materialized_bucket(bucket):
            converted.append((positions, child, encounter))
            encounter += 1
    converted = _coalesce_materialized_children(converted)
    converted.sort(key=lambda item: (str(item[1].type), item[2]))
    if len(converted) > 127:
        raise ValueError("mixed materialized column exceeds Arrow union type limit")

    if len(converted) == 1:
        positions, child, _encounter = converted[0]
        if not plan.null_positions and positions == list(range(len(series))):
            return child
        offsets_by_position = {position: offset for offset, position in enumerate(positions)}
        return child.take(
            pa.array(
                [offsets_by_position.get(position) for position in range(len(series))],
                type=pa.int32(),
            )
        )

    type_ids = np.empty(len(series), dtype=np.int8)
    offsets = np.empty(len(series), dtype=np.int32)
    children: list[pa.Array] = []
    for child_index, (positions, child, _encounter) in enumerate(converted):
        children.append(child)
        for offset, position in enumerate(positions):
            type_ids[position] = child_index
            offsets[position] = offset

    if plan.null_positions:
        null_child = children[0]
        null_start = len(null_child)
        children[0] = pa.concat_arrays(
            [null_child, pa.nulls(len(plan.null_positions), type=null_child.type)]
        )
        for null_offset, position in enumerate(plan.null_positions):
            type_ids[position] = 0
            offsets[position] = null_start + null_offset

    type_id_array = pa.Array.from_buffers(
        pa.int8(),
        len(type_ids),
        [None, pa.py_buffer(type_ids)],
    )
    offset_array = pa.Array.from_buffers(
        pa.int32(),
        len(offsets),
        [None, pa.py_buffer(offsets)],
    )
    return pa.UnionArray.from_dense(
        type_id_array,
        offset_array,
        children,
        field_names=[f"variant_{index}_{child.type}" for index, child in enumerate(children)],
        type_codes=list(range(len(children))),
    )


def _plan_materialized_conversion(series: pd.Series) -> _MaterializedConversionPlan:
    """Discover, validate, group, and cap variants before calling Arrow."""
    buckets: dict[tuple[object, ...], _MaterializedBucket] = {}
    null_positions: list[int] = []
    canonical_values: list[object | None] = []
    for position, value in enumerate(series):
        if _is_materialized_null(value):
            null_positions.append(position)
            canonical_values.append(None)
            continue
        canonical = _canonical_materialized_scalar(value)
        canonical_values.append(canonical)
        key = _materialized_bucket_key(canonical)
        bucket = buckets.get(key)
        if bucket is None:
            bucket = _MaterializedBucket(key=key, values=[], positions=[])
            buckets[key] = bucket
        bucket.values.append(canonical)
        bucket.positions.append(position)

    child_count = 0
    for bucket in buckets.values():
        if bucket.key[:1] == ("decimal",):
            child_count += len(_decimal_bucket_groups(bucket))
        elif bucket.key[:1] in {("timestamp",), ("duration",)}:
            child_count += _materialized_temporal_child_count(bucket)
        else:
            child_count += 1
        if child_count > 127:
            raise ValueError("mixed materialized column exceeds Arrow union type limit")
    return _MaterializedConversionPlan(
        buckets=tuple(buckets.values()),
        null_positions=tuple(null_positions),
        canonical_values=tuple(canonical_values),
        child_count=child_count,
    )


def _materialized_bucket_key(value: object) -> tuple[object, ...]:
    """Classify exact supported scalars without singleton Arrow probes."""
    if isinstance(value, PandasTemporalPayload):
        if value.family == "timestamp":
            return (
                "timestamp",
                _materialized_timezone_semantic_key(value.timezone),
            )
        return ("duration",)
    value_type = type(value)
    if value_type is datetime:
        timezone_value = cast(datetime, value).tzinfo
        return ("timestamp", _materialized_timezone_semantic_key(timezone_value))
    if value_type is pd.Timestamp:
        timezone_value = cast(pd.Timestamp, value).tzinfo
        return ("timestamp", _materialized_timezone_semantic_key(timezone_value))
    if value_type is timedelta or value_type is pd.Timedelta:
        return ("duration",)
    if value_type is time:
        return ("time",)
    if value_type is date:
        return ("date",)
    if value_type is Decimal:
        return ("decimal",)
    if value_type in {bool, int, float, str, bytes}:
        return (value_type.__name__,)
    return ("other", value_type)


def _canonical_materialized_scalar(value: object) -> object:
    """Project supported subclasses through inert base-class descriptors."""
    temporal = pandas_temporal_payload(value)
    if temporal is not None:
        return temporal
    value_type = type(value)
    if value_type in {
        datetime,
        pd.Timestamp,
        timedelta,
        pd.Timedelta,
        time,
        date,
        Decimal,
        bool,
        int,
        float,
        str,
        bytes,
    }:
        return value
    mro = type.__getattribute__(value_type, "__mro__")
    if str in mro:
        return str.__str__(cast(str, value))
    if bytes in mro:
        return bytes.__bytes__(cast(bytes, value))
    if int in mro:
        return int.__int__(cast(int, value))
    if float in mro:
        return float.__float__(cast(float, value))
    if Decimal in mro:
        return Decimal(Decimal.__str__(cast(Decimal, value)))
    if datetime in mro:
        datetime_value = cast(datetime, value)
        return datetime(
            datetime.year.__get__(datetime_value, datetime),
            datetime.month.__get__(datetime_value, datetime),
            datetime.day.__get__(datetime_value, datetime),
            datetime.hour.__get__(datetime_value, datetime),
            datetime.minute.__get__(datetime_value, datetime),
            datetime.second.__get__(datetime_value, datetime),
            datetime.microsecond.__get__(datetime_value, datetime),
            tzinfo=datetime.tzinfo.__get__(datetime_value, datetime),
            fold=datetime.fold.__get__(datetime_value, datetime),
        )
    if timedelta in mro:
        duration = cast(timedelta, value)
        return timedelta(
            days=timedelta.days.__get__(duration, timedelta),
            seconds=timedelta.seconds.__get__(duration, timedelta),
            microseconds=timedelta.microseconds.__get__(duration, timedelta),
        )
    if time in mro:
        time_value = cast(time, value)
        return time(
            time.hour.__get__(time_value, time),
            time.minute.__get__(time_value, time),
            time.second.__get__(time_value, time),
            time.microsecond.__get__(time_value, time),
            tzinfo=time.tzinfo.__get__(time_value, time),
            fold=time.fold.__get__(time_value, time),
        )
    if date in mro:
        date_value = cast(date, value)
        return date(
            date.year.__get__(date_value, date),
            date.month.__get__(date_value, date),
            date.day.__get__(date_value, date),
        )
    return value


def _coalesce_materialized_children(
    converted: list[tuple[list[int], pa.Array, int]],
) -> list[tuple[list[int], pa.Array, int]]:
    """Merge only fragments whose exact Arrow child types are identical."""
    coalesced: list[tuple[list[int], pa.Array, int]] = []
    indexes: dict[pa.DataType, int] = {}
    for positions, child, encounter in converted:
        existing_index = indexes.get(child.type)
        if existing_index is None:
            indexes[child.type] = len(coalesced)
            coalesced.append((list(positions), child, encounter))
            continue
        old_positions, old_child, old_encounter = coalesced[existing_index]
        coalesced[existing_index] = (
            [*old_positions, *positions],
            pa.concat_arrays([old_child, child]),
            old_encounter,
        )
    return coalesced


def _materialized_timezone_semantic_key(value: object) -> tuple[object, ...]:
    """Return hook-free timezone semantics, with identity for custom zones."""
    if value is None:
        return ("naive",)
    projection, safe = _timezone_label_projection(value)
    if safe and type(projection) is tuple:
        if projection[:1] == ("timezone",):
            offset = cast(tuple[object, ...], projection[1])
            return ("fixed", *offset[1:])
        if projection[:1] == ("zoneinfo",):
            return projection
    return ("identity", id(value))


def _materialized_arrow_timezone(value: object) -> str | None:
    """Return Arrow's stable timezone text for a trusted timezone."""
    if value is None:
        return None
    projection, safe = _timezone_label_projection(value)
    if not safe or type(projection) is not tuple:
        return None
    if projection[:1] == ("zoneinfo",):
        return cast(str, projection[1])
    if projection[:1] != ("timezone",):
        return None
    offset = cast(tuple[object, ...], projection[1])
    days, seconds, microseconds = cast(tuple[int, int, int], offset[1:])
    total_microseconds = ((days * 86_400 + seconds) * 1_000_000) + microseconds
    if total_microseconds == 0:
        return "UTC"
    minute_microseconds = 60 * 1_000_000
    if total_microseconds % minute_microseconds:
        return None
    total_minutes = total_microseconds // minute_microseconds
    sign = "+" if total_minutes >= 0 else "-"
    hours, minutes = divmod(abs(total_minutes), 60)
    return f"{sign}{hours:02d}:{minutes:02d}"


def _convert_materialized_bucket(
    bucket: _MaterializedBucket,
) -> list[tuple[list[int], pa.Array]]:
    family = cast(str, bucket.key[0])
    if family in {"timestamp", "duration"}:
        return _convert_materialized_temporal_bucket(bucket)
    if family == "decimal":
        return _convert_decimal_bucket(bucket)
    return [(bucket.positions, pa.array(bucket.values, from_pandas=True))]


def _convert_materialized_temporal_bucket(
    bucket: _MaterializedBucket,
) -> list[tuple[list[int], pa.Array]]:
    planned = _materialized_temporal_groups(bucket)
    if planned is None:
        return [(bucket.positions, pa.array(bucket.values, from_pandas=True))]
    return [
        (
            positions,
            arrow_temporal_array(
                cast("list[PandasTemporalPayload | None]", payloads),
                target,
            ),
        )
        for positions, payloads, target in planned
    ]


def _materialized_temporal_child_count(bucket: _MaterializedBucket) -> int:
    planned = _materialized_temporal_groups(bucket)
    return 1 if planned is None else len(planned)


def _materialized_temporal_groups(
    bucket: _MaterializedBucket,
) -> (
    list[
        tuple[
            list[int],
            list[PandasTemporalPayload],
            pa.DataType,
        ]
    ]
    | None
):
    payloads: list[PandasTemporalPayload] = []
    for value in bucket.values:
        payload = _materialized_temporal_payload(value)
        if payload is None:
            return None
        payloads.append(payload)
    common_type = common_temporal_arrow_type(payloads)
    if common_type is not None:
        return [(list(bucket.positions), payloads, common_type)]
    grouped: dict[str, tuple[list[int], list[PandasTemporalPayload]]] = {}
    for position, payload in zip(bucket.positions, payloads, strict=True):
        positions, values = grouped.setdefault(payload.unit, ([], []))
        positions.append(position)
        values.append(payload)
    planned = []
    for unit in ("ns", "us", "ms", "s"):
        group = grouped.get(unit)
        if group is None:
            continue
        positions, values = group
        target = common_temporal_arrow_type(values)
        if target is None:
            raise ValueError("pandas temporal values have no lossless Arrow representation")
        planned.append((positions, values, target))
    return planned


def _materialized_temporal_payload(
    value: object,
) -> PandasTemporalPayload | None:
    return temporal_payload(value)


def _convert_decimal_bucket(
    bucket: _MaterializedBucket,
) -> list[tuple[list[int], pa.Array]]:
    """Convert the prevalidated deterministic decimal subgroups."""
    converted: list[tuple[list[int], pa.Array]] = []
    for integer_digits, scale, values, positions in _decimal_bucket_groups(bucket):
        precision = integer_digits + scale
        target = (
            pa.decimal128(precision, scale) if precision <= 38 else pa.decimal256(precision, scale)
        )
        converted.append((positions, pa.array(values, type=target, from_pandas=True)))
    return converted


def _decimal_bucket_groups(
    bucket: _MaterializedBucket,
) -> list[tuple[int, int, list[object], list[int]]]:
    """Plan compatible Decimal child types without invoking Arrow."""
    shapes = [_decimal_required_shape(cast(Decimal, value)) for value in bucket.values]
    if any(precision > 76 for precision, _scale in shapes):
        raise ValueError("decimal value exceeds Arrow decimal256 precision")
    groups: list[tuple[int, int, list[object], list[int]]] = []
    for value, position, (precision, scale) in zip(
        bucket.values,
        bucket.positions,
        shapes,
        strict=True,
    ):
        integer_digits = precision - scale
        selected = None
        for index, (max_integer, max_scale, _values, _positions) in enumerate(groups):
            if max(max_integer, integer_digits) + max(max_scale, scale) <= 76:
                selected = index
                break
        if selected is None:
            groups.append((integer_digits, scale, [value], [position]))
            continue
        max_integer, max_scale, values, positions = groups[selected]
        values.append(value)
        positions.append(position)
        groups[selected] = (
            max(max_integer, integer_digits),
            max(max_scale, scale),
            values,
            positions,
        )
    return groups


def _decimal_required_shape(value: Decimal) -> tuple[int, int]:
    _sign, digits, exponent = Decimal.as_tuple(value)
    if type(exponent) is not int:
        raise ValueError("non-finite Decimal values are not supported")
    if exponent >= 0:
        return max(1, len(digits) + exponent), 0
    scale = -exponent
    return max(len(digits), scale), scale


def _materialized_scalar_types(
    values: list[object],
) -> tuple[list[pa.DataType | None], set[pa.DataType]]:
    """Retain the former private inspection seam over family-vector inference."""
    inferred = _materialized_union_array(pd.Series(values, dtype=object))
    if not pa.types.is_union(inferred.type):
        if pa.types.is_null(inferred.type):
            return [None] * len(values), set()
        return (
            [None if _is_materialized_null(value) else inferred.type for value in values],
            {inferred.type},
        )
    union = cast(pa.UnionArray, inferred)
    child_types = {field.type for field in union.type}
    scalar_types = [
        (
            None
            if _is_materialized_null(value)
            else union.type.field(int(union.type_codes[position].as_py())).type
        )
        for position, value in enumerate(values)
    ]
    return scalar_types, child_types


def _is_materialized_null(value: object) -> bool:
    if value is None or value is pd.NA or value is pd.NaT:
        return True
    return type(value) is float and pd.isna(value)


class _ActiveOperationError(RuntimeError):
    """Identify a rejected concurrent or re-entrant workbook operation."""


def _operation_error(message: str) -> _ActiveOperationError:
    error = _ActiveOperationError(message)
    _mark_fallback_blocked(
        error,
        _FallbackBlockReason.CONFIGURATION,
    )
    return error


def _contains_active_operation_error(error: BaseException) -> bool:
    candidates, _complete = _bounded_exception_graph(error)
    return any(isinstance(candidate, _ActiveOperationError) for candidate in candidates)


def _sheet_error(name: str, error: BaseException) -> SheetError:
    """Project one ordinary per-sheet failure without retaining its traceback."""
    context = getattr(error, "context", {})
    if not isinstance(context, dict):
        context = {}
    return SheetError(
        sheet_name=name,
        error_type=type(error).__name__,
        message=str(error),
        context=context.copy(),
    )


def _failure_sheet_result(name: str, error: BaseException) -> SheetResult:
    """Return the public frozen failure representation for one sheet."""
    return SheetResult(name=name, error=_sheet_error(name, error))


def _release_operation_and_confirm(
    workbook: "MessyWorkbook",
    token: object,
) -> bool:
    release_operation: Callable[[object], Any] = workbook._end_operation
    return not release_operation(token)


_OwnedResourceT = TypeVar("_OwnedResourceT")
_PreparedResourceT = TypeVar("_PreparedResourceT")
_BoundStreamT = TypeVar("_BoundStreamT")


class _CloseOnceOwner(Generic[_OwnedResourceT]):
    """Proxy one closeable reader through a shared close-once boundary."""

    def __init__(self) -> None:
        self._resource: _OwnedResourceT | None = None
        self._attached = False
        self._closed = False
        self._process_close_pending = False

    def attach(self, resource: _OwnedResourceT) -> "_CloseOnceOwner[_OwnedResourceT]":
        """Fill a previously reserved ownership slot exactly once."""
        if self._closed or self._attached:
            raise RuntimeError("Stream ownership slot is no longer available")
        try:
            self._resource = resource
            self._attached = True
        except BaseException:
            if not self._attached and self._resource is resource:
                self._resource = None
            raise
        return self

    def __getattr__(self, name: str) -> Any:
        resource = self._resource
        if resource is None:
            raise AttributeError(name)
        return getattr(resource, name)

    def _owns(self, resource: object) -> bool:
        """Check an attachment by identity without invoking resource hooks."""
        return self._resource is resource

    def replace(self, resource: _OwnedResourceT, replacement: _OwnedResourceT) -> None:
        """Promote an already-owned resource to its final wrapper."""
        if self._closed or not self._attached or self._resource is not resource:
            raise RuntimeError("Stream ownership slot does not own the expected resource")
        self._resource = replacement

    def close(self) -> None:
        if self._closed and self._resource is None:
            return
        self._closed = True
        resource = self._resource
        if resource is None:
            return
        try:
            self._resource = _close_if_present(resource)
        except BaseException as error:
            self._process_close_pending = _contains_process_failure(error)
            if not self._process_close_pending:
                self._resource = None
            raise
        self._process_close_pending = False


class _SampleResourceOwner:
    """Keep every sample resource reachable until dependency-ordered cleanup."""

    def __init__(self, source: SourceHandle) -> None:
        self.rows: Iterator[tuple[Any, ...]] | None = None
        self.workbook: Any | None = None
        self.source_context: _RetryableSourceContext | None = _RetryableSourceContext(source)

    def enter_source(self, context: Any) -> BackendSource:
        owner = self.source_context
        if owner is None:
            raise RuntimeError("sample source owner is closed")
        return cast(BackendSource, _enter_prearmed_context(owner, context))

    def close(self) -> None:
        cleanups: list[tuple[str, Callable[[], object]]] = []
        if self.rows is not None:
            cleanups.append(("sample row iterator cleanup", self._close_rows))
        if self.workbook is not None:
            cleanups.append(("sample workbook cleanup", self._close_workbook))
        if self.source_context is not None:
            cleanups.append(("sample source borrow cleanup", self._close_source_context))
        _run_cleanups(cleanups)

    def _close_rows(self) -> None:
        rows = self.rows
        if rows is None:
            return
        try:
            self.rows = _close_if_present(rows)
        except BaseException as error:
            if not _contains_process_failure(error):
                self.rows = None
            raise

    def _close_workbook(self) -> None:
        workbook = self.workbook
        if workbook is None:
            return
        try:
            self.workbook = _close_if_present(workbook)
        except BaseException as error:
            if not _contains_process_failure(error):
                self.workbook = None
            raise

    def _close_source_context(self) -> None:
        source_context = self.source_context
        if source_context is None:
            return
        try:
            self.source_context = _close_if_present(source_context)
        except BaseException as error:
            if not _contains_process_failure(error):
                self.source_context = None
            raise


class _MaterializedOperationLease:
    """Own one exact non-streaming operation token through retryable release."""

    def __init__(self, workbook: "MessyWorkbook", token: object) -> None:
        self._workbook: MessyWorkbook | None = workbook
        self._token = token
        self._started = False
        self._release_pending = True
        self._released = False
        self._body_active = False

    def __enter__(self) -> "_MaterializedOperationLease":
        workbook = self._workbook
        if workbook is None or self._released:
            raise _operation_error("Materialized operation lease is no longer available")
        try:
            workbook._register_materialized_lease(self)
            workbook._begin_operation(self._token)
            self._started = True
        except BaseException as error:
            self._started = workbook._active_operation_token is self._token
            self._release_pending = True
            self._release(
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise
        return self

    def _body_started(self) -> None:
        workbook = self._workbook
        if (
            workbook is None
            or self._released
            or not self._started
            or workbook._active_materialized_lease is not self
            or workbook._active_operation_token is not self._token
            or self._body_active
        ):
            raise _operation_error("Materialized operation lease is not active")
        self._body_active = True

    def _body_complete(self) -> None:
        self._body_active = False

    def _release_once(self) -> None:
        if self._released:
            return
        workbook = self._workbook
        if workbook is None:
            self._released = True
            return
        if workbook._active_operation_token is not self._token:
            self._finish_release(workbook)
            return
        workbook._end_operation(self._token)
        if workbook._active_operation_token is not self._token:
            self._finish_release(workbook)

    def _finish_release(self, workbook: "MessyWorkbook") -> None:
        if workbook._active_materialized_lease is self:
            workbook._active_materialized_lease = None
        self._release_pending = False
        self._released = True
        self._workbook = None

    def _release(
        self,
        *,
        primary_error: BaseException | None = None,
        primary_traceback: TracebackType | None = None,
    ) -> None:
        if self._released:
            return
        _run_cleanups(
            [
                ("materialized operation release", self._release_once),
                ("materialized operation release retry", self._release_once),
            ],
            primary_error=primary_error,
            primary_traceback=primary_traceback,
        )

    def _release_from_owner(self) -> None:
        self._release()

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        self._body_active = False
        del exc_type
        self._release(
            primary_error=exc_value,
            primary_traceback=traceback,
        )


def _finalize_stream_operation_lease(
    workbook_ref: weakref.ReferenceType["MessyWorkbook"],
    token: object,
) -> None:
    """Release a reservation whose lease never crossed its return boundary."""
    workbook = workbook_ref()
    if workbook is not None:
        workbook._end_operation(token)


class _StreamOperationLease:
    """Workbook-visible construction placeholder and child-stream owner."""

    def __init__(self, workbook: "MessyWorkbook", token: object) -> None:
        self._workbook: MessyWorkbook | None = workbook
        self._token = token
        self._abandonment_finalizer = weakref.finalize(
            self,
            _finalize_stream_operation_lease,
            weakref.ref(workbook),
            token,
        )
        self._partial: _CloseOnceOwner[Any] | None = None
        self._stream_ref: weakref.ReferenceType[Any] | None = None
        self._bound = False
        self._operation_released = False
        self._released = False

    def reserve(self) -> _CloseOnceOwner[Any]:
        """Allocate and record an empty ownership slot before opening a reader."""
        if self._released or self._bound or self._partial is not None:
            raise RuntimeError("Stream operation lease already owns a resource")
        owner: _CloseOnceOwner[Any] = _CloseOnceOwner()
        self._partial = owner
        workbook = self._workbook
        if workbook is None:
            raise _operation_error("MessyWorkbook is closed")
        workbook._register_stream(self._token, self)
        return owner

    def own(self, partial: _OwnedResourceT) -> _CloseOnceOwner[_OwnedResourceT]:
        """Compatibility helper for tests and already-opened resources."""
        try:
            owner = self.reserve()
        except BaseException as error:
            _run_cleanups(
                [("unrecorded stream reader cleanup", lambda: _close_if_present(partial))],
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise
        return cast(_CloseOnceOwner[_OwnedResourceT], owner.attach(partial))

    def adopt(
        self,
        factory: Callable[[_CloseOnceOwner[_OwnedResourceT]], _PreparedResourceT],
        resource: Callable[[_PreparedResourceT], _OwnedResourceT],
    ) -> tuple[_CloseOnceOwner[_OwnedResourceT], _PreparedResourceT]:
        """Reserve before construction and own every factory-return boundary."""
        reserved_owner = self.reserve()
        owner = cast(_CloseOnceOwner[_OwnedResourceT], reserved_owner)
        prepared: _PreparedResourceT | None = None
        produced: _OwnedResourceT | None = None
        try:
            prepared = factory(owner)
            produced = resource(prepared)
            if not owner._owns(produced):
                owner.attach(produced)
        except BaseException as error:
            cleanups: list[tuple[str, Callable[[], object]]] = []
            if produced is not None:
                rollback_target = owner if owner._owns(produced) else produced
                cleanups.append(
                    (
                        "unadopted stream reader cleanup",
                        lambda: _close_if_present(rollback_target),
                    )
                )
            elif prepared is not None:
                cleanups.append(
                    (
                        "unextracted stream reader cleanup",
                        lambda: _close_if_present(resource(prepared)),
                    )
                )
            _run_cleanups(
                cleanups,
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise
        assert prepared is not None
        return owner, prepared

    def bind(self, stream: _BoundStreamT) -> _BoundStreamT:
        """Register a stream while retaining lease-owned reader cleanup."""
        if self._released or self._bound:
            raise RuntimeError("Stream operation lease is no longer available")
        stream_ref = weakref.ref(stream)
        self._stream_ref = stream_ref
        workbook = self._workbook
        if workbook is None:
            raise _operation_error("MessyWorkbook is closed")
        self._bound = workbook._replace_stream(self._token, self, stream)
        return stream

    def invalidate_from_owner(self) -> None:
        """Let parent close retry an in-progress construction owner."""
        self.release()

    def release(self) -> None:
        self._release(skip_pending_process_close=False)

    def release_after_source_cleanup(self) -> bool:
        """Release a stream without immediately retrying its failed source close."""
        self._release(skip_pending_process_close=True)
        return self._released

    def _release(self, *, skip_pending_process_close: bool) -> None:
        """Release the matching reservation at most once."""
        if self._released:
            return
        cleanups: list[tuple[str, Any]] = []
        if self._partial is not None and not (
            skip_pending_process_close and self._partial._process_close_pending
        ):

            def close_partial() -> None:
                partial = self._partial
                if partial is None:
                    return
                try:
                    self._partial = _close_if_present(partial)
                except BaseException as error:
                    if not _contains_process_failure(error):
                        self._partial = None
                    raise

            cleanups.append(("owned stream reader cleanup", close_partial))
        if not self._operation_released:

            def release_operation() -> None:
                if self._partial is not None:
                    return
                active_workbook = self._workbook
                if active_workbook is not None:
                    self._operation_released = _release_operation_and_confirm(
                        active_workbook,
                        self._token,
                    )
                else:
                    self._operation_released = True

            cleanups.append(("stream reservation release", release_operation))
        try:
            _run_cleanups(cleanups)
        finally:
            if self._partial is None and self._operation_released:
                self._released = True
                self._stream_ref = None
                self._workbook = None
                self._abandonment_finalizer.detach()

    def __enter__(self) -> "_StreamOperationLease":
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        del exc_type
        if self._bound and exc_value is None:
            return

        stream = self._stream_ref() if self._stream_ref is not None else None
        cleanups: list[tuple[str, Any]] = []
        if stream is not None:
            cleanups.append(
                ("partially registered stream cleanup", lambda: _close_if_present(stream))
            )
        cleanups.append(("stream reservation release", self.release))
        _run_cleanups(
            cleanups,
            primary_error=exc_value,
            primary_traceback=traceback,
        )


class MessyWorkbook:
    """Main entry point for parsing Excel files."""

    def __init__(
        self,
        file_path_or_buffer: str | Path | BinaryIO,
        sheet_config: SheetConfig | None = None,
        formula_config: FormulaConfig | None = None,
        filename: str | None = None,
        registry: HandlerRegistry | None = None,
    ):
        """Open an Excel file for parsing.

        Args:
            file_path_or_buffer: Path to file, or file-like object (BytesIO, etc.)
            sheet_config: Configuration for parsing sheets
            formula_config: Configuration for formula evaluation
            filename: Optional filename hint when using file-like objects (for format detection)
            registry: Optional format-handler registry for custom parsing behavior
        """
        self._closed = False
        self._active_operation_token: object | None = None
        self._active_stream_slot: Any | weakref.ReferenceType[Any] | None = None
        self._active_stream_cleanup: Any | None = None
        self._active_stream_lease_ref: weakref.ReferenceType[_StreamOperationLease] | None = None
        self._active_stream: Any | None = None
        self._active_materialized_lease: _MaterializedOperationLease | None = None
        self._sheets: dict[str, MessySheet] = {}
        self._formula_loaded = False
        self._wb: openpyxl.Workbook | None = None
        self._cached_wb: openpyxl.Workbook | None = None
        self._wb_source: BinaryIO | None = None
        self._cached_wb_source: BinaryIO | None = None
        self._fastexcel_session: FastexcelSession | None = None
        self._manifest_reader: ManifestReader | None = None
        self._stream_structure_sampler: StructureSampler | None = None
        self._sample_owner: _SampleResourceOwner | None = None
        self._source_handle_close_pending = False
        self._parse_metrics = ParseMetrics()

        self._sheet_config = sheet_config or SheetConfig()
        self._formula_config = formula_config or FormulaConfig()

        self._registry = registry if registry is not None else HandlerRegistry()
        self._analyzer = StructureAnalyzer(get_structure_cache())
        self._formula_engine = FormulaEngine(self._formula_config)

        source_handle: SourceHandle | None = None
        try:
            self._source_handle_close_pending = True
            prepare_source = SourceHandle.__dict__.get("prepare")
            if prepare_source is not None:
                source_handle = SourceHandle.prepare(
                    file_path_or_buffer,
                    filename=filename,
                )
            else:
                source_handle = SourceHandle(file_path_or_buffer, filename=filename)
            self._source_handle = source_handle
            self._source_handle_close_pending = True
            start_source = getattr(source_handle, "start", None)
            if callable(start_source):
                start_source()
        except BaseException as error:
            acquired_handle = getattr(self, "_source_handle", source_handle)
            if acquired_handle is not None:
                self._source_handle = acquired_handle
                self._source_handle_close_pending = True
                _run_cleanups(
                    [
                        ("source handle construction rollback", self._close_source_handle),
                        (
                            "source handle construction rollback retry",
                            self._close_source_handle,
                        ),
                    ],
                    primary_error=error,
                    primary_traceback=_exception_traceback(error),
                )
            else:
                self._source_handle_close_pending = False
            if _contains_process_failure(error):
                raise
            file_desc = describe_source(file_path_or_buffer, filename)
            raise FormatError(
                f"Cannot read from file object: {error}",
                file_path=file_desc,
            ) from error
        try:
            self._initialize_source()
        except BaseException as error:
            _run_cleanups(
                [
                    ("workbook construction cleanup", self._close),
                    ("workbook construction cleanup retry", self._close),
                ],
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise

    @property
    def _active_stream(self) -> Any | None:
        """Resolve the active child without keeping public streams alive."""
        slot = self.__dict__.get("_active_stream_slot")
        if isinstance(slot, weakref.ReferenceType):
            return slot()
        return slot

    @_active_stream.setter
    def _active_stream(self, stream: Any | None) -> None:
        cleanup_state = getattr(stream, "_cleanup_state", None)
        if stream is not None and cleanup_state is not None:
            self.__dict__["_active_stream_slot"] = weakref.ref(stream)
            self.__dict__["_active_stream_cleanup"] = cleanup_state
            return
        self.__dict__["_active_stream_slot"] = stream
        self.__dict__["_active_stream_cleanup"] = None

    def _recover_abandoned_stream_operation(self) -> None:
        """Finish cleanup for a stream or lease lost at a return boundary."""
        token = getattr(self, "_active_operation_token", None)
        if token is None or self._active_stream is not None:
            return
        cleanup_state = self.__dict__.get("_active_stream_cleanup")
        if cleanup_state is not None:
            cleanup_state.close()
            if getattr(self, "_active_operation_token", None) is not token:
                return
        lease_ref = getattr(self, "_active_stream_lease_ref", None)
        if lease_ref is not None and lease_ref() is None:
            self._end_operation(token)

    def _begin_operation(self, token: object | None = None) -> object:
        """Reserve the workbook for one parse or child stream."""
        if getattr(self, "_closed", False):
            raise _operation_error("MessyWorkbook is closed")
        self._recover_abandoned_stream_operation()
        retained_lease = getattr(self, "_active_materialized_lease", None)
        if (
            retained_lease is not None
            and retained_lease._token is not token
            and retained_lease._release_pending
            and not retained_lease._body_active
        ):
            retained_lease._release()
        if getattr(self, "_active_operation_token", None) is not None:
            raise _operation_error("MessyWorkbook already has an active parse or stream")
        if token is None:
            token = object()
        self._active_operation_token = token
        return token

    def _end_operation(self, token: object) -> None:
        """Release only the exact current token; stale callbacks are harmless."""
        if getattr(self, "_active_operation_token", None) is not token:
            return
        self._active_stream = None
        self._active_stream_lease_ref = None
        self._active_operation_token = None

    def _register_materialized_lease(
        self,
        lease: _MaterializedOperationLease,
    ) -> None:
        """Retain a lease before its exact operation token can be committed."""
        current = getattr(self, "_active_materialized_lease", None)
        if current is lease:
            return
        if current is not None:
            if current._release_pending and not current._body_active:
                current._release()
            if getattr(self, "_active_materialized_lease", None) is not None:
                raise _operation_error("MessyWorkbook already has an active parse or stream")
        self._active_materialized_lease = lease

    def _register_stream(self, token: object, stream: Any) -> None:
        """Register a child only while its exact operation token is current."""
        if getattr(self, "_closed", False):
            raise _operation_error("MessyWorkbook is closed")
        if getattr(self, "_active_operation_token", None) is not token:
            raise _operation_error("MessyWorkbook stream reservation is no longer active")
        if getattr(self, "_active_stream", None) is not None:
            raise _operation_error("MessyWorkbook already has an active parse or stream")
        self._active_stream = stream

    def _replace_stream(self, token: object, expected: object, stream: Any) -> bool:
        """Atomically promote the exact construction placeholder to its stream."""
        if getattr(self, "_closed", False):
            raise _operation_error("MessyWorkbook is closed")
        if getattr(self, "_active_operation_token", None) is not token:
            raise _operation_error("MessyWorkbook stream reservation is no longer active")
        active_stream = getattr(self, "_active_stream", None)
        if active_stream is not None and active_stream is not expected:
            raise _operation_error("MessyWorkbook stream construction owner changed")
        self._active_stream = stream
        return True

    def _stream_operation(self) -> _StreamOperationLease:
        """Reserve a future child stream with construction-failure cleanup."""
        token = object()
        lease = _StreamOperationLease(self, token)
        try:
            self._begin_operation(token)
            self._active_stream_lease_ref = weakref.ref(lease)
            return lease
        except BaseException as error:
            _run_cleanups(
                [("stream operation commit rollback", lambda: self._end_operation(token))],
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise

    def _materialized_operation(self) -> _MaterializedOperationLease:
        """Create a retryable lease around one preallocated exact token."""
        return _MaterializedOperationLease(self, object())

    @contextmanager
    def _registry_source(self) -> Iterator[SourceHandle | BackendSource]:
        """Adapt the internal handle for legacy registry subclasses."""
        accepts_handle = bool(type(self._registry).__dict__.get("_accepts_source_handle", False))
        if accepts_handle:
            yield self._source_handle
            return
        with self._source_handle.open_legacy() as source:
            yield source

    def _initialize_source(self) -> None:
        """Detect, inspect, and validate the source without taking ownership."""

        if self._source_handle.path is not None and not self._source_handle.path.exists():
            raise FileError(
                f"File not found: {self._source_handle.path}",
                file_path=str(self._source_handle.path),
            )

        with self._registry_source() as source:
            self._format_info = self._registry.detect_format(
                source,
                filename=self._source_handle.filename_hint,
            )

        if self._format_info.format_type == "unknown":
            file_desc = self._source_handle.description
            raise FormatError(
                f"Unknown file format: {file_desc}",
                file_path=str(file_desc),
            )

        if self._format_info.format_type == "xlsb":
            file_desc = self._source_handle.description
            raise FormatError(
                "XLSB (Excel Binary) format is not supported. "
                "Please convert the file to XLSX format.",
                file_path=str(file_desc),
                detected_format="xlsb",
            )

        # Validate extension matches detected format for Excel files
        # This catches files with .xlsx extension but different content
        if self._source_handle.path is not None:
            file_ext = self._source_handle.path.suffix.lower()
            excel_extensions = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
            if file_ext in excel_extensions and self._format_info.format_type not in (
                "xlsx",
                "xlsm",
                "xls",
                "xltx",
                "xltm",
            ):
                raise FormatError(
                    f"File extension {file_ext} suggests Excel format, but content is {self._format_info.format_type}",
                    file_path=str(self._source_handle.path),
                    detected_format=self._format_info.format_type,
                )

        # Get sheet names and validate file is readable
        with self._registry_source() as source:
            self._sheet_names = self._registry.get_sheet_names(
                source,
            )

        # Validate that the file is actually readable (not just format-detected)
        # This catches corrupted files that pass format detection but can't be opened
        if self._format_info.format_type in ("xlsx", "xlsm", "xltx", "xltm", "xls"):
            with self._registry_source() as source:
                is_valid, error = self._registry.validate(
                    source,
                    self._format_info.format_type,
                )
            if not is_valid:
                file_desc = self._source_handle.description
                raise FormatError(
                    f"File appears corrupted or invalid: {error}",
                    file_path=str(file_desc),
                    detected_format=self._format_info.format_type,
                )

    @property
    def file_path(self) -> Path | None:
        """Path to the Excel file, or None if reading from buffer."""
        return cast(Path | None, self._source_handle.path)

    @property
    def source(self) -> Path | BinaryIO:
        """The source file path or buffer."""
        return cast(Path | BinaryIO, self._source_handle.original)

    @property
    def sheet_names(self) -> list[str]:
        """List of sheet names in the workbook."""
        return cast(list[str], self._sheet_names.copy())

    @property
    def format_type(self) -> str:
        """Detected file format (xlsx, xls, csv, etc.)."""
        return cast(str, self._format_info.format_type)

    @property
    def parse_metrics(self) -> ParseMetrics:
        """Return cumulative parser-work counters owned by this workbook."""
        metrics = getattr(self, "_parse_metrics", None)
        if metrics is None:
            metrics = ParseMetrics()
            self._parse_metrics = metrics
        return metrics

    def get_sheet(self, name: str | None = None) -> MessySheet:
        """Get a sheet by name."""
        if name is None:
            name = self._sheet_names[0]

        if name not in self._sheet_names:
            file_desc = self._source_handle.description
            raise FormatError(
                f"Sheet '{name}' not found",
                file_path=str(file_desc),
            )

        if name not in self._sheets:
            self._sheets[name] = MessySheet(self, name)

        return self._sheets[name]

    def to_dataframe(
        self,
        sheet: str | None = None,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Convert a sheet to a pandas DataFrame."""
        warn_legacy("MessyWorkbook.to_dataframe")
        return self._to_dataframe_compat(sheet, config)

    def _to_dataframe_compat(
        self,
        sheet: str | None = None,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        sheet_name = sheet or self._sheet_names[0]
        return self._parse_sheet(sheet_name, config)

    def to_arrow(
        self,
        sheet: str | None = None,
        config: SheetConfig | None = None,
    ) -> pa.Table:
        """Materialize one sheet under the existing global normalization contract."""
        self._validate_public_config(config)
        sheet_name = self._resolve_public_sheet(sheet)
        lease = self._materialized_operation()
        with lease:
            try:
                lease._body_started()
                frame = self._parse_sheet_unreserved(sheet_name, config)
                return _arrow_table_from_dataframe(frame)
            finally:
                lease._body_complete()

    def iter_batches(
        self,
        sheet: str | None = None,
        batch_size: int = 65_536,
        config: SheetConfig | None = None,
    ) -> BatchStream:
        """Return a schema-stable, one-shot stream of Arrow record batches."""
        self._validate_public_batch_size(batch_size)
        with self._stream_operation() as lease:
            owned_reader, _prepared = lease.adopt(
                lambda construction_owner: self._prepare_streaming_operation(
                    sheet,
                    batch_size,
                    config,
                    construction_owner=construction_owner,
                ),
                lambda result: result.reader,
            )
            stream = BatchStream(
                reader_batches(owned_reader),
                owned_reader.schema,
                lease.release_after_source_cleanup,
            )
            return lease.bind(stream)

    def iter_dataframe_chunks(
        self,
        sheet: str | None = None,
        batch_size: int = 65_536,
        config: SheetConfig | None = None,
    ) -> DataFrameChunkStream:
        """Return Arrow-backed pandas chunks with one global RangeIndex."""
        self._validate_public_batch_size(batch_size)
        with self._stream_operation() as lease:
            owned_reader, prepared = lease.adopt(
                lambda construction_owner: self._prepare_streaming_operation(
                    sheet,
                    batch_size,
                    config,
                    construction_owner=construction_owner,
                ),
                lambda result: result.reader,
            )
            display_names = prepared.display_names
            offset = 0

            def frames() -> Iterator[pd.DataFrame]:
                nonlocal offset
                for batch in reader_batches(owned_reader):
                    frame = _dataframe_from_record_batch(batch, display_names)
                    frame.index = pd.RangeIndex(offset, offset + len(frame))
                    offset += len(frame)
                    yield frame

            stream = DataFrameChunkStream(frames(), lease.release_after_source_cleanup)
            return lease.bind(stream)

    def iter_sheets(
        self,
        config: SheetConfig | None = None,
    ) -> SheetStream:
        """Return one ordered success or ordinary failure result per sheet."""
        self._validate_public_config(config)
        with self._stream_operation() as lease:
            pending = deque(self._plan_workbook_sheets(config))

            def results() -> Iterator[SheetResult]:
                result: SheetResult | None = None
                frame: pd.DataFrame | None = None
                current: PlannedSheet | None = None
                while pending:
                    # Release the prior yielded frame before opening the next
                    # sheet-local materializer.
                    result = None
                    frame = None
                    current = pending.popleft()
                    if current.state is PlannedSheetState.ERROR:
                        assert current.error is not None
                        result = _failure_sheet_result(current.name, current.error)
                    elif current.state is PlannedSheetState.READY:
                        assert current.parse_plan is not None
                        try:
                            frame = self._materialize_compiled_plan(
                                current.name,
                                current.parse_plan,
                            )
                        except BaseException as error:
                            if (
                                not isinstance(error, Exception)
                                or _contains_active_operation_error(error)
                                or _contains_process_failure(error)
                            ):
                                raise
                            result = _failure_sheet_result(current.name, error)
                        else:
                            result = SheetResult(name=current.name, dataframe=frame)
                    else:
                        current = None
                        continue
                    assert result is not None
                    current = None
                    yield result
                    result = None
                    frame = None

            stream = SheetStream(results(), lease.release_after_source_cleanup)
            return lease.bind(stream)

    def to_dataframes(
        self,
        config: SheetConfig | None = None,
        include_errors: bool = False,
    ) -> dict[str, pd.DataFrame] | tuple[dict[str, pd.DataFrame], list[SheetError]]:
        """Convert all sheets to DataFrames.

        Args:
            config: Optional sheet configuration.
            include_errors: If True, return a tuple of (results, errors) instead
                of just results. Each error contains structured information about
                which sheet failed and why.

        Returns:
            If include_errors is False (default): dict mapping sheet name to DataFrame.
            If include_errors is True: tuple of (results_dict, errors_list).
        """
        warn_legacy("MessyWorkbook.to_dataframes")
        return self._to_dataframes_compat(config, include_errors)

    def _to_dataframes_compat(
        self,
        config: SheetConfig | None = None,
        include_errors: bool = False,
    ) -> dict[str, pd.DataFrame] | tuple[dict[str, pd.DataFrame], list[SheetError]]:
        lease = self._materialized_operation()
        with lease:
            try:
                lease._body_started()
                result = {}
                errors: list[SheetError] = []
                for item in self._plan_workbook_sheets(config):
                    name = item.name
                    if item.state is PlannedSheetState.ERROR:
                        assert item.error is not None
                        if include_errors:
                            errors.append(_sheet_error(name, item.error))
                        continue
                    if item.state is not PlannedSheetState.READY:
                        continue
                    assert item.parse_plan is not None
                    try:
                        result[name] = self._materialize_compiled_plan(
                            name,
                            item.parse_plan,
                        )
                    except _ActiveOperationError:
                        raise
                    except Exception as e:
                        if _contains_active_operation_error(e) or _contains_process_failure(e):
                            raise
                        logger.warning("Failed to parse sheet %r, skipping", name, exc_info=True)
                        if include_errors:
                            errors.append(_sheet_error(name, e))
                if include_errors:
                    return result, errors
                return result
            finally:
                lease._body_complete()

    def get_structure(self, sheet: str | None = None) -> StructureInfo:
        """Get detected structure for a sheet."""
        sheet_name = sheet or self._sheet_names[0]
        return self._analyze_structure(sheet_name, self._sheet_config)

    def get_cell(
        self,
        sheet: str,
        row: int,
        col: int,
    ) -> CellValue:
        """Get a single cell value."""
        self._ensure_workbook()

        if self._wb is None:
            raise FileError("Workbook not loaded — call _ensure_workbook() first")
        ws = self._wb[sheet]
        cell = ws.cell(row, col)

        resolved_value = cell.value

        formula = None
        is_formula = False
        if hasattr(cell, "data_type") and cell.data_type == "f":
            is_formula = True
            if (
                hasattr(cell, "value")
                and isinstance(cell.value, str)
                and cell.value.startswith("=")
            ):
                formula = cell.value

        if is_formula and self._formula_config.mode != FormulaEvaluationMode.DISABLED:
            cached_value = self._get_cached_cell_value(sheet, row, col)
            resolved_value = cached_value
            self._ensure_formula_engine()
            try:
                resolved_value = self._formula_engine.evaluate(sheet, row, col, cached_value)
            except (ValueError, TypeError, KeyError) as e:
                logger.debug(
                    "Formula evaluation failed for cell (%s, %d, %d): %s", sheet, row, col, e
                )

        data_type = self._get_data_type(resolved_value)

        is_merged = self._is_cell_merged(ws, row, col)

        is_hidden = self._is_cell_hidden(ws, row, col)

        return CellValue(
            value=resolved_value,
            formula=formula,
            is_merged=is_merged,
            is_hidden=is_hidden,
            data_type=data_type,
            original_format=cell.number_format if hasattr(cell, "number_format") else None,
        )

    def get_cell_by_ref(self, ref: str) -> CellValue:
        """Get a cell by A1-style reference."""
        from messy_xlsx.utils import cell_ref_to_coords

        sheet, row, col = cell_ref_to_coords(ref)
        sheet = sheet or self._sheet_names[0]
        return self.get_cell(sheet, row, col)

    @staticmethod
    def _validate_public_batch_size(batch_size: object) -> None:
        if not isinstance(batch_size, int) or isinstance(batch_size, bool) or batch_size < 1:
            raise ValueError("batch_size must be >= 1")

    @staticmethod
    def _validate_public_config(config: object) -> None:
        if config is not None and not isinstance(config, SheetConfig):
            raise TypeError("config must be a SheetConfig or None")

    def _resolve_public_sheet(self, sheet: str | None) -> str:
        if sheet is not None and not isinstance(sheet, str):
            raise TypeError("sheet must be a str or None")
        name = sheet if sheet is not None else self._sheet_names[0]
        if name not in self._sheet_names:
            raise FormatError(
                f"Sheet '{name}' not found",
                file_path=self._source_handle.description,
            )
        return name

    def _uses_builtin_ooxml_planner(self) -> bool:
        """Return whether manifest/sampler planning may bypass no extensions."""
        if self.format_type not in {"xlsx", "xlsm"}:
            return False
        if type(self._registry) is not HandlerRegistry:
            return False
        if not self._registry._uses_builtin_components():
            return False
        return type(self._registry.get_handler(self.format_type)) is XLSXHandler

    def _uses_builtin_xls_planner(self) -> bool:
        """Return whether bounded xlrd inspection may bypass no extensions."""
        if self.format_type != "xls":
            return False
        if type(self._registry) is not HandlerRegistry:
            return False
        if not self._registry._uses_builtin_components():
            return False
        return type(self._registry.get_handler(self.format_type)) is XLSHandler

    def _uses_builtin_csv_planner(self) -> bool:
        """Return whether native text streaming may bypass no extensions."""
        if self.format_type not in {"csv", "tsv", "txt"}:
            return False
        if type(self._registry) is not HandlerRegistry:
            return False
        if not self._registry._uses_builtin_components():
            return False
        return type(self._registry.get_handler(self.format_type)) is CSVHandler

    @staticmethod
    def _should_propagate_sheet_error(error: BaseException) -> bool:
        """Return whether an adapter must not convert a per-sheet failure."""
        return (
            not isinstance(error, Exception)
            or _contains_active_operation_error(error)
            or _contains_process_failure(error)
        )

    def _plan_workbook_sheets(
        self,
        config: SheetConfig | None,
    ) -> tuple[PlannedSheet, ...]:
        """Compile immutable plans for all sheets before any output is exposed."""
        active_config = config if config is not None else self._sheet_config
        format_type = self.format_type
        use_ooxml = self._uses_builtin_ooxml_planner()
        has_custom_registry = type(self._registry) is not HandlerRegistry
        if not has_custom_registry:
            has_custom_registry = not self._registry._uses_builtin_components()
        planning_config = (
            replace(active_config, auto_detect=False)
            if has_custom_registry and active_config.auto_detect
            else active_config
        )
        structures: dict[str, StructureInfo] = {}

        def analyze(name: str) -> SheetInfo:
            structure: StructureInfo | None = None
            manifest = self._get_sheet_manifest(name) if use_ooxml else None
            has_ooxml_values = manifest is not None and manifest.semantic_nonempty_rows.contains(
                manifest.semantic_data_region[0]
            )
            if (
                not has_custom_registry
                and (not use_ooxml or has_ooxml_values)
                and requires_structure_analysis(planning_config, format_type)
            ):
                structure = (
                    self._analyze_stream_structure(name, planning_config)
                    if use_ooxml
                    else self._analyze_structure(name, planning_config)
                )
                structures[name] = structure

            if not use_ooxml:
                return SheetInfo(
                    name=name,
                    row_count=0,
                    col_count=0,
                    header_row=0,
                )

            assert manifest is not None
            start_row, end_row, start_col, end_col = manifest.semantic_data_region
            header_row = 0
            if structure is not None:
                header_row = max(0, (structure.header_row or 1) - 1)
            is_empty = not has_ooxml_values
            return SheetInfo(
                name=name,
                row_count=end_row - start_row + 1 if has_ooxml_values else 0,
                col_count=end_col - start_col + 1 if has_ooxml_values else 0,
                header_row=header_row,
                is_empty=is_empty,
            )

        def compile_selected(name: str, info: SheetInfo) -> ParsePlan:
            sheet_config = planning_config
            if (
                use_ooxml
                and info.is_empty
                and requires_structure_analysis(planning_config, format_type)
            ):
                sheet_config = replace(planning_config, auto_detect=False)
            return compile_parse_plan(
                sheet_config,
                structures.get(name),
                format_type,
            )

        planner = SheetPlanner(
            analyze,
            compile_selected,
            should_propagate=lambda error: (
                _contains_active_operation_error(error) or _contains_process_failure(error)
            ),
            analysis_failure_info=lambda name, error: SheetInfo(
                name=name,
                row_count=0,
                col_count=0,
                header_row=0,
                is_empty=True,
                skip_reason=f"Parse error: {error}",
            ),
        )
        return planner.plan(self._sheet_names, select_all=True)

    def _prepare_streaming_operation(
        self,
        sheet: str | None,
        batch_size: int,
        config: SheetConfig | None,
        *,
        construction_owner: _CloseOnceOwner[Any] | None = None,
    ) -> _PreparedStreamingOperation:
        """Compile all bounded evidence and open one reader before return."""
        self._validate_public_config(config)
        sheet_name = self._resolve_public_sheet(sheet)
        active_config = config if config is not None else self._sheet_config
        format_type = self.format_type
        has_custom_registry = type(self._registry) is not HandlerRegistry
        if not has_custom_registry:
            has_custom_registry = not self._registry._uses_builtin_components()
        planning_config = (
            replace(active_config, auto_detect=False)
            if has_custom_registry and active_config.auto_detect
            else active_config
        )
        decision = BackendRouter().select(
            WorkbookContext(
                format_type=format_type,
                output_mode=OutputMode.STREAMING,
                evaluate_formulas=active_config.evaluate_formulas,
                has_custom_registry=has_custom_registry,
            )
        )

        structure = None
        if requires_structure_analysis(planning_config, format_type):
            if decision.backend is BackendKind.OPENPYXL_STREAMING:
                structure = self._analyze_stream_structure(sheet_name, planning_config)
            else:
                structure = self._analyze_structure(sheet_name, planning_config)
        plan = compile_parse_plan(
            planning_config,
            structure,
            format_type,
            OutputMode.STREAMING,
            batch_size,
        )
        if decision.backend is BackendKind.OPENPYXL_STREAMING:
            return self._prepare_ooxml_stream(
                sheet_name,
                plan,
                construction_owner=construction_owner,
            )
        csv_execution: tuple[CSVExecutionKind, CSVExecutionReason] | None = None
        if decision.backend is BackendKind.CSV_STREAMING and self._uses_builtin_csv_planner():
            reason = csv_native.capability_reason()
            if reason is None:
                csv_execution = (
                    CSVExecutionKind.NATIVE,
                    CSVExecutionReason.NATIVE_SELECTED,
                )
                prepared = prepare_csv_streaming_reader(
                    self._source_handle,
                    plan,
                    self.parse_metrics,
                    construction_owner=construction_owner,
                )
                self.parse_metrics.record_csv_execution(*csv_execution)
                return prepared
            csv_execution = (
                CSVExecutionKind.MATERIALIZED_FALLBACK,
                reason,
            )

        elif (
            decision.backend is BackendKind.CUSTOM_DATAFRAME
            and format_type in {"csv", "tsv", "txt"}
        ):
            csv_execution = (
                CSVExecutionKind.CUSTOM_SPI,
                CSVExecutionReason.CUSTOM_SPI,
            )

        # Task 15 replaces the remaining XLS compatibility adapter. Custom
        # registries intentionally remain materialized because their SPI
        # returns a complete DataFrame.
        frame = self._materialize_raw_frame(sheet_name, format_type, plan)
        date_system = (
            self._get_manifest_reader().workbook.date_system
            if not has_custom_registry and format_type in {"xlsx", "xlsm", "xltx", "xltm"}
            else "1900"
        )
        prepared = prepare_materialized_streaming_reader(
            frame,
            plan,
            batch_size,
            date_system=date_system,
        )
        if csv_execution is not None:
            self.parse_metrics.record_csv_execution(*csv_execution)
        return prepared

    def _prepare_ooxml_stream(
        self,
        sheet: str,
        plan: ParsePlan,
        *,
        construction_owner: _CloseOnceOwner[Any] | None = None,
    ) -> _PreparedStreamingOperation:
        manifest = self._get_sheet_manifest(sheet)
        transform = CoordinateTransform.from_manifest(manifest)
        sample, layout = self._produce_normalization_sample(
            sheet,
            manifest,
            plan,
            transform,
        )
        normalization_plan = compile_normalization_plan(sample, plan)
        raw_reader: Any | None = None
        owned_reader: _CloseOnceReader | None = None
        try:
            if construction_owner is None:
                raw_reader = OpenpyxlStreamingReader(
                    self._source_handle,
                    manifest,
                    plan,
                    layout,
                    transform,
                )
                owned_reader = _CloseOnceReader(raw_reader)
            else:
                owned_reader = _CloseOnceReader()
                construction_owner.attach(owned_reader)
                prepare_reader = getattr(OpenpyxlStreamingReader, "prepare", None)
                if callable(prepare_reader):
                    raw_reader = prepare_reader(
                        self._source_handle,
                        manifest,
                        plan,
                        layout,
                        transform,
                    )
                    owned_reader.attach(raw_reader)
                else:
                    owned_reader._reader = OpenpyxlStreamingReader(
                        self._source_handle,
                        manifest,
                        plan,
                        layout,
                        transform,
                    )
                    raw_reader = owned_reader._reader
                assert raw_reader is not None
                start_reader = getattr(raw_reader, "start", None)
                if callable(prepare_reader) and callable(start_reader):
                    start_reader()
            assert owned_reader is not None
            prepared = wrap_normalized_streaming_reader(
                owned_reader,
                normalization_plan,
                normalize=plan.normalize,
                rollback_on_error=construction_owner is None,
            )
            if construction_owner is not None:
                construction_owner.replace(owned_reader, prepared.reader)
            return prepared
        except BaseException as error:
            if construction_owner is not None:
                raise
            rollback_reader = owned_reader if owned_reader is not None else raw_reader
            cleanups = (
                []
                if rollback_reader is None
                else [("OOXML raw reader rollback", lambda: _close_if_present(rollback_reader))]
            )
            _run_cleanups(
                cleanups,
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise

    def _produce_normalization_sample(
        self,
        sheet: str,
        manifest: SheetManifest,
        plan: ParsePlan,
        transform: CoordinateTransform,
    ) -> tuple[NormalizationSample, StreamingWorksheetLayout]:
        """Read and close one bounded sample before opening the full reader."""
        bounds = _read_bounds(manifest, plan)
        width = 0 if bounds is None else bounds[3] - bounds[2] + 1
        raw_schema = pa.schema([pa.field(str(ordinal), pa.string()) for ordinal in range(width)])
        raw_column_numbers = () if bounds is None else tuple(range(bounds[2], bounds[3] + 1))
        layout = StreamingWorksheetLayout.compile(
            manifest,
            plan,
            raw_schema,
            transform,
            raw_column_numbers=raw_column_numbers,
        )
        if bounds is None:
            return (
                NormalizationSample(
                    schema=layout.output_schema,
                    column_identities=(),
                    columns=(),
                    row_numbers=pa.array([], type=pa.int64()),
                    date_system=self._get_manifest_reader().workbook.date_system,
                ),
                layout,
            )

        if width > _NORMALIZATION_SAMPLE_CELLS:
            raise ValueError(
                f"sample raw window may retain at most {_NORMALIZATION_SAMPLE_CELLS} cells"
            )
        capacity, capacity_reason = _sample_capacity(width, plan.header_rows)
        projection = CoordinateOperation._parse_projection(plan.cell_range)
        physical_max_row = max(
            (
                manifest.observed_max_row,
                *(merged.max_row for merged in manifest.merged_ranges),
            )
        )
        merge_fill = MergeStrategy(plan.merge_strategy) is MergeStrategy.FILL
        anchor_rows: set[int] = set()
        synthetic_row: int | None = None

        if projection is not None:
            retained_count = max(
                0,
                projection.max_row - projection.min_row + 1 - plan.skip_footer,
            )
            sample_count = min(retained_count, capacity)
            core_intervals: tuple[tuple[int, int], ...] = ()
            if sample_count:
                sample_end = projection.min_row + sample_count - 1
                physical_end = min(sample_end, physical_max_row)
                if projection.min_row <= physical_end:
                    core_intervals = ((projection.min_row, physical_end),)
                if sample_end > physical_end:
                    synthetic_row = sample_end
                if merge_fill:
                    anchor_rows = {
                        merged.min_row
                        for merged in manifest.merged_ranges
                        if merged.min_row < projection.min_row
                        and CoordinateOperation._merge_intersects_projection(
                            merged,
                            projection,
                        )
                    }
            sample_plan = replace(plan, skip_rows=0, skip_footer=0)
            windows = _coalesce_row_windows(core_intervals, anchor_rows)
        else:
            visible = _visible_row_intervals(
                bounds[0],
                bounds[1],
                manifest.hidden_rows,
                ignore_hidden=plan.ignore_hidden,
            )
            visible_count = _interval_row_count(visible)
            retained_count = max(
                0,
                visible_count - plan.skip_rows - plan.skip_footer,
            )
            sample_count = min(retained_count, capacity)
            core_intervals = _slice_row_intervals(
                visible,
                skip=plan.skip_rows,
                count=sample_count,
            )
            if core_intervals and merge_fill:
                anchor_rows = {
                    merged.min_row
                    for merged in manifest.merged_ranges
                    if _row_range_intersects_intervals(
                        merged.min_row,
                        merged.max_row,
                        core_intervals,
                    )
                    and not _row_is_in_intervals(
                        merged.min_row,
                        core_intervals,
                    )
                }
            core_start = core_intervals[0][0] if core_intervals else bounds[1] + 1
            supplied_visible_prefix = sum(
                row < core_start
                and (not plan.ignore_hidden or not manifest.hidden_rows.contains(row))
                for row in anchor_rows
            )
            sample_plan = replace(
                plan,
                skip_rows=supplied_visible_prefix,
                skip_footer=0,
            )
            windows = _coalesce_row_windows(core_intervals, anchor_rows)

        sample_truncated = retained_count > sample_count
        raw_budget = _RawBatchBudget(
            max_rows=capacity,
            max_cells=_NORMALIZATION_SAMPLE_CELLS,
            max_bytes=_NORMALIZATION_SAMPLE_BYTES,
            truncate_to_budget=True,
        )
        if retained_count and capacity == 0:
            raise _sample_budget_error(raw_budget, capacity_reason)

        try:
            self._close_sample_owner()
            sample_owner = _SampleResourceOwner(self._source_handle)
            self._sample_owner = sample_owner
            backend_context = self._source_handle.open_backend()
            backend = sample_owner.enter_source(backend_context)
            sample_owner.workbook = openpyxl.load_workbook(
                backend,
                read_only=True,
                data_only=plan.data_only,
                keep_links=False,
                keep_vba=False,
            )
            workbook = sample_owner.workbook
            assert workbook is not None
            worksheet = workbook[sheet]
            reset_dimensions = getattr(worksheet, "reset_dimensions", None)
            if callable(reset_dimensions):
                reset_dimensions()
            operation = transform.open(sample_plan, layout.prepared_schema)
            accumulator = _NormalizationSampleAccumulator(
                layout.output_schema,
                date_system=self._get_manifest_reader().workbook.date_system,
                preserve_native=not plan.normalize,
                max_rows=_NORMALIZATION_SAMPLE_ROWS,
                max_cells=_NORMALIZATION_SAMPLE_CELLS,
                max_bytes=_NORMALIZATION_SAMPLE_BYTES,
            )
            raw: CoordinateBatch | None = None

            stop_sampling = False
            iterator_row: int | None = None
            if windows:
                iterator_row = windows[0][0]
                sample_owner.rows = iter(
                    worksheet.iter_rows(
                        values_only=True,
                        min_row=iterator_row,
                        max_row=windows[-1][1],
                        min_col=bounds[2],
                        max_col=bounds[3],
                    )
                )
                row_iterator = sample_owner.rows
            for window_start, window_stop in windows:
                row_iterator = sample_owner.rows
                assert row_iterator is not None
                assert iterator_row is not None
                while iterator_row < window_start:
                    try:
                        next(row_iterator)
                    except StopIteration:
                        stop_sampling = True
                        break
                    iterator_row += 1
                if stop_sampling:
                    break
                next_row = window_start
                while next_row <= window_stop:
                    if raw_budget.exhausted_reason is not None or accumulator.full:
                        stop_sampling = True
                        break
                    stop_row = min(
                        window_stop,
                        next_row + _NORMALIZATION_SAMPLE_ROWS - 1,
                    )
                    requested = stop_row - next_row + 1
                    raw = _raw_coordinate_batch(
                        islice(row_iterator, requested),
                        layout,
                        next_row,
                        budget=raw_budget,
                    )
                    if raw.batch.num_rows == 0:
                        stop_sampling = True
                        break
                    accumulator.consume(operation.push(raw))
                    if raw.batch.num_rows < requested:
                        stop_sampling = True
                        break
                    next_row += raw.batch.num_rows
                    iterator_row += raw.batch.num_rows
                if stop_sampling:
                    break

            if (
                synthetic_row is not None
                and raw_budget.exhausted_reason is None
                and not accumulator.full
            ):
                raw = _raw_coordinate_batch(
                    iter(((),)),
                    layout,
                    synthetic_row,
                    budget=raw_budget,
                )
                if raw.batch.num_rows:
                    accumulator.consume(operation.push(raw))

            if raw_budget.exhausted_reason is not None:
                raise _sample_budget_error(raw_budget)

            identities = operation.identity_snapshot()
            if (
                plan.header_rows > 0
                and not identities
                and len(layout.output_schema) > 0
                and (raw_budget.exhausted_reason is not None or sample_truncated)
            ):
                raise _sample_budget_error(
                    raw_budget,
                    raw_budget.exhausted_reason or capacity_reason,
                )
            raw = None
            del operation
            sample = accumulator.finish(identities)
        except BaseException as error:
            _run_cleanups(
                [("sample resource cleanup", self._close_sample_owner)],
                primary_error=error,
                primary_traceback=_exception_traceback(error),
            )
            raise

        _run_cleanups([("sample resource cleanup", self._close_sample_owner)])
        return sample, layout

    def _parse_sheet(
        self,
        sheet: str,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Parse a sheet to DataFrame with normalization."""
        lease = self._materialized_operation()
        with lease:
            try:
                lease._body_started()
                return self._parse_sheet_unreserved(sheet, config)
            finally:
                lease._body_complete()

    def _parse_sheet_unreserved(
        self,
        sheet: str,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Parse one sheet while the caller owns the workbook reservation."""
        config = config or self._sheet_config
        format_type = self.format_type

        structure = None
        if requires_structure_analysis(config, format_type):
            structure = self._analyze_structure(sheet, config)
        plan = compile_parse_plan(config, structure, format_type)
        return self._materialize_compiled_plan(sheet, plan)

    def _materialize_compiled_plan(
        self,
        sheet: str,
        plan: ParsePlan,
    ) -> pd.DataFrame:
        """Materialize and normalize one already-frozen plan."""
        df = self._materialize_raw_frame_counted(sheet, self.format_type, plan)

        if plan.normalize:
            semantic_hints = plan.thaw_type_hints()
            _validate_materialized_label_targets(
                df.columns,
                list(semantic_hints.items()),
            )
            pipeline = NormalizationPipeline(
                decimal_separator=plan.decimal_separator,
                thousands_separator=plan.thousands_separator,
                use_extended_missing_list=plan.use_extended_missing_list,
                preserve_types=plan.preserve_types,
            )

            df = pipeline.normalize(
                df,
                semantic_hints=semantic_hints,
                skip_steps=list(plan.skip_normalization_steps),
            )

        # Sanitize column names if requested
        if plan.sanitize_column_names:
            df = self._sanitize_columns(df)

        # Apply user renames (user overrides take precedence)
        if plan.column_renames:
            rename_items = list(plan.thaw_column_renames().items())
            df = _rename_materialized_columns(
                df,
                rename_items,
            )

        # Preserve the legacy behavior where disabling normalization also
        # bypasses row filters. S15 owns any change to that public contract.
        if not plan.normalize:
            return df

        # Drop rows matching regex pattern
        if plan.drop_regex and not df.empty:
            pattern = re.compile(plan.drop_regex)
            mask = df.apply(
                lambda row: any(
                    bool(pattern.search(str(v)))
                    for v in row
                    if v is not None and not (isinstance(v, float) and pd.isna(v))
                ),
                axis=1,
            )
            df = df[~mask].reset_index(drop=True)

        # Drop rows matching column-value conditions
        if plan.drop_conditions and not df.empty:
            for col, value in plan.thaw_drop_conditions():
                if col is None:
                    continue
                ordinals = _materialized_condition_ordinals(df.columns, col)
                if len(ordinals) == 1:
                    df = df[df.iloc[:, ordinals[0]] != value].reset_index(drop=True)
                elif ordinals:
                    selected = df.iloc[:, list(ordinals)]
                    df = df[selected != value].reset_index(drop=True)

        return df

    def _materialize_raw_frame_counted(
        self,
        sheet: str,
        format_type: str,
        plan: ParsePlan,
    ) -> pd.DataFrame:
        """Count one raw backend result unless its coordinator already did."""
        metrics = self.parse_metrics
        full_before = metrics.full_materializations
        failures_before = metrics.failed_attempts
        try:
            frame = self._materialize_raw_frame(sheet, format_type, plan)
        except BaseException:
            if metrics.failed_attempts == failures_before:
                metrics.failed_attempts += 1
            raise
        if metrics.full_materializations == full_before:
            metrics.full_materializations += 1
        return frame

    def _materialize_raw_frame(
        self,
        sheet: str,
        format_type: str,
        plan: ParsePlan,
    ) -> pd.DataFrame:
        """Materialize only handler/coordinate parsing, before global normalization."""
        built_in = self._parse_builtin_materialized(sheet, format_type, plan)
        if built_in is _NO_BUILTIN_MATERIALIZATION:
            with self._registry_source() as source:
                parse = self._registry.parse
                if (
                    getattr(parse, "__func__", None) is HandlerRegistry.parse
                    and getattr(parse, "__self__", None) is self._registry
                ):
                    return self._registry._parse_counted(
                        source,
                        sheet=sheet,
                        options=plan.to_parse_options(),
                        format_type=format_type,
                        metrics=self.parse_metrics,
                    )
                return self._registry.parse(
                    source,
                    sheet=sheet,
                    options=plan.to_parse_options(),
                    format_type=format_type,
                )
        assert isinstance(built_in, pd.DataFrame)
        return built_in

    def _parse_builtin_materialized(
        self,
        sheet: str,
        format_type: str,
        plan: ParsePlan,
    ) -> pd.DataFrame | object:
        """Use the bound-plan seam only for the untouched built-in XLSX stack."""
        if format_type not in {"xlsx", "xlsm"}:
            return _NO_BUILTIN_MATERIALIZATION
        if type(self._registry) is not HandlerRegistry:
            return _NO_BUILTIN_MATERIALIZATION
        if not self._registry._uses_builtin_components():
            return _NO_BUILTIN_MATERIALIZATION
        handler = self._registry.get_handler(format_type)
        if type(handler) is not XLSXHandler:
            return _NO_BUILTIN_MATERIALIZATION
        assert handler is not None
        if not _is_fastexcel_materialized_plan(plan):
            return _NO_BUILTIN_MATERIALIZATION
        transform: CoordinateTransform | None = None
        coordinate_features = (
            plan.merge_strategy != "skip" or plan.ignore_hidden or bool(plan.cell_range)
        )
        if coordinate_features:
            if not self._coordinate_range_is_supported(plan):
                return _NO_BUILTIN_MATERIALIZATION
            if sheet not in self._sheet_names:
                return _NO_BUILTIN_MATERIALIZATION
            manifest = self._get_sheet_manifest(sheet)
            if not self._manifest_supports_coordinate_plan(manifest, plan):
                return _NO_BUILTIN_MATERIALIZATION
            transform = CoordinateTransform.from_manifest(manifest)
        try:
            return handler._parse_materialized_plan(
                self._source_handle,
                sheet,
                plan,
                self._get_fastexcel_session,
                metrics=self.parse_metrics,
                transform=transform,
            )
        except Exception as error:
            if _blocks_backend_retry(error):
                raise

        file_desc = self._source_handle.description
        name = self._source_handle.path.name if self._source_handle.path is not None else file_desc
        raise FormatError(
            f"All handlers failed for {name}",
            file_path=file_desc,
            detected_format=format_type,
            attempted_formats=[type(handler).__name__],
        )

    def _get_fastexcel_session(self) -> FastexcelSession:
        """Return the workbook-owned session shared by eligible sheet reads."""
        session = self._fastexcel_session
        if session is None:
            session = FastexcelSession(self._source_handle)
            self._fastexcel_session = session
        return session

    def _get_manifest_reader(self) -> ManifestReader:
        reader = self._manifest_reader
        if reader is None:
            reader = ManifestReader(self._source_handle)
            self._manifest_reader = reader
            self.parse_metrics.manifest_builds += 1
        return reader

    def _get_sheet_manifest(self, sheet: str) -> SheetManifest:
        return self._get_manifest_reader().sheet(sheet)

    def _analyze_stream_structure(
        self,
        sheet: str,
        config: SheetConfig,
    ) -> StructureInfo:
        sampler = self._stream_structure_sampler
        if sampler is None:
            sampler = StructureSampler(
                self._get_fastexcel_session(),
                self._get_manifest_reader(),
                metrics=self.parse_metrics,
            )
            self._stream_structure_sampler = sampler
        return cast(StructureInfo, sampler.analyze(sheet, config.header_patterns))

    @staticmethod
    def _coordinate_range_is_supported(plan: ParsePlan) -> bool:
        if not plan.cell_range:
            return True
        try:
            CoordinateTransform(
                hidden_rows=IntervalIndex(()),
                hidden_columns=IntervalIndex(()),
                merged_ranges=(),
            ).open(plan)
        except ValueError:
            return False
        return True

    @staticmethod
    def _manifest_supports_coordinate_plan(
        manifest: SheetManifest,
        plan: ParsePlan,
    ) -> bool:
        if not plan.ignore_hidden or bool(plan.cell_range):
            return True
        if any(interval.start != interval.end for interval in manifest.hidden_columns.intervals):
            return False
        if manifest.observed_max_col == 0:
            return True
        intervals = manifest.hidden_columns.intervals
        return not (
            len(intervals) == 1
            and intervals[0].start == 1
            and intervals[0].end >= manifest.observed_max_col
        )

    def _analyze_structure(self, sheet: str, config: SheetConfig | None = None) -> StructureInfo:
        """Analyze sheet structure."""
        header_patterns = config.header_patterns if config else None
        return cast(
            StructureInfo,
            self._analyzer.analyze(
                self._source_handle,
                sheet,
                header_patterns=header_patterns,
            ),
        )

    def _sanitize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize column names for BigQuery compatibility."""
        from .utils import sanitize_column_name

        new_columns = []
        seen: dict[str, int] = {}

        for col in df.columns:
            clean = sanitize_column_name(col)

            # Handle duplicates by appending counter
            if clean in seen:
                seen[clean] += 1
                clean = f"{clean}_{seen[clean]}"
            else:
                seen[clean] = 0

            new_columns.append(clean)

        df.columns = new_columns
        return df

    def _ensure_formula_engine(self) -> None:
        """Lazily load formula engine on first get_cell() call."""
        if self._formula_loaded:
            return
        self._formula_loaded = True

        if (
            self._formula_config.mode != FormulaEvaluationMode.DISABLED
            and self._formula_engine.is_available
        ):
            try:
                if self._source_handle.path is not None:
                    self._formula_engine.load_workbook(self._source_handle.path)
            except (OSError, ValueError, TypeError) as e:
                logger.debug("Formula engine load failed: %s", e)

    def _ensure_workbook(self) -> None:
        """Ensure openpyxl workbook is loaded."""
        if self._wb is None:
            source: Path | BinaryIO
            owned_source: BinaryIO | None
            if self._source_handle.path is None:
                owned_source = self._source_handle.detached_binary()
                source = owned_source
            else:
                owned_source = None
                source = self._source_handle.path
            try:
                self._wb = openpyxl.load_workbook(
                    source,
                    read_only=False,
                    data_only=False,
                )
            except BaseException:
                if owned_source is not None:
                    owned_source.close()
                raise
            self._wb_source = owned_source

    def _get_cached_cell_value(self, sheet: str, row: int, col: int) -> Any:
        """Read a formula's cached result from a data-only workbook view."""
        if self._cached_wb is None:
            source: Path | BinaryIO
            owned_source: BinaryIO | None
            if self._source_handle.path is None:
                owned_source = self._source_handle.detached_binary()
                source = owned_source
            else:
                owned_source = None
                source = self._source_handle.path
            try:
                self._cached_wb = openpyxl.load_workbook(
                    source,
                    read_only=False,
                    data_only=True,
                )
            except BaseException:
                if owned_source is not None:
                    owned_source.close()
                raise
            self._cached_wb_source = owned_source

        return self._cached_wb[sheet].cell(row, col).value

    def _get_data_type(self, value: Any) -> str:
        """Determine data type string for a value."""
        if value is None:
            return "empty"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, str):
            if value.startswith("#") and value.endswith("!"):
                return "error"
            return "text"
        if hasattr(value, "date"):
            return "date"
        return "text"

    def _is_cell_merged(self, ws: Any, row: int, col: int) -> bool:
        """Check if cell is part of a merged range."""
        try:
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    return True
        except (AttributeError, TypeError):
            pass
        return False

    def _is_cell_hidden(self, ws: Any, row: int, col: int) -> bool:
        """Check if cell is in a hidden row or column."""
        try:
            if row in ws.row_dimensions and ws.row_dimensions[row].hidden:
                return True
            from openpyxl.utils import get_column_letter

            col_letter = get_column_letter(col)
            if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden:
                return True
        except (AttributeError, TypeError):
            pass
        return False

    def close(self) -> None:
        """Close the workbook and release resources."""
        self._close()

    def _close(
        self,
        *,
        primary_error: BaseException | None = None,
        primary_traceback: TracebackType | None = None,
    ) -> None:
        """Close owned slots, retaining exact owners after process failures."""
        self._closed = True
        self._manifest_reader = None
        self._stream_structure_sampler = None

        cleanups: list[tuple[str, Any]] = []
        if getattr(self, "_sample_owner", None) is not None:
            cleanups.append(
                (
                    "sample resource cleanup",
                    self._close_sample_owner,
                )
            )
        if (
            getattr(self, "_active_stream", None) is not None
            or self.__dict__.get("_active_stream_cleanup") is not None
        ):
            cleanups.append(("active stream invalidation", self._invalidate_active_stream))
        elif getattr(self, "_active_materialized_lease", None) is not None:
            cleanups.append(
                (
                    "materialized operation release",
                    self._release_materialized_lease,
                )
            )
        elif getattr(self, "_active_operation_token", None) is not None:
            self._active_operation_token = None
        for label, attribute in (
            ("fastexcel session cleanup", "_fastexcel_session"),
            ("workbook cleanup", "_wb"),
            ("cached workbook cleanup", "_cached_wb"),
            ("workbook source cleanup", "_wb_source"),
            ("cached workbook source cleanup", "_cached_wb_source"),
        ):
            if getattr(self, attribute, None) is not None:
                cleanups.append(
                    (
                        label,
                        lambda attribute=attribute: self._close_owned_attribute(attribute),
                    )
                )
        source_handle = getattr(self, "_source_handle", None)
        source_pending = getattr(
            self,
            "_source_handle_close_pending",
            source_handle is not None,
        )
        if source_handle is not None and source_pending:
            cleanups.append(("source handle cleanup", self._close_source_handle))
        _run_cleanups(
            cleanups,
            primary_error=primary_error,
            primary_traceback=primary_traceback,
        )

    def _release_materialized_lease(self) -> None:
        lease = getattr(self, "_active_materialized_lease", None)
        if lease is None:
            return
        lease._release_from_owner()

    def _invalidate_active_stream(self) -> None:
        stream = getattr(self, "_active_stream", None)
        cleanup_state = self.__dict__.get("_active_stream_cleanup")
        if stream is None and cleanup_state is None:
            return
        token = getattr(self, "_active_operation_token", None)
        invalidate = getattr(stream, "invalidate_from_owner", None)
        try:
            if callable(invalidate):
                invalidate()
            elif cleanup_state is not None:
                cleanup_state.close()
            if token is not None:
                self._end_operation(token)
            else:
                self._active_stream = None
        except BaseException as error:
            if not _contains_process_failure(error):
                if token is not None:
                    self._end_operation(token)
                else:
                    self._active_stream = None
            raise

    def _close_owned_attribute(self, attribute: str) -> None:
        owner = getattr(self, attribute, None)
        if owner is None:
            return
        try:
            setattr(self, attribute, _close_if_present(owner))
        except BaseException as error:
            if not _contains_process_failure(error):
                setattr(self, attribute, None)
            raise

    def _close_source_handle(self) -> None:
        source_handle = getattr(self, "_source_handle", None)
        if not getattr(
            self,
            "_source_handle_close_pending",
            source_handle is not None,
        ):
            return
        if source_handle is None:
            self._source_handle_close_pending = False
            return
        if (
            getattr(self, "_sample_owner", None) is not None
            or getattr(self, "_active_stream", None) is not None
            or self.__dict__.get("_active_stream_cleanup") is not None
        ):
            return
        try:
            _close_if_present(source_handle)
        except BaseException:
            self._source_handle_close_pending = not bool(getattr(source_handle, "closed", False))
            raise
        self._source_handle_close_pending = False

    def _close_sample_owner(self) -> None:
        owner = getattr(self, "_sample_owner", None)
        if owner is None:
            return
        try:
            self._sample_owner = _close_if_present(owner)
        except BaseException as error:
            if not _contains_process_failure(error):
                self._sample_owner = None
            raise

    def __enter__(self) -> "MessyWorkbook":
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        del exc_type
        if isinstance(exc_val, BaseException):
            self._close(primary_error=exc_val, primary_traceback=exc_tb)
            return
        self.close()

    def __repr__(self) -> str:
        name = (
            self._source_handle.path.name
            if self._source_handle.path is not None
            else self._source_handle.description
        )
        return f"MessyWorkbook({name!r}, sheets={self._sheet_names})"
