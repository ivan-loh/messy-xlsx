"""Structure analysis for Excel sheets."""

# ============================================================================
# Imports
# ============================================================================

from pathlib import Path
from typing import Any, BinaryIO

import openpyxl

from messy_xlsx.cache import StructureCache, get_structure_cache
from messy_xlsx.detection.locale_detector import LocaleDetector
from messy_xlsx.exceptions import StructureError
from messy_xlsx.models import StructureInfo, TableInfo

# ============================================================================
# Type Aliases
# ============================================================================

FileSource = Path | BinaryIO


# ============================================================================
# Config
# ============================================================================

MAX_ANALYSIS_ROWS = 10_000
MAX_SAMPLE_ROWS = 1_000


# ============================================================================
# Core
# ============================================================================


class StructureAnalyzer:
    """Analyze Excel sheet structure."""

    def __init__(self, cache: StructureCache | None = None):
        """Initialize analyzer."""
        self.cache = cache or get_structure_cache()
        self.locale_detector = LocaleDetector()

    def analyze(
        self,
        file_source: FileSource,
        sheet: str,
        force: bool = False,
        header_patterns: list[str] | None = None,
    ) -> StructureInfo:
        """Analyze sheet structure."""
        is_fileobj = hasattr(file_source, "read")

        # Only use cache for file paths (not file-like objects)
        if not is_fileobj:
            file_path = Path(file_source)  # type: ignore[arg-type]
            if not force:
                cached = self.cache.get(file_path, sheet)
                if cached:
                    return cached
        else:
            file_path = None
            # Reset buffer position
            if hasattr(file_source, "seek"):
                file_source.seek(0)

        try:
            # Note: Cannot use read_only=True as ReadOnlyWorksheet doesn't have merged_cells
            wb = openpyxl.load_workbook(
                file_source,
                read_only=False,
                data_only=True,
            )
        except Exception as e:
            raise StructureError(
                f"Cannot open file for analysis: {e}",
                sheet=sheet,
                detection_phase="open",
            ) from e

        try:
            if sheet not in wb.sheetnames:
                raise StructureError(
                    f"Sheet '{sheet}' not found",
                    sheet=sheet,
                    detection_phase="sheet_lookup",
                )

            ws = wb[sheet]

            data_region = self._detect_data_region(ws)
            merged = self._detect_merged_cells(ws)
            hidden_rows, hidden_cols = self._detect_hidden_content(ws)
            header_info = self._detect_headers(ws, data_region, merged, header_patterns)
            metadata = self._detect_metadata_rows(ws, data_region, header_info)
            tables = self._detect_multiple_tables(ws, data_region, header_info)
            locale_info = self.locale_detector.detect(ws, data_region)
            blank_rows = self._detect_blank_rows(ws, data_region)
            sparse_columns = self._detect_sparse_columns(ws, data_region)

            # Collect all results that depend on the open workbook
            suggested_skip_rows = self._suggest_skip_rows(metadata, header_info)
            suggested_skip_footer = self._suggest_skip_footer(ws, data_region)

        finally:
            wb.close()

        # Formula detection needs a separate data_only=False pass.
        # Must run AFTER the first workbook is closed to avoid stream corruption.
        has_formulas = self._detect_formulas_from_source(file_source, sheet, data_region)

        result = StructureInfo(
            data_start_row=data_region["start_row"],
            data_end_row=data_region["end_row"],
            data_start_col=data_region["start_col"],
            data_end_col=data_region["end_col"],
            header_row=header_info.get("header_row"),
            header_rows_count=header_info.get("header_rows_count", 1),
            header_confidence=header_info.get("confidence", 0.0),
            metadata_rows=metadata,
            merged_ranges=merged,
            merged_in_headers=self._check_merged_in_headers(merged, header_info),
            merged_in_data=self._check_merged_in_data(merged, header_info),
            hidden_rows=hidden_rows,
            hidden_columns=hidden_cols,
            detected_locale=locale_info.locale,
            decimal_separator=locale_info.decimal_separator,
            thousands_separator=locale_info.thousands_separator,
            num_tables=len(tables),
            table_ranges=[self._table_to_dict(t) for t in tables],
            blank_rows=blank_rows,
            has_formulas=has_formulas,
            sparse_columns=sparse_columns,
            suggested_skip_rows=suggested_skip_rows,
            suggested_skip_footer=suggested_skip_footer,
        )

        # Only cache for file paths
        if file_path:
            self.cache.put(file_path, sheet, result)

        return result

    def _detect_data_region(self, ws: Any) -> dict[str, int]:
        """Find the boundaries of actual data."""
        max_row_hint = ws.max_row or 1

        scan_limit = min(MAX_ANALYSIS_ROWS, max_row_hint)

        min_row, max_row = None, None
        min_col, max_col = None, None
        empty_row_streak = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=scan_limit), start=1):
            row_has_data = False
            for col_idx, value in enumerate(row, start=1):
                if value is not None:
                    row_has_data = True
                    if min_row is None:
                        min_row = row_idx
                    max_row = row_idx
                    if min_col is None or col_idx < min_col:
                        min_col = col_idx
                    if max_col is None or col_idx > max_col:
                        max_col = col_idx

            if row_has_data:
                empty_row_streak = 0
            else:
                empty_row_streak += 1

            if min_row and empty_row_streak > 100:
                break

        if max_row and scan_limit < max_row_hint and empty_row_streak < 10:
            max_row = max_row_hint

        if min_row is None:
            return {"start_row": 1, "end_row": 1, "start_col": 1, "end_col": 1}

        return {
            "start_row": min_row,
            "end_row": max_row or 1,
            "start_col": min_col or 1,
            "end_col": max_col or 1,
        }

    def _detect_merged_cells(self, ws: Any) -> list[tuple[int, int, int, int]]:
        """Detect all merged cell ranges."""
        merged = []
        try:
            for merged_range in ws.merged_cells.ranges:
                merged.append(
                    (
                        merged_range.min_row,
                        merged_range.min_col,
                        merged_range.max_row,
                        merged_range.max_col,
                    )
                )
        except Exception:
            # Silently fail if worksheet doesn't support merged cells
            pass
        return merged

    def _detect_hidden_content(self, ws: Any) -> tuple[list[int], list[int]]:
        """Detect hidden rows and columns."""
        hidden_rows = []
        hidden_cols = []

        try:
            for row_idx, dim in ws.row_dimensions.items():
                if dim.hidden:
                    hidden_rows.append(row_idx)

            for col_key, dim in ws.column_dimensions.items():
                if dim.hidden:
                    from openpyxl.utils import column_index_from_string

                    hidden_cols.append(column_index_from_string(col_key))
        except Exception:
            pass

        return sorted(hidden_rows), sorted(hidden_cols)

    def _detect_headers(
        self,
        ws: Any,
        data_region: dict[str, int],
        merged_ranges: list[tuple[int, int, int, int]],
        header_patterns: list[str] | None = None,
    ) -> dict[str, Any]:
        """Detect header row(s) with confidence scoring."""
        import re

        start_row = data_region["start_row"]
        end_row = min(start_row + 15, data_region["end_row"])
        start_col = data_region["start_col"]
        end_col = data_region["end_col"]
        total_cols = end_col - start_col + 1

        best_header_row = None
        best_confidence = 0.0
        best_underscore_count = 0

        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                try:
                    cell = ws.cell(row_idx, col_idx)
                    row_values.append(cell.value)
                except Exception:
                    row_values.append(None)

            non_empty = [v for v in row_values if v is not None]

            if not non_empty:
                continue

            all_strings = all(isinstance(v, str) for v in non_empty)

            if not all_strings:
                continue

            # Calculate column coverage (what % of columns are filled)
            coverage = len(non_empty) / total_cols if total_cols > 0 else 0

            # Sparse rows (1-2 cells) are likely metadata, not headers
            # Penalize heavily if coverage is very low
            if coverage < 0.3 and len(non_empty) <= 2:
                continue  # Skip metadata-like rows entirely

            # Check for numbers in rows below (look ahead up to 3 rows)
            has_numbers_below = False
            for look_ahead in range(1, 4):
                check_row = row_idx + look_ahead
                if check_row > data_region["end_row"]:
                    break
                for col_idx in range(start_col, end_col + 1):
                    try:
                        cell = ws.cell(check_row, col_idx)
                        if isinstance(cell.value, (int, float)):
                            has_numbers_below = True
                            break
                    except Exception:
                        pass
                if has_numbers_below:
                    break

            # Base confidence
            confidence = 0.3

            # Boost for all strings
            if all_strings:
                confidence += 0.1

            # Boost for good column coverage
            if coverage >= 0.5:
                confidence += 0.2
            elif coverage >= 0.3:
                confidence += 0.1

            # Boost for numbers in data rows below
            if has_numbers_below:
                confidence += 0.2

            # Boost for header-like naming patterns (snake_case, contains common header words)
            header_like_count = 0
            for val in non_empty:
                val_str = str(val).lower()
                # Check for snake_case or common header patterns
                if re.match(r"^[a-z][a-z0-9_]*$", val_str) or re.search(
                    r"\b(id|name|date|time|code|type|status|number|amount|qty|count|total)\b",
                    val_str,
                    re.I,
                ):  # snake_case
                    header_like_count += 1

            if header_like_count > 0:
                header_ratio = header_like_count / len(non_empty)
                confidence += min(0.2, header_ratio * 0.3)

            # Boost for merged cells in header
            has_merged = any(mr[0] == row_idx for mr in merged_ranges)
            if has_merged:
                confidence += 0.05

            # Pattern matching boost
            if header_patterns:
                row_text = " ".join(str(v) for v in row_values if v is not None).lower()
                pattern_matches = sum(
                    1 for pattern in header_patterns if re.search(pattern, row_text, re.IGNORECASE)
                )
                if pattern_matches > 0:
                    confidence += min(0.15, 0.05 * pattern_matches)

            # Tiebreaker for multi-row headers: actual column-name rows
            # typically contain underscores (snake_case) while sub-category
            # rows ("Timing", "Identifiers") don't.
            underscore_count = sum(1 for v in non_empty if "_" in str(v))

            if confidence > best_confidence or (
                confidence == best_confidence and underscore_count > best_underscore_count
            ):
                best_confidence = confidence
                best_header_row = row_idx
                best_underscore_count = underscore_count

        if best_header_row is not None:
            return {
                "header_row": best_header_row,
                "header_rows_count": 1,
                "confidence": best_confidence,
            }

        return {
            "header_row": None,
            "header_rows_count": 0,
            "confidence": 0.0,
        }

    def _detect_metadata_rows(
        self,
        ws: Any,
        data_region: dict[str, int],
        header_info: dict[str, Any],
    ) -> list[int]:
        """Detect metadata/title rows before the header."""
        metadata_rows = []
        header_row = header_info.get("header_row") or data_region["start_row"]

        for row_idx in range(data_region["start_row"], header_row):
            row_values = []
            for col_idx in range(data_region["start_col"], data_region["end_col"] + 1):
                try:
                    cell = ws.cell(row_idx, col_idx)
                    row_values.append(cell.value)
                except Exception:
                    pass

            non_empty = [v for v in row_values if v is not None]

            if 1 <= len(non_empty) <= 2:
                metadata_rows.append(row_idx)

        return metadata_rows

    def _detect_multiple_tables(
        self,
        ws: Any,
        data_region: dict[str, int],
        header_info: dict[str, Any],
    ) -> list[TableInfo]:
        """Detect multiple tables separated by blank rows."""
        total_rows = data_region["end_row"] - data_region["start_row"] + 1

        if total_rows > MAX_ANALYSIS_ROWS:
            return [
                TableInfo(
                    start_row=data_region["start_row"],
                    end_row=data_region["end_row"],
                    start_col=data_region["start_col"],
                    end_col=data_region["end_col"],
                    has_header=header_info.get("header_row") is not None,
                    header_row=header_info.get("header_row"),
                    confidence=header_info.get("confidence", 1.0),
                )
            ]

        blank_rows = self._detect_blank_rows(ws, data_region)

        if not blank_rows:
            return [
                TableInfo(
                    start_row=data_region["start_row"],
                    end_row=data_region["end_row"],
                    start_col=data_region["start_col"],
                    end_col=data_region["end_col"],
                    has_header=header_info.get("header_row") is not None,
                    header_row=header_info.get("header_row"),
                    confidence=header_info.get("confidence", 1.0),
                )
            ]

        groups = self._group_consecutive(blank_rows)

        separators = [g for g in groups if len(g) >= 2]

        if not separators:
            return [
                TableInfo(
                    start_row=data_region["start_row"],
                    end_row=data_region["end_row"],
                    start_col=data_region["start_col"],
                    end_col=data_region["end_col"],
                    has_header=header_info.get("header_row") is not None,
                    header_row=header_info.get("header_row"),
                )
            ]

        tables = []
        prev_end = data_region["start_row"] - 1

        for sep_group in separators:
            sep_start = min(sep_group)

            if sep_start > prev_end + 1:
                tables.append(
                    TableInfo(
                        start_row=prev_end + 1,
                        end_row=sep_start - 1,
                        start_col=data_region["start_col"],
                        end_col=data_region["end_col"],
                    )
                )

            prev_end = max(sep_group)

        if prev_end < data_region["end_row"]:
            tables.append(
                TableInfo(
                    start_row=prev_end + 1,
                    end_row=data_region["end_row"],
                    start_col=data_region["start_col"],
                    end_col=data_region["end_col"],
                )
            )

        return tables

    def _detect_blank_rows(self, ws: Any, data_region: dict[str, int]) -> list[int]:
        """Find blank rows within data region."""
        total_rows = data_region["end_row"] - data_region["start_row"] + 1

        if total_rows > MAX_SAMPLE_ROWS:
            return self._detect_blank_rows_sampled(ws, data_region)

        blank_rows = []

        for row_idx, row in enumerate(
            ws.iter_rows(
                min_row=data_region["start_row"],
                max_row=data_region["end_row"],
                min_col=data_region["start_col"],
                max_col=data_region["end_col"],
                values_only=True,
            ),
            start=data_region["start_row"],
        ):
            if all(v is None for v in row):
                blank_rows.append(row_idx)

        return blank_rows

    def _detect_blank_rows_sampled(self, ws: Any, data_region: dict[str, int]) -> list[int]:
        """Sample-based blank row detection for large files."""
        start = data_region["start_row"]
        end = data_region["end_row"]

        sample_rows: list[int] = []
        # Sample first and last 300 rows
        sample_rows.extend(range(start, min(start + 300, end + 1)))
        sample_rows.extend(range(max(start, end - 300), end + 1))
        # Also sample evenly throughout the middle to catch separators
        total = end - start + 1
        if total > 600:
            step = max(1, (total - 600) // 20)
            sample_rows.extend(range(start + 300, end - 300, step))
            # For each sampled point, also check a few rows around it
            # to detect multi-row blank separators
            extra = []
            for r in range(start + 300, end - 300, step):
                for offset in range(-2, 3):
                    candidate = r + offset
                    if start <= candidate <= end:
                        extra.append(candidate)
            sample_rows.extend(extra)

        blank_rows = []
        for row_idx in sorted(set(sample_rows)):
            row = next(
                ws.iter_rows(
                    min_row=row_idx,
                    max_row=row_idx,
                    min_col=data_region["start_col"],
                    max_col=data_region["end_col"],
                    values_only=True,
                )
            )

            if all(v is None for v in row):
                blank_rows.append(row_idx)

        return blank_rows

    def _detect_formulas_from_source(
        self,
        file_source: FileSource,
        sheet: str,
        data_region: dict[str, int],
    ) -> bool:
        """Check if sheet contains formulas using a data_only=False pass."""
        try:
            is_fileobj = hasattr(file_source, "read")
            if is_fileobj and hasattr(file_source, "seek"):
                file_source.seek(0)

            wb2 = openpyxl.load_workbook(file_source, read_only=True, data_only=False)
            try:
                ws2 = wb2[sheet]
                total_rows = data_region["end_row"] - data_region["start_row"] + 1
                sample_rows = min(50, total_rows)
                sample_cols = min(10, data_region["end_col"] - data_region["start_col"] + 1)

                for row_idx in range(
                    data_region["start_row"], data_region["start_row"] + sample_rows
                ):
                    for col_idx in range(
                        data_region["start_col"], data_region["start_col"] + sample_cols
                    ):
                        try:
                            cell = ws2.cell(row_idx, col_idx)
                            if cell.data_type == "f":
                                return True
                        except Exception:
                            pass
                return False
            finally:
                wb2.close()
        except Exception:
            return False

    def _detect_sparse_columns(self, ws: Any, data_region: dict[str, int]) -> list[int]:
        """Identify columns that are >90% empty within the data region."""
        total_rows = data_region["end_row"] - data_region["start_row"] + 1
        if total_rows == 0:
            return []

        sample_limit = min(MAX_SAMPLE_ROWS, total_rows)
        num_cols = data_region["end_col"] - data_region["start_col"] + 1
        filled = [0] * num_cols

        for row in ws.iter_rows(
            min_row=data_region["start_row"],
            max_row=data_region["start_row"] + sample_limit - 1,
            min_col=data_region["start_col"],
            max_col=data_region["end_col"],
            values_only=True,
        ):
            for col_offset, value in enumerate(row):
                if value is not None:
                    filled[col_offset] += 1

        threshold = sample_limit * 0.1  # >90% empty means <10% filled
        sparse = []
        for col_offset, count in enumerate(filled):
            if count < threshold:
                sparse.append(data_region["start_col"] + col_offset)

        return sparse

    def _group_consecutive(self, numbers: list[int]) -> list[list[int]]:
        """Group consecutive numbers."""
        if not numbers:
            return []

        groups = []
        current_group = [numbers[0]]

        for num in numbers[1:]:
            if num == current_group[-1] + 1:
                current_group.append(num)
            else:
                groups.append(current_group)
                current_group = [num]

        groups.append(current_group)
        return groups

    def _check_merged_in_headers(
        self,
        merged_ranges: list,
        header_info: dict,
    ) -> bool:
        """Check if any merged cells are in header rows."""
        header_row = header_info.get("header_row")
        if header_row is None:
            return False

        return any(mr[0] <= header_row <= mr[2] for mr in merged_ranges)

    def _check_merged_in_data(
        self,
        merged_ranges: list,
        header_info: dict,
    ) -> bool:
        """Check if any merged cells are in data rows."""
        header_row = header_info.get("header_row") or 0

        return any(mr[0] > header_row for mr in merged_ranges)

    def _table_to_dict(self, table: TableInfo) -> dict:
        """Convert TableInfo to dictionary."""
        return {
            "start_row": table.start_row,
            "end_row": table.end_row,
            "start_col": table.start_col,
            "end_col": table.end_col,
            "has_header": table.has_header,
            "header_row": table.header_row,
            "confidence": table.confidence,
        }

    def _suggest_skip_rows(self, metadata_rows: list[int], header_info: dict) -> int:
        """Suggest number of rows to skip."""
        if metadata_rows:
            return len(metadata_rows)
        return 0

    def _suggest_skip_footer(self, ws: Any, data_region: dict[str, int]) -> int:
        """Suggest number of footer rows to skip.

        Only strips rows that are clearly footer metadata (attribution, notes,
        generation info). Does NOT strip "Total" or "Grand Total" rows since
        those often contain legitimate summary data that users expect to see.
        Scans past blank rows to catch footers separated by gaps.
        """
        import re

        footer_patterns = [
            re.compile(r"\bprepared\s+by\b", re.I),
            re.compile(r"\bnotes?\s*:", re.I),
            re.compile(r"\bgenerated\s+(on|at|by)\b", re.I),
            re.compile(r"\breport\s+generated\b", re.I),
            re.compile(r"\bsource\s*:", re.I),
            re.compile(r"\bdisclaimer\b", re.I),
            re.compile(r"\bconfidential\b", re.I),
            re.compile(r"\bvalues\s+subject\s+to\b", re.I),
        ]

        skip = 0
        blank_streak = 0

        for row_idx in range(data_region["end_row"], max(1, data_region["end_row"] - 10), -1):
            row_values = []
            for col_idx in range(data_region["start_col"], data_region["end_col"] + 1):
                try:
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                except Exception:
                    pass

            if not row_values:
                blank_streak += 1
                if blank_streak <= 3:
                    skip += 1
                    continue
                break

            row_text = " ".join(row_values)
            if any(p.search(row_text) for p in footer_patterns):
                skip += 1
                blank_streak = 0
            else:
                break

        return skip
