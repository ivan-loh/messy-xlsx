"""Bounded structural analysis over retained worksheet evidence."""

from __future__ import annotations

import re
from dataclasses import dataclass
from types import SimpleNamespace
from typing import Any

import pandas as pd
from openpyxl.styles.numbers import is_date_format

from messy_xlsx.cache import StructureCache
from messy_xlsx.detection.locale_detector import LocaleDetector
from messy_xlsx.detection.structure_analyzer import StructureAnalyzer
from messy_xlsx.models import StructureInfo
from messy_xlsx.ooxml.models import CellEvidence, IntervalIndex, SheetManifest

_NUMERIC_EVIDENCE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?$")
_MISSING_EVIDENCE = object()
_DATE_EVIDENCE = object()


@dataclass(frozen=True)
class StructureEvidence:
    """A bounded set of worksheet rows keyed by one-based coordinates."""

    row_numbers: tuple[int, ...]
    values: pd.DataFrame

    def row(self, row_number: int) -> tuple[object, ...]:
        """Return retained values for one worksheet row, or an empty tuple."""
        if row_number not in self.values.index:
            return ()
        selected = self.values.loc[row_number]
        if isinstance(selected, pd.DataFrame):
            selected = selected.iloc[0]
        return tuple(None if pd.isna(value) else value for value in selected.tolist())


@dataclass(frozen=True)
class SampleWindow:
    """One contiguous one-based worksheet sampling window."""

    start_row: int
    n_rows: int

    def __post_init__(self) -> None:
        if self.start_row < 1 or self.n_rows < 1:
            raise ValueError("sample windows require positive one-based bounds")


def blank_row_sample_positions(start: int, end: int) -> tuple[int, ...]:
    """Return the existing analyzer's bounded blank-row sample coordinates."""
    if start < 1 or end < start:
        return ()
    sample_rows: list[int] = []
    sample_rows.extend(range(start, min(start + 300, end + 1)))
    sample_rows.extend(range(max(start, end - 300), end + 1))
    total = end - start + 1
    if total > 600:
        step = max(1, (total - 600) // 20)
        middle = range(start + 300, end - 300, step)
        sample_rows.extend(middle)
        for row in middle:
            sample_rows.extend(
                candidate for candidate in range(row - 2, row + 3) if start <= candidate <= end
            )
    return tuple(sorted(set(sample_rows)))


def coalesce_rows(rows: tuple[int, ...]) -> tuple[SampleWindow, ...]:
    """Coalesce sorted or unsorted row coordinates into contiguous windows."""
    ordered = sorted(set(rows))
    if not ordered:
        return ()
    windows: list[SampleWindow] = []
    start = previous = ordered[0]
    for row in ordered[1:]:
        if row == previous + 1:
            previous = row
            continue
        windows.append(SampleWindow(start, previous - start + 1))
        start = previous = row
    windows.append(SampleWindow(start, previous - start + 1))
    return tuple(windows)


def structure_sample_windows(
    max_row: int,
    start_row: int = 1,
) -> tuple[SampleWindow, ...]:
    """Build the bounded union of head, blank-sample, and footer rows."""
    if max_row < 1 or start_row < 1 or start_row > max_row:
        return ()
    head = range(start_row, min(max_row, start_row + 9_999) + 1)
    tail = range(max(start_row, max_row - 9), max_row + 1)
    sampled_blanks = blank_row_sample_positions(start_row, max_row)
    rows = tuple(sorted(set(head) | set(tail) | set(sampled_blanks)))
    return coalesce_rows(rows)


class _EvidenceWorksheet:
    """Read-only worksheet-shaped adapter over bounded evidence."""

    def __init__(self, evidence: StructureEvidence, manifest: SheetManifest) -> None:
        self._evidence = evidence
        self._retained_rows = frozenset(evidence.row_numbers)
        self._cell_evidence = {(cell.row, cell.column): cell for cell in manifest.cell_evidence}
        self._semantic_nonempty_rows = manifest.semantic_nonempty_rows
        region = manifest.semantic_data_region
        self.max_row = region[1]
        self.max_column = region[3]

    def cell(self, row: int, column: int) -> SimpleNamespace:
        value = self._value(row, column)
        metadata = self._cell_evidence.get((row, column))
        number_format = "General" if metadata is None else metadata.number_format
        return SimpleNamespace(value=value, number_format=number_format)

    def _value(self, row: int, column: int) -> object:
        if not self._semantic_nonempty_rows.contains(row):
            return None
        if row not in self._retained_rows:
            return _MISSING_EVIDENCE
        values = self._evidence.row(row)
        value = values[column - 1] if column <= len(values) else None
        return _worksheet_scalar(value, self._cell_evidence.get((row, column)))

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
        values_only: bool = False,
    ) -> Any:
        final_row = self.max_row if max_row is None else max_row
        final_col = self.max_column if max_col is None else max_col
        for row_number in range(min_row, final_row + 1):
            row = tuple(self._value(row_number, column) for column in range(min_col, final_col + 1))
            if values_only:
                yield row
            else:
                yield tuple(SimpleNamespace(value=value) for value in row)


def _worksheet_scalar(value: object, evidence: CellEvidence | None) -> object:
    """Recover a scalar only when exact OOXML provenance permits coercion."""
    if evidence is None:
        return value
    if not evidence.has_value:
        return None
    if evidence.data_type == "b":
        if isinstance(value, str):
            return value.casefold() in {"1", "true"}
        return bool(value)
    if evidence.data_type == "d" or is_date_format(evidence.number_format):
        return value if not isinstance(value, str) else _DATE_EVIDENCE
    if (
        evidence.data_type in {"n", ""}
        and isinstance(value, str)
        and _NUMERIC_EVIDENCE.fullmatch(value)
    ):
        try:
            return float(value) if any(marker in value for marker in ".eE") else int(value)
        except ValueError:
            return value
    return value


def _expand_intervals(index: IntervalIndex) -> list[int]:
    return [
        coordinate
        for interval in index.intervals
        for coordinate in range(interval.start, interval.end + 1)
    ]


def _locale_evidence(
    worksheet: _EvidenceWorksheet,
    data_region: dict[str, int],
) -> tuple[list[str], list[str]]:
    text_values: list[str] = []
    format_codes: list[str] = []
    end_row = min(data_region["end_row"], data_region["start_row"] + 50)
    end_col = min(data_region["end_col"], data_region["start_col"] + 20)
    for row in range(data_region["start_row"], end_row + 1):
        for column in range(data_region["start_col"], end_col + 1):
            cell = worksheet.cell(row, column)
            if isinstance(cell.value, str):
                text_values.append(cell.value)
            if cell.number_format and cell.number_format != "General":
                format_codes.append(cell.number_format)
    return text_values, format_codes


def analyze_structure_evidence(
    evidence: StructureEvidence,
    manifest: SheetManifest,
    header_patterns: tuple[str, ...],
) -> StructureInfo:
    """Apply the established analyzer heuristics to bounded evidence."""
    analyzer = StructureAnalyzer(cache=StructureCache())
    worksheet = _EvidenceWorksheet(evidence, manifest)
    start_row, end_row, start_col, end_col = manifest.semantic_data_region
    data_region = {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
    }
    merged = [
        (item.min_row, item.min_col, item.max_row, item.max_col) for item in manifest.merged_ranges
    ]
    hidden_rows = _expand_intervals(manifest.hidden_rows)
    hidden_columns = _expand_intervals(manifest.hidden_columns)
    header = analyzer._detect_headers(
        worksheet,
        data_region,
        merged,
        list(header_patterns) or None,
    )
    metadata = analyzer._detect_metadata_rows(worksheet, data_region, header)
    tables = analyzer._detect_multiple_tables(worksheet, data_region, header)
    blank_rows = analyzer._detect_blank_rows(worksheet, data_region)
    sparse_columns = analyzer._detect_sparse_columns(worksheet, data_region)
    text_values, format_codes = _locale_evidence(worksheet, data_region)
    locale = LocaleDetector().detect_from_evidence(text_values, format_codes)
    return StructureInfo(
        data_start_row=data_region["start_row"],
        data_end_row=data_region["end_row"],
        data_start_col=data_region["start_col"],
        data_end_col=data_region["end_col"],
        header_row=header.get("header_row"),
        header_rows_count=header.get("header_rows_count", 1),
        header_confidence=header.get("confidence", 0.0),
        metadata_rows=metadata,
        merged_ranges=merged,
        merged_in_headers=analyzer._check_merged_in_headers(merged, header),
        merged_in_data=analyzer._check_merged_in_data(merged, header),
        hidden_rows=hidden_rows,
        hidden_columns=hidden_columns,
        detected_locale=locale.locale,
        decimal_separator=locale.decimal_separator,
        thousands_separator=locale.thousands_separator,
        num_tables=len(tables),
        table_ranges=[analyzer._table_to_dict(table) for table in tables],
        blank_rows=blank_rows,
        has_formulas=manifest.legacy_has_formulas,
        sparse_columns=sparse_columns,
        suggested_skip_rows=analyzer._suggest_skip_rows(metadata, header),
        suggested_skip_footer=analyzer._suggest_skip_footer(worksheet, data_region),
    )


class StructureSampler:
    """Cache bounded structure analysis by sheet and header-pattern variant."""

    def __init__(
        self,
        excel_reader: Any,
        manifest_reader: Any,
        metrics: Any | None = None,
    ) -> None:
        self._excel_reader = excel_reader
        self._manifest_reader = manifest_reader
        self._metrics = metrics
        self._cache: dict[tuple[str, tuple[str, ...]], StructureInfo] = {}

    def analyze(
        self,
        sheet: str,
        header_patterns: list[str] | tuple[str, ...] | None = None,
    ) -> StructureInfo:
        """Analyze one sheet without retaining a complete discarded frame."""
        key = (sheet, tuple(header_patterns or ()))
        cached = self._cache.get(key)
        if cached is not None:
            return cached
        manifest = self._manifest_reader.sheet(sheet)
        start_row, end_row, _start_col, _end_col = manifest.semantic_data_region
        windows = structure_sample_windows(end_row, start_row)
        evidence = self._excel_reader.sample_windows(
            sheet,
            windows=windows,
            min_column=manifest.observed_min_col,
            max_column=manifest.observed_max_col,
        )
        result = analyze_structure_evidence(evidence, manifest, key[1])
        self._cache[key] = result
        if self._metrics is not None:
            self._metrics.sample_reads += 1
        return result
